'''
Author: zoodehao
Date: 2021-08-27 16:04:26
LastEditTime: 2021-12-03 14:38:21
FilePath: \Py3.9\5_萍姐【彩盒发货】发送版.py
Description: 逝者如斯夫, 不舍昼夜.
'''
import re
import os
import glob
import time
import hmac
import json
import xlrd
import pyhdb
import base64
import pymysql
import hashlib
import requests
import calendar
import win32com
import pythoncom
import numpy as np
import pandas as pd
import urllib.parse
import urllib.request
from math import isnan
import openpyxl as opxl
from decimal import Decimal
from termcolor import cprint
from datetime import datetime
from PIL import ImageGrab, Image
from time import strftime, gmtime
from win32com.client import Dispatch, DispatchEx
from qiniu import Auth, put_file, etag, BucketManager
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection, NamedStyle, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, IconSetRule, Rule, IconSet, FormatObject


# ----------------------------------------------------------------------------
# 加载格式框架
# ----------------------------------------------------------------------------
framepath = 'C:/Users/Zeus/Desktop/autoSend/5_萍姐/目标/萍姐_格式框架.xlsx'
workbook  = opxl.load_workbook(framepath)


# ----------------------------------------------------------------------------
# 全局日期设置
# ----------------------------------------------------------------------------
globalYear  = datetime.now().year
globalMonth = datetime.now().month
globalDay   = datetime.now().day


# ----------------------------------------------------------------------------
# 数据区
# ----------------------------------------------------------------------------
allStartTime = datetime.now()
print(" > 数据获取中, 请稍等片刻")

# 内网: IP: 192.168.20.183  账号: HANA1107318 密码: Zeus@1107311
# 公网: IP: 119.145.248.183  账号: HANAHBPOUTTEST 密码: Zeus@test147
def get_HANA_Connection():
    connObj = pyhdb.connect(host='192.168.20.183',
                                  port     = 30015,
                                  user     = 'HANA1107318',
                                  password = 'Zeus@1107311')
    return connObj


# 9-12月发货纯销
def get_matFHCX(FHCX):
    cursorFHCX = FHCX.cursor()
    cursorFHCX.execute("""
                       WITH AA AS (SELECT KUNNR, NAME1
                       FROM "ECC_BI"."KNA1"), -- 五级名称
                       AAA AS (SELECT MATNR, MAKTX
                       FROM "ECC_BI"."MAKT"
                       WHERE MANDT = 800 AND SPRAS = 1), -- 物料名称
                       AAAA AS (SELECT MATNR, KBETR
                       FROM "_SYS_BIC"."HD-HAND.SD.POWER_BI.DATA/ZAN_A304"
                       WHERE VKORG = '2008' AND VTWEG = 'A3' AND KSCHL = 'VKP0'), -- 标准零售价
                       AAAAA AS (SELECT ZGRKUNNR, ZGKUNNR, ZMODE, ZKHLX
                       FROM "ECC_BI"."ZTSD017" -- 匹配HX CX 折扣
                       WHERE ZORG = '2008' AND ZCHBOX = '' AND (TO_NUMBER(ZENDDATA) >= TO_NUMBER((LEFT(NOW(),4) || RIGHT(LEFT(NOW(),7),2) || RIGHT(LEFT(NOW(),10),2))))), -- 折扣/计算方式
                       BB AS (SELECT 'CX' AS TYPE, ZDATE_DEAL, MONTH1, ZZTP06, PARTNER, ZQDCY, ZQDCY_DES, ZMATNR, SUM(MENGE) AS AMOUNT
                       FROM "_SYS_BIC"."HD-HAND.SD.DATA/ZAN_SD_CX"
                       WHERE ERP_SALES_ORG IN ('2001', '2008') AND ZYEAR = '2021' AND MONTH1 IN ('10', '11', '12')
                       GROUP BY ZDATE_DEAL, MONTH1, ZZTP06, PARTNER, ZQDCY, ZQDCY_DES, ZMATNR), -- 纯销
                       CC AS (SELECT BB.TYPE, BB.ZDATE_DEAL, BB.MONTH1, BB.ZZTP06, BB.PARTNER, BB.ZQDCY, BB.ZQDCY_DES, BB.ZMATNR, AAA.MAKTX, BB.AMOUNT
                       FROM BB
                       LEFT JOIN AAA ON BB.ZMATNR = AAA.MATNR),
                       DD AS (SELECT CC.TYPE, CC.ZDATE_DEAL, CC.MONTH1, CC.ZZTP06, CC.PARTNER, AA.NAME1 AS KEHU, CC.ZQDCY, CC.ZQDCY_DES, CC.ZMATNR, CC.MAKTX, CC.AMOUNT
                       FROM CC
                       LEFT JOIN AA ON CC.PARTNER = AA.KUNNR),
                       EE AS (SELECT DD.TYPE, DD.ZDATE_DEAL, DD.MONTH1, DD.ZZTP06, AA.NAME1, DD.PARTNER, DD.KEHU, DD.ZQDCY, DD.ZQDCY_DES, DD.ZMATNR, DD.MAKTX, DD.AMOUNT
                       FROM DD
                       LEFT JOIN AA ON DD.ZZTP06 = AA.KUNNR),
                       FF AS (SELECT 'FH' AS TYPE, LFDAT, MONTH(LFDAT) AS MON, KUNNR_GR, NAME_GR, KUNNR, NAME1, '' AS STOREID, '' AS STRORE, ('000000000' || MATNR) AS ZMATNR, MAKTX, SUM(LFIMG1) AS AMOUNT
                       FROM "ECC_BI"."ZOUTBOUND"
                       WHERE ZZPART1_T_S LIKE N'%草晶华销售事业部%' AND LFDAT != '' AND BEZEI2 != '领用订单' AND YEAR(LFDAT) = '2021'
                       GROUP BY LFDAT, KUNNR_GR, NAME_GR, KUNNR, NAME1, MATNR, MAKTX), -- 发货
                       HH AS (SELECT FF.TYPE, FF.LFDAT, FF.MON, FF.KUNNR_GR, FF.NAME_GR, FF.KUNNR, FF.NAME1, FF.STOREID, FF.STRORE, FF.ZMATNR, FF.MAKTX, AAAA.KBETR, FF.AMOUNT
                       FROM FF
                       LEFT JOIN AAAA ON FF.ZMATNR = AAAA.MATNR),
                       JJ AS (SELECT EE.TYPE, EE.ZDATE_DEAL, EE.MONTH1, EE.ZZTP06, EE.NAME1, EE.PARTNER, EE.KEHU, EE.ZQDCY, EE.ZQDCY_DES, EE.ZMATNR, EE.MAKTX, AAAA.KBETR AS KBETR, EE.AMOUNT
                       FROM EE
                       LEFT JOIN AAAA ON EE.ZMATNR = AAAA.MATNR),
                       LL AS (SELECT HH.TYPE, AAAAA.ZMODE, HH.LFDAT, HH.MON, HH.KUNNR_GR, HH.NAME_GR, HH.KUNNR, HH.NAME1, HH.STOREID, HH.STRORE, HH.ZMATNR, HH.MAKTX, HH.KBETR, AAAAA.ZKHLX, HH.AMOUNT
                       FROM HH
                       LEFT JOIN AAAAA ON (HH.KUNNR_GR = AAAAA.ZGRKUNNR AND HH.KUNNR = AAAAA.ZGKUNNR)) -- 发货
                       SELECT JJ.TYPE, AAAAA.ZMODE, JJ.ZDATE_DEAL, JJ.MONTH1, JJ.ZZTP06, JJ.NAME1, JJ.PARTNER, JJ.KEHU, JJ.ZQDCY, JJ.ZQDCY_DES, JJ.ZMATNR, JJ.MAKTX, JJ.KBETR, AAAAA.ZKHLX, JJ.AMOUNT
                       FROM JJ
                       LEFT JOIN AAAAA ON (JJ.ZZTP06 = AAAAA.ZGRKUNNR AND JJ.PARTNER = AAAAA.ZGKUNNR) -- 纯销
                       WHERE JJ.MAKTX LIKE '%16袋%'
                       UNION ALL
                       SELECT LL.TYPE, LL.ZMODE, LL.LFDAT, LL.MON, LL.KUNNR_GR, LL.NAME_GR, LL.KUNNR, LL.NAME1, LL.STOREID, LL.STRORE, LL.ZMATNR, LL.MAKTX, LL.KBETR, LL.ZKHLX, LL.AMOUNT
                       FROM LL -- 发货
                       WHERE LL.MAKTX LIKE '%16袋%'
                       """)
    matFHCX = cursorFHCX.fetchall()
    return matFHCX

# 门店别名表
def get_matRestore(Restore):
    cursorRestore = Restore.cursor()
    cursorRestore.execute("""
                        WITH A1 AS (SELECT ZNUMBER, PARTNER, ZGR
                        FROM "ECC_BI"."ZSD007"
                        WHERE ZVALID = ''), -- StoreID + ClientID => GRID
                        A2 AS (SELECT ZNUMBER, ZSUOSHUNUMBER
                        FROM "ECC_BI"."ZTCRM003"
                        WHERE ZSUOSHUTYPE = '经销商' AND ZVALID = '' AND (TO_NUMBER(ZJIESHUDATE) >= TO_NUMBER((LEFT(NOW(),4) || RIGHT(LEFT(NOW(),7),2) || RIGHT(LEFT(NOW(),10),2))))
                        ), -- StoreID => ClientID
                        BB AS (SELECT ZQDCY_FROM, ZMDBM, ZMDMC, PARTNER
                        FROM "ECC_BI"."ZCRTBMAP002"
                        WHERE VKORG = '2008' AND ZSFZF = '' AND ZMDBM != ''),  -- 别名表
                        CC AS (SELECT BB.ZQDCY_FROM, BB.ZMDBM, BB.ZMDMC, A2.ZSUOSHUNUMBER AS ZSUOSHUNUMBER
                        FROM BB
                        LEFT JOIN A2 ON BB.ZMDBM = A2.ZNUMBER)
                        SELECT CC.ZQDCY_FROM, CC.ZMDBM, CC.ZMDMC, CC.ZSUOSHUNUMBER, A1.ZGR, (A1.ZGR || CC.ZSUOSHUNUMBER) AS GR_CLIENT
                        FROM CC
                        LEFT JOIN A1 ON (CC.ZMDBM = A1.ZNUMBER AND CC.ZSUOSHUNUMBER = A1.PARTNER)
                        """)
    matRestore = cursorRestore.fetchall()
    return matRestore

# 计算方式 折扣
def get_matDiscount(Discount):
    cursorDiscount = Discount.cursor()
    cursorDiscount.execute("""
                           SELECT (ZGRKUNNR || ZGKUNNR) AS KEY, ZGRKUNNR, ZGKUNNR, ZMODE, ZKHLX
                           FROM "ECC_BI"."ZTSD017"
                           WHERE ZORG = '2008' AND ZCHBOX = '' AND (TO_NUMBER(ZENDDATA) >= TO_NUMBER((LEFT(NOW(),4) || RIGHT(LEFT(NOW(),7),2) || RIGHT(LEFT(NOW(),10),2))))
                           """)
    matDiscount = cursorDiscount.fetchall()
    return matDiscount

# （草晶华）物料名称 标准单价信息表
def get_matMaterialInfo(MaterialInfo):
    cursorMaterialInfo = MaterialInfo.cursor()
    cursorMaterialInfo.execute("""WITH A1 AS(SELECT MATNR, KBETR
                                   FROM "_SYS_BIC"."HD-HAND.SD.POWER_BI.DATA/ZAN_A304"
                                   WHERE VKORG = '2008' AND VTWEG = 'A3' AND KSCHL = 'VKP0'), -- 物料标准单价
                                   A2 AS(SELECT MATNR, MAKTX
                                   FROM "ECC_BI"."MAKT"
                                   WHERE MANDT = 800 AND SPRAS = 1)
                                   SELECT A2.MATNR, A2.MAKTX, A1.KBETR
                                   FROM A2
                                   LEFT JOIN A1 ON A1.MATNR = A2.MATNR
                                   """)
    matMaterialInfo = cursorMaterialInfo.fetchall()
    return matMaterialInfo

# 五级架构中文名称
def get_matLevel5(Level5):
    cursorLevel5 = Level5.cursor()
    cursorLevel5.execute("""SELECT KUNNR, NAME1
                            FROM "ECC_BI"."KNA1"
                            """)
    matLevel5 = cursorLevel5.fetchall()
    return matLevel5

conn = get_HANA_Connection()

# 发货纯销
dataFHCX = pd.DataFrame(get_matFHCX(conn), columns=['数据源', '计算方式', '日期', '月份', 'GR编码', 'GR名称', '客户ID', '客户名称', '门店ID', '门店名称', '物料ID', '物料名称', '标准单价', '折扣', '数量'])

# 门店别名表
dataRestore = pd.DataFrame(get_matRestore(conn), columns=['网上门店名称', '门店ID', '门店名称', '客户ID', 'GR编码', 'GR客户ID'])

# 计算方式 折扣
dataDiscount = pd.DataFrame(get_matDiscount(conn), columns=['主键', 'GR编码', '客户ID', '计算方式', '折扣'])

# 标准物料信息
dataMatetialStandardInfo = pd.DataFrame(get_matMaterialInfo(conn), columns=['物料ID', '物料名称', '标准单价'])

# 五级架构中文名称
dataLevel5 = pd.DataFrame(get_matLevel5(conn), columns=['五级ID', '五级中文名称'])


conn = pymysql.connect(host='192.168.20.241',
                        port    = 3306,
                        user    = 'root',
                        passwd  = 'Powerbi#1217',
                        db      = 'dkh',
                        charset = 'utf8')

cursor = conn.cursor()

executeCode = """SELECT dept_5, DKH_materiel_id, 'DKH' AS SOURCE, date, MONTH(date) AS MONTH, sfa_desc, SUM(amount)
                 FROM dkhfact
                 WHERE YEAR(date) = '2021'
                 GROUP BY dept_5, DKH_materiel_id, date, sfa_desc
                 """

cursor.execute(executeCode)  # 执行查询
rowNum = cursor.rowcount  # 数据条数
getDKH = cursor.fetchall()  # 获取全部数据
conn.commit()  # 提交确认
cursor.close()  # 关闭光标
conn.close()  # 关闭连接

dataDKH = pd.DataFrame(getDKH, columns=['dept_5', 'materialID', '数据源', '日期', '月份', '门店名称', '数量'])
getdataEndTime = datetime.now()
print(" >>获取数据耗时: " + strftime("%H:%M:%S", gmtime((getdataEndTime - allStartTime).seconds)))

# ----------------------------------------------------------------------------
# 字典区
# ----------------------------------------------------------------------------
print(" > 数据清洗中, 请稍等片刻")

cleandataStartTime = datetime.now()

storeID_Dict = {}  # 标准门店ID
store_Info = pd.read_excel('D:/FilesCenter/DKH-BottomTable/DKH对照表.xlsx', sheet_name='SFA_Hierarchy', header=0, dtype=str)  # 网上门店信息表
for irow in range(store_Info.shape[0]):
    storeID_Dict[store_Info.loc[irow, 'OUT_SFA_bianma']] = str(store_Info.loc[irow, 'SFA_id']).replace('.0', '')
    
GRID_Dict = {}  # GR编码
clientID_Dict = {}  # 客户ID
GRclientID_Dict = {} # GR客户ID
for irow in range(dataRestore.shape[0]):
    GRID_Dict[dataRestore.loc[irow, '门店ID']] = dataRestore.loc[irow, 'GR编码']
    clientID_Dict[dataRestore.loc[irow, '门店ID']] = dataRestore.loc[irow, '客户ID']
    GRclientID_Dict[dataRestore.loc[irow, '门店ID']] = dataRestore.loc[irow, 'GR客户ID']

materialID_Dict = {}  # 标准物料ID
material_Info = pd.read_excel('D:/FilesCenter/DKH-BottomTable/商品编码对照字典.xlsx', sheet_name=0, header=0, dtype=str)  # 物料信息表
for irow in range(material_Info.shape[0]):
    materialID_Dict[str(material_Info.loc[irow, '编码']).replace('.0', '')] = material_Info.loc[irow, '标准物料ID']

materialStandardDesc_Dict = {}  # 标准物料名称
materialStandardPrice_Dict = {}  # 标准物料单价
for irow in range(dataMatetialStandardInfo.shape[0]):
    materialStandardDesc_Dict[dataMatetialStandardInfo.loc[irow, '物料ID']] = dataMatetialStandardInfo.loc[irow, '物料名称']
    materialStandardPrice_Dict[dataMatetialStandardInfo.loc[irow, '物料ID']] = dataMatetialStandardInfo.loc[irow, '标准单价']
    
Method_Dict = {}  # 计算方式
Discount_Dict = {}  # 折扣
for irow in range(dataDiscount.shape[0]):
    Method_Dict[dataDiscount.loc[irow, '主键']] = dataDiscount.loc[irow, '计算方式']
    Discount_Dict[dataDiscount.loc[irow, '主键']] = dataDiscount.loc[irow, '折扣']

Level5_Dict = {}  # 五级中文名称
for irow in range(dataLevel5.shape[0]):
    Level5_Dict[dataLevel5.loc[irow, '五级ID']] = dataLevel5.loc[irow, '五级中文名称']

def newDF():
    return pd.DataFrame(columns=['dept_5', 'materialID', 'GR客户ID', '数据源', '计算方式', '日期', '月份', 'GR编码', 'GR名称', '客户ID', '客户名称', '门店ID', '门店名称', '物料ID', '物料名称', '标准单价', '折扣', '数量'])

dfDKH = newDF()
dfDKH['dept_5'] = dataDKH['dept_5']
dfDKH['materialID'] = dataDKH['materialID']
dfDKH['数据源'] = dataDKH['数据源']
dfDKH['日期'] = dataDKH['日期']
dfDKH['月份'] = dataDKH['月份']
dfDKH['门店名称'] = dataDKH['门店名称']
dfDKH['数量'] = dataDKH['数量']

dfDKH['门店ID'] = dfDKH.apply(lambda x: storeID_Dict.setdefault(x['dept_5'], ''), axis=1)
dfDKH['物料ID'] = dfDKH.apply(lambda x: materialID_Dict.setdefault(x['materialID'], ''), axis=1)
dfDKH['GR编码'] = dfDKH.apply(lambda x: GRID_Dict.setdefault(x['门店ID'], ''), axis=1)
dfDKH['客户ID'] = dfDKH.apply(lambda x: clientID_Dict.setdefault(x['门店ID'], ''), axis=1)
dfDKH['GR客户ID'] = dfDKH.apply(lambda x: GRclientID_Dict.setdefault(x['门店ID'], ''), axis=1)
dfDKH['GR名称'] = dfDKH.apply(lambda x: Level5_Dict.setdefault(x['GR编码'], ''), axis=1)
dfDKH['客户名称'] = dfDKH.apply(lambda x: Level5_Dict.setdefault(x['客户ID'], ''), axis=1)
dfDKH['物料名称'] = dfDKH.apply(lambda x: materialStandardDesc_Dict.setdefault(x['物料ID'], ''), axis=1)
dfDKH['标准单价'] = dfDKH.apply(lambda x: materialStandardPrice_Dict.setdefault(x['物料ID'], ''), axis=1)
dfDKH['计算方式'] = dfDKH.apply(lambda x: Method_Dict.setdefault(x['GR客户ID'], ''), axis=1)
dfDKH['折扣'] = dfDKH.apply(lambda x: Discount_Dict.setdefault(x['GR客户ID'], ''), axis=1)

# dfDKH.to_excel('C:/Users/zoodehao/Desktop/网上-数据源1.xlsx', index=False)

dfDKH = dfDKH[(dfDKH['计算方式'] == 'CX') & (dfDKH['物料名称'].str.contains('16袋'))]  # 只取纯销中彩盒
dfNew = dfDKH.drop(['dept_5', 'materialID', 'GR客户ID'], axis=1)

# dfNew.to_excel('C:/Users/zoodehao/Desktop/网上-数据源2.xlsx', index=False)

# 数据源为FH GR为1100003217 计算方式全部改为HX
# for irow in range(dataFHCX.shape[0]):
#     if dataFHCX.loc[irow, '数据源'] == 'FH' and dataFHCX.loc[irow, 'GR编码'] == '1100003217':
#         dataFHCX.loc[irow, '计算方式'] = 'HX'

dataFHCX['折扣'] = pd.to_numeric(dataFHCX['折扣'], errors='coerce')

for irow in range(dataFHCX.shape[0]):
    if dataFHCX.loc[irow, '数据源'] == 'FH' and dataFHCX.loc[irow, 'GR编码'] == '1100003217':
        dataFHCX.loc[irow, '计算方式'] = 'HX'
        if dataFHCX.loc[irow, 'GR编码'] in ['1100003357', '1100003217', '1100005256'] and isnan(dataFHCX.loc[irow, '折扣']):
            dataFHCX.loc[irow, '折扣'] = 20
    else:
        if dataFHCX.loc[irow, 'GR编码'] in ['1100003357', '1100003217', '1100005256'] and isnan(dataFHCX.loc[irow, '折扣']):
            dataFHCX.loc[irow, '折扣'] = 20
            
dataFHCX = dataFHCX[(dataFHCX['计算方式'] == 'HX')]  # 只取核销HX(SQL框定彩盒了)

allData = dataFHCX.append([dfNew])  # 整合数据
allData['标准单价'] = pd.to_numeric(allData['标准单价'], errors='coerce')
allData['折扣'] = pd.to_numeric(allData['折扣'], errors='coerce')
allData['数量'] = pd.to_numeric(allData['数量'], errors='coerce')


cleandataEndTime = datetime.now()
print(" >>清洗数据耗时: " + strftime("%H:%M:%S", gmtime((cleandataEndTime - cleandataStartTime).seconds)))

allData.to_excel('C:/Users/Zeus/Desktop/autoSend/5_萍姐/百日PK数据源(彩盒部分)-' + str(globalYear) + str(globalMonth).zfill(2) + str(globalDay).zfill(2) + '.xlsx', index=False)

allEndTime = datetime.now()
cprint(">>>总耗时: " + strftime("%H:%M:%S", gmtime((allEndTime - allStartTime).seconds)), 'cyan', attrs=['bold', 'reverse', 'blink'])


# # ----------------------------------------------------------------------------
# # HANA数据库获取数据
# # ----------------------------------------------------------------------------
# # 内网: IP: 192.168.20.183  账号: HANA1107318 密码: Zeus@1107311
# # 公网: IP: 119.145.248.183  账号: HANAHBPOUTTEST 密码: Zeus@test147
# def get_HANA_Connection():
#     connectionObj = pyhdb.connect(host='192.168.20.183',
#                                   port     = 30015,
#                                   user     = 'HANA1107318',
#                                   password = 'Zeus@1107311')
#     return connectionObj


# # 获取本年彩盒发货流水账ZSD043
# def get_matZSD043(ZSD043):
#     cursorZSD043 = ZSD043.cursor()
#     cursorZSD043.execute("""SELECT LFDAT, TO_NUMBER(MONTH (LFDAT)) AS YUE, TO_NUMBER(RIGHT(LFDAT, 2)) AS TIAN, ZZPART2_T_S, BEZEI, KUNNR_GR, NAME_GR, MAKTX, TO_NUMBER(SUM (LFIMG_J)) AS J_AMOUNT
#                             FROM "ECC_BI"."ZOUTBOUND"
#                             WHERE ZZPART1_T_S LIKE N'%草晶华销售事业部%' AND LFDAT != '' AND BEZEI2 != '领用订单' AND MAKTX LIKE '%16袋%' AND YEAR(LFDAT) = :1
#                             GROUP BY LFDAT, BEZEI, ZZPART2_T_S, KUNNR_GR, NAME_GR, MAKTX""",
#         [str(globalYear)])
#     matZSD043 = cursorZSD043.fetchall()
#     return matZSD043


# conn = get_HANA_Connection()

# # 彩盒发货流水账ZSD043
# dataCaiHe = pd.DataFrame(get_matZSD043(conn), columns=['发货日期', '月份', '天', '大区', '省份', 'GR编码', 'GR名称', '物料描述', '交货件数'])

# dataMeiGui = dataCaiHe[((dataCaiHe['物料描述'] == '玫瑰桑椹破壁草本-60盒-3克×16袋') | (dataCaiHe['物料描述'] == '玫瑰桑椹破壁草本-3克×16袋'))]

# print('【发货】最新日期:' + str(max(dataCaiHe['发货日期'])))

'''
# ----------------------------------------------------------------------------
# 第一页 ①彩盒
# ----------------------------------------------------------------------------
worksheet1 = workbook['①彩盒']

GR = [
    '1100003886', '1100004615', '1100003677', '1100003926', '1100003924',
    '1100005301', '1200000586', '1100004612', '1100004614', '1200000575',
    '1200000576', '1100004772', '1100003937', '1200000559', '1200000269',
    '1100002947', '1100005264', '1100005274', '1200000577', '1200000290',
    '1200000310', '1200000285', '1200000587', '1200000560', '1100003348',
    '1100005968', '1100003355', '1100004065', '1100004532', '1200000191',
    '1100003911', '1100004499', '1100004705', '1100004662', '1100005077',
    '1200000284', '1100003228', '1100003840', '1100005827', '1100004625',
    '1100003400', '1100004433', '1100004610', '1200000543', '1100004256',
    '1100005785', '1200000370', '1100004423', '1200000375', '1100003341',
    '1200000231', '1100004322', '1100003600', '1200000306', '1100004447',
    '1100006053', '1100005158', '1100005649', '1100002889', '1200000261',
    '1100003357', '1200000264', '1200000267', '1100005256', '1200000263',
    '1100004683', '1100003541', '1100005330', '1100005289', '1100004694',
    '1100005010', '1100005170', '1100006055', '1100005250', '1100005270',
    '1100006129', '1100003996', '1200000278', '1200000273', '1100005154',
    '1100005265', '1100005431', '1100005414', '1100003306', '1100005164',
    '1100005269', '1100005564', '1100003548', '1100003630', '1100005142',
    '1100005503', '1100004175', '1100001710', '1100004559', '1100000104',
    '1100002602', '1100003215', '1100003209', '1100002969', '1100003378',
    '1100003217', '1100003347', '1100003087', '1100005012', '1100004646',
    '1100005268', '1100005155', '1100005231', '1100004588', '1100003665',
    '1100005594', '1100005380', '1100004858', '1200000253', '1100001570'
]

Row = [i for i in range(5, 72)] + [j for j in range(73, 121)]

# 非汇总部分
for jcol in range(8, 11):
    for irow, iname in zip(Row, GR):
        if jcol == 8:
            worksheet1.cell(row=irow, column=jcol).value = sum(allData[(allData['月份'] == 8) & (allData['GR编码'] == iname)]['数量'])
        elif jcol == 9:
            worksheet1.cell(row=irow, column=jcol).value = sum(allData[(allData['月份'] == 9) & (allData['GR编码'] == iname)]['数量'])
        elif jcol == 10:
            worksheet1.cell(row=irow, column=jcol).value = sum(allData[(allData['GR编码'] == iname)]['数量'])

# 汇总部分
for jcol in range(8, 11):
    worksheet1.cell(row=72, column=jcol).value = '=SUM(' + str(get_column_letter(jcol)) + '5:' + str(get_column_letter(jcol)) + '71)'
    worksheet1.cell(row=121, column=jcol).value = '=SUM(' + str(get_column_letter(jcol)) + '73:' + str(get_column_letter(jcol)) + '120)'
    worksheet1.cell(row=122, column=jcol).value = '=' + str(get_column_letter(jcol)) + '72+' + str(get_column_letter(jcol)) + '121'
 

# ----------------------------------------------------------------------------
# 带时间文字填充
# ----------------------------------------------------------------------------
# 第 1 页 - ①彩盒
worksheet1.cell(row=3, column=1).value = '2、发货数据截止至' + str(globalMonth) + '月' + str(globalDay) + '日，网上流向数据截止至' + str(globalMonth) + '月' + str(globalDay - 2) + '日(大部分网上流向提供前两天数据)'
worksheet1.cell(row=3, column=10).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)

worksheet1.column_dimensions['F'].hidden = True  # 隐藏F列 - GR编码
worksheet1.column_dimensions['G'].hidden = True  # 隐藏G列 - GR名称

for irow in range(4, 123):
    if worksheet1.cell(row=irow, column=8).value == 0 and worksheet1.cell(row=irow, column=9).value == 0 and worksheet1.cell(row=irow, column=10).value == 0:
        worksheet1.row_dimensions[irow].hidden = True


# ----------------------------------------------------------------------------
# 存储表格
# ----------------------------------------------------------------------------
workbook.save('C:/Users/zoodehao/Desktop/AutoSend/5_萍姐/萍姐_正式发送文件.xlsx')

'''

# # ----------------------------------------------------------------------------
# # 以下是发送板块
# # ----------------------------------------------------------------------------
# # 清空指定文件夹
# def deleteOldFiles(path):
#     deleteFileList = os.listdir(path)
#     all_PNG = glob.glob(path + "*.PNG")
#     print("该目录下文件有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中, PNG: " +
#           str(len(all_PNG)) + "个")
#     if len(all_PNG) != 0:
#         for deletefile in deleteFileList:
#             isDeleteFile = os.path.join(path, deletefile)
#             if os.path.isfile(isDeleteFile):
#                 os.remove(isDeleteFile)
#         all_DelPNG = glob.glob(path + "*.*")
#         if len(all_DelPNG) == 0:
#             print("已清空文件夹！！！")
#         else:
#             print("存在未删除文件, 请检查是否存在非PNG格式文件")
#     else:
#         print("不存在PNG文件")


# # screenArea——格式类似"A1:J10"
# def excelCatchScreen(file_name, sheet_name, name, save_path):
#     pythoncom.CoInitialize()  # excel多线程相关
#     Application = win32com.client.gencache.EnsureDispatch(
#         "Excel.Application")  # 启动excel
#     Application.Visible = False  # 是否可视化
#     Application.DisplayAlerts = False  # 是否显示警告
#     wb = Application.Workbooks.Open(file_name, ReadOnly=False)  # 打开excel
#     # ws = wb.Sheets(sheet_name)  # 选择Sheet
#     ws = wb.Worksheets(sheet_name)  # 选择Sheet
#     ws.Activate()  # 激活当前工作表
#     userange = ws.UsedRange
#     # 注意：要从A1开始的表格
#     screen_area = 'A1:' + str(
#         opxl.utils.get_column_letter(userange.Columns.Count)) + str(
#             userange.Rows.Count)
#     ws.Range(screen_area).CopyPicture()  # 复制图片区域
#     time.sleep(1)
#     ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
#     Application.Selection.ShapeRange.Name = name  # 将刚刚选择的Shape重命名, 避免与已有图片混淆
#     ws.Shapes(name).Copy()  # 选择图片
#     time.sleep(1)
#     img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
#     img_name = name + ".PNG"
#     img.save(save_path + img_name)  # 保存图片
#     # time.sleep(1)
#     # wb.Save()
#     # time.sleep(1)
#     wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
#     time.sleep(1)
#     Application.Quit()  # 退出excel
#     pythoncom.CoUninitialize()


# # 定义钉钉功能
# class dingdingFunction(object):
#     def __init__(self, roboturl, robotsecret, appkey, appsecret):
#         """
#         :param roboturl: 群机器人WebHook_url
#         :param robotsecret: 安全设置的加签秘钥
#         :param appkey: 企业开发平台小程序AppKey
#         :param appsecret: 企业开发平台小程序AppSecret
#         """
#         self.roboturl = roboturl
#         self.robotsecret = robotsecret
#         self.appkey = appkey
#         self.appsecret = appsecret
#         timestamp = round(time.time() * 1000)  # 时间戳
#         secret_enc = robotsecret.encode('utf-8')
#         string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
#         string_to_sign_enc = string_to_sign.encode('utf-8')
#         hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
#         sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
#         self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(timestamp, sign)  # 最终url,url+时间戳+签名

#     # 发送文件
#     def getAccess_token(self):
#         url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (AppKey, AppSecret)
#         headers = {'Content-Type': "application/x-www-form-urlencoded"}
#         data = {'appkey': self.appkey, 'appsecret': self.appsecret}
#         r = requests.request('GET', url, data=data, headers=headers)
#         access_token = r.json()["access_token"]
#         return access_token

#     def getMedia_id(self, filespath):
#         access_token = self.getAccess_token()  # 拿到接口凭证
#         url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
#         files = {'media': open(filespath, 'rb')}
#         data = {'access_token': access_token, 'type': 'file'}
#         response = requests.post(url, files=files, data=data)
#         json = response.json()
#         return json["media_id"]

#     def sendFile(self, chatid, filespath):
#         access_token = self.getAccess_token()
#         media_id = self.getMedia_id(filespath)
#         url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
#         header = {'Content-Type': 'application/json'}
#         data = {
#             'access_token': access_token,
#             'chatid': chatid,
#             'msg': {
#                 'msgtype': 'file',
#                 'file': {
#                     'media_id': media_id
#                 }
#             }
#         }
#         r = requests.request('POST', url, data=json.dumps(data), headers=header)
#         print(r.json()["errmsg"])

#     # 发送消息
#     def sendMessage(self, content):
#         """
#         :param content: 发送内容
#         """
#         header = {"Content-Type": "application/json", "Charset": "UTF-8"}
#         sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
#         sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
#         request = urllib.request.Request(url=self.webhook_url, data=sendContent, headers=header)  # 发送请求
#         opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
#         print(opener.read().decode())  # 打印返回的结果


# # 上传本地图片获取网上图片URL
# def get_image_url(imagePath, pictureName):
#     if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
#         filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
#     elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
#         filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
#     else:
#         print("请检查图片格式！！！")
#     # 七牛云密钥管理：https://portal.qiniu.com/user/key
#     # 【账号：13267854059  密码：z****】
#     access_key = "fjlWDEbF1fqBU98UsdDJRcSSKODT9Gq7tA3gu8eY"
#     secret_key = "thiWFpO881GfhlaAz1Wkk2yEcvV3ue2OHnY_5D9V"
#     keyq = Auth(access_key, secret_key)
#     bucket = "zues3737img"  # 七牛云盘名
#     # 删除
#     butm = BucketManager(keyq)
#     reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
#     # 上传
#     time.sleep(1)
#     token = keyq.upload_token(bucket, filename)  # 上传新图片
#     reformUp, informUp = put_file(token, filename, imagePath)
#     if reformUp is not None:
#         print('已成功上传 {}'.format(filename))
#         time.sleep(1)
#         baseURL = "https://cjh3737.zeus.cn/"  # 加速域名
#         # subURL = baseURL + '/' + filename + '?imageView2/0/quality/100!/sharpen/1/interlace/1'
#         subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/sharpen/50/interlace/1/ignore-error/1'
#         pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
#         time.sleep(1)
#         return pictureURL
#     else:
#         print(filename + '上传失败！！！')


# if __name__ == '__main__':

#     deleteOldFiles('C:/Users/zoodehao/Desktop/AutoSend/5_萍姐/Pictures/')  # 清空文件夹历史文件

#     AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序AppKey
#     AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序AppSecret

#     RobotWebHookURL1 = 'https://oapi.dingtalk.com/robot/send?access_token=f55407f99b80521faf30aa8e78035d006e41bddd358073f373551d0b91f36601'  # 销管：日销售数据共享群
#     RobotWebHookURL2 = 'https://oapi.dingtalk.com/robot/send?access_token=7b601d4772a4d367af6e9bba268caabaed36ed62e12a5eb6f5da29f451bfcbfc'  # 草晶华销售事业部（领导群）

#     RobotSecret = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'  # 群机器人加签秘钥secret(默认数运小助手)

#     fileFullPath = 'C:/Users/zoodehao/Desktop/AutoSend/5_萍姐/萍姐_正式发送文件.xlsx'
#     savePictuePath = 'C:/Users/zoodehao/Desktop/AutoSend/5_萍姐/Pictures/'

#     workbook = opxl.load_workbook(fileFullPath)
#     worksheetnames = workbook.sheetnames
    
#     sendtitle = [
#         "###### **① 草晶华事业部本日发货明细表**",
#         "###### **② 草晶华事业部本日回款明细表**",
#         "###### **③ 草晶华-发货汇总表**",
#         "###### **④ 草晶华-大客户发货汇总表（含分部）**",
#         "###### **⑤ 草晶华-未发货明细表**",
#         "###### **⑥ 草晶华-新品出货情况**",
#         "###### **⑦ 草晶华-欠货表**"
#     ]

#     sendTypes = int(input('>>>0sendAll-1sendSingle:'))

#     if sendTypes == 0:  # 发送形式 - 全部

#         pictureURL = []
#         for sheetname, picturename in zip(worksheetnames, worksheetnames):
#             try:
#                 excelCatchScreen(fileFullPath, sheetname, picturename, savePictuePath)
#             except BaseException:
#                 print(picturename + '截图出错！！！')
#             try:
#                 getURL = get_image_url(savePictuePath + picturename + '.PNG', picturename)
#                 pictureURL.append(getURL)
#             except BaseException:
#                 print(picturename + '图片URL出错！！！')

#         for ititle, iurl in zip(sendtitle, pictureURL):
#             ddMessage = {  # 发布消息内容
#                 "msgtype": "markdown",
#                 "markdown": {
#                     "title":"【发货回款】进度",  # @某人 才会显示标题
#                     "text":ititle + "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
#                     "\n###### ----------------------------------------"
#                     "\n###### 发布时间：" + str(datetime.now()).split('.')[0]
#                 },  # 发布时间
#                 "at": {
#                     # "atMobiles": [15817552982],  # 指定@某人
#                     "isAtAll": False  # 是否@所有人[False:否, True:是]
#                 }
#             }

#             # 发送消息
#             dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 销管：日销售数据共享群
#             dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 草晶华销售事业部（领导群）

#         # 消息数据
#         summaryMessage = {  # 发布消息内容
#             "msgtype": "markdown",
#             "markdown": {
#                 "title":"发货核销汇总分析",  # @某人 才会显示标题
#                 "text":"##### **各位领导：**"
#                 "\n ##### **草晶华事业部**"
#                 "\n ##### **① " + str(globalYear) + "年" + str(globalMonth) + "月" + str(globalDay) + "日**"
#                 "\n > ##### 发货（发货价）**" + str("{:.2f}".format(1)) + "** 万元"
#                 "\n > ##### 发货（结算价）**" + str("{:.2f}".format(2)) + "** 万元"
#                 "\n > ##### 回款（发货价）**" + str("{:.2f}".format(3)) + "** 万元"
#                 "\n > ##### 回款（结算价）**" + str("{:.2f}".format(4)) + "** 万元"
#                 "\n ##### **② " + str(globalMonth) + "月累计**"
#                 "\n > ##### 发货（发货价）**" + str("{:.2f}".format(5)) + "** 万元"
#                 "\n > ##### 发货（结算价）**" + str("{:.2f}".format(6)) + "** 万元"
#                 "\n > ##### 回款（发货价）**" + str("{:.2f}".format(7)) + "** 万元"
#                 "\n > ##### 回款（结算价）**" + str("{:.2f}".format(8)) + "** 万元"
#                 "\n ##### **③ " + str(globalYear) + "年总况**"
#                 "\n > ##### 1-" + str(globalMonth) + "月发货完成率 **" + str("{:.2f}%".format(8)) + "**, 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(9)) + "**"
#             },
#             "at": {
#                 # "atMobiles": [15817552982],  # 指定@某人
#                 "isAtAll": False  # 是否@所有人[False:否, True:是]
#             }
#         }

#         dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(summaryMessage)  # 销管：日销售数据共享群
#         dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(summaryMessage)  # 草晶华销售事业部（领导群）

#     elif sendTypes == 1:  # 发送形式 - 选择性单张 从 1 开始

#         send_NO_Picture = int(input('发送第几张图片？'))

#         print('***单独发送: ' + str(worksheetnames[send_NO_Picture - 1]) + '.PNG')

#         excelCatchScreen(fileFullPath, worksheetnames[send_NO_Picture - 1], worksheetnames[send_NO_Picture - 1], savePictuePath)

#         ddMessage = {  # 发布消息内容
#             "msgtype": "markdown",
#             "markdown": {
#                 "title":"【发货回款】进度",  # @某人 才会显示标题
#                 "text":sendtitle[send_NO_Picture - 1] + 
#                 "\n![Image被拦截, 请使用非公司网络查看](" + get_image_url(savePictuePath + worksheetnames[send_NO_Picture - 1] + '.PNG', worksheetnames[send_NO_Picture - 1]) + ")"
#                 "\n###### ----------------------------------------"
#                 "\n###### 发布时间：" + str(datetime.now()).split('.')[0]
#             },  # 发布时间
#             "at": {
#                 # "atMobiles": [15817552982],  # 指定@某人
#                 "isAtAll": False  # 是否@所有人[False:否, True:是]
#             }
#         }

#         # 发送消息
#         dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 销管：日销售数据共享群
#         dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 草晶华销售事业部（领导群）

#     else:
#         print('请输入0-1正确的发送方式！！！')
