'''
Author: zoodehao
Date: 2021-08-31 14:51:26
LastEditTime: 2021-09-26 20:10:11
FilePath: \PyCode\6_PK10定制报表.py
Description: 逝者如斯夫, 不舍昼夜.
'''
import re
import os
import glob
import time
import hmac
import json
import xlrd
import pyhdb
import base64
import pymysql
import hashlib
import requests
import calendar
import win32com
import pythoncom
import numpy as np
import pandas as pd
import urllib.parse
import urllib.request
from math import isnan
import openpyxl as opxl
from decimal import Decimal
from termcolor import cprint
from datetime import datetime
from PIL import ImageGrab, Image
from time import strftime, gmtime
from sqlalchemy import create_engine
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch, DispatchEx
from sqlalchemy.types import NVARCHAR, Float, Integer
from qiniu import Auth, put_file, etag, BucketManager
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection, NamedStyle, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, IconSetRule, Rule, IconSet, FormatObject

# ----------------------------------------------------------------------------
# 加载格式框架
# ----------------------------------------------------------------------------
# framepath = 'C:/Users/zoodehao/Desktop/萍姐_格式框架.xlsx'
# workbook = opxl.load_workbook(framepath)


# ----------------------------------------------------------------------------
# 全局日期设置
# ----------------------------------------------------------------------------
globalYear  = datetime.now().year
globalMonth = datetime.now().month
globalDay   = datetime.now().day


# ----------------------------------------------------------------------------
# 数据区
# ----------------------------------------------------------------------------
allStartTime = datetime.now()
print(" > 数据获取中, 请稍等片刻")

# 内网: IP: 192.168.20.183  账号: HANA1107318 密码: Zeus@1107311
# 公网: IP: 119.145.248.183  账号: HANAHBPOUTTEST 密码: Zeus@test147
def get_HANA_Connection():
    connObj = pyhdb.connect(host='192.168.20.183',
                                  port     = 30015,
                                  user     = 'HANA1107318',
                                  password = 'Zeus@1107311')
    return connObj

# 9-12月发货纯销
def get_matFHCX(FHCX):
    cursorFHCX = FHCX.cursor()
    cursorFHCX.execute("""
                       WITH AA AS (SELECT KUNNR, NAME1
                       FROM "ECC_BI"."KNA1"), -- 五级名称
                       AAA AS (SELECT MATNR, MAKTX
                       FROM "ECC_BI"."MAKT"
                       WHERE MANDT = 800 AND SPRAS = 1), -- 物料名称
                       AAAA AS (SELECT MATNR, KBETR
                       FROM "_SYS_BIC"."HD-HAND.SD.POWER_BI.DATA/ZAN_A304"
                       WHERE VKORG = '2008' AND VTWEG = 'A3' AND KSCHL = 'VKP0'), -- 标准零售价
                       AAAAA AS (SELECT ZGRKUNNR, ZGKUNNR, ZMODE, ZKHLX
                       FROM "ECC_BI"."ZTSD017" -- 匹配HX CX 折扣
                       WHERE ZORG = '2008' AND ZCHBOX = '' AND (TO_NUMBER(ZENDDATA) >= TO_NUMBER((LEFT(NOW(),4) || RIGHT(LEFT(NOW(),7),2) || RIGHT(LEFT(NOW(),10),2))))), -- 折扣/计算方式
                       BB AS (SELECT 'CX' AS TYPE, ZDATE_DEAL, MONTH1, ZZTP06, PARTNER, ZQDCY, ZQDCY_DES, ZMATNR, SUM(MENGE) AS AMOUNT
                       FROM "_SYS_BIC"."HD-HAND.SD.DATA/ZAN_SD_CX"
                       WHERE ERP_SALES_ORG IN ('2001', '2008') AND ZYEAR = '2021' AND MONTH1 IN ('08', '09', '10', '11', '12') -- 纯销日期筛选范围
                       GROUP BY ZDATE_DEAL, MONTH1, ZZTP06, PARTNER, ZQDCY, ZQDCY_DES, ZMATNR), -- 纯销
                       CC AS (SELECT BB.TYPE, BB.ZDATE_DEAL, BB.MONTH1, BB.ZZTP06, BB.PARTNER, BB.ZQDCY, BB.ZQDCY_DES, BB.ZMATNR, AAA.MAKTX, BB.AMOUNT
                       FROM BB
                       LEFT JOIN AAA ON BB.ZMATNR = AAA.MATNR),
                       DD AS (SELECT CC.TYPE, CC.ZDATE_DEAL, CC.MONTH1, CC.ZZTP06, CC.PARTNER, AA.NAME1 AS KEHU, CC.ZQDCY, CC.ZQDCY_DES, CC.ZMATNR, CC.MAKTX, CC.AMOUNT
                       FROM CC
                       LEFT JOIN AA ON CC.PARTNER = AA.KUNNR),
                       EE AS (SELECT DD.TYPE, DD.ZDATE_DEAL, DD.MONTH1, DD.ZZTP06, AA.NAME1, DD.PARTNER, DD.KEHU, DD.ZQDCY, DD.ZQDCY_DES, DD.ZMATNR, DD.MAKTX, DD.AMOUNT
                       FROM DD
                       LEFT JOIN AA ON DD.ZZTP06 = AA.KUNNR),
                       FF AS (SELECT 'FH' AS TYPE, LFDAT, MONTH(LFDAT) AS MON, KUNNR_GR, NAME_GR, KUNNR, NAME1, '' AS STOREID, '' AS STRORE, ('000000000' || MATNR) AS ZMATNR, MAKTX, SUM(LFIMG1) AS AMOUNT
                       FROM "ECC_BI"."ZOUTBOUND"
                       WHERE ZZPART1_T_S LIKE N'%草晶华销售事业部%' AND LFDAT != '' AND BEZEI2 != '领用订单' AND YEAR(LFDAT) = '2021' AND MONTH(LFDAT) IN ('08', '09', '10', '11', '12') -- 发货日期筛选范围
                       GROUP BY LFDAT, KUNNR_GR, NAME_GR, KUNNR, NAME1, MATNR, MAKTX), -- 发货
                       HH AS (SELECT FF.TYPE, FF.LFDAT, FF.MON, FF.KUNNR_GR, FF.NAME_GR, FF.KUNNR, FF.NAME1, FF.STOREID, FF.STRORE, FF.ZMATNR, FF.MAKTX, AAAA.KBETR, FF.AMOUNT
                       FROM FF
                       LEFT JOIN AAAA ON FF.ZMATNR = AAAA.MATNR),
                       JJ AS (SELECT EE.TYPE, EE.ZDATE_DEAL, EE.MONTH1, EE.ZZTP06, EE.NAME1, EE.PARTNER, EE.KEHU, EE.ZQDCY, EE.ZQDCY_DES, EE.ZMATNR, EE.MAKTX, AAAA.KBETR AS KBETR, EE.AMOUNT
                       FROM EE
                       LEFT JOIN AAAA ON EE.ZMATNR = AAAA.MATNR),
                       NN AS (SELECT JJ.TYPE, AAAAA.ZMODE, JJ.ZDATE_DEAL, JJ.MONTH1, JJ.ZZTP06, JJ.NAME1, JJ.PARTNER, JJ.KEHU, JJ.ZQDCY, JJ.ZQDCY_DES, JJ.ZMATNR, JJ.MAKTX, JJ.KBETR, AAAAA.ZKHLX, JJ.AMOUNT
                       FROM JJ
                       LEFT JOIN AAAAA ON (JJ.ZZTP06 = AAAAA.ZGRKUNNR AND JJ.PARTNER = AAAAA.ZGKUNNR)) -- 纯销
                       SELECT NN.TYPE, NN.ZMODE, NN.ZDATE_DEAL, NN.MONTH1, NN.ZZTP06, NN.NAME1, NN.PARTNER, NN.KEHU, NN.ZQDCY, NN.ZQDCY_DES, NN.ZMATNR, NN.MAKTX, NN.KBETR, NN.ZKHLX, NN.AMOUNT
                       FROM NN
                       WHERE LEFT(NN.AMOUNT, 1) != '' -- 去掉数量为空的
                       UNION ALL
                       SELECT HH.TYPE, AAAAA.ZMODE, HH.LFDAT, HH.MON, HH.KUNNR_GR, HH.NAME_GR, HH.KUNNR, HH.NAME1, HH.STOREID, HH.STRORE, HH.ZMATNR, HH.MAKTX, HH.KBETR, AAAAA.ZKHLX, HH.AMOUNT
                       FROM HH
                       LEFT JOIN AAAAA ON (HH.KUNNR_GR = AAAAA.ZGRKUNNR AND HH.KUNNR = AAAAA.ZGKUNNR) -- 发货
                       """)
    matFHCX = cursorFHCX.fetchall()
    return matFHCX

# 门店别名表
def get_matRestore(Restore):
    cursorRestore = Restore.cursor()
    cursorRestore.execute("""
                        WITH A1 AS (SELECT ZNUMBER, PARTNER, ZGR
                        FROM "ECC_BI"."ZSD007"
                        WHERE ZVALID = ''), -- StoreID + ClientID => GRID
                        A2 AS (SELECT ZNUMBER, ZSUOSHUNUMBER
                        FROM "ECC_BI"."ZTCRM003"
                        WHERE ZSUOSHUTYPE = '经销商' AND ZVALID = '' AND (TO_NUMBER(ZJIESHUDATE) >= TO_NUMBER((LEFT(NOW(),4) || RIGHT(LEFT(NOW(),7),2) || RIGHT(LEFT(NOW(),10),2))))
                        ), -- StoreID => ClientID
                        BB AS (SELECT ZQDCY_FROM, ZMDBM, ZMDMC, PARTNER
                        FROM "ECC_BI"."ZCRTBMAP002"
                        WHERE VKORG = '2008' AND ZSFZF = '' AND ZMDBM != ''),  -- 别名表
                        CC AS (SELECT BB.ZQDCY_FROM, BB.ZMDBM, BB.ZMDMC, A2.ZSUOSHUNUMBER AS ZSUOSHUNUMBER
                        FROM BB
                        LEFT JOIN A2 ON BB.ZMDBM = A2.ZNUMBER)
                        SELECT CC.ZQDCY_FROM, CC.ZMDBM, CC.ZMDMC, CC.ZSUOSHUNUMBER, A1.ZGR, (A1.ZGR || CC.ZSUOSHUNUMBER) AS GR_CLIENT
                        FROM CC
                        LEFT JOIN A1 ON (CC.ZMDBM = A1.ZNUMBER AND CC.ZSUOSHUNUMBER = A1.PARTNER)
                        """)
    matRestore = cursorRestore.fetchall()
    return matRestore

# 计算方式 折扣
def get_matDiscount(Discount):
    cursorDiscount = Discount.cursor()
    cursorDiscount.execute("""
                           SELECT (ZGRKUNNR || ZGKUNNR) AS KEY, ZGRKUNNR, ZGKUNNR, ZMODE, ZKHLX
                           FROM "ECC_BI"."ZTSD017"
                           WHERE ZORG = '2008' AND ZCHBOX = '' AND (TO_NUMBER(ZENDDATA) >= TO_NUMBER((LEFT(NOW(),4) || RIGHT(LEFT(NOW(),7),2) || RIGHT(LEFT(NOW(),10),2))))
                           """)
    matDiscount = cursorDiscount.fetchall()
    return matDiscount

# （草晶华）物料名称 标准单价信息表
def get_matMaterialInfo(MaterialInfo):
    cursorMaterialInfo = MaterialInfo.cursor()
    cursorMaterialInfo.execute("""WITH A1 AS(SELECT MATNR, KBETR
                                   FROM "_SYS_BIC"."HD-HAND.SD.POWER_BI.DATA/ZAN_A304"
                                   WHERE VKORG = '2008' AND VTWEG = 'A3' AND KSCHL = 'VKP0'), -- 物料标准单价
                                   A2 AS(SELECT MATNR, MAKTX
                                   FROM "ECC_BI"."MAKT"
                                   WHERE MANDT = 800 AND SPRAS = 1)
                                   SELECT A2.MATNR, A2.MAKTX, A1.KBETR
                                   FROM A2
                                   LEFT JOIN A1 ON A1.MATNR = A2.MATNR
                                   """)
    matMaterialInfo = cursorMaterialInfo.fetchall()
    return matMaterialInfo

# 五级架构中文名称
def get_matLevel5(Level5):
    cursorLevel5 = Level5.cursor()
    cursorLevel5.execute("""SELECT KUNNR, NAME1
                            FROM "ECC_BI"."KNA1"
                            """)
    matLevel5 = cursorLevel5.fetchall()
    return matLevel5

conn = get_HANA_Connection()

# 发货纯销
dataFHCX = pd.DataFrame(get_matFHCX(conn), columns=['数据源', '计算方式', '日期', '月份', 'GR编码', 'GR名称', '客户ID', '客户名称', '门店ID', '门店名称', '物料ID', '物料名称', '标准单价', '折扣', '数量'])

# 门店别名表
dataRestore = pd.DataFrame(get_matRestore(conn), columns=['网上门店名称', '门店ID', '门店名称', '客户ID', 'GR编码', 'GR客户ID'])

# 计算方式 折扣
dataDiscount = pd.DataFrame(get_matDiscount(conn), columns=['主键', 'GR编码', '客户ID', '计算方式', '折扣'])

# 标准物料信息
dataMatetialStandardInfo = pd.DataFrame(get_matMaterialInfo(conn), columns=['物料ID', '物料名称', '标准单价'])

# 五级架构中文名称
dataLevel5 = pd.DataFrame(get_matLevel5(conn), columns=['五级ID', '五级中文名称'])


conn = pymysql.connect(host='192.168.20.241',
                        port    = 3306,
                        user    = 'root',
                        passwd  = 'Powerbi#1217',
                        db      = 'dkh',
                        charset = 'utf8')

cursor = conn.cursor()

executeCode = """SELECT dept_5, DKH_materiel_id, 'DKH' AS SOURCE, date, MONTH(date) AS MONTH, sfa_desc, SUM(amount)
                 FROM dkhfact
                 WHERE YEAR(date) = '2021' AND MONTH(date) IN ('08', '09', '10', '11', '12')
                 GROUP BY dept_5, DKH_materiel_id, date, sfa_desc"""

cursor.execute(executeCode)  # 执行查询
rowNum = cursor.rowcount  # 数据条数
getDKH = cursor.fetchall()  # 获取全部数据
conn.commit()  # 提交确认
cursor.close()  # 关闭光标
conn.close()  # 关闭连接

dataDKH = pd.DataFrame(getDKH, columns=['dept_5', 'materialID', '数据源', '日期', '月份', '门店名称', '数量'])
getdataEndTime = datetime.now()
print(" >>获取数据耗时: " + strftime("%H:%M:%S", gmtime((getdataEndTime - allStartTime).seconds)))

# ----------------------------------------------------------------------------
# 字典区
# ----------------------------------------------------------------------------
print(" > 数据清洗中, 请稍等片刻")

cleandataStartTime = datetime.now()

storeID_Dict = {}  # 标准门店ID
store_Info = pd.read_excel('D:/FilesCenter/DKH-BottomTable/DKH对照表.xlsx', sheet_name='SFA_Hierarchy', header=0, dtype=str)  # 网上门店信息表
for irow in range(store_Info.shape[0]):
    storeID_Dict[store_Info.loc[irow, 'OUT_SFA_bianma']] = str(store_Info.loc[irow, 'SFA_id']).replace('.0', '')
    
GRID_Dict = {}  # GR编码
clientID_Dict = {}  # 客户ID
GRclientID_Dict = {} # GR客户ID
for irow in range(dataRestore.shape[0]):
    GRID_Dict[dataRestore.loc[irow, '门店ID']] = dataRestore.loc[irow, 'GR编码']
    clientID_Dict[dataRestore.loc[irow, '门店ID']] = dataRestore.loc[irow, '客户ID']
    GRclientID_Dict[dataRestore.loc[irow, '门店ID']] = dataRestore.loc[irow, 'GR客户ID']

materialID_Dict = {}  # 标准物料ID
material_Info = pd.read_excel('D:/FilesCenter/DKH-BottomTable/商品编码对照字典.xlsx', sheet_name=0, header=0, dtype=str)  # 物料信息表
for irow in range(material_Info.shape[0]):
    materialID_Dict[str(material_Info.loc[irow, '编码']).replace('.0', '')] = material_Info.loc[irow, '标准物料ID']

materialStandardDesc_Dict = {}  # 标准物料名称
materialStandardPrice_Dict = {}  # 标准物料单价
for irow in range(dataMatetialStandardInfo.shape[0]):
    materialStandardDesc_Dict[dataMatetialStandardInfo.loc[irow, '物料ID']] = dataMatetialStandardInfo.loc[irow, '物料名称']
    materialStandardPrice_Dict[dataMatetialStandardInfo.loc[irow, '物料ID']] = dataMatetialStandardInfo.loc[irow, '标准单价']
    
Method_Dict = {}  # 计算方式
Discount_Dict = {}  # 折扣
for irow in range(dataDiscount.shape[0]):
    Method_Dict[dataDiscount.loc[irow, '主键']] = dataDiscount.loc[irow, '计算方式']
    Discount_Dict[dataDiscount.loc[irow, '主键']] = dataDiscount.loc[irow, '折扣']

Level5_Dict = {}  # 五级中文名称
for irow in range(dataLevel5.shape[0]):
    Level5_Dict[dataLevel5.loc[irow, '五级ID']] = dataLevel5.loc[irow, '五级中文名称']

def newDF():
    return pd.DataFrame(columns=['dept_5', 'materialID', 'GR客户ID', '数据源', '计算方式', '日期', '月份', 'GR编码', 'GR名称', '客户ID', '客户名称', '门店ID', '门店名称', '物料ID', '物料名称', '标准单价', '折扣', '数量'])

print('debug1')
dfDKH = newDF()
dfDKH['dept_5'] = dataDKH['dept_5']
dfDKH['materialID'] = dataDKH['materialID']
dfDKH['数据源'] = dataDKH['数据源']
dfDKH['日期'] = dataDKH['日期']
dfDKH['月份'] = dataDKH['月份']
dfDKH['门店名称'] = dataDKH['门店名称']
dfDKH['数量'] = dataDKH['数量']

dfDKH['门店ID'] = dfDKH.apply(lambda x: storeID_Dict.setdefault(x['dept_5'], ''), axis=1)
dfDKH['物料ID'] = dfDKH.apply(lambda x: materialID_Dict.setdefault(x['materialID'], ''), axis=1)
dfDKH['GR编码'] = dfDKH.apply(lambda x: GRID_Dict.setdefault(x['门店ID'], ''), axis=1)
dfDKH['客户ID'] = dfDKH.apply(lambda x: clientID_Dict.setdefault(x['门店ID'], ''), axis=1)
dfDKH['GR客户ID'] = dfDKH.apply(lambda x: GRclientID_Dict.setdefault(x['门店ID'], ''), axis=1)
dfDKH['GR名称'] = dfDKH.apply(lambda x: Level5_Dict.setdefault(x['GR编码'], ''), axis=1)
dfDKH['客户名称'] = dfDKH.apply(lambda x: Level5_Dict.setdefault(x['客户ID'], ''), axis=1)
dfDKH['物料名称'] = dfDKH.apply(lambda x: materialStandardDesc_Dict.setdefault(x['物料ID'], ''), axis=1)
dfDKH['标准单价'] = dfDKH.apply(lambda x: materialStandardPrice_Dict.setdefault(x['物料ID'], ''), axis=1)
dfDKH['计算方式'] = dfDKH.apply(lambda x: Method_Dict.setdefault(x['GR客户ID'], ''), axis=1)
dfDKH['折扣'] = dfDKH.apply(lambda x: Discount_Dict.setdefault(x['GR客户ID'], ''), axis=1)

print('debug2')
dfDKH = dfDKH[dfDKH['计算方式'] == 'CX']  # 只取纯销, 行数太多
dfNew = dfDKH.drop(['dept_5', 'materialID', 'GR客户ID'], axis=1)

print('debug3')
# 数据源为FH GR为1100003217 计算方式全部改为HX
dataFHCX['折扣'] = pd.to_numeric(dataFHCX['折扣'], errors='coerce')
for irow in range(dataFHCX.shape[0]):
    if dataFHCX.loc[irow, '数据源'] == 'FH' and dataFHCX.loc[irow, 'GR编码'] == '1100003217':
        dataFHCX.loc[irow, '计算方式'] = 'HX'
        if dataFHCX.loc[irow, 'GR编码'] in ['1100003357', '1100003217', '1100005256'] and isnan(dataFHCX.loc[irow, '折扣']):
            dataFHCX.loc[irow, '折扣'] = 20
    else:
        if dataFHCX.loc[irow, 'GR编码'] in ['1100003357', '1100003217', '1100005256'] and isnan(dataFHCX.loc[irow, '折扣']):
            dataFHCX.loc[irow, '折扣'] = 20
 
print('debug4')
dataCX = dataFHCX[(dataFHCX['数据源'] == 'CX') & (dataFHCX['计算方式'] == 'CX')]  # 只取纯销CX
dataFH = dataFHCX[(dataFHCX['数据源'] == 'FH') & (dataFHCX['计算方式'] == 'HX')]  # 只取发货HX

print('debug5')
allData = dfNew.append([dataCX, dataFH])  # 整合数据
allData['标准单价'] = pd.to_numeric(allData['标准单价'], errors='coerce')
allData['折扣'] = pd.to_numeric(allData['折扣'], errors='coerce')
allData['数量'] = pd.to_numeric(allData['数量'], errors='coerce')

cleandataEndTime = datetime.now()
print(" >>清洗数据耗时: " + strftime("%H:%M:%S", gmtime((cleandataEndTime - cleandataStartTime).seconds)))

print(" > 上传数据中, 请稍等片刻")
exportStartTime = datetime.now()

def mapping_df_types(df):
    dtypedict = {}
    for i, j in zip(df.columns, df.dtypes):
        if "object" in str(j):
            dtypedict.update({i: NVARCHAR(length=100)})
        if "float" in str(j):
            dtypedict.update({i: Float(precision=2, asdecimal=True)})
        if "int" in str(j):
            dtypedict.update({i: Integer()})
    return dtypedict

dtypedictPK100 = mapping_df_types(allData)
enginePK100 = create_engine('mysql+pymysql://root:Powerbi#1217@192.168.20.241:3306/dkh', encoding='utf-8', echo=False,
                        pool_size=100, max_overflow=10, pool_timeout=100, pool_recycle=7200)
allData.to_sql('PK10', con=enginePK100, dtype=dtypedictPK100, index=False, if_exists='replace')
# allData.to_excel('C:/Users/Zeus/Desktop/百日PK数据源-' + str(globalYear) + str(globalMonth).zfill(2) + str(globalDay).zfill(2) + '.xlsx', index=False)

allEndTime = datetime.now()
print(" >>上传数据耗时: " + strftime("%H:%M:%S", gmtime((allEndTime - exportStartTime).seconds)))
cprint(">>>总耗时: " + strftime("%H:%M:%S", gmtime((allEndTime - allStartTime).seconds)), 'cyan', attrs=['bold', 'reverse', 'blink'])

       




 




# # ----------------------------------------------------------------------------
# # 第一页 ①彩盒
# # ----------------------------------------------------------------------------
# worksheet1 = workbook['①彩盒']


# # ----------------------------------------------------------------------------
# # 带时间文字填充
# # ----------------------------------------------------------------------------
# # 第 1 页 - ①彩盒
# worksheet1.cell(row=2, column=2).value = str(globalYear) + '.01.01-' + str(globalYear) + '.' + str(globalMonth).zfill(2) + '.' + str(globalDay).zfill(2)


# # ----------------------------------------------------------------------------
# # 存储表格
# # ----------------------------------------------------------------------------
# try:
#     workbook.save('C:/Users/zoodehao/Desktop/萍姐_正式发送文件.xlsx')
#     print('文件储存完毕！！！')
# except Exception as e:
#     print('保存文件失败！！！', e)

# # ----------------------------------------------------------------------------
# # 以下是发送板块
# # ----------------------------------------------------------------------------


# # 清空指定文件夹
# def deleteOldFiles(path):
#     deleteFileList = os.listdir(path)
#     all_PNG = glob.glob(path + "*.PNG")
#     print("该目录下文件有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中, PNG: " +
#           str(len(all_PNG)) + "个")
#     if len(all_PNG) != 0:
#         for deletefile in deleteFileList:
#             isDeleteFile = os.path.join(path, deletefile)
#             if os.path.isfile(isDeleteFile):
#                 os.remove(isDeleteFile)
#         all_DelPNG = glob.glob(path + "*.*")
#         if len(all_DelPNG) == 0:
#             print("已清空文件夹！！！")
#         else:
#             print("存在未删除文件, 请检查是否存在非PNG格式文件")
#     else:
#         print("不存在PNG文件")


# # screenArea——格式类似"A1:J10"
# def excelCatchScreen(file_name, sheet_name, name, save_path):
#     pythoncom.CoInitialize()  # excel多线程相关
#     Application = win32com.client.gencache.EnsureDispatch(
#         "Excel.Application")  # 启动excel
#     Application.Visible = False  # 是否可视化
#     Application.DisplayAlerts = False  # 是否显示警告
#     wb = Application.Workbooks.Open(file_name, ReadOnly=False)  # 打开excel
#     # ws = wb.Sheets(sheet_name)  # 选择Sheet
#     ws = wb.Worksheets(sheet_name)  # 选择Sheet
#     ws.Activate()  # 激活当前工作表
#     userange = ws.UsedRange
#     # 注意：要从A1开始的表格
#     screen_area = 'A1:' + str(
#         opxl.utils.get_column_letter(userange.Columns.Count)) + str(
#             userange.Rows.Count)
#     ws.Range(screen_area).CopyPicture()  # 复制图片区域
#     time.sleep(1)
#     ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
#     Application.Selection.ShapeRange.Name = name  # 将刚刚选择的Shape重命名, 避免与已有图片混淆
#     ws.Shapes(name).Copy()  # 选择图片
#     time.sleep(1)
#     img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
#     img_name = name + ".PNG"
#     img.save(save_path + img_name)  # 保存图片
#     # time.sleep(1)
#     # wb.Save()
#     # time.sleep(1)
#     wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
#     time.sleep(1)
#     Application.Quit()  # 退出excel
#     pythoncom.CoUninitialize()


# # 定义钉钉功能
# class dingdingFunction(object):
#     def __init__(self, roboturl, robotsecret, appkey, appsecret):
#         """
#         :param roboturl: 群机器人WebHook_url
#         :param robotsecret: 安全设置的加签秘钥
#         :param appkey: 企业开发平台小程序AppKey
#         :param appsecret: 企业开发平台小程序AppSecret
#         """
#         self.roboturl = roboturl
#         self.robotsecret = robotsecret
#         self.appkey = appkey
#         self.appsecret = appsecret
#         timestamp = round(time.time() * 1000)  # 时间戳
#         secret_enc = robotsecret.encode('utf-8')
#         string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
#         string_to_sign_enc = string_to_sign.encode('utf-8')
#         hmac_code = hmac.new(secret_enc,
#                              string_to_sign_enc,
#                              digestmod=hashlib.sha256).digest()
#         sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
#         self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(
#             timestamp, sign)  # 最终url,url+时间戳+签名

#     # 发送文件
#     def getAccess_token(self):
#         url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (
#             AppKey, AppSecret)
#         headers = {'Content-Type': "application/x-www-form-urlencoded"}
#         data = {'appkey': self.appkey, 'appsecret': self.appsecret}
#         r = requests.request('GET', url, data=data, headers=headers)
#         access_token = r.json()["access_token"]
#         return access_token

#     def getMedia_id(self, filespath):
#         access_token = self.getAccess_token()  # 拿到接口凭证
#         url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
#         files = {'media': open(filespath, 'rb')}
#         data = {'access_token': access_token, 'type': 'file'}
#         response = requests.post(url, files=files, data=data)
#         json = response.json()
#         return json["media_id"]

#     def sendFile(self, chatid, filespath):
#         access_token = self.getAccess_token()
#         media_id = self.getMedia_id(filespath)
#         url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
#         header = {'Content-Type': 'application/json'}
#         data = {
#             'access_token': access_token,
#             'chatid': chatid,
#             'msg': {
#                 'msgtype': 'file',
#                 'file': {
#                     'media_id': media_id
#                 }
#             }
#         }
#         r = requests.request('POST',
#                              url,
#                              data=json.dumps(data),
#                              headers=header)
#         print(r.json()["errmsg"])

#     # 发送消息
#     def sendMessage(self, content):
#         """
#         :param content: 发送内容
#         """
#         header = {"Content-Type": "application/json", "Charset": "UTF-8"}
#         sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
#         sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
#         request = urllib.request.Request(url=self.webhook_url,
#                                          data=sendContent,
#                                          headers=header)  # 发送请求
#         opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
#         print(opener.read().decode())  # 打印返回的结果


# # 上传本地图片获取网上图片URL
# def get_image_url(imagePath, pictureName):
#     if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split(
#             '.')[-1] == 'JPG':
#         filename = pictureName + '.' + str(imagePath).split('.')[
#             -1]  # 七牛云网盘文件名
#     elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split(
#             '.')[-1] == 'PNG':
#         filename = pictureName + '.' + str(imagePath).split('.')[
#             -1]  # 七牛云网盘文件名
#     else:
#         print("请检查图片格式！！！")
#     # 七牛云密钥管理：https://portal.qiniu.com/user/key
#     # 【账号：13267854059  密码：z****】
#     access_key = "fjlWDEbF1fqBU98UsdDJRcSSKODT9Gq7tA3gu8eY"
#     secret_key = "thiWFpO881GfhlaAz1Wkk2yEcvV3ue2OHnY_5D9V"
#     keyq = Auth(access_key, secret_key)
#     bucket = "zues3737img"  # 七牛云盘名
#     # 删除
#     butm = BucketManager(keyq)
#     reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
#     # 上传
#     time.sleep(1)
#     token = keyq.upload_token(bucket, filename)  # 上传新图片
#     reformUp, informUp = put_file(token, filename, imagePath)
#     if reformUp is not None:
#         print('已成功上传 {}'.format(filename))
#         time.sleep(1)
#         baseURL = "https://cjh3737.zeus.cn/"  # 加速域名
#         # subURL = baseURL + '/' + filename + '?imageView2/0/quality/100!/sharpen/1/interlace/1'
#         subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/sharpen/50/interlace/1/ignore-error/1'
#         pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
#         time.sleep(1)
#         return pictureURL
#     else:
#         print(filename + '上传失败！！！')


# if __name__ == '__main__':

#     deleteOldFiles('C:/Users/zoodehao/Desktop/Picture/')  # 清空文件夹历史文件

#     AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序AppKey
#     AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序AppSecret

#     RobotWebHookURL1 = 'https://oapi.dingtalk.com/robot/send?access_token=f55407f99b80521faf30aa8e78035d006e41bddd358073f373551d0b91f36601'  # 销管：日销售数据共享群
#     RobotWebHookURL2 = 'https://oapi.dingtalk.com/robot/send?access_token=7b601d4772a4d367af6e9bba268caabaed36ed62e12a5eb6f5da29f451bfcbfc'  # 草晶华销售事业部（领导群）

#     RobotSecret = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'  # 群机器人加签秘钥secret(默认数运小助手)

#     fileFullPath = 'C:/Users/zoodehao/Desktop/萍姐_正式发送文件.xlsx'
#     savePictuePath = 'C:/Users/zoodehao/Desktop/Picture/'

#     workbook = opxl.load_workbook(fileFullPath)
#     worksheetnames = workbook.sheetnames
    
#     sendtitle = [
#         "###### **① 草晶华事业部本日发货明细表**", "###### **② 草晶华事业部本日回款明细表**",
#         "###### **③ 草晶华-发货汇总表**", "###### **④ 草晶华-大客户发货汇总表（含分部）**",
#         "###### **⑤ 草晶华-未发货明细表**", "###### **⑥ 草晶华-新品出货情况**",
#         "###### **⑦ 草晶华-欠货表**"
#     ]

#     sendTypes = int(input('>>>0sendAll-1sendSingle:'))

#     if sendTypes == 0:  # 发送形式 - 全部

#         pictureURL = []
#         for sheetname, picturename in zip(worksheetnames, worksheetnames):
#             try:
#                 excelCatchScreen(fileFullPath, sheetname, picturename,
#                                  savePictuePath)
#             except BaseException:
#                 print(picturename + '截图出错！！！')
#             try:
#                 getURL = get_image_url(savePictuePath + picturename + '.PNG',
#                                        picturename)
#                 pictureURL.append(getURL)
#             except BaseException:
#                 print(picturename + '图片URL出错！！！')

#         for ititle, iurl in zip(sendtitle, pictureURL):
#             ddMessage = {  # 发布消息内容
#                 "msgtype": "markdown",
#                 "markdown": {
#                     "title":
#                     "【发货回款】进度",  # @某人 才会显示标题
#                     "text":
#                     ititle + "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
#                     "\n###### ----------------------------------------"
#                     "\n###### 发布时间：" + str(datetime.now()).split('.')[0]
#                 },  # 发布时间
#                 "at": {
#                     # "atMobiles": [15817552982],  # 指定@某人
#                     "isAtAll": False  # 是否@所有人[False:否, True:是]
#                 }
#             }

#             # 发送消息
#             dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 销管：日销售数据共享群
#             dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 草晶华销售事业部（领导群）

#         # 消息数据
#         summaryMessage = {  # 发布消息内容
#             "msgtype": "markdown",
#             "markdown": {
#                 "title":
#                 "发货核销汇总分析",  # @某人 才会显示标题
#                 "text":
#                 "##### **各位领导：**"
#                 "\n ##### **草晶华事业部**"
#                 "\n ##### **① " + str(globalYear) + "年" + str(globalMonth) +
#                 "月" + str(globalDay) + "日**"
#                 "\n > ##### 发货（发货价）**" + str("{:.2f}".format(1)) + "** 万元"
#                 "\n > ##### 发货（结算价）**" + str("{:.2f}".format(2)) + "** 万元"
#                 "\n > ##### 回款（发货价）**" + str("{:.2f}".format(3)) + "** 万元"
#                 "\n > ##### 回款（结算价）**" + str("{:.2f}".format(4)) + "** 万元"
#                 "\n ##### **② " + str(globalMonth) + "月累计**"
#                 "\n > ##### 发货（发货价）**" + str("{:.2f}".format(5)) + "** 万元"
#                 "\n > ##### 发货（结算价）**" + str("{:.2f}".format(6)) + "** 万元"
#                 "\n > ##### 回款（发货价）**" + str("{:.2f}".format(7)) + "** 万元"
#                 "\n > ##### 回款（结算价）**" + str("{:.2f}".format(8)) + "** 万元"
#                 "\n ##### **③ " + str(globalYear) + "年总况**"
#                 "\n > ##### 1-" + str(globalMonth) + "月发货完成率 **" + str("{:.2f}%".format(8)) + "**, 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(9)) + "**"
#             },
#             "at": {
#                 # "atMobiles": [15817552982],  # 指定@某人
#                 "isAtAll": False  # 是否@所有人[False:否, True:是]
#             }
#         }

#         dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(summaryMessage)  # 销管：日销售数据共享群
#         dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(summaryMessage)  # 草晶华销售事业部（领导群）

#     elif sendTypes == 1:  # 发送形式 - 选择性单张 从 1 开始

#         send_NO_Picture = int(input('发送第几张图片？'))

#         print('***单独发送: ' + str(worksheetnames[send_NO_Picture - 1]) + '.PNG')

#         excelCatchScreen(fileFullPath, worksheetnames[send_NO_Picture - 1], worksheetnames[send_NO_Picture - 1], savePictuePath)

#         ddMessage = {  # 发布消息内容
#             "msgtype": "markdown",
#             "markdown": {
#                 "title":
#                 "【发货回款】进度",  # @某人 才会显示标题
#                 "text":
#                 sendtitle[send_NO_Picture - 1] + "\n![Image被拦截, 请使用非公司网络查看](" +
#                 get_image_url(
#                     savePictuePath + worksheetnames[send_NO_Picture - 1] +
#                     '.PNG', worksheetnames[send_NO_Picture - 1]) + ")"
#                 "\n###### ----------------------------------------"
#                 "\n###### 发布时间：" + str(datetime.now()).split('.')[0]
#             },  # 发布时间
#             "at": {
#                 # "atMobiles": [15817552982],  # 指定@某人
#                 "isAtAll": False  # 是否@所有人[False:否, True:是]
#             }
#         }

#         # 发送消息
#         dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 销管：日销售数据共享群
#         dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 草晶华销售事业部（领导群）

#     else:
#         print('请输入0-1正确的发送方式！！！')
