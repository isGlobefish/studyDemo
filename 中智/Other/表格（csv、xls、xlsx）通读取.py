# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.2.2
@projectName   : pythonProjectPy3.9 
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2020/12/4 15:04
'''

import csv
import chardet
import xlrd
from openpyxl import load_workbook


class TableRead(object):
    CSV = 'csv'
    XLSX = 'xlsx'
    XLS = 'xls'

    def __init__(self, path):
        self.path = path
        self.fileType = path.split('.')[-1]
        if self.fileType not in [TableRead.CSV, TableRead.XLS, TableRead.XLSX]:
            raise NameError('文件类型错误')
        self._intFile()

    def _intFile(self):
        if self.fileType == TableRead.CSV:
            self._initCsv()
        elif self.fileType == TableRead.XLSX:
            self._initXlsx()
        else:
            self._initXls()

    def _initCsv(self):
        f = open(self.path, 'r', encoding=self._detectCoding())
        dialect = self._detectDelimiters()
        self.file = csv.reader(f, dialect)
        self.sheet = self.file
        self.rows = self.file.__iter__()

    def _initXlsx(self):
        file = load_workbook(
            self.path,
            read_only=True,
            data_only=True
        )
        self.file = file
        self.sheet = self.file[file.sheetnames[0]]
        self.rows = self.sheet.rows.__iter__()

    def _initXls(self):
        self.file = xlrd.open_workbook(self.path)
        self.sheet = self.file.sheet_by_index(0)
        self.rows = iter(range(self.sheet.nrows))

    def _detectCoding(self):
        with open(self.path, 'rb') as fin:
            encoding_type = chardet.detect(fin.readline())['encoding']
        return encoding_type

    def _detectDelimiters(self, delimiters="\t|,;"):
        fin = open(self.path, 'r', encoding=self._detectCoding())
        dialect = csv.Sniffer().sniff(fin.readline(), delimiters=delimiters)
        return dialect

    def getNextLine(self):
        if self.fileType == TableRead.CSV:
            return self.rows.__next__()
        elif self.fileType == TableRead.XLSX:
            return [i.value for i in self.rows.__next__()]
        else:
            return self.sheet.row_values(self.rows.__next__())


if __name__ == '__main__':
    file_path = 'C:/Users/Long/Desktop/老百姓sheet2自动填充表.csv'
    a = TableRead(file_path)
    print(a.getNextLine())

    file_path = 'C:/Users/Long/Desktop/网上流向账号跟密码.xlsx'
    a = TableRead(file_path)
    print(a.getNextLine())

    file_path = '1.csv'
    a = TableRead(file_path)
    print(a.getNextLine())

import xlwt
import xlrd

data = xlrd.open_workbook('C:/Users/Long/Desktop/321.xlsx')
sheet1Data = data.sheet_by_index(0)
workbook = xlwt.Workbook(encoding='utf-8')
booksheet = workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
nrows = sheet1Data.nrows
cols = sheet1Data.ncols
for i in range(nrows):
    for j in range(cols):
        booksheet.write(i, j, sheet1Data.cell_value(rowx=i, colx=j))
workbook.save('C:/Users/Long/Desktop/000.xls')

import os
import pandas as pd


# 文件迁移（仅是适用excel表格）
def migrateFiles(orginalPath, goalPath, fileNameList):
    fileList = os.listdir(orginalPath)  # 该文件夹下所有的文件（包括文件夹）
    print("转换" + str(fileList) + "文件格式")
    for file, filename in zip(fileList, fileNameList):  # 遍历所有文件
        fileName = os.path.splitext(file)[0]  # 获取文件名
        fileType = os.path.splitext(file)[1]  # 获取文件扩展名
        if fileType == '.xls' or fileType == '.csv':
            data = xlrd.open_workbook(orginalPath + fileName + fileType)
            sheet1Data = data.sheet_by_index(0)
            workbook = xlwt.Workbook(encoding='utf-8')
            booksheet = workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
            nrows = sheet1Data.nrows
            cols = sheet1Data.ncols
            for i in range(nrows):
                for j in range(cols):
                    booksheet.write(i, j, sheet1Data.cell_value(rowx=i, colx=j))
            workbook.save(goalPath + filename + fileType)
        elif fileType == '.xlsx':
            data = pd.read_excel(orginalPath + fileName + fileType, header=0, index=False)
            data.to_excel(goalPath + filename + fileType)
        else:
            print('不支持该格式文件迁移')
            pass
    print("文件位置迁移完成！！！")


migrateFiles('C:\\Users\\Long\\Desktop\\000\\', 'C:\\Users\\Long\\Desktop\\111\\', ['111', '222', '333'])


