# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.2.2
@projectName   : pythonProjectPy3.9 
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2020/11/30 18:12
'''
import os
import re
import cv2
import glob
import json
import time
import xlrd
import xlwt
import csv
import openpyxl
import datetime
import xlsxwriter
import pytesseract  # 用于图片转文字
import numpy as np
import pandas as pd
import urllib.request
from PIL import Image  # 用于打开图片和对图片处理
from termcolor import cprint
from selenium import webdriver
import win32com.client as win32
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains


def get_exce(wei_zhi):
    all_exce = glob.glob(wei_zhi + "*.xls")
    print("该目录下有" + str(len(all_exce)) + "个excel文件：")
    if (len(all_exce) == 0):
        return 0
    else:
        for i in range(len(all_exce)):
            print(all_exce[i])
        return all_exce


def input_a(id_1, value):
    input_box = driver.find_element_by_id(id_1)
    try:
        input_box.send_keys(value)
    except Exception as e:
        print('fail', e)


def input_b(id_1, value):
    input_box = driver.find_element_by_xpath(id_1)
    try:
        input_box.send_keys(value)
    except Exception as e:
        print('fail', e)


def button_a(xpath):
    global wait
    wait = WebDriverWait(driver, 3)
    button = driver.find_element_by_xpath(xpath)
    try:
        button.click()
    except Exception as e:
        print('fail搜索', e)


def button_b(xpath):
    global wait
    wait = WebDriverWait(driver, 3)
    button = driver.find_element_by_xpath(xpath)
    try:
        driver.execute_script("$(arguments[0]).click()", button)
    except Exception as e:
        print('fail搜索', e)


def clear_a(id_1):
    input_box = driver.find_element_by_xpath(id_1)
    try:
        input_box.clear()
    except Exception as e:
        print('fail清空输入框', e)


def switch(xpath):
    xf = driver.find_element_by_xpath(xpath)
    try:
        driver.switch_to.frame(xf)  # 切换
    except Exception as e:
        print('切换失败', e)


def xuanfu(xpath):
    element = driver.find_element_by_xpath(xpath)
    try:
        ActionChains(driver).move_to_element(element).perform()  # 鼠标悬浮
    except Exception as e:
        print('悬浮失败', e)


# 获取exce文件下的所有sheet
def get_sheet(fh):
    sheets = fh.sheets()
    return sheets


# 获取sheet下有多少行数据
def get_sheetrow_num(sheet):
    return sheet.nrows


# 海王表头及填充省份
def get_sheet_data(sheet, row, j):
    for i in range(row - 1):
        if (i == 0):
            global biao_tou
            biao_tou = ['类别', '商品SAP编码', '商品名称', '规格', '单位', '店号/区域ID', '店名/区域', '销量', '过账日期', '合同价', '省份']
            continue
        values = sheet.row_values(i)
        values.append(sf[j])
        all_data.append(values)
    return all_data


def time_0(time):
    if time > 9:
        time1 = str(time)
    else:
        time1 = '0' + str(time)
    return time1


# 每月最后一天日期
def last_day_of_month(any_day):
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)


# 清空指定文件夹
def deleteOldFiles(path):
    # 删除指定文件夹下的文件
    # path = "E:/大客户数据/LBX/mergeTableLBX/"
    deleteFileList = os.listdir(path)
    all_Xls = glob.glob(path + "*.xls")
    all_Xlsx = glob.glob(path + "*.xlsx")
    all_Csv = glob.glob(path + "*.csv")
    print("该目录下有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中【xls:" + str(len(all_Xls)) + ", xlsx:" + str(
        len(all_Xlsx)) + ", csv:" + str(len(all_Csv)) + "】")
    if (len(all_Xls) != 0 or len(all_Xlsx) != 0 or len(all_Csv) != 0):
        for deletefile in deleteFileList:
            isDeleteFile = os.path.join(path, deletefile)
            if os.path.isfile(isDeleteFile):
                os.remove(isDeleteFile)
        all_DelXls = glob.glob(path + "*.xls")
        all_DelXlsx = glob.glob(path + "*.xlsx")
        all_DelCsv = glob.glob(path + "*.csv")
        if (len(all_DelXls) == 0 and len(all_DelXlsx) == 0 and len(all_DelCsv) == 0):
            print("已清空文件夹！！！")
        else:
            print("存在未删除文件")
    else:
        print("不存在excel文件")


# Excel格式转换：.csv ---> .xls
def getFormat(openPath, savePath):
    deleteOldFiles(savePath)  # 清空文件夹
    fileList = os.listdir(openPath)  # 该文件夹下所有的文件（包括文件夹）
    print("转换" + str(fileList) + "文件格式")
    for file in fileList:  # 遍历所有文件
        fileName = os.path.splitext(file)[0]  # 获取文件名
        fileType = os.path.splitext(file)[1]  # 获取文件扩展名
        try:
            data = xlrd.open_workbook(openPath + fileName + fileType)  # 读取文件
            sheet1Data = data.sheet_by_index(0)  # 选择第一个子页
            workbook = xlwt.Workbook(encoding='utf-8')
            booksheet = workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
            nrows = sheet1Data.nrows
            cols = sheet1Data.ncols
            for i in range(nrows):
                for j in range(cols):
                    booksheet.write(i, j, sheet1Data.cell_value(rowx=i, colx=j))
            workbook.save(savePath + fileName + '.xls')
        except:
            with open(openPath + fileName + fileType, 'r') as csvfile:
                openCsvFile = csv.reader(csvfile)  # 读取文件
                workbook = openpyxl.Workbook()  # 打开一个文件
                sheet = workbook.create_sheet(index=0)  # 在文件上创建Sheet1
                for rows, lines in enumerate(openCsvFile, start=1):
                    for cols, values in enumerate(lines, start=1):
                        sheet.cell(row=rows, column=cols, value=values)  # 写入内容
                workbook.save(savePath + fileName + '.xls')  # 保存Excel
    print("（csv->xls）已转换格式完成！！！")


# Excel格式转换：.xls ---> .xlsx
def convFormat(openPath, savePath, finalPath, dkhName):
    fileList = os.listdir(openPath)  # 该文件夹下所有的文件（包括文件夹）
    print("转换" + str(fileList) + "文件格式")
    for file in fileList:  # 遍历所有文件
        fileName = os.path.splitext(file)[0]  # 获取文件名
        fileType = os.path.splitext(file)[1]  # 获取文件扩展名
        openFiles = openPath + fileName + fileType
        saveFiles = savePath + fileName + fileType
        finalFiles = finalPath + dkhName + str(month) + fileType
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(openFiles)
        wb.SaveAs(saveFiles + "x", FileFormat=51)  # FileFormat = 51转为.xlsx、FileFormat = 56转为.xls
        wb.SaveAs(finalFiles + "x", FileFormat=51)  # FileFormat = 51转为.xlsx、FileFormat = 56转为.xls
        wb.Close()
        excel.Application.Quit()
    print(dkhName + "（xls->xlsx）已转换格式完成！！！")


# 文件迁移（仅是适用excel表格）
def migrateFiles(orginalPath, goalPath, fileNameList):
    fileList = os.listdir(orginalPath)  # 该文件夹下所有的文件（包括文件夹）
    print("转换" + str(fileList) + "文件格式")
    for file, filename in zip(fileList, fileNameList):  # 遍历所有文件
        fileName = os.path.splitext(file)[0]  # 获取文件名
        fileType = os.path.splitext(file)[1]  # 获取文件扩展名
        if fileType == '.xls' or fileType == '.csv':
            data = xlrd.open_workbook(orginalPath + fileName + fileType)
            sheet1Data = data.sheet_by_index(0)
            workbook = xlwt.Workbook(encoding='utf-8')
            booksheet = workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
            nrows = sheet1Data.nrows
            cols = sheet1Data.ncols
            for i in range(nrows):
                for j in range(cols):
                    booksheet.write(i, j, sheet1Data.cell_value(rowx=i, colx=j))
            workbook.save(goalPath + filename + fileType)
        elif fileType == '.xlsx':
            data = pd.read_excel(orginalPath + fileName + fileType, header=0)
            data.to_excel(goalPath + filename + str(month) + fileType)
        else:
            print('不支持该格式文件迁移')
            pass
    print("文件位置迁移完成！！！")


# 合并老百姓
def mergeXlsTable(downloadPath, savePath):
    allXlsNum = glob.glob(downloadPath + "*.xls")
    print("下载了" + str(len(allXlsNum)) + "个xls文件")

    def mergeTableLBX():
        # 新建一个dataframe文件
        dfLBX = pd.DataFrame(columns=["日期", "商品编码", "商品名称", "单位", "规格", "数量", "业务部门", "厂家"])
        return dfLBX

        # 排序文件并命名

    oldFileList = os.listdir(downloadPath)  # 该文件夹下所有的文件（包括文件夹）
    createTimeList = []
    for oldfile in oldFileList:
        fileName = os.path.splitext(oldfile)[0]  # 获取文件名
        createTime = int(fileName[20:34])  # 获取时间部分数字
        createTimeList.append(createTime)  # 合并为一个列表

    createTimeListSort = sorted(createTimeList, reverse=False)  # reverse=False升序【从小到大，也即是先下载的在最前面，也即是1号到N号】

    # 按下载顺序（日期顺序）重命名文件
    for indexfile, valuefile in enumerate(createTimeListSort):  # 下角标：start = 1
        for file in oldFileList:  # 遍历所有文件
            '''
            if os.path.isdir(Olddir):   #如果是文件夹则跳过
                continue
            '''
            Olddir = os.path.join(downloadPath, file)  # 原来的文件路径
            fileMergeName = 'storeProductSummary_' + str(valuefile)
            indexFile = indexfile + 1
            fileName = os.path.splitext(file)[0]  # 获取文件名
            fileType = os.path.splitext(file)[1]  # 获取文件扩展名
            fileNewName = str(year) + '-' + str(month).zfill(2) + '-' + str(indexFile).zfill(2)
            if fileMergeName == fileName:
                Newdir = os.path.join(downloadPath, fileNewName + fileType)
                os.rename(Olddir, Newdir)  # 文件重命名
    # 老百姓合并数据
    dataLBX = mergeTableLBX()
    newFileList = os.listdir(downloadPath)  # 该文件夹下所有的文件（包括文件夹）
    for refile in newFileList:
        refileName = os.path.splitext(refile)[0]  # 获取文件名
        refileType = os.path.splitext(refile)[1]  # 获取文件扩展名
        openFile = pd.read_excel(str(downloadPath) + str(refileName) + str(refileType))
        openFile.insert(0, "日期", '')
        openFile["日期"] = refileName
        dataLBX = dataLBX.append(openFile)
    dataLBX.to_excel(
        savePath + 'dataLBX' + str(year) + '.' + str(month).zfill(2) + '.01-' + str(lastDay).zfill(2) + '.xlsx',
        index=False)
    print("老百姓数据合并输出完成！！！")


# 合并漱玉
def mergeTableSY(openPath, savePath, transitPath, finalPath, dkhName):
    # 漱玉 Excel格式转换：.xls ---> .xlsx
    def convFormatSY(open0Path, save0Path):
        deleteOldFiles(save0Path)  # 清空文件夹
        fileList = os.listdir(open0Path)  # 该文件夹下所有的文件（包括文件夹）
        print("转换" + str(fileList) + "文件格式（xls->xlsx）")
        for file in fileList:  # 遍历所有文件
            fileName = os.path.splitext(file)[0]  # 获取文件名
            fileType = os.path.splitext(file)[1]  # 获取文件扩展名
            openFiles = open0Path + fileName + fileType
            saveFiles = save0Path + fileName + fileType
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(openFiles)
            wb.SaveAs(saveFiles + "x", FileFormat=51)  # FileFormat = 51转为.xlsx、FileFormat = 56转为.xls
            wb.Close()
            excel.Application.Quit()
        print("已转换格式完成！！！")

    convFormatSY(openPath, savePath)  # 转换文件格式

    # 新建表
    def dfNewSY():
        dfSY = pd.DataFrame(columns=["销售日期", "公司编码", "公司名称", "实际区域", "门店名称", "货号", "品名", "规格", "生产单位", "数量", "零售总额"])
        return dfSY

    # 合并漱玉数据
    dataSY = dfNewSY()
    fileList = os.listdir(savePath)  # 该文件夹下所有的文件（包括文件夹）
    print("合并" + str(fileList) + "文件")
    for refile in fileList:
        fileName = os.path.splitext(refile)[0]  # 获取文件名
        fileType = os.path.splitext(refile)[1]  # 获取文件扩展名
        openFile = pd.read_excel(str(savePath) + str(fileName) + str(fileType))
        dataSY = dataSY.append(openFile)
    dataSY.to_excel(transitPath + 'data' + dkhName + str(year) + '.' + str(month) + '.01-' + str(lastDay) + '.xlsx',
                    index=False)
    dataSY.to_excel(finalPath + dkhName + str(month) + '.xlsx', index=False)
    print("漱玉数据合并输出完成！！！")


# 时间
year = int(time.strftime("%Y", time.localtime()))  # 本年
month = int(time.strftime("%m", time.localtime())) - 1  # 上个月
day = int(time.strftime("%d", time.localtime()))  # 本日
time0 = str(year) + "-" + time_0(month) + "-" + str('01')
time15 = str(year) + "-" + time_0(month) + "-" + str('15')
time16 = str(year) + "-" + time_0(month) + "-" + str('16')
lasttime = last_day_of_month(datetime.date(year, month, day))
lastDay = int(str(lasttime)[8:10])  # 上个月最后一天
dateDSL = time0 + ' - ' + str(lasttime)  # 上月大参林下载日期范围
time1 = str(lasttime)
time2 = str(lasttime)


def get_track(distance):
    """
    根据偏移量获取移动轨迹
    :param distance: 偏移量
    :return: 移动轨迹
    """
    # 移动轨迹
    track = []
    # 当前位移
    current = 0
    # 减速阈值
    mid = distance * 4 / 5
    # 计算间隔
    t = 0.2
    # 初速度
    v = 0
    while current < distance:
        if current < mid:
            # 加速度为正2
            a = 2
        else:
            # 加速度为负3
            a = -1
        # 初速度v0
        v0 = v
        # 当前速度v = v0 + at
        v = v0 + a * t
        # 移动距离x = v0t + 1/2 * a * t^2
        move = v0 * t + 1 / 2 * a * t * t
        # 当前位移
        current += move
        # 加入轨迹
        track.append(round(move))
    return track


def GJ():  # 高济
    global driver
    print('\n' + "开始导出>高济<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\GJ\\downloadGJ\\"  # 高济下载路径
    deleteOldFiles(wei_zhi)  # 高济文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(5)  # 隐式等待
    driver.get('http://srm.gaojihealth.cn/user/login')  # 进入网址
    for cishu in range(1, 6):  # 尝试5次
        try:
            try:
                time.sleep(1)
                input_b('/html/body/div/div/div[1]/div[2]/div[1]/form/div[1]/div/div/span/input', '13590889987')  # 账号
                input_b('/html/body/div/div/div[1]/div[2]/div[1]/form/div[2]/div/div/span/input', 'cb&123456')  # 密码
                time.sleep(1)
                png = driver.find_element_by_xpath(
                    '/html/body/div/div/div[1]/div[2]/div[1]/form/div[3]/div/div[1]/img[1]')  # 图片位置
                png_src = png.get_attribute('src')  # 图片链接
                response = urllib.request.urlopen(png_src)  # 登录图片链接
                cat_img = response.read()  # 读取图片
                nparr = np.frombuffer(cat_img, dtype=np.uint8)  # 转格式
                img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)  # 图片处理
                gray = img
                circles1 = cv2.HoughCircles(gray, cv2.HOUGH_GRADIENT, 1, 130, param1=10, param2=50, minRadius=0,
                                            maxRadius=60)  # 寻找圆形
                circles = circles1[0, :, :]  # 取一个圆形
                circles = np.uint16(np.around(circles))
                print("圆心坐标", circles[0][0], circles[0][1])
                time.sleep(1)
                # 获取拖拽的圆球
                slideblock = driver.find_element_by_xpath(
                    '/html/body/div/div/div[1]/div[2]/div[1]/form/div[3]/div/div[2]/div[3]')  # 拖动按钮
                track = get_track(circles[0][0] - 18)  # 设置速度
                ActionChains(driver).click_and_hold(slideblock).perform()  # 按住
                for x in track:
                    ActionChains(driver).move_by_offset(xoffset=x, yoffset=0).perform()  # 拖动
                time.sleep(0.5)
                ActionChains(driver).release().perform()  # 放手
                time.sleep(1)
                button_a('/html/body/div/div/div[1]/div[2]/div[1]/form/div[4]/div/div/span/div')  # 登录
                time.sleep(1)
                button_a('/html/body/div/div/section/aside/div/ul/li[2]/div[1]')  # 数据中心
                break
            except:
                slideblock = driver.find_element_by_xpath(
                    '/html/body/div/div/div[1]/div[2]/div[1]/form/div[3]/div/div[2]/div[3]')  # 拖动按钮
                track = get_track(circles[0][0] - 18)  # 设置速度
                ActionChains(driver).click_and_hold(slideblock).perform()  # 按住
                for x in track:
                    ActionChains(driver).move_by_offset(xoffset=150, yoffset=0).perform()  # 拖动
                time.sleep(0.5)
                ActionChains(driver).release().perform()  # 放手
                time.sleep(1)
                button_a('/html/body/div/div/div[1]/div[2]/div[1]/form/div[4]/div/div/span/div')  # 登录
                time.sleep(1)
                button_a('/html/body/div/div/section/aside/div/ul/li[2]/div[1]')  # 数据中心
                break
        except:
            print('登陆失败', cishu)
    time.sleep(2)
    button_a('/html/body/div/div/section/aside/div/ul/li[2]/ul/li')  # 数据下载
    time.sleep(3)
    button_a(
        '/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[1]/span[2]/div/input')  # 打开月份
    time.sleep(1)
    button_a('/html/body/div[2]/div/div/div/div/div/div/div[2]/div/div[2]/table/tbody/tr[' + str(
        month // 3 + 1) + ']/td[' + str(month % 3) + ']')  # 选择月份
    time.sleep(2)
    button_a(
        '/html/body/div[1]/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span/span')  # 下载门店销售明细
    time.sleep(5)


def GD():  # 国大
    global driver
    print('\n' + "开始导出>国大<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\GD\downloadGD\\"  # 国大下载路径
    deleteOldFiles(wei_zhi)  # 国大文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(5)  # 隐式等待
    for cishu in range(1, 6):  # 尝试5次
        try:
            driver.get('http://221.133.237.227:801/')
            input_b('//input[@id="loginId"]', '601317')  # 账号
            input_b('//input[@id="password"]', 'zb888888')  # 密码
            # 获取验证码
            # button_a('/html/body/form/div/div[2]/p[3]/span[2]/a')  # 看不清楚，换一张
            driver.save_screenshot('picture.png')  # 全屏截图
            page_snap_obj = Image.open('picture.png')
            img = driver.find_element_by_xpath('/html/body/form/div/div[2]/p[4]/span[2]/img')  # 验证码元素位置
            time.sleep(1)
            location = img.location
            size = img.size  # 获取验证码的大小参数
            left = location['x']
            top = location['y']
            right = left + size['width']
            bottom = top + size['height']
            image_obj = page_snap_obj.crop((left, top, right, bottom))  # 按照验证码的长宽，切割验证码
            img = image_obj.convert("L")  # 转灰度
            pixdata = img.load()
            w, h = img.size
            threshold = 160
            # 遍历所有像素，大于阈值的为黑色
            for y in range(h):
                for x in range(w):
                    if pixdata[x, y] < threshold:
                        pixdata[x, y] = 0
                    else:
                        pixdata[x, y] = 255
            data = img.getdata()
            w, h = img.size
            black_point = 0
            for x in range(1, w - 1):
                for y in range(1, h - 1):
                    mid_pixel = data[w * y + x]  # 中央像素点像素值
                    if mid_pixel < 50:  # 找出上下左右四个方向像素点像素值
                        top_pixel = data[w * (y - 1) + x]
                        left_pixel = data[w * y + (x - 1)]
                        down_pixel = data[w * (y + 1) + x]
                        right_pixel = data[w * y + (x + 1)]
                        # 判断上下左右的黑色像素点总个数
                        if top_pixel < 10:
                            black_point += 1
                        if left_pixel < 10:
                            black_point += 1
                        if down_pixel < 10:
                            black_point += 1
                        if right_pixel < 10:
                            black_point += 1
                        if black_point < 1:
                            img.putpixel((x, y), 255)
                        black_point = 0
            img.save('image.jpg')
            pytesseract.pytesseract.tesseract_cmd = r"E:\tesseract-v5.0.0\tesseract.exe"  # 设置pyteseract路径
            result = pytesseract.image_to_string(Image.open('image.jpg'))  # 识别图片里面的文字
            dropSpecialString = re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])", "",
                                       result)  # 去除识别出来的特殊字符
            num = dropSpecialString[0:4]  # 只获取前4个字符
            print("本次识别验证码为：", num)
            input_b('//input[@id="imagecode"]', num)  # 输入验证码
            time.sleep(0.5)
            button_a('//*[@onclick="SumitLogin();"]')  # 登陆
            time.sleep(2)
            button_a('/html/body/div[2]/div[2]/div/ul/li/ul/li[5]/div/span[4]')  # 门店零售
            break
        except:
            print('登陆失败', cishu)
            pass
    time.sleep(2)
    switch('/html/body/div[3]/div/div/div[2]/div[2]/div/iframe')  # 切入
    js_begin = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-text.validatebox-text").removeAttribute("readonly");'
    driver.execute_script(js_begin)
    # 用js方法输入日期
    js_begin1_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-text.validatebox-text").value="' + time0 + '"'  # 表面输入上月第一天
    driver.execute_script(js_begin1_value)
    js_begin2_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-value").value="' + time0 + '"'  # 实际有效的输入上月第一天
    driver.execute_script(js_begin2_value)
    time.sleep(2)
    js_end = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-text.validatebox-text").removeAttribute("readonly");'
    driver.execute_script(js_end)
    js_end1_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-text.validatebox-text").value="' + time1 + '"'  # 表面输入上个月最后一天
    driver.execute_script(js_end1_value)
    js_end2_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-value").value="' + time1 + '"'  # 实际有效的输入上个月最后一天
    driver.execute_script(js_end2_value)
    time.sleep(2)
    button_a('/html/body/div[1]/div/div[1]/table/tbody/tr/td[3]/a/span/span')  # 导出数据
    time.sleep(3)
    print("输出成功！！！")
    driver.switch_to.parent_frame()  # 切出
    time.sleep(10)


def QY():  # 全亿
    global driver
    print('\n' + "开始导出>全亿<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\QY\\downloadQY\\"  # 全亿下载路径
    deleteOldFiles(wei_zhi)  # 全亿文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(100)  # 隐式等待
    driver.get('http://oa.myquanyi.com/login/Login.jsp?logintype=2')  # 进入网址
    input_b('/html/body/div[1]/div[2]/form/div[2]/table/tbody/tr[1]/td/div/input', 'JT007')  # 账号
    time.sleep(0.5)
    input_b('/html/body/div[1]/div[2]/form/div[2]/table/tbody/tr[3]/td/div/input', '123456')  # 密码
    time.sleep(0.5)
    button_a('/html/body/div[1]/div[2]/form/div[2]/table/tbody/tr[8]/td/input')  # 登录
    time.sleep(3)
    button_a('/html/body/div[2]/div[1]/div/div[3]/ul/li[4]/a/span')  # 销售查询
    time.sleep(2)
    driver.switch_to.parent_frame()  # 切出
    switch('/html/body/div[2]/div[2]/div/iframe')  # 切入流程
    switch('/html/body/div[4]/div[5]/iframe')  # 切入查询
    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[3]/td[2]/div/div/div[1]/span/span/button')  # 公司名称
    time.sleep(0.5)
    driver.switch_to.parent_frame()  # 切出
    driver.switch_to.parent_frame()  # 切出
    switch('/html/body/div[4]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/iframe')  # 切入
    switch('/html/body/div[1]/div[2]/div/iframe')  # 切入
    button_a('/html/body/div[1]/form/table[2]/tbody/tr[2]/td/div/div/div/ul/li/ul/li[2]/div/button[2]')  # 全选
    time.sleep(0.5)
    button_a('/html/body/div[2]/table/tbody/tr/td/input[1]')  # 确定
    switch('/html/body/div[2]/div[2]/div/iframe')  # 切入流程
    switch('/html/body/div[4]/div[5]/iframe')  # 切入查询
    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[2]/div/button')  # 点击日期低峰日历
    time.sleep(0.5)
    switch('/html/body/div[5]/iframe')  # 切入时间
    button_a('/html/body/div/div[1]/div[2]/a')  # 点击上个月
    for i in range(1, 8):
        if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[2]/td[" + str(i) + "]")[
            0].text == '1':
            button_a("/html/body/div/div[3]/table/tbody/tr[2]/td[" + str(i) + "]")
            break
    driver.switch_to.parent_frame()  # 切出

    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[4]/div/button')  # 点击日期高峰日历
    switch('/html/body/div[5]/iframe')  # 切入时间
    button_a('/html/body/div/div[1]/div[2]/a')  # 点击上个月
    for i in range(1, 8):
        for j in range(5, 8):
            if driver.find_elements_by_xpath(
                    "/html/body/div/div[3]/table/tbody/tr[" + str(j) + "]/td[" + str(i) + "]")[0].text == str(
                lasttime.day):
                button_a("/html/body/div/div[3]/table/tbody/tr[" + str(j) + "]/td[" + str(i) + "]")
                break
    driver.switch_to.parent_frame()  # 切出
    time.sleep(2)
    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[7]/td/input')  # 数据下载
    time.sleep(2)
    # 等待alert弹出框可见
    # 从html页面切换到alert弹框
    alert = driver.switch_to.alert
    # 获取alert的文本内容
    print(alert.text)
    # 接受--选择“确定”
    alert.accept()
    # alert.dismiss()#取消
    time.sleep(0)
    # driver.quit()


def SY():  # 漱玉
    global driver
    print('\n' + "开始导出>漱玉<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\SY\\downloadSY\\"  # 漱玉下载路径
    deleteOldFiles(wei_zhi)  # 漱玉文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(30)  # 隐式等待
    driver.get('http://60.217.250.254:8765/')  # 进入网址
    button_a('/html/body/form/div[3]/div[1]/table/tbody/tr/td[3]/span/a[1]')  # 登录
    time.sleep(3)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[1]/input', '20000618')  # 账号
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[2]/input', 'cjh123')  # 密码
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[2]/div/div/p/input')  # 登录
    time.sleep(2)
    xuanfu('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/a')  # 商品查询
    time.sleep(1)
    button_a('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/ul/li[2]/a')  # 销售查询
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[6]/input', time0)  # 开始时间
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[8]/input', time15)  # 结束时间
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[9]/input')  # 查询
    time.sleep(30)
    button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[2]/table/tbody/tr/td/input')  # 导出
    time.sleep(30)
    button_a('/html/body/form/div[3]/fieldset/table/tbody/tr/td/input[1]')  # 导出
    time.sleep(5)
    driver.get('http://60.217.250.254:8765/')  # 进入网址
    button_a('/html/body/form/div[3]/div[1]/table/tbody/tr/td[3]/span/a[1]')  # 登录
    time.sleep(1)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[1]/input', '20000618')  # 账号
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[2]/input', 'cjh123')  # 密码
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[2]/div/div/p/input')  # 登录
    time.sleep(3)
    xuanfu('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/a')  # 商品查询
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/ul/li[2]/a')  # 销售查询
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[6]/input', time16)
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[8]/input', time1)
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[9]/input')  # 查询
    time.sleep(30)
    button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[2]/table/tbody/tr/td/input')  # 导出
    time.sleep(15)
    button_a('/html/body/form/div[3]/fieldset/table/tbody/tr/td/input[1]')  # 导出
    time.sleep(5)


def HW():  # 海王
    print('\n' + "开始导出>海王<数据ing......")
    global sf
    sf = ['北京', '长春', '成都', '大连', '电商', '福州', '广州', '河南', '湖北', '湖南', '深圳总部', '杭州', '江苏', '辽宁', '宁波', '青岛', '潍坊', '上海',
          '沈阳', '深圳', '天津', '泰州']  # 海王省份列表
    biao_tou = "NULL"
    global all_data
    all_data = []
    wei_zhi = "G:\\销管dkh每月数据\\HW\\downloadHW\\"  # 海王下载路径
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    global driver
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(5)  # 隐式等待
    with open('G:/销管dkh每月数据/HW/cookiesHW.txt', 'r', encoding='utf8') as f:
        listCookies = json.loads(f.read())  # 读取cookies
    driver.get('http://srm.nepstar.cn')  # 进入登录界面
    driver.delete_all_cookies()  # 删旧cookies
    for cookie in listCookies:
        if 'expiry' in cookie:
            del cookie['expiry']
        driver.add_cookie(cookie)  # 新增cookies
    time.sleep(3)
    driver.get(
        'http://srm.nepstar.cn/ELSServer_HWXC/default2.jsp?account=110829_1001&loginChage=N&telphone1=18279409642')
    # 读取完cookie刷新页面
    button_a('//*[@id="treeMenu"]/li[7]/a')  # 数据查询
    time.sleep(0.5)
    button_a('//*[@id="salesOutSourcingInfoManage"]')  # 销售查询
    time.sleep(0.5)
    xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')  # 先通过xpath定位到iframe
    driver.switch_to.frame(xf)  # 再将定位对象传给switch_to.frame()方法
    button_a('/html/body/div[1]/div[2]/form/div[1]/div/div/span')  # 商品编码
    time.sleep(5)
    driver.switch_to.parent_frame()
    xf = driver.find_element_by_xpath('/html/body/div[4]/div/table/tbody/tr[2]/td/div/iframe')
    driver.switch_to.frame(xf)  # 切换
    button_a('/html/body/div/main/div[1]/div[1]/div[1]/table/thead/tr/th[2]/div/span/input')  # 全选
    time.sleep(0.5)
    button_a('/html/body/div/main/div[2]/button[1]')  # 确定
    time.sleep(0.5)
    driver.switch_to.parent_frame()
    xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')
    driver.switch_to.frame(xf)  # 切换
    input_b('/html/body/div[1]/div[2]/form/div[2]/div/div/input', time0)  # 查询日期
    time.sleep(0.5)
    input_b('/html/body/div[1]/div[2]/form/div[3]/div/div/input', time2)
    time.sleep(0.5)
    button_a('/html/body/div[1]/div[2]/form/div[4]/div/div/div/p')  # 联采合同
    time.sleep(0.5)
    button_a('/html/body/div[1]/div[2]/form/div[4]/div/div/div/div/ul/li[3]')
    time.sleep(0.5)
    for i in range(2, 24):
        if i == 12:
            cprint("跳过深圳总部", 'cyan', attrs=['bold', 'reverse', 'blink'])
        else:
            button_a('/html/body/div[1]/div[2]/form/div[5]/div/div/div/p')  # 选地区
            time.sleep(0.5)
            ix = '/html/body/div[1]/div[2]/form/div[5]/div/div/div/div/ul/li[' + str(i) + ']'
            button_a(ix)  # 选好地区
            time.sleep(0.5)
            button_a('/html/body/div[1]/div[2]/form/button[2]')  # 查询
            time.sleep(7)
            button_a('/html/body/div[1]/div[3]/div[1]/div[1]/table/thead/tr/th[2]/div/span')  # 全选
            time.sleep(1)
            button_a('/html/body/div[1]/div[1]/nav/div/div/ul/li[1]/a')  # 查看明细
            time.sleep(1)
            driver.switch_to.parent_frame()  # 回退
            try:
                driver.find_element_by_xpath('/html/body/div[2]/div/nav[1]/ul/li[4]/a')  # 查看明细是否存在
                xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[4]/iframe')  # 切入明细
                driver.switch_to.frame(xf)
                time.sleep(8)
                button_a('/html/body/div/div[1]/nav/div/div/ul/li[2]/a')  # 导出
                driver.switch_to.parent_frame()  # 切出
                time.sleep(6)
                element = driver.find_element_by_xpath('/html/body/div[2]/div/nav[1]/ul/li[4]/a/i[2]')
                ActionChains(driver).move_to_element(element).perform()  # 鼠标悬浮
                button_a('/html/body/div[2]/div/nav[1]/ul/li[4]/a/i[2]')  # 关闭明细
                time.sleep(8)
                all_exce = get_exce(wei_zhi)
                # 得到要合并的所有exce表格数据
                if (all_exce == 0):
                    cprint(sf[i - 2] + "下载出错！！！", 'magenta', attrs=['bold', 'reverse', 'blink'])
                else:
                    for exce in all_exce:
                        fh = xlrd.open_workbook(exce)
                        # 打开文件
                        sheets = get_sheet(fh)
                        # 获取文件下的sheet数量
                        for sheet in range(len(sheets)):
                            row = get_sheetrow_num(sheets[sheet])
                            # 获取一个sheet下的所有的数据的行数
                            all_data = get_sheet_data(sheets[sheet], row, i - 2)
                            os.remove(exce)

                print("导出完成！", sf[i - 2])
                xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')  # 切入查询
                time.sleep(1)
                driver.switch_to.frame(xf)  # 切换
                button_a('/html/body/div[1]/div[3]/div[1]/div[1]/table/thead/tr/th[2]/div/span')  # 全不选

            except:
                cprint(sf[i - 2] + "->列无数据", 'magenta', attrs=['bold', 'reverse', 'blink'])
                xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')  # 切入查询
                driver.switch_to.frame(xf)  # 切换
                button_a('/html/body/div[1]/div[3]/div[1]/div[1]/table/thead/tr/th[2]/div/span')  # 全不选
            # finally:
    dictCookies = driver.get_cookies()  # 获取cookies
    jsonCookies = json.dumps(dictCookies)
    with open('G:/销管dkh每月数据/HW/cookiesHW.txt', 'w') as f:
        f.write(jsonCookies)  # 保存新cookies
    # driver.quit()
    biao_tou = ['类别', '商品SAP编码', '商品名称', '规格', '单位', '店号/区域ID', '店名/区域', '销量', '过账日期', '合同价', '省份']
    all_data.insert(0, biao_tou)  # 表头写入
    # 下面开始文件数据的写入
    new_excel = "G:\\销管dkh每月数据\\HW\\downloadHW\\" + "HW" + str(month) + ".xlsx"  # 新建的excel文件名字
    fh1 = xlsxwriter.Workbook(new_excel)  # 新建一个exce表
    new_sheet = fh1.add_worksheet()  # 新建一个sheet表
    for i in range(len(all_data)):
        for j in range(len(all_data[i])):
            c = all_data[i][j]
            new_sheet.write(i, j, c)
    fh1.close()  # 关闭该excel表


def LBX():  # 老百姓
    global driver
    print('\n' + "开始导出>老百姓<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\LBX\\downloadLBX\\"  # 老百姓下载路径
    deleteOldFiles(wei_zhi)  # 清空老百姓文件夹文件
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(30)  # 隐式等待
    driver.get('http://srm.lbxdrugs.com:6066/srm/coframe/auth/login/login.jsp')  # 进入网址
    input_b('/html/body/div/div[1]/div/form/div[1]/span/span/input', '900022800005')  # 账号
    input_b('/html/body/div/div[1]/div/form/div[2]/span/span/input', 'zz106126')  # 密码
    button_a('/html/body/div/div[1]/div/form/div[4]/input')  # 登录
    time.sleep(1)
    try:  # 点击取消每天第一次打开通知框
        button_b('//*[@id="wrapper"]/div[5]/span')
    except:
        pass
    time.sleep(3)
    button_a('/html/body/div[2]/div/div[2]/div/div/div[1]/div[1]/ul/li[1]/dl/dt')  # 销售管理
    time.sleep(0.5)
    button_a('/html/body/div[2]/div/div[2]/div/div/div[1]/div[1]/ul/li[1]/dl/dd[1]/ul/li/a')  # 门店销售明细
    time.sleep(3)
    switch('/html/body/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div/iframe')  # 切换明细
    input_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[2]/span/span/input', '')  # 激活文本输入框
    time.sleep(0.5)
    xuanfu('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[2]/span')
    button_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[2]/span')  # 鼠标点击省公司
    time.sleep(2)
    button_b('/html/body/div[7]/div/div[1]/div[1]/table/tbody/tr/td[1]/input')  # 全选
    button_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[2]/span/span/span/span[2]/span')  # 鼠标点击省公司
    time.sleep(0.5)
    input_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[4]/span/span/input', '')  # 激活文本输入框
    time.sleep(0.5)
    button_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[4]/span/span/span/span[2]/span')  # 商品名称
    driver.switch_to.parent_frame()  # 切出
    switch('/html/body/div[4]/div/div[2]/div[2]/iframe')  # 切换
    time.sleep(5)
    button_b('/html/body/div[3]/div/div[2]/div[2]/div[2]/table/tbody/tr[2]/td[3]/div/div[1]/input')  # 双击
    time.sleep(1)
    button_b('/html/body/div[3]/div/div[2]/div[2]/div[2]/table/tbody/tr[2]/td[3]/div/div[1]/input')  # 全选
    time.sleep(1)
    button_b('/html/body/a[1]/span')  # 确定
    driver.switch_to.parent_frame()  # 切出
    for dayi in range(1, lastDay + 1):
        lbsTime = str(year) + "-" + str(month).zfill(2) + "-" + str(dayi).zfill(2)
        switch('/html/body/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div/iframe')  # 切换明细
        driver.find_element_by_xpath(
            '/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input').clear()  # 清空开始时间
        time.sleep(0.5)
        input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input', lbsTime)  # 输入开始日期
        time.sleep(0.5)
        driver.find_element_by_xpath(
            '/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span[1]/input').clear()  # 清空结尾时间
        time.sleep(0.5)
        input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span[1]/input', lbsTime)  # 输入日期
        time.sleep(0.5)
        button_b('/html/body/div[2]/a[1]/span')  # 查询
        time.sleep(3)
        button_b('/html/body/div[2]/a[2]/span')  # 导出
        time.sleep(3)
        driver.switch_to.parent_frame()  # 切出
        time.sleep(3)
    deleteOldFiles('G:\\销管dkh每月数据\\LBX\\xlsxLBX\\')  # 清空老百姓文件夹文件
    mergeXlsTable('G:\\销管dkh每月数据\\LBX\\downloadLBX\\', 'G:\\销管dkh每月数据\\LBX\\xlsxLBX\\')  # 合并表格


def YF():  # 益丰
    global driver
    print('\n' + "开始导出>益丰<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\YF\\downloadYF\\"  # 益丰下载路径
    deleteOldFiles(wei_zhi)  # 清空益丰文件夹文件
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(20)  # 隐式等待
    driver.get('http://vendor.yfdyf.cn/sup/login')  # 进入网址
    input_b('/html/body/form/input[1]', '中山市中智')  # 账号
    time.sleep(0.5)
    input_b('/html/body/form/input[2]', 'zz106126')  # 密码
    # input_b('/html/body/form/input[2]','zsszz')#密码
    time.sleep(0.5)
    button_a('/html/body/form/input[3]')  # 登录
    time.sleep(3)
    button_a('/html/body/div[1]/div[1]/div/div[2]/ul/li[3]/a/span')  # 供应商查询
    time.sleep(1)
    button_a('/html/body/div[1]/div[2]/div[1]/div[1]/div[2]/div/div[2]/div/ul/li[3]/a')  # 销售记录
    time.sleep(3)
    switch('/html/body/div[1]/div[2]/div[1]/div[3]/div/div[2]/div/div[4]/iframe')  # 切入
    time.sleep(2)
    js_begin = 'document.getElementById("beginTima").removeAttribute("readonly");'
    driver.execute_script(js_begin)
    # 用js方法输入日期
    js_begin_value = 'document.getElementById("beginTima").value="' + str(time0) + '"'  # 输入上月第一天
    driver.execute_script(js_begin_value)
    time.sleep(2)
    js_end_value = 'document.getElementById("endTima").value="' + str(time1) + '"'  # 上个月最后一天
    driver.execute_script(js_end_value)
    js_end = 'document.getElementById("endTima").removeAttribute("readonly");'
    driver.execute_script(js_end)
    time.sleep(1)
    button_a('/html/body/form/ul/li[4]/input[1]')  # 查询
    time.sleep(20)
    button_a('/html/body/form/ul/li[4]/input[2]')  # 导出
    time.sleep(3)
    driver.switch_to.parent_frame()  # 切出
    time.sleep(2)
    button_a('/html/body/div[3]/div[4]/table/tbody/tr[2]/td[2]/div/div[2]/div/div[2]/button[1]')  # 确定
    time.sleep(3)


def DSL():  # 大参林
    global driver
    print('\n' + "开始导出>大参林<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\DSL\\downloadDSL\\"  # 大参林下载路径
    deleteOldFiles(wei_zhi)  # 清空大参林文件夹文件
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(20)  # 隐式等待
    driver.get('http://120.79.137.132/chain/login')  # 进入网址
    time.sleep(3)
    input_b('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[1]/div/div[1]/input',
            '13267854059')  # 手机号【数运中心邹德豪】
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[2]/div/div/div[2]/button/span')  # 获取验证码
    verificationCode = input("输入手机验证码：")  # 输入接收的验证码
    input_b('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[2]/div/div/div[1]/div/input',
            verificationCode)  # 手机验证码
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[3]/div/button')  # 登录
    time.sleep(2)
    button_a('/html/body/div[1]/div/div[1]/div[1]/div/ul/li[4]/div/span')  # 流向服务
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[1]/div[1]/div/ul/li[4]/ul/li[1]/span')  # 数据查询
    time.sleep(1)
    button_a(
        '/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div/div/div/div[1]/div/span')  # 授权编号
    time.sleep(1)
    button_a(
        '/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div/div/div/div[2]/ul[2]/li[1]')  # 门店销售
    time.sleep(1)
    input_b('//*[@id="app"]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div/div/div/div[1]/div/input',
            dateDSL)  # 查询日期
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[3]/div/div/button/span')  # 确认查询
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[3]/div/div/button/span')  # 确认查询
    time.sleep(2)
    try:
        button_a('/html/body/div[12]/div[2]/div/div/div/div/div[3]/button[2]/span')  # 弹出框 确定
    except:
        pass
    time.sleep(3)
    button_a('/html/body/div[1]/div/div[1]/div[1]/div/ul/li[4]/ul/li[2]/span')  # 流向下载
    print("等待刷新查询数据ing......")
    time.sleep(200)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div/div/div/div/button/span')  # 刷新
    time.sleep(2)
    button_a(
        '/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/div[1]/div/div[2]/table/tbody/tr[1]/td[10]/div/button/span')  # 下载结果
    # 需要等待一段时间才能有下载文件的反应


def AHLF():  # 安徽立方连锁药房有限公司
    global driver
    print('\n' + "开始导出>安徽立方连锁药房有限公司<数据ing......")
    wei_zhi = "G:\\销管dkh每月数据\\AHLF\\downloadAHLF\\"  # 安徽立方连锁药房有限公司下载路径
    deleteOldFiles(wei_zhi)  # 清空安徽立方连锁药房有限公司文件夹文件
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(20)  # 隐式等待
    driver.get('http://lfyy.lifeon.cn/flow/flow_sale_browse.asp')  # 进入网址
    time.sleep(1)
    # 点击确定
    alert = driver.switch_to.alert
    alert.accept()
    time.sleep(2)
    input_b('/html/body/div[2]/form[1]/div[2]/div/div/input', 'ZSSZZYY')  # 账号
    time.sleep(0.5)
    input_b('/html/body/div[2]/form[1]/div[3]/div/div/input', '123456')  # 密码
    time.sleep(0.5)
    button_a('/html/body/div[2]/form[1]/div[4]/button')  # 登录
    time.sleep(3)
    button_a('/html/body/div[2]/div[1]/ul/li[3]/a/span[1]')  # 流向查询系统
    time.sleep(1)
    button_a('/html/body/div[2]/div[1]/ul/li[3]/ul/li[2]/a')  # 销售流向查询
    time.sleep(3)
    js_begin = 'document.querySelector("#flowForm > div:nth-child(5) > div > div > input").removeAttribute("readonly");'
    driver.execute_script(js_begin)
    # 用js方法输入日期
    js_begin_value = 'document.querySelector("#flowForm > div:nth-child(5) > div > div > input").value="' + str(
        time0) + '"'  # 输入上月第一天
    driver.execute_script(js_begin_value)
    time.sleep(2)
    js_end = 'document.querySelector("#flowForm > div:nth-child(6) > div > div > input").removeAttribute("readonly");'
    driver.execute_script(js_end)
    js_end_value = 'document.querySelector("#flowForm > div:nth-child(6) > div > div > input").value="' + str(
        time1) + '"'  # 上个月最后一天
    driver.execute_script(js_end_value)
    time.sleep(2)
    button_a('/html/body/div[2]/div[2]/div/div[2]/div/div[2]/div[2]/form/div[10]/button[1]')  # 查询
    time.sleep(5)
    button_a(
        '/html/body/div[2]/div[2]/div/div[2]/div/div[1]/div[2]/div/div[1]/div[1]/div/label/div/a/div/b')  # 每页显示数据条数
    time.sleep(2)
    button_a('/html/body/div[4]/ul/li[7]/div')  # 所有
    time.sleep(2)
    button_a('/html/body/div[2]/div[2]/div/div[2]/div/div[1]/div[1]/div[2]/div[2]/a')  # 显示列
    time.sleep(2)
    clicksList = [2, 6, 10, 11, 12, 13]  # 点击单据编码、单位、省份、城市、县区和单位编码
    for i in clicksList:
        time.sleep(1)
        button_a('/html/body/div[2]/div[2]/div/div[2]/div/div[1]/div[1]/div[2]/div[2]/div/label[' + str(i) + ']/input')
    time.sleep(5)
    button_a('/html/body/div[2]/div[2]/div/div[2]/div/div[1]/div[1]/div[2]/div[1]/a')  # 导出


'''

try:
    GJ()  # 高济
except Exception as e:
    print("高济导出出错！！！", e)


try:
    GD()  # 国大
except Exception as e:
    print ("国大导出出错！！！",e)


try:
    QY()  # 全亿
except Exception as e:
    print ("全亿导出出错！！！",e)


try:
    SY()  # 漱玉
except Exception as e:
    print ("漱玉导出出错！！！",e)


try:
    HW()  # 海王
except Exception as e:
    print ("海王导出出错！！！",e)


try:
    LBX()  # 老百姓
except Exception as e:
    print("老百姓导出出错！！！", e)


try:
    YF()  # 益丰
except Exception as e:
    print("益丰导出出错！！！", e)


try:
    DSL()  # 大参林
except Exception as e:
    print("大参林导出出错！！！", e)

try:
    AHLF()  # 安徽立方连锁药房有限公司
except Exception as e:
    print("安徽立方导出出错！！！", e)

'''

# 在确保所有文件都已经成功下载好之后在运行下面代码
try:  # 格式转换与文件迁移
    print('\n' + ">>>高济ing......")
    convFormat('G:\\销管dkh每月数据\\GJ/downloadGJ\\', 'G:\\销管dkh每月数据\\GJ\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', 'GJ')  # 高济

    print('\n' + ">>>国大ing......")
    getFormat('G:\\销管dkh每月数据\\GD\\downloadGD\\', 'G:\\销管dkh每月数据\\GD\\xlsFormatGD\\')  # 国大(csv -> xls)
    convFormat('G:\\销管dkh每月数据\\GD\\xlsFormatGD\\', 'G:\\销管dkh每月数据\\GD\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', 'GD')  # 国大

    print('\n' + ">>>全亿ing......")
    convFormat('G:\\销管dkh每月数据\\QY\\downloadQY\\', 'G:\\销管dkh每月数据\\QY\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', 'QY')  # 全亿

    print('\n' + ">>>漱玉ing......")
    mergeTableSY('G:\\销管dkh每月数据\\SY\\downloadSY\\', 'G:\\销管dkh每月数据\\SY\\xlsxFormatSY\\',
                 'G:\\销管dkh每月数据\\SY\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', 'SY')  # 漱玉

    print('\n' + ">>>老百姓ing......")
    migrateFiles('G:\\销管dkh每月数据\\LBX\\xlsxLBX\\', 'G:\\销管dkh每月数据\\LBX\\', ['LBX'])
    migrateFiles('G:\\销管dkh每月数据\\LBX\\xlsxLBX\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', ['LBX'])  # 老百姓

    print('\n' + ">>>海王ing......")
    deleteOldFiles('G:\\销管dkh每月数据\\HW\\downloadHW\\')  # 清空海王文件夹文件
    migrateFiles('G:\\销管dkh每月数据\\HW\\downloadHW\\', 'G:\\销管dkh每月数据\\HW\\', ['HW'])  # 海王
    migrateFiles('G:\\销管dkh每月数据\\HW\\downloadHW\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', ['HW'])  # 海王

    print('\n' + ">>>益丰ing......")
    migrateFiles('G:\\销管dkh每月数据\\YF\\downloadYF\\', 'G:\\销管dkh每月数据\\YF\\', ['YF'])  # 益丰
    migrateFiles('G:\\销管dkh每月数据\\HW\\downloadHW\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', ['YF'])  # 益丰

    print('\n' + ">>>大参林ing......")
    getFormat('G:\\销管dkh每月数据\\DSL\\downloadDSL\\', 'G:\\销管dkh每月数据\\DSL\\xlsFormatDSL\\')  # 大参林(csv -> xls)
    convFormat('G:\\销管dkh每月数据\\DSL\\xlsFormatDSL\\', 'G:\\销管dkh每月数据\\DSL\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', 'DSL')  # 大参林

    print('\n' + ">>>安徽立方ing......")
    convFormat('G:\\销管dkh每月数据\\AHLF\\downloadAHLF\\', 'G:\\销管dkh每月数据\\AHLF\\', 'G:\\销管dkh每月数据\\AAA_上月数据集合\\', 'AHLF')  # 安徽立方

except Exception as e:
    print("格式转换或文件迁移出错", e)
