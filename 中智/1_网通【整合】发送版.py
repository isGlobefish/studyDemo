'''
Author: zoodehao
Date: 2021-05-21 15:54:35
LastEditTime: 2021-08-02 08:44:19
FilePath: \PyCode\1_网通【整合】发送版.py
Description: 逝者如斯夫, 不舍昼夜.
'''
# ----------------------------------------------------------------------------
# 需求说明：每日9：00程序运行, 自动生成一个多子页的Excel表格, 自动截图发送到指定钉钉群
# ----------------------------------------------------------------------------
import os
import re
import xlwt
import glob
import xlrd
import uuid
import time
import hmac
import json
import pyhdb
import base64
import shutil
import pymysql
import hashlib
import requests
import win32com
import calendar
import pythoncom
import numpy as np
import urllib.parse
import pandas as pd
import urllib.request
import openpyxl as opxl
from termcolor import cprint
from datetime import datetime
from PIL import ImageGrab, Image
from time import strftime, gmtime
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch, DispatchEx
from qiniu import Auth, put_file, etag, BucketManager
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection, NamedStyle, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule,IconSetRule, Rule, IconSet, FormatObject


def export_file():
    # ----------------------------------------------------------------------------
    # 加载格式框架
    # ----------------------------------------------------------------------------
    filepath = 'C:/Users/Zeus/Desktop/autoSend/1_网通/目标/网通_格式框架.xlsx'
    # workbook = opxl.load_workbook(filepath, data_only=True)
    workbook = opxl.load_workbook(filepath)


    # ----------------------------------------------------------------------------
    # 全局日期设置 (直接影响数据获取范围)
    # ----------------------------------------------------------------------------
    # globalYear = datetime.now().year
    # globalMonth = datetime.now().month - 1
    # globalDay = 30

    if datetime.now().day == 1 and datetime.now().month != 1:
        globalYear = datetime.now().year
        globalMonth = datetime.now().month - 1
        globalDay = calendar.monthrange(globalYear, globalMonth)[1]
    elif datetime.now().day == 1 and datetime.now().month == 1:
        globalYear = datetime.now().year - 1
        globalMonth = 12
        globalDay = calendar.monthrange(globalYear, globalMonth)[1]
    else:
        globalYear = datetime.now().year
        # globalMonth = datetime.now().month - 1
        # globalDay = calendar.monthrange(globalYear, globalMonth)[1]
        globalMonth = datetime.now().month
        globalDay = datetime.now().day - 1


    # ----------------------------------------------------------------------------
    # 全局【目标 匹配表】数据准备
    # ----------------------------------------------------------------------------
    target_data      = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/1_网通/目标/网通_数据源.xlsx', sheet_name=0, header=0)
    wangtong_target  = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/1_网通/目标/网通_数据源.xlsx', sheet_name=1, header=0)
    wangtong_shidian = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/1_网通/目标/网通_数据源.xlsx', sheet_name=2, header=0)
    sanji_pipei      = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/1_网通/目标/网通_数据源.xlsx', sheet_name=3, header=0)
    sanji_baiqiang   = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/1_网通/目标/网通_数据源.xlsx', sheet_name=4, header=0, dtype=str)

    # fahuo_data = pd.read_excel('C:/Users/zoodehao/Desktop/网通_数据源.xlsx', sheet_name=4, header=0)
    # huikuan_data = pd.read_excel('C:/Users/zoodehao/Desktop/网通_数据源.xlsx', sheet_name=5, header=0)
    # wangtong_data = pd.read_excel('C:/Users/zoodehao/Desktop/网通_数据源.xlsx', sheet_name=6, header=0)


    # ----------------------------------------------------------------------------
    # HANA数据库获取数据
    # ----------------------------------------------------------------------------
    # 获取 Connection 对象
    # 内网: IP: 192.168.20.183  账号: HANA1107318 密码: Zeus@1107311
    # 公网: IP: 119.145.248.183  账号: HANAHBPOUTTEST 密码: Zeus@test147
    def get_HANA_Connection(): 
        connectionObj = pyhdb.connect(
            host     = '192.168.20.183',
            port     = 30015,
            user     = 'HANA1107318',
            password = 'Zeus@1107311'
        )
        return connectionObj


    # 获取发货核销近两年数据
    def get_matFHHX(FHHX):
        cursor = FHHX.cursor()
        cursor.execute("""SELECT GJAHR AS Year,TO_NUMBER(MONAT) AS Month,VRGAR,ZZPLZ,WW020,WW021,WW023,SUM(VV001)/10000 AS KHJSJE
                        FROM "ECC_BI"."ZTFI302"
                        WHERE VRGAR = 'A3' AND RIGHT (WW020, 4) = '5702' AND LEFT (MATNR, 1) <> 'Z' AND ZSJLH = 11 AND GJAHR IN (:1, :2)
                        AND LEFT (ZZPLZ, 1) = 'T'
                        GROUP BY GJAHR,MONAT,VRGAR,ZSJLH,ZZPLZ,WW020,WW021,WW023
                        UNION ALL
                        SELECT GJAHR AS Year,TO_NUMBER(MONAT) AS Month,VRGAR,ZZPLZ,WW020,WW021,WW023,SUM(VV001)/10000 AS KHJSJE
                        FROM "ECC_BI"."ZTFI302"
                        WHERE VRGAR = 'A3' AND RIGHT (WW020, 4) <> '5702' AND ZSJLH = 11 AND GJAHR IN (:1, :2) AND LEFT (ZZPLZ, 1) = 'T'
                        GROUP BY GJAHR,MONAT,VRGAR,ZSJLH,ZZPLZ,WW020,WW021,WW023
                        UNION ALL
                        SELECT GJAHR AS Year,TO_NUMBER(MONAT) AS Month,VRGAR,ZZPLZ,WW020,WW021,WW023,SUM(VV001)/10000 AS KHJSJE
                        FROM "ECC_BI"."ZTFI302"
                        WHERE VRGAR = 'A2' AND ZSJLH = 11 AND GJAHR IN (:1, :2) AND LEFT (ZZPLZ, 1) = 'T'
                        GROUP BY GJAHR, MONAT, VRGAR, ZSJLH, ZZPLZ, WW020, WW021,WW023""",
                    [str(globalYear - 1), str(globalYear)])
        matFHHX = cursor.fetchall()
        return matFHHX


    # 获取海典近两年数据
    def get_matHD(HD):
        cursor = HD.cursor()
        cursor.execute("""SELECT YEAR(ACCDATE) AS NIAN,
                        MONTH(ACCDATE) AS YUE,
                        SUM(NETSUM)/10000 AS SJJE
                        FROM "ECC_BI"."V_PRT_SALE_POS_BI"
                        WHERE YEAR (ACCDATE) IN (:1, :2) AND (LEFT (BUSNO, 2) = 40 OR BUSNO = 7900 OR LEFT (BUSNO, 1) = 9) AND LEFT(ACCDATE,10) != LEFT(NOW(),10)
                        GROUP BY ACCDATE""",
                    [str(globalYear - 1), str(globalYear)])
        matHD = cursor.fetchall()
        return matHD


    # 获取网通近两年数据
    def get_matWT(WT):
        cursor = WT.cursor()
        cursor.execute("""SELECT GJAHR AS Year,TO_NUMBER(MONAT) AS Month,WW021 AS ErJi,WW022 AS SanJi,WW023 AS SiJi,MATNR,KNDNR,SUM(VV001)/10000 AS KHJSJE
                        FROM "ECC_BI"."ZTFI302"
                        WHERE VRGAR = 'A2' AND RIGHT (WW020, 4) = '1003'  AND ZSJLH = 11 AND GJAHR IN (:1, :2) AND LEFT (ZZPLZ, 1) = 'T'
                        GROUP BY GJAHR,MONAT,WW020,WW021,WW022,WW023,MATNR,KNDNR""",
                    [str(globalYear - 1), str(globalYear)])
        matWT = cursor.fetchall()
        return matWT

    conn = get_HANA_Connection()

    # 发货核销
    dataFHHX = pd.DataFrame(get_matFHHX(conn), columns=['Year', 'Month', 'Ftype', 'Category', 'YiJi', 'ErJi', 'SiJi', 'KHJSJE']) 
    dataFHHX['KHJSJE'] = pd.to_numeric(dataFHHX['KHJSJE'], errors='coerce')

    # 海典
    dataHD = pd.DataFrame(get_matHD(conn), columns=['Year', 'Month', 'SJJE'])
    dataHD['SJJE'] = pd.to_numeric(dataHD['SJJE'], errors='coerce')

    # 网通
    dataWT = pd.DataFrame(get_matWT(conn), columns=['Year', 'Month', 'ErJi', 'SanJi', 'SiJi', 'Material', 'SanDaF', 'KHJSJE'])
    dataWT['KHJSJE'] = pd.to_numeric(dataWT['KHJSJE'], errors='coerce')


    # dataWT.to_excel('C:/Users/zoodehao/Desktop/wangtong1.xlsx', index=False)
    # ----------------------------------------------------------------------------
    # 数据清洗整理部分
    # ----------------------------------------------------------------------------
    # <发货核销> 把2021年网通事业部 TG-推广线和LT-流通线 里面河南的归到试点区
    # for irow in range(len(dataFHHX)):
    #     if dataFHHX.loc[irow, 'Year'] == '2021' and (dataFHHX.loc[irow, 'Month'] == '1' or dataFHHX.loc[irow, 'Month'] == '2') and (dataFHHX.loc[irow, 'SiJi'] == '0000001202' or dataFHHX.loc[irow, 'SiJi'] == '0000001221'):
    #         dataFHHX.loc[irow, 'ErJi'] = '0000006038'
    for irow in range(len(dataFHHX)):
        if dataFHHX.loc[irow, 'YiJi'] == '' and dataFHHX.loc[irow, 'ErJi'] == '' and dataFHHX.loc[irow, 'Ftype'] == 'A3':
            dataFHHX.loc[irow, 'YiJi'] = '0000001003'
            dataFHHX.loc[irow, 'ErJi'] = '0000006038'
        elif dataFHHX.loc[irow, 'Year'] == '2021' and dataFHHX.loc[irow, 'YiJi'] == '0000001003' and (dataFHHX.loc[irow, 'SiJi'] == '0000001202' or dataFHHX.loc[irow, 'SiJi'] == '0000001221'):
            dataFHHX.loc[irow, 'ErJi'] = '0000006038'


    # <网通> 把2021年网通事业部 TG-推广线和LT-流通线 里面河南的归到试点区并划分 流通线和推广线
    for irow in range(len(dataWT)):
        if dataWT.loc[irow, 'Year'] == '2021' and (dataWT.loc[irow, 'SiJi'] == '0000001202' or dataWT.loc[irow, 'SiJi'] == '0000001221'):
            dataWT.loc[irow, 'ErJi'] = '0000006038'
            
    # <网通> 划分区域 重新调整百强历史三级
    for irow in range(len(dataWT)):
        # <网通> 试点区划分 LT-流通线 和 TG-推广线
        if int(dataWT.loc[irow, 'ErJi']) == 6038:
            for jrow in range(len(wangtong_shidian)):
                if str(dataWT.loc[irow, 'Material'])[-9:] == str(wangtong_shidian.loc[jrow, 'ID']):
                    dataWT.loc[irow, 'Material'] = wangtong_shidian.loc[jrow, 'HuaFen']
        # <网通> BQ-百强连锁线的三级重新整理为 BQ-黄河大区 BQ-珠江大区 BQ-长江大区
        elif int(dataWT.loc[irow, 'ErJi']) == 2031:
            for krow in range(len(sanji_baiqiang)):
                if dataWT.loc[irow, 'SanDaF'] == sanji_baiqiang.loc[krow, 'SongdaFID']:
                    dataWT.loc[irow, 'SanJi'] = sanji_baiqiang.loc[krow, 'SanJiID']

    # dataFHHX.to_excel('C:/Users/zoodehao/Desktop/dataFHHX_2.xlsx', index=False)

    # <网通> 三级编码为空补全---重新定义三级编码
    for irow in range(len(dataWT)):
        if dataWT.loc[irow, 'SanJi'] == '':
            for jrow in range(len(sanji_pipei)):
                if str(dataWT.loc[irow, 'SanDaF']) == str(sanji_pipei.loc[jrow, 'SongDaFang']):
                    dataWT.loc[irow, 'SanJi'] = '000000' + str(sanji_pipei.loc[jrow, 'SanJi'])

    # dataWT.to_excel('C:/Users/Zeus/Desktop/wangtong1.xlsx', index=False)
    # dataFHHX.to_excel('C:/Users/Zeus/Desktop/dataFHHX_2.xlsx', index=False)
    # ----------------------------------------------------------------------------
    # 第一页 D表（此表发01高效协同作战指挥部）
    # ----------------------------------------------------------------------------
    worksheet1 = workbook['D表']

    # familysub = ['草晶华', '冉1P（底）', '冉2C（底）', '网1（推广）', '网2（流通）', '网3（百强）', '网其他', '电A1（云P）', '电A2（云C)', '电B（林）', '电C', '第三终端二部（茹）', '大智', 'Total']
    # for irow, ifamily in zip(range(5, 19), familysub):
    #     if irow != 18: # 细分填充数值
    #         worksheet1.cell(row=irow, column=2).value = sum(huikuan_data[(huikuan_data['大家庭'] == ifamily) & (huikuan_data['月份'] == int(globalMonth))]['回款金额']) / sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] == int(globalMonth))]['目标'])
    #         worksheet1.cell(row=irow, column=3).value = sum(fahuo_data[(fahuo_data['大家庭'] == ifamily) & (fahuo_data['月份'] == int(globalMonth))]['发货金额']) / sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] == int(globalMonth))]['目标'])
    #         worksheet1.cell(row=irow, column=4).value = sum(huikuan_data[(huikuan_data['大家庭'] == ifamily) & (huikuan_data['月份'] <= int(globalMonth))]['回款金额']) / sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] <= int(globalMonth))]['目标'])
    #         worksheet1.cell(row=irow, column=5).value = sum(fahuo_data[(fahuo_data['大家庭'] == ifamily) & (fahuo_data['月份'] <= int(globalMonth))]['发货金额']) / sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] <= int(globalMonth))]['目标'])
    #         worksheet1.cell(row=irow, column=6).value = sum(huikuan_data[huikuan_data['大家庭'] == ifamily]['回款金额']) / sum(target_data[target_data['事业部'] == ifamily]['目标'])
    #         worksheet1.cell(row=irow, column=7).value = sum(fahuo_data[fahuo_data['大家庭'] == ifamily]['发货金额']) / sum(target_data[target_data['事业部'] == ifamily]['目标'])
    #     else: # 汇总填充数值
    #         worksheet1.cell(row=irow, column=2).value = sum(huikuan_data[huikuan_data['月份'] == int(globalMonth)]['回款金额']) / sum(target_data[target_data['月份'] == int(globalMonth)]['目标'])
    #         worksheet1.cell(row=irow, column=3).value = sum(fahuo_data[fahuo_data['月份'] == int(globalMonth)]['发货金额']) / sum(target_data[target_data['月份'] == int(globalMonth)]['目标'])
    #         worksheet1.cell(row=irow, column=4).value = sum(huikuan_data[huikuan_data['月份'] <= int(globalMonth)]['回款金额']) / sum(target_data[target_data['月份'] <= int(globalMonth)]['目标'])
    #         worksheet1.cell(row=irow, column=5).value = sum(fahuo_data[fahuo_data['月份'] <= int(globalMonth)]['发货金额']) / sum(target_data[target_data['月份'] <= int(globalMonth)]['目标'])
    #         worksheet1.cell(row=irow, column=6).value = sum(huikuan_data['回款金额']) / sum(target_data['目标'])
    #         worksheet1.cell(row=irow, column=7).value = sum(fahuo_data['发货金额']) / sum(target_data['目标'])

    Ftype1 = ['A3', 'A2', 'A3', 'A2', 'A3', 'A2']
    Fcol1 = [2, 3, 4, 5, 6, 7]

    for icol, itype in zip(Fcol1, Ftype1):
        if icol == 2 or icol == 3:  # n月目标完成率
            # 草晶华
            caojinghua = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
            worksheet1.cell(row=5, column=icol).value= caojinghua / sum(target_data[(target_data['事业部'] == '草晶华') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 冉1P（底）
            ran1P = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
            worksheet1.cell(row=6, column=icol).value= ran1P / sum(target_data[(target_data['事业部'] == '冉1P（底）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 冉2C（底）
            ran2C = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
            worksheet1.cell(row=7, column=icol).value= ran2C / sum(target_data[(target_data['事业部'] == '冉2C（底）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 网1（推广）
            wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
            tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
            wang1Tatget = sum(target_data[(target_data['事业部'] == '网1（推广）') & (target_data['月份'] == int(globalMonth))]['目标'])
            tryTGTatget = sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西')) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            worksheet1.cell(row=8, column=icol).value= (wang1 + tryTG) / (wang1Tatget + tryTGTatget)
            # worksheet1.cell(row=8, column=icol).value= wang1 / sum(target_data[(target_data['事业部'] == '网1（推广）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 网2（流通）
            wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
            tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
            wang2Tatget = sum(target_data[(target_data['事业部'] == '网2（流通）') & (target_data['月份'] == int(globalMonth))]['目标'])
            tryLTTatget = sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西')) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            worksheet1.cell(row=9, column=icol).value= (wang2 + tryLT) / (wang2Tatget + tryLTTatget)
            # worksheet1.cell(row=9, column=icol).value= wang2 / sum(target_data[(target_data['事业部'] == '网2（流通）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 网3（百强）
            wang3 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
            worksheet1.cell(row=10, column=icol).value= wang3 / sum(target_data[(target_data['事业部'] == '网3（百强）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 网其他
            other = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
            worksheet1.cell(row=11, column=icol).value= other / sum(target_data[(target_data['事业部'] == '网其他') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 电A1（云P）
            dianA1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
            worksheet1.cell(row=12, column=icol).value= dianA1 / sum(target_data[(target_data['事业部'] == '电A1（云P）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 电A2（云C)
            dianA2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
            worksheet1.cell(row=13, column=icol).value= dianA2 / sum(target_data[(target_data['事业部'] == '电A2（云C)') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 电B（林）
            dianB = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
            worksheet1.cell(row=14, column=icol).value= dianB / sum(target_data[(target_data['事业部'] == '电B（林）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 电C
            dianC = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
            worksheet1.cell(row=15, column=icol).value= dianC / sum(target_data[(target_data['事业部'] == '电C') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 第三终端一部
            yibu = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
            # 第三终端二部（茹）
            erbu = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
            worksheet1.cell(row=16, column=icol).value= erbu / sum(target_data[(target_data['事业部'] == '第三终端二部（茹）') & (target_data['月份'] == int(globalMonth))]['目标'])
            # 大智
            dazhi = sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])
            worksheet1.cell(row=17, column=icol).value= dazhi / sum(target_data[(target_data['事业部'] == '大智') & (target_data['月份'] == int(globalMonth))]['目标'])
            # Total = 草晶华 + 大智
            all = caojinghua + ran1P + ran2C + wang1 + wang2 + wang3 + other + dianA1 + dianA2 + dianB + dianC +  yibu + erbu + dazhi
            # worksheet1.cell(row=18, column=icol).value= (sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype)]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])) / sum(target_data[target_data['月份'] == int(globalMonth)]['目标'])
            worksheet1.cell(row=18, column=icol).value= all / sum(target_data[target_data['月份'] == int(globalMonth)]['目标'])
        elif icol == 4 or icol == 5:  # 1 - n月目标完成率
            # 草晶华
            caojinghua = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
            worksheet1.cell(row=5, column=icol).value= caojinghua / sum(target_data[(target_data['事业部'] == '草晶华') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 冉1P（底）
            ran1P = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
            worksheet1.cell(row=6, column=icol).value= ran1P  / sum(target_data[(target_data['事业部'] == '冉1P（底）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 冉2C（底）
            ran2C = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
            worksheet1.cell(row=7, column=icol).value= ran2C / sum(target_data[(target_data['事业部'] == '冉2C（底）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 网1（推广）
            wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
            tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
            wang1Tatget = sum(target_data[(target_data['事业部'] == '网1（推广）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            tryTGTatget = sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            worksheet1.cell(row=8, column=icol).value = (wang1 + tryTG) / (wang1Tatget + tryTGTatget)
            # wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
            # worksheet1.cell(row=8, column=icol).value= wang1 / sum(target_data[(target_data['事业部'] == '网1（推广）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 网2（流通）
            wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
            tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
            wang2Tatget = sum(target_data[(target_data['事业部'] == '网2（流通）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            tryLTTatget = sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            worksheet1.cell(row=9, column=icol).value= (wang2 + tryLT) / (wang2Tatget + tryLTTatget)
            # wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
            # worksheet1.cell(row=9, column=icol).value= wang2 / sum(target_data[(target_data['事业部'] == '网2（流通）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 网3（百强）
            wang3 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
            worksheet1.cell(row=10, column=icol).value= wang3 / sum(target_data[(target_data['事业部'] == '网3（百强）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 网其他
            other = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
            worksheet1.cell(row=11, column=icol).value= other / sum(target_data[(target_data['事业部'] == '网其他') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 电A1（云P）
            dianA1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
            worksheet1.cell(row=12, column=icol).value= dianA1 / sum(target_data[(target_data['事业部'] == '电A1（云P）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 电A2（云C)
            dianA2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
            worksheet1.cell(row=13, column=icol).value= dianA2 / sum(target_data[(target_data['事业部'] == '电A2（云C)') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 电B（林）
            dianB = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
            worksheet1.cell(row=14, column=icol).value= dianB / sum(target_data[(target_data['事业部'] == '电B（林）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 电C
            dianC = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
            worksheet1.cell(row=15, column=icol).value= dianC / sum(target_data[(target_data['事业部'] == '电C') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 第三终端一部
            yibu = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
            # 第三终端二部（茹）
            erbu = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
            worksheet1.cell(row=16, column=icol).value= erbu / sum(target_data[(target_data['事业部'] == '第三终端二部（茹）') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # 大智
            dazhi = sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])
            worksheet1.cell(row=17, column=icol).value= dazhi / sum(target_data[(target_data['事业部'] == '大智') & (target_data['月份'] <= int(globalMonth))]['目标'])
            # Total = 草晶华 + 大智
            all = caojinghua + ran1P + ran2C + wang1 + wang2 + wang3 + other + dianA1 + dianA2 + dianB + dianC + yibu + erbu + dazhi
            # worksheet1.cell(row=18, column=icol).value= (sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype)]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])) / sum(target_data[target_data['月份'] <= int(globalMonth)]['目标'])
            worksheet1.cell(row=18, column=icol).value= all / sum(target_data[target_data['月份'] <= int(globalMonth)]['目标'])
        elif icol == 6 or icol == 7:  # 全年目标完成率
            # 草晶华
            caojinghua = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
            worksheet1.cell(row=5, column=icol).value= caojinghua / sum(target_data[(target_data['事业部'] == '草晶华')]['目标'])
            # 冉1P（底）
            ran1P = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
            worksheet1.cell(row=6, column=icol).value= ran1P / sum(target_data[(target_data['事业部'] == '冉1P（底）')]['目标'])
            # 冉2C（底）
            ran2C = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
            worksheet1.cell(row=7, column=icol).value= ran2C / sum(target_data[(target_data['事业部'] == '冉2C（底）')]['目标'])
            # 网1（推广）
            wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
            tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
            wang1Tatget = sum(target_data[(target_data['事业部'] == '网1（推广）')]['目标'])
            tryTGTatget = sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西'))]['目标'])
            worksheet1.cell(row=8, column=icol).value = (wang1 + tryTG) / (wang1Tatget + tryTGTatget)
            # wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
            # worksheet1.cell(row=8, column=icol).value= wang1 / sum(target_data[(target_data['事业部'] == '网1（推广）')]['目标'])
            # 网2（流通）
            wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
            tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
            wang2Tatget = sum(target_data[(target_data['事业部'] == '网2（流通）')]['目标'])
            tryLTTatget = sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西'))]['目标'])
            worksheet1.cell(row=9, column=icol).value= (wang2 + tryLT) / (wang2Tatget + tryLTTatget)
            # wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
            # worksheet1.cell(row=9, column=icol).value= wang2 / sum(target_data[(target_data['事业部'] == '网2（流通）')]['目标'])
            # 网3（百强）
            wang3 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
            worksheet1.cell(row=10, column=icol).value= wang3 / sum(target_data[(target_data['事业部'] == '网3（百强）')]['目标'])
            # 网其他
            other = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
            worksheet1.cell(row=11, column=icol).value= other / sum(target_data[(target_data['事业部'] == '网其他')]['目标'])
            # 电A1（云P）
            dianA1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
            worksheet1.cell(row=12, column=icol).value= dianA1 / sum(target_data[(target_data['事业部'] == '电A1（云P）')]['目标'])
            # 电A2（云C)
            dianA2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
            worksheet1.cell(row=13, column=icol).value= dianA2 / sum(target_data[(target_data['事业部'] == '电A2（云C)')]['目标'])
            # 电B（林）
            dianB = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
            worksheet1.cell(row=14, column=icol).value= dianB / sum(target_data[(target_data['事业部'] == '电B（林）')]['目标'])
            # 电C
            dianC = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
            worksheet1.cell(row=15, column=icol).value= dianC / sum(target_data[(target_data['事业部'] == '电C')]['目标'])
            # 第三终端一部
            yibu = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
            # 第三终端二部（茹）
            erbu = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
            worksheet1.cell(row=16, column=icol).value= erbu / sum(target_data[(target_data['事业部'] == '第三终端二部（茹）')]['目标'])
            # 大智
            dazhi = sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])
            worksheet1.cell(row=17, column=icol).value= dazhi / sum(target_data[target_data['事业部'] == '大智']['目标'])
            # Total = 草晶华 + 大智
            all = caojinghua + ran1P + ran2C + wang1 + wang2 + wang3 + other + dianA1 + dianA2 + dianB + dianC + yibu + erbu + dazhi
            # worksheet1.cell(row=18, column=icol).value= (sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Ftype'] == itype)]['KHJSJE']) + sum(dataHD[dataHD['Year'] == int(globalYear)]['SJJE'])) / sum(target_data['目标'])
            worksheet1.cell(row=18, column=icol).value = all / sum(target_data['目标'])

    worksheet1.row_dimensions[11].hidden = True  # 隐藏行 - 网其他

    # 设置图标集-条件格式
    for icol in range(2, 8):
        first = FormatObject(type='num', val=0)
        second = FormatObject(type='formula', val='=$' + get_column_letter(icol) + '$18')
        third = FormatObject(type='num', val=1)
        iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third], showValue=None, percent=None, reverse=None)
        rule = Rule(type='iconSet', iconSet=iconset)
        worksheet1.conditional_formatting.add(get_column_letter(icol) + '5:' + get_column_letter(icol) + '17', rule)


    # ----------------------------------------------------------------------------
    # 第二页 A表（全年发货）
    # ----------------------------------------------------------------------------
    worksheet2 = workbook['A表']

    family = ['草晶华', '冉1P（底）', '冉2C（底）', '网1（推广）', '网2（流通）', '网3（百强）', '网其他', '电A1（云P）', '电A2（云C)', '电B（林）', '电C', '第三终端一部', '第三终端二部（茹）', '大智', 'Total']
    Ftype2 = ['A3', 'A2']
    Fcol2 = [4, 8]
    Lcol2 = [3, 7]

    for irow, ifamily in zip(range(4, 19), family):
        worksheet2.cell(row=irow, column=5).value = '=B' + str(irow) + '-D' + str(irow)
        worksheet2.cell(row=irow, column=6).value = '=IFERROR(D' + str(irow) + '/' + 'B' + str(irow) + ',"-")'
        worksheet2.cell(row=irow, column=9).value = '=B' + str(irow) + '-H' + str(irow)
        worksheet2.cell(row=irow, column=10).value = '=IFERROR(H' + str(irow) + '/' + 'B' + str(irow) + ',"-")'
        if irow <= 17:
            if irow == 7:
                worksheet2.cell(row=irow, column=2).value = sum(target_data[target_data['事业部'] == ifamily]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西'))]['目标'])
            elif irow == 8:
                worksheet2.cell(row=irow, column=2).value = sum(target_data[target_data['事业部'] == ifamily]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西'))]['目标'])
            else:
                worksheet2.cell(row=irow, column=2).value = sum(target_data[target_data['事业部'] == ifamily]['目标'])
            # worksheet2.cell(row=irow, column=4).value = sum(huikuan_data[huikuan_data['大家庭'] == ifamily]['回款金额'])
            # worksheet2.cell(row=irow, column=8).value = sum(fahuo_data[fahuo_data['大家庭'] == ifamily]['发货金额'])
        else:
            worksheet2.cell(row=irow, column=2).value = sum(target_data['目标'])
            # worksheet2.cell(row=irow, column=4).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Ftype'] == 'A3')]['KHJSJE']) + sum(dataHD[dataHD['Year'] == int(globalYear)]['SJJE'])
            # worksheet2.cell(row=irow, column=8).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Ftype'] == 'A2')]['KHJSJE']) + sum(dataHD[dataHD['Year'] == int(globalYear)]['SJJE'])
            # worksheet2.cell(row=irow, column=3).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == 'A3')]['KHJSJE']) + sum(dataHD[dataHD['Year'] == int(globalYear - 1)]['SJJE'])
            # worksheet2.cell(row=irow, column=7).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == 'A2')]['KHJSJE']) + sum(dataHD[dataHD['Year'] == int(globalYear - 1)]['SJJE'])
            worksheet2.cell(row=irow, column=3).value = '=SUM(C4:C17)'
            worksheet2.cell(row=irow, column=4).value = '=SUM(D4:D17)'
            worksheet2.cell(row=irow, column=7).value = '=SUM(G4:G17)'
            worksheet2.cell(row=irow, column=8).value = '=SUM(H4:H17)'
            # worksheet2.cell(row=irow, column=8).value = '=SUM(H4:H17)'
            # worksheet2.cell(row=irow, column=4).value = sum(huikuan_data['回款金额'])
            # worksheet2.cell(row=irow, column=8).value = sum(fahuo_data['发货金额'])

    for icol,itype in zip(Fcol2, Ftype2): # 本年全年(实际: 1-n月)回款/全年发货
        # 草晶华
        worksheet2.cell(row=4, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
        # 冉1P（底）
        worksheet2.cell(row=5, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 冉2C（底）
        worksheet2.cell(row=6, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 网1（推广）
        wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
        worksheet2.cell(row=7, column=icol).value= (wang1 + tryTG)
        # worksheet2.cell(row=7, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        # 网2（流通）
        wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
        worksheet2.cell(row=8, column=icol).value= (wang2 + tryLT)
        # worksheet2.cell(row=8, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        # 网3（百强）
        worksheet2.cell(row=9, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
        # 网其他
        worksheet2.cell(row=10, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
        # 电A1（云P）
        worksheet2.cell(row=11, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电A2（云C)
        worksheet2.cell(row=12, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电B（林）
        worksheet2.cell(row=13, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
        # 电C
        worksheet2.cell(row=14, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
        # 第三终端一部
        worksheet2.cell(row=15, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
        # 第三终端二部（茹）
        worksheet2.cell(row=16, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
        # 大智
        worksheet2.cell(row=17, column=icol).value= sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])

    for icol,itype in zip(Lcol2, Ftype2): # 上年全年回款/全年发货
        # 草晶华
        worksheet2.cell(row=4, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
        # 冉1P（底）
        worksheet2.cell(row=5, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 冉2C（底）
        worksheet2.cell(row=6, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 网1（推广）
        wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear -1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
        worksheet2.cell(row=7, column=icol).value= (wang1 + tryTG)
        # worksheet2.cell(row=7, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        # 网2（流通）
        wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
        worksheet2.cell(row=8, column=icol).value= (wang2 + tryLT)
        # worksheet2.cell(row=8, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        # 网3（百强）
        worksheet2.cell(row=9, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
        # 网其他
        worksheet2.cell(row=10, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
        # 电A1（云P）
        worksheet2.cell(row=11, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电A2（云C)
        worksheet2.cell(row=12, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电B（林）
        worksheet2.cell(row=13, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
        # 电C
        worksheet2.cell(row=14, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
        # 第三终端一部
        worksheet2.cell(row=15, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
        # 第三终端二部（茹）
        worksheet2.cell(row=16, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
        # 大智
        worksheet2.cell(row=17, column=icol).value= sum(dataHD[dataHD['Year'] == int(globalYear - 1)]['SJJE'])

    worksheet2.row_dimensions[10].hidden = True  # 隐藏行 - 网其他


    # ----------------------------------------------------------------------------
    # 第三页 C表（当月发货）
    # ----------------------------------------------------------------------------
    worksheet3 = workbook['C表']

    Ftype3 = ['A3', 'A2']
    Fcol3  = [3, 9]
    Lcol3  = [5, 6]

    for irow, ifamily in zip(range(5, 20), family):
        worksheet3.cell(row=irow, column=7).value = '=IFERROR(C' + str(irow) + '/' + 'B' + str(irow) + ',"-")'
        worksheet3.cell(row=irow, column=8).value = '=IFERROR(C' + str(irow) + '/E' + str(irow) + '-1,"-")'
        worksheet3.cell(row=irow, column=10).value = '=B' + str(irow) + '-I' + str(irow)
        worksheet3.cell(row=irow, column=11).value = '=IFERROR(I' + str(irow) + '/' + 'B' + str(irow) + ',"-")'
        worksheet3.cell(row=irow, column=12).value = '=IFERROR(I' + str(irow) + '/F' + str(irow) + '-1,"-")'
        if irow <= 18:
            if irow == 8:
                worksheet3.cell(row=irow, column=2).value = sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] == int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西')) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            elif irow ==9:
                worksheet3.cell(row=irow, column=2).value = sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] == int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西')) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            else:
                worksheet3.cell(row=irow, column=2).value = sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] == int(globalMonth))]['目标'])
            worksheet3.cell(row=irow, column=4).value = '=B' + str(irow) + '-C' + str(irow)
            # worksheet3.cell(row=irow, column=3).value = sum(huikuan_data[(huikuan_data['大家庭'] == ifamily) & (huikuan_data['月份'] == int(globalMonth))]['回款金额'])
            # worksheet3.cell(row=irow, column=9).value = sum(fahuo_data[(fahuo_data['大家庭'] == ifamily) & (fahuo_data['月份'] == int(globalMonth))]['发货金额'])
        else:
            worksheet3.cell(row=irow, column=2).value = sum(target_data[target_data['月份'] == int(globalMonth)]['目标'])
            # worksheet3.cell(row=irow, column=3).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == 'A3')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])
            # worksheet3.cell(row=irow, column=9).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == 'A2')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])
            # worksheet3.cell(row=irow, column=5).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == 'A3')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear - 1)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])
            # worksheet3.cell(row=irow, column=6).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == 'A2')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear - 1)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])
            worksheet3.cell(row=irow, column=3).value = '=SUM(C5:C18)'
            worksheet3.cell(row=irow, column=4).value = '=SUM(D5:D18)'
            worksheet3.cell(row=irow, column=5).value = '=SUM(E5:E18)'
            worksheet3.cell(row=irow, column=6).value = '=SUM(F5:F18)'
            worksheet3.cell(row=irow, column=9).value = '=SUM(I5:I18)'
            worksheet3.cell(row=irow, column=10).value = '=SUM(J5:J18)'
            # worksheet3.cell(row=irow, column=3).value = '=SUM(C5:C18)'
            # worksheet3.cell(row=irow, column=9).value = '=SUM(I5:I18)'
            # worksheet3.cell(row=irow, column=3).value = sum(huikuan_data[huikuan_data['月份'] == int(globalMonth)]['回款金额'])
            # worksheet3.cell(row=irow, column=9).value = sum(fahuo_data[huikuan_data['月份'] == int(globalMonth)]['发货金额'])

    for icol,itype in zip(Fcol3, Ftype3): # 本年-3、9列
        # 草晶华
        worksheet3.cell(row=5, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
        # 冉1P（底）
        worksheet3.cell(row=6, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 冉2C（底）
        worksheet3.cell(row=7, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 网1（推广）
        wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
        worksheet3.cell(row=8, column=icol).value = (wang1 + tryTG)
        # worksheet3.cell(row=8, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        # 网2（流通）
        wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
        worksheet3.cell(row=9, column=icol).value = (wang2 + tryLT)
        # worksheet3.cell(row=9, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        # 网3（百强）
        worksheet3.cell(row=10, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
        # 网其他
        worksheet3.cell(row=11, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
        # 电A1（云P）
        worksheet3.cell(row=12, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电A2（云C)
        worksheet3.cell(row=13, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电B（林）
        worksheet3.cell(row=14, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
        # 电C
        worksheet3.cell(row=15, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
        # 第三终端一部
        worksheet3.cell(row=16, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
        # 第三终端二部（茹）
        worksheet3.cell(row=17, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
        # 大智
        worksheet3.cell(row=18, column=icol).value = sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])

    for icol,itype in zip(Lcol3, Ftype3): # 上年-5、6列
        # 草晶华
        worksheet3.cell(row=5, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
        # 冉1P（底）
        worksheet3.cell(row=6, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 冉2C（底）
        worksheet3.cell(row=7, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 网1（推广）
        wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
        worksheet3.cell(row=8, column=icol).value = (wang1 + tryTG)
        # worksheet3.cell(row=8, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        # 网2（流通）
        wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
        worksheet3.cell(row=9, column=icol).value = (wang2 + tryLT)
        # worksheet3.cell(row=9, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        # 网3（百强）
        worksheet3.cell(row=10, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
        # 网其他
        worksheet3.cell(row=11, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
        # 电A1（云P）
        worksheet3.cell(row=12, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电A2（云C)
        worksheet3.cell(row=13, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电B（林）
        worksheet3.cell(row=14, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
        # 电C
        worksheet3.cell(row=15, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
        # 第三终端一部
        worksheet3.cell(row=16, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
        # 第三终端二部（茹）
        worksheet3.cell(row=17, column=icol).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] == globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
        # 大智
        worksheet3.cell(row=18, column=icol).value = sum(dataHD[(dataHD['Year'] == int(globalYear - 1)) & (dataHD['Month'] == int(globalMonth))]['SJJE'])

    worksheet3.row_dimensions[11].hidden = True  # 隐藏行 - 网其他


    # ----------------------------------------------------------------------------
    # 第四页 B表（累计月份发货）
    # ----------------------------------------------------------------------------
    worksheet4 = workbook['B表']

    Ftype4 = ['A3', 'A2']
    Fcol4  = [3, 9]
    Lcol4  = [7, 8]

    for irow, ifamily in zip(range(5, 20), family):
        worksheet4.cell(row=irow, column=4).value = '=B' + str(irow) + '-C' + str(irow)
        worksheet4.cell(row=irow, column=5).value = '=IFERROR(C' + str(irow) + '/' + 'B' + str(irow) + ',"-")'
        worksheet4.cell(row=irow, column=6).value = '=IFERROR(C' + str(irow) + '/G' + str(irow) + '-1,"-")'
        worksheet4.cell(row=irow, column=10).value = '=B' + str(irow) + '-I' + str(irow)
        worksheet4.cell(row=irow, column=11).value = '=IFERROR(I' + str(irow) + '/' + 'B' + str(irow) + ',"-")'
        worksheet4.cell(row=irow, column=12).value = '=IFERROR(I' + str(irow) + '/H' + str(irow) + '-1,"-")'
        if irow <= 18:
            if irow == 8:
                worksheet4.cell(row=irow, column=2).value = sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] <= int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            elif irow == 9:
                worksheet4.cell(row=irow, column=2).value = sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] <= int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            else:
                worksheet4.cell(row=irow, column=2).value = sum(target_data[(target_data['事业部'] == ifamily) & (target_data['月份'] <= int(globalMonth))]['目标'])
            # worksheet4.cell(row=irow, column=3).value = sum(huikuan_data[(huikuan_data['大家庭'] == ifamily) & (huikuan_data['月份'] <= int(globalMonth))]['回款金额'])
            # worksheet4.cell(row=irow, column=9).value = sum(fahuo_data[(fahuo_data['大家庭'] == ifamily) & (fahuo_data['月份'] <= int(globalMonth))]['发货金额'])
        else:
            worksheet4.cell(row=irow, column=2).value = sum(target_data[target_data['月份'] <= int(globalMonth)]['目标'])
            # worksheet4.cell(row=irow, column=3).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == 'A3')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])
            # worksheet4.cell(row=irow, column=9).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == 'A2')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])
            # worksheet4.cell(row=irow, column=7).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == 'A3')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear - 1)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])
            # worksheet4.cell(row=irow, column=8).value = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == 'A2')]['KHJSJE']) + sum(dataHD[(dataHD['Year'] == int(globalYear - 1)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])
            worksheet4.cell(row=irow, column=3).value = '=SUM(C5:C18)'
            worksheet4.cell(row=irow, column=7).value = '=SUM(G5:G18)'
            worksheet4.cell(row=irow, column=8).value = '=SUM(H5:H18)'
            worksheet4.cell(row=irow, column=9).value = '=SUM(I5:I18)'
            # worksheet4.cell(row=irow, column=3).value = '=SUM(C5:C18)'
            # worksheet4.cell(row=irow, column=9).value = '=SUM(I5:I18)'
            # worksheet4.cell(row=irow, column=3).value = sum(huikuan_data[huikuan_data['月份'] <= int(globalMonth)]['回款金额'])
            # worksheet4.cell(row=irow, column=9).value = sum(fahuo_data[huikuan_data['月份'] <= int(globalMonth)]['发货金额'])

    for icol,itype in zip(Fcol4, Ftype4): # 本年
        # 草晶华
        worksheet4.cell(row=5, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
        # 冉1P（底）
        worksheet4.cell(row=6, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 冉2C（底）
        worksheet4.cell(row=7, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 网1（推广）
        wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
        worksheet4.cell(row=8, column=icol).value = (wang1 + tryTG)
        # worksheet4.cell(row=8, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        # 网2（流通）
        wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
        worksheet4.cell(row=9, column=icol).value = (wang2 + tryLT)
        # worksheet4.cell(row=9, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        # 网3（百强）
        worksheet4.cell(row=10, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
        # 网其他
        worksheet4.cell(row=11, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
        # 电A1（云P）
        worksheet4.cell(row=12, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电A2（云C)
        worksheet4.cell(row=13, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电B（林）
        worksheet4.cell(row=14, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
        # 电C
        worksheet4.cell(row=15, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
        # 第三终端一部
        worksheet4.cell(row=16, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
        # 第三终端二部（茹）
        worksheet4.cell(row=17, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
        # 大智
        worksheet4.cell(row=18, column=icol).value= sum(dataHD[(dataHD['Year'] == int(globalYear)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])

    for icol,itype in zip(Lcol4, Ftype4): # 上年
        # 草晶华
        worksheet4.cell(row=5, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000005702')]['KHJSJE'])
        # 冉1P（底）
        worksheet4.cell(row=6, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T2') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 冉2C（底）
        worksheet4.cell(row=7, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['Category'] == 'T3') & (dataFHHX['YiJi'] == '0000001004')]['KHJSJE'])
        # 网1（推广）
        wang1 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        tryTG = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000001221') | (dataFHHX['SiJi'] == '0000001231'))]['KHJSJE'])
        worksheet4.cell(row=8, column=icol).value = (wang1 + tryTG)
        # worksheet4.cell(row=8, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001157')]['KHJSJE'])
        # 网2（流通）
        wang2 = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        tryLT = sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003') & (dataFHHX['ErJi'] == '0000006038') & ((dataFHHX['SiJi'] == '0000005990') | (dataFHHX['SiJi'] == '0000006046'))]['KHJSJE'])
        worksheet4.cell(row=9, column=icol).value = (wang2 + tryLT)
        # worksheet4.cell(row=9, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000001156')]['KHJSJE'])
        # 网3（百强）
        worksheet4.cell(row=10, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & (dataFHHX['ErJi'] == '0000002031')]['KHJSJE'])
        # 网其他
        worksheet4.cell(row=11, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001003')  & ((dataFHHX['ErJi'] == '0000001155') | (dataFHHX['ErJi'] == '0000006038'))]['KHJSJE'])
        # 电A1（云P）
        worksheet4.cell(row=12, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T2') | (dataFHHX['Category'] == 'T4')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电A2（云C)
        worksheet4.cell(row=13, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & ((dataFHHX['Category'] == 'T3') | (dataFHHX['Category'] == 'T9')) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000001790')]['KHJSJE'])
        # 电B（林）
        worksheet4.cell(row=14, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & (dataFHHX['ErJi'] == '0000005435')]['KHJSJE'])
        # 电C
        worksheet4.cell(row=15, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000001006') & ((dataFHHX['ErJi'] == '0000001641') | (dataFHHX['ErJi'] == '0000005819'))]['KHJSJE'])
        # 第三终端一部
        worksheet4.cell(row=16, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005409')]['KHJSJE'])
        # 第三终端二部（茹）
        worksheet4.cell(row=17, column=icol).value= sum(dataFHHX[(dataFHHX['Year'] == str(globalYear - 1)) & (dataFHHX['Month'] <= globalMonth) & (dataFHHX['Ftype'] == itype) & (dataFHHX['YiJi'] == '0000005177') & (dataFHHX['ErJi'] == '0000005410')]['KHJSJE'])
        # 大智
        worksheet4.cell(row=18, column=icol).value= sum(dataHD[(dataHD['Year'] == int(globalYear - 1)) & (dataHD['Month'] <= int(globalMonth))]['SJJE'])

    worksheet4.row_dimensions[11].hidden = True  # 隐藏行 - 网其他


    # ----------------------------------------------------------------------------
    # 第五页 网A
    # ----------------------------------------------------------------------------
    worksheet5 = workbook['网A']

    wAB = ['一区', '二区', '三区', '北区', '南区', '中区', '珠江大区', '黄河大区', '长江大区', '流通河南', '推广河南', '流通陕西', '推广陕西', '基药']
    wABrow = [4, 5, 6, 8, 9, 10, 12, 13, 14, 16, 17, 18, 19, 21]
    sumdif = ['推广', '流通', '百强连锁', '试点区', '合计']
    sumdifrow = [7, 11, 15, 20, 22]

    for irow in range(4, 23): # 常规公式列
        worksheet5.cell(row=irow, column=4).value = '=IFERROR(C' + str(irow) + '/B' + str(irow) + ',"-")'
        worksheet5.cell(row=irow, column=6).value = '=IFERROR(C' + str(irow) + '/E' + str(irow) + '-1,"-")'
        worksheet5.cell(row=irow, column=9).value = '=IFERROR(H' + str(irow) + '/G' + str(irow) + ',"-")'
        worksheet5.cell(row=irow, column=11).value = '=IFERROR(H' + str(irow) + '/J' + str(irow) + '-1,"-")'

    for irow, conent in zip(wABrow, wAB): # 分类填充数值 - 目标
        worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[(wangtong_target['区域'] == conent) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
        worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[(wangtong_target['区域'] == conent) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
        # worksheet5.cell(row=irow, column=3).value = sum(wangtong_data[(wangtong_data['大区'] == conent) & (wangtong_data['月份'] == int(globalMonth))]['发货金额(含税）'])
        # worksheet5.cell(row=irow, column=8).value = sum(wangtong_data[(wangtong_data['大区'] == conent) & (wangtong_data['月份'] <= int(globalMonth))]['发货金额(含税）'])

    for irow, conent in zip(sumdifrow, sumdif): # 小计填充数值
        if irow == 7:
            # worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] == int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西')) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            # worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] <= int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
        elif irow == 11:
            # worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] == int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            # worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] <= int(globalMonth))]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西')) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
        elif irow == 22: # 合计填充数值
            worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[wangtong_target['月份'] == int(globalMonth)]['目标'])
            worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[wangtong_target['月份'] <= int(globalMonth)]['目标'])
            # worksheet5.cell(row=irow, column=3).value = sum(wangtong_data[wangtong_data['月份'] == int(globalMonth)]['发货金额(含税）'])
            # worksheet5.cell(row=irow, column=8).value = sum(wangtong_data[wangtong_data['月份'] <= int(globalMonth)]['发货金额(含税）'])
        else:
            worksheet5.cell(row=irow, column=2).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] == int(globalMonth))]['目标'])
            worksheet5.cell(row=irow, column=7).value = sum(wangtong_target[(wangtong_target['小计'] == conent) & (wangtong_target['月份'] <= int(globalMonth))]['目标'])
            # worksheet5.cell(row=irow, column=3).value = sum(wangtong_data[(wangtong_data['类型'] == conent) & (wangtong_data['月份'] == int(globalMonth))]['发货金额(含税）'])
            # worksheet5.cell(row=irow, column=8).value = sum(wangtong_data[(wangtong_data['类型'] == conent) & (wangtong_data['月份'] <= int(globalMonth))]['发货金额(含税）'])

    for icol in [3, 8]: # 本年
        if icol == 3:
            # 推广
            worksheet5.cell(row=4, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001165')]['KHJSJE'])
            worksheet5.cell(row=5, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000005481')]['KHJSJE'])
            worksheet5.cell(row=6, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001164')]['KHJSJE'])
            # TG = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # tryTGHN = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # tryTGSX = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # worksheet5.cell(row=7, column=icol).value = TG + tryTGHN + tryTGSX
            worksheet5.cell(row=7, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # 流通
            worksheet5.cell(row=8, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005751')]['KHJSJE'])
            worksheet5.cell(row=9, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005749')]['KHJSJE'])
            worksheet5.cell(row=10, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005750')]['KHJSJE'])
            # LT = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # tryLTHN = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # tryLTSX = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # worksheet5.cell(row=11, column=icol).value = LT + tryLTHN + tryLTSX
            worksheet5.cell(row=11, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # 百强连锁
            worksheet5.cell(row=12, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006044')]['KHJSJE'])
            worksheet5.cell(row=13, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006043')]['KHJSJE'])
            worksheet5.cell(row=14, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006011')]['KHJSJE'])
            worksheet5.cell(row=15, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031')]['KHJSJE'])
            # 试点区
            worksheet5.cell(row=16, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=17, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=18, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=19, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=20, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038')]['KHJSJE'])
            # 基药
            worksheet5.cell(row=21, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001155') & (dataWT['SanJi'] == '0000001158')]['KHJSJE'])
            # 合计
            worksheet5.cell(row=22, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] == globalMonth)]['KHJSJE'])
        else:
            # 推广
            worksheet5.cell(row=4, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001165')]['KHJSJE'])
            worksheet5.cell(row=5, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000005481')]['KHJSJE'])
            worksheet5.cell(row=6, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001164')]['KHJSJE'])
            # TG = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # tryTGHN = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # tryTGSX = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # worksheet5.cell(row=7, column=icol).value = TG + tryTGHN + tryTGSX
            worksheet5.cell(row=7, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # 流通
            worksheet5.cell(row=8, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005751')]['KHJSJE'])
            worksheet5.cell(row=9, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005749')]['KHJSJE'])
            worksheet5.cell(row=10, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005750')]['KHJSJE'])
            # LT = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # tryLTHN = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # tryLTSX = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # worksheet5.cell(row=11, column=icol).value = LT + tryLTHN + tryLTSX
            worksheet5.cell(row=11, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # 百强连锁
            worksheet5.cell(row=12, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006044')]['KHJSJE'])
            worksheet5.cell(row=13, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006043')]['KHJSJE'])
            worksheet5.cell(row=14, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006011')]['KHJSJE'])
            worksheet5.cell(row=15, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031')]['KHJSJE'])
            # 试点区
            worksheet5.cell(row=16, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=17, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=18, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=19, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=20, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038')]['KHJSJE'])
            # 基药
            worksheet5.cell(row=21, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001155') & (dataWT['SanJi'] == '0000001158')]['KHJSJE'])
            # 合计
            worksheet5.cell(row=22, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth)]['KHJSJE'])

    for icol in [5, 10]: # 上年
        if icol == 5:
            # 推广
            worksheet5.cell(row=4, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001165')]['KHJSJE'])
            worksheet5.cell(row=5, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000005481')]['KHJSJE'])
            worksheet5.cell(row=6, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001164')]['KHJSJE'])
            # TG = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # tryTGHN = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # tryTGSX = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # worksheet5.cell(row=7, column=icol).value = TG + tryTGHN + tryTGSX
            worksheet5.cell(row=7, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # 流通
            worksheet5.cell(row=8, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005751')]['KHJSJE'])
            worksheet5.cell(row=9, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005749')]['KHJSJE'])
            worksheet5.cell(row=10, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005750')]['KHJSJE'])
            # LT = sum(dataWT[(dataWT['Year'] == str(globalYear -1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # tryLTHN = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # tryLTSX = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # worksheet5.cell(row=11, column=icol).value = LT + tryLTHN + tryLTSX
            worksheet5.cell(row=11, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # 百强连锁
            worksheet5.cell(row=12, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006044')]['KHJSJE'])
            worksheet5.cell(row=13, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006043')]['KHJSJE'])
            worksheet5.cell(row=14, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006011')]['KHJSJE'])
            worksheet5.cell(row=15, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000002031')]['KHJSJE'])
            # 试点区
            worksheet5.cell(row=16, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=17, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=18, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=19, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=20, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000006038')]['KHJSJE'])
            # 基药
            worksheet5.cell(row=21, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth) & (dataWT['ErJi'] == '0000001155') & (dataWT['SanJi'] == '0000001158')]['KHJSJE'])
            # 合计
            worksheet5.cell(row=22, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] == globalMonth)]['KHJSJE'])
        else:
            # 推广
            worksheet5.cell(row=4, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001165')]['KHJSJE'])
            worksheet5.cell(row=5, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000005481')]['KHJSJE'])
            worksheet5.cell(row=6, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001164')]['KHJSJE'])
            # TG = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # tryTGHN = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # tryTGSX = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            # worksheet5.cell(row=7, column=icol).value = TG + tryTGHN + tryTGSX
            worksheet5.cell(row=7, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
            # 流通
            worksheet5.cell(row=8, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005751')]['KHJSJE'])
            worksheet5.cell(row=9, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005749')]['KHJSJE'])
            worksheet5.cell(row=10, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005750')]['KHJSJE'])
            # LT = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # tryLTHN = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # tryLTSX = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            # worksheet5.cell(row=11, column=icol).value = LT + tryLTHN + tryLTSX
            worksheet5.cell(row=11, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
            # 百强连锁
            worksheet5.cell(row=12, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006044')]['KHJSJE'])
            worksheet5.cell(row=13, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006043')]['KHJSJE'])
            worksheet5.cell(row=14, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006011')]['KHJSJE'])
            worksheet5.cell(row=15, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031')]['KHJSJE'])
            # 试点区
            worksheet5.cell(row=16, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=17, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=18, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
            worksheet5.cell(row=19, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
            worksheet5.cell(row=20, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038')]['KHJSJE'])
            # 基药
            worksheet5.cell(row=21, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001155') & (dataWT['SanJi'] == '0000001158')]['KHJSJE'])
            # 合计
            worksheet5.cell(row=22, column=icol).value = sum(dataWT[(dataWT['Year'] == str(globalYear - 1)) & (dataWT['Month'] <= globalMonth)]['KHJSJE'])

    worksheet5.row_dimensions[21].hidden = True  # 隐藏行

    # 设置图标集-条件格式
    for icol in [4, 9]:
        first = FormatObject(type='num', val=0)
        second = FormatObject(type='num', val= 1)
        third = FormatObject(type='num', val= 1)
        iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third], showValue=None, percent=None, reverse=None)
        rule = Rule(type='iconSet', iconSet=iconset)
        worksheet5.conditional_formatting.add(get_column_letter(icol) + '4:' + get_column_letter(icol) + '22', rule)


    # ----------------------------------------------------------------------------
    # 第六页 网B
    # ----------------------------------------------------------------------------
    worksheet6 = workbook['网B']

    for irow in range(4, 23): # 常规公式列
        worksheet6.cell(row=irow, column=4).value = '=IFERROR(C' + str(irow) + '/B' + str(irow) + ',"-")'

    for irow, conent in zip(wABrow, wAB): # 分类填充数值
        worksheet6.cell(row=irow, column=2).value = sum(wangtong_target[wangtong_target['区域'] == conent]['目标'])
        # worksheet6.cell(row=irow, column=3).value = sum(wangtong_data[(wangtong_data['大区'] == conent) & (wangtong_data['月份'] <= int(globalMonth))]['发货金额(含税）'])

    for irow, conent in zip(sumdifrow, sumdif): # 合计填充数值
        # if irow == 7:
        #     worksheet6.cell(row=irow, column=2).value = sum(wangtong_target[wangtong_target['小计'] == conent]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '推广河南') | (wangtong_target['区域'] == '推广陕西'))]['目标'])
        # elif irow == 11:
        #     worksheet6.cell(row=irow, column=2).value = sum(wangtong_target[wangtong_target['小计'] == conent]['目标']) + sum(wangtong_target[((wangtong_target['区域'] == '流通河南') | (wangtong_target['区域'] == '流通陕西'))]['目标'])
        # worksheet6.cell(row=irow, column=3).value = sum(wangtong_data[(wangtong_data['类型'] == conent) & (wangtong_data['月份'] <= int(globalMonth))]['发货金额(含税）'])
        if irow == 22:
            worksheet6.cell(row=irow, column=2).value = sum(wangtong_target['目标'])
            # worksheet6.cell(row=irow, column=3).value = sum(wangtong_data[wangtong_data['月份'] <= int(globalMonth)]['发货金额(含税）'])
        else:
            worksheet6.cell(row=irow, column=2).value = sum(wangtong_target[wangtong_target['小计'] == conent]['目标'])
            
    # 推广
    worksheet6.cell(row=4, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001165')]['KHJSJE'])
    worksheet6.cell(row=5, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000005481')]['KHJSJE'])
    worksheet6.cell(row=6, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157') & (dataWT['SanJi'] == '0000001164')]['KHJSJE'])
    # TG = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
    # tryTGHN = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
    # tryTGSX = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
    # worksheet6.cell(row=7, column=3).value = TG + tryTGHN + tryTGSX
    worksheet6.cell(row=7, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001157')]['KHJSJE'])
    # 流通
    worksheet6.cell(row=8, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005751')]['KHJSJE'])
    worksheet6.cell(row=9, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005749')]['KHJSJE'])
    worksheet6.cell(row=10, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156') & (dataWT['SanJi'] == '0000005750')]['KHJSJE'])
    # LT = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
    # tryLTHN = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
    # tryLTSX = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
    # worksheet6.cell(row=11, column=3).value = LT + tryLTHN + tryLTSX
    worksheet6.cell(row=11, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001156')]['KHJSJE'])
    # 百强连锁
    worksheet6.cell(row=12, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006044')]['KHJSJE'])
    worksheet6.cell(row=13, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006043')]['KHJSJE'])
    worksheet6.cell(row=14, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031') & (dataWT['SanJi'] == '0000006011')]['KHJSJE'])
    worksheet6.cell(row=15, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000002031')]['KHJSJE'])
    # 试点区
    worksheet6.cell(row=16, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
    worksheet6.cell(row=17, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000005990') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
    worksheet6.cell(row=18, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'LT-流通线')]['KHJSJE'])
    worksheet6.cell(row=19, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038') & (dataWT['SiJi'] == '0000006046') & (dataWT['Material'] == 'TG-推广线')]['KHJSJE'])
    worksheet6.cell(row=20, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000006038')]['KHJSJE'])
    # 基药
    worksheet6.cell(row=21, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth) & (dataWT['ErJi'] == '0000001155') & (dataWT['SanJi'] == '0000001158')]['KHJSJE'])
    # 合计
    worksheet6.cell(row=22, column=3).value = sum(dataWT[(dataWT['Year'] == str(globalYear)) & (dataWT['Month'] <= globalMonth)]['KHJSJE'])

    worksheet6.row_dimensions[21].hidden = True  # 隐藏行 - 基药

    # 设置图标集-条件格式
    for icol in [4]:
        first = FormatObject(type='num', val=0)
        second = FormatObject(type='num', val= (globalMonth / 12))
        third = FormatObject(type='num', val= (globalMonth / 12))
        iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third], showValue=None, percent=None, reverse=None)
        rule = Rule(type='iconSet', iconSet=iconset)
        worksheet6.conditional_formatting.add(get_column_letter(icol) + '4:' + get_column_letter(icol) + '22', rule)
    
    
    # ----------------------------------------------------------------------------
    # 带时间文字填充
    # ----------------------------------------------------------------------------
    # 第 1 页
    worksheet1.cell(row=2, column=1).value = str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
    worksheet1.cell(row=20, column=1).value = '1、红绿灯警示说明：' + str(globalMonth) + '月完成率对标的是当月目标100%, 1-' + str(globalMonth) + '月完成率对标的是1-' + str(globalMonth) + '月累计目标100%, 全年完成率对标的是全年目标' + str("{:.1f}%".format(globalMonth/12*100)) + ';'
    worksheet1.cell(row=3, column=2).value = str(globalMonth) + '月目标完成率'
    worksheet1.cell(row=3, column=4).value = '1-' + str(globalMonth) + '月目标完成率'

    # 第 2 页
    worksheet2.cell(row=2, column=1).value = str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)

    # 第 3 页
    worksheet3.cell(row=2, column=1).value = str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
    worksheet3.cell(row=3, column=2).value = str(globalMonth) + '月'
    worksheet3.cell(row=3, column=5).value = '去年' +  str(globalMonth) + '月'
    worksheet3.cell(row=3, column=7).value = str(globalMonth) + '月回款'
    worksheet3.cell(row=3, column=9).value = str(globalMonth) + '月'
    worksheet3.cell(row=3, column=11).value = str(globalMonth) + '月发货'

    # 第 4 页
    worksheet4.cell(row=2, column=1).value = str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
    worksheet4.cell(row=3, column=2).value = '1-' + str(globalMonth) + '月'
    worksheet4.cell(row=3, column=3).value = '1-' + str(globalMonth) + '月回款'
    worksheet4.cell(row=3, column=7).value = '去年同期1-' + str(globalMonth) + '月'
    worksheet4.cell(row=3, column=9).value = '1-' + str(globalMonth) + '月发货'

    # 第 5 页
    worksheet5.cell(row=2, column=1).value = str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
    worksheet5.cell(row=3, column=2).value = str(globalMonth) + '月目标'
    worksheet5.cell(row=3, column=3).value = str(globalMonth) + '月发货'
    worksheet5.cell(row=3, column=4).value = str(globalMonth) + '月发货完成率'
    worksheet5.cell(row=3, column=5).value = '同期去年' + str(globalMonth) + '月发货'
    worksheet5.cell(row=3, column=6).value = str(globalMonth) + '月发货同比'
    worksheet5.cell(row=3, column=7).value = '1-' + str(globalMonth) + '月目标'
    worksheet5.cell(row=3, column=8).value = '1-' + str(globalMonth) + '月发货'
    worksheet5.cell(row=3, column=9).value = '1-' + str(globalMonth) + '月发货完成率'
    worksheet5.cell(row=3, column=10).value = '同期去年1-' + str(globalMonth) + '月发货'
    worksheet5.cell(row=3, column=11).value = '1-' + str(globalMonth) + '月发货同比'
    worksheet5.cell(row=26, column=1).value = '3、红绿灯警示说明：' + str(globalMonth) + '月完成率对标的是当月目标100%, 1-' + str(globalMonth) + '月完成率对标的是1-' + str(globalMonth) + '月累计目标100%;'

    # 第 6 页
    worksheet6.cell(row=2, column=1).value = str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
    worksheet6.cell(row=27, column=1).value = '3、红路灯警示说明：全年完成率对标的是全年目标' + str("{:.1f}%".format(globalMonth/12*100)) + ';'

    
    # ----------------------------------------------------------------------------
    # 存储表格
    # ----------------------------------------------------------------------------
    workbook.save('C:/Users/Zeus/Desktop/autoSend/1_网通/网通_正式发送文件.xlsx')


# ----------------------------------------------------------------------------
# 以下是发送部分
# ----------------------------------------------------------------------------
# 清空指定文件夹
def deleteOldFiles(path):
    deleteFileList = os.listdir(path)
    all_PNG = glob.glob(path + "*.PNG")
    print("该目录下文件有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中, PNG: " + str(len(all_PNG)) + "个")
    if len(all_PNG) != 0:
        for deletefile in deleteFileList:
            isDeleteFile = os.path.join(path, deletefile)
            if os.path.isfile(isDeleteFile):
                os.remove(isDeleteFile)
        all_DelPNG = glob.glob(path + "*.*")
        if len(all_DelPNG) == 0:
            print("已清空文件夹！！！")
        else:
            print("存在未删除文件, 请检查是否存在非PNG格式文件")
    else:
        print("不存在PNG文件")


# screenArea——格式类似"A1:J10"
def excelCatchScreen(file_name, sheet_name, name, save_path):
    pythoncom.CoInitialize()  # excel多线程相关
    Application = win32com.client.gencache.EnsureDispatch("Excel.Application")  # 启动excel
    Application.Visible = False  # 可视化
    Application.DisplayAlerts = False  # 是否显示警告
    wb = Application.Workbooks.Open(file_name, ReadOnly=False)  # 打开excel
    # ws = wb.Sheets(sheet_name)  # 选择Sheet
    ws = wb.Worksheets(sheet_name)  # 选择Sheet
    ws.Activate()  # 激活当前工作表
    userange = ws.UsedRange
    # 注意：要从A1开始的表格
    screen_area = 'A1:' + str(opxl.utils.get_column_letter(userange.Columns.Count)) + str(userange.Rows.Count)
    ws.Range(screen_area).CopyPicture()  # 复制图片区域
    time.sleep(1)
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
    Application.Selection.ShapeRange.Name = name  # 将刚刚选择的Shape重命名, 避免与已有图片混淆
    ws.Shapes(name).Copy()  # 选择图片
    time.sleep(1)
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    img_name = name + ".PNG"
    img.save(save_path + img_name)  # 保存图片
    # time.sleep(1)
    # wb.Save()
    # time.sleep(1)
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    time.sleep(1)
    Application.Quit()  # 退出excel
    pythoncom.CoUninitialize()


# 定义钉钉功能
class dingdingFunction(object):
    def __init__(self, roboturl, robotsecret, appkey, appsecret):
        """
        :param roboturl: 群机器人WebHook_url
        :param robotsecret: 安全设置的加签秘钥
        :param appkey: 企业开发平台小程序AppKey
        :param appsecret: 企业开发平台小程序AppSecret
        """
        self.roboturl = roboturl
        self.robotsecret = robotsecret
        self.appkey = appkey
        self.appsecret = appsecret
        timestamp = round(time.time() * 1000)  # 时间戳
        secret_enc = robotsecret.encode('utf-8')
        string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
        string_to_sign_enc = string_to_sign.encode('utf-8')
        hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
        self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(timestamp, sign)  # 最终url,url+时间戳+签名

    # 发送文件
    def getAccess_token(self):
        url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (AppKey, AppSecret)
        headers = {
            'Content-Type': "application/x-www-form-urlencoded"
        }
        data = {'appkey': self.appkey,
                'appsecret': self.appsecret}
        r = requests.request('GET', url, data=data, headers=headers)
        access_token = r.json()["access_token"]
        return access_token

    def getMedia_id(self, filespath):
        access_token = self.getAccess_token()  # 拿到接口凭证
        url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
        files = {'media': open(filespath, 'rb')}
        data = {'access_token': access_token,
                'type': 'file'}
        response = requests.post(url, files=files, data=data)
        json = response.json()
        return json["media_id"]

    def sendFile(self, chatid, filespath):
        access_token = self.getAccess_token()
        media_id = self.getMedia_id(filespath)
        url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
        header = {
            'Content-Type': 'application/json'
        }
        data = {'access_token': access_token,
                'chatid': chatid,
                'msg': {
                    'msgtype': 'file',
                    'file': {'media_id': media_id}
                }}
        r = requests.request('POST', url, data=json.dumps(data), headers=header)
        print(r.json()["errmsg"])

    # 发送消息
    def sendMessage(self, content):
        """
        :param content: 发送内容
        """
        header = {
            "Content-Type": "application/json",
            "Charset": "UTF-8"
        }
        sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
        sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
        request = urllib.request.Request(url=self.webhook_url, data=sendContent, headers=header)  # 发送请求
        opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
        print(opener.read().decode())  # 打印返回的结果


# 上传本地图片获取网上图片URL
# def get_image_url(imagePath, pictureName):
#     if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
#         filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
#     elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
#         filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
#     else:
#         print("请检查图片格式！！！")
#     # 七牛云密钥管理：https://portal.qiniu.com/user/key
#     # 【账号：1326754059  密码：z****】
#     access_key = "DZnCErimkn2yQrn4aYel3JX7vPXKRonlvDFoVh1e"
#     secret_key = "FBEHIFyMG28nWZrn316df-ny5bmIz_LanRWtabCi"
#     keyq = Auth(access_key, secret_key)
#     bucket = "qiniu730173201"  # 七牛云盘名
#     # 删除
#     butm = BucketManager(keyq)
#     reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
#     # 上传
#     time.sleep(5)
#     token = keyq.upload_token(bucket, filename)  # 上传新图片
#     reformUp, informUp = put_file(token, filename, imagePath)
#     if reformUp is not None:
#         print('已经成功上传 {}'.format(filename))
#     else:
#         print(filename + '上传失败！！！')
#     time.sleep(5)
#     baseURL = "http://zzsy.zeus.cn/"  # 中智二级域名
#     subURL = baseURL + '/' + filename
#     pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
#     time.sleep(6)
#     return pictureURL

def get_image_url(imagePath, pictureName):
    if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    else:
        print("请检查图片格式！！！")
    # 七牛云密钥管理：https://portal.qiniu.com/user/key
    # 【账号：13267854059  密码：z****】
    access_key = "fjlWDEbF1fqBU98UsdDJRcSSKODT9Gq7tA3gu8eY"
    secret_key = "thiWFpO881GfhlaAz1Wkk2yEcvV3ue2OHnY_5D9V"
    keyq = Auth(access_key, secret_key)
    bucket = "zues3737img"  # 七牛云盘名
    # 删除
    butm = BucketManager(keyq)
    reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
    # 上传
    time.sleep(1)
    token = keyq.upload_token(bucket, filename)  # 上传新图片
    reformUp, informUp = put_file(token, filename, imagePath)
    if reformUp is not None:
        print('已成功上传 {}'.format(filename))
        time.sleep(1)
        baseURL = "https://cjh3737.zeus.cn/"  # 加速域名
        # subURL = baseURL + '/' + filename + '?imageView2/0/quality/100!/sharpen/1/interlace/1'
        subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/shapen/50/interlace/1/ignore-error/1'
        pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
        time.sleep(1)
        return pictureURL
    else:
        print(filename + '上传失败！！！')


if __name__ == '__main__':

    AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序 - AppKey
    AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序 - AppSecret
    
    webhook1 = 'https://oapi.dingtalk.com/robot/send?access_token=a4bb3555c97ce4b3e61c7ce02e45e519d0d8f81e720913e1bf565cac2abc0b46'  # 01高效协同作战指挥部
    webhook2 = 'https://oapi.dingtalk.com/robot/send?access_token=d6a952d514e5c3a6c46e4be8514874955ad784bef6c169d8392a6261d0b36257'  # 销售数据日报群
    webhook3 = 'https://oapi.dingtalk.com/robot/send?access_token=28b33a8cce676b6e54f0b977de44e88ea43f11a527237ff779e9721139b62f43'  # 网通数据汇报群
    
    # webhook1 = 'https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c'  # 机器人测试群
    # webhook2 = 'https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c'  # 机器人测试群
    # webhook3 = 'https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c'  # 机器人测试群
    
    secret0 = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'  # 群机器人加签秘钥secret(默认草晶华小助手)
    secret1 = 'SEC2514a86af502cb08a90e2c135ff746921b15e8b47bbe6c65e6b3b1ebcc36b903'  # 01高效协同作战指挥部(普通群)
    
    RobotWebHookURL = [webhook1, webhook2, webhook2, webhook2, webhook3, webhook3]
    RobotSecret = [secret1, secret0, secret0, secret0, secret0, secret0]
    # RobotSecret = [secret0, secret0, secret0, secret0, secret0, secret0]  # 机器人测试群
    
    fileFullPath = 'C:/Users/Zeus/Desktop/autoSend/1_网通/网通_正式发送文件.xlsx'
    savePictuePath = 'C:/Users/Zeus/Desktop/autoSend/1_网通/Pictures/'

    worksheetnames = ['D表', 'A表', 'C表', 'B表', '网A', '网B']

    sendtitle = ["###### **① D表-发货回款目标完成情况**",
                 "###### **② A表-全年目标完成情况**",
                 "###### **③ C表-本月目标完成情况**",
                 "###### **④ B表-同期累计目标完成情况**",
                 "###### **⑤ 网A-同期目标完成情况**", 
                 "###### **⑥ 网B-全年目标完成情况**"]
    
    sendTypes = int(input('>>>0sendAll-1sendSingle:'))
    
    if sendTypes == 0: # 发送形式 - 全部

        export_file()
        print(' > 文件保存成功！！！')

        deleteOldFiles('C:/Users/Zeus/Desktop/autoSend/1_网通/Pictures/')  # 清空文件夹历史文件

        pictureURL = []
        for sheetname, picturename in zip(worksheetnames, worksheetnames):
            try:
                excelCatchScreen(fileFullPath, sheetname, picturename, savePictuePath)
            except BaseException:
                print(picturename + '截图出错！！！')
            try:
                getURL = get_image_url(savePictuePath + picturename + '.PNG', picturename)
                pictureURL.append(getURL)
            except BaseException:
                print(picturename + '图片URL出错！！！')
        if len(pictureURL) == len(sendtitle):
            for inum, iurl, itext in zip(range(6), pictureURL, sendtitle):
                ddMessage = {  # 发布消息内容
                    "msgtype": "markdown",
                    "markdown": {"title": "销售日报",  # @某人 才会显示标题
                                    "text": itext + 
                                    "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
                                    "\n###### ----------------------------------------"
                                    "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                    "at": {
                        # "atMobiles": [15817552982],  # 指定@某人
                        "isAtAll": False  # 是否@所有人[False:否, True:是]
                    }
                }

                # 发送消息
                dingdingFunction(RobotWebHookURL[inum], RobotSecret[inum], AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
                
                if inum == 3:
                    atWhoMessage = {  # 发布消息内容
                            "msgtype": "markdown",
                            "markdown": {"title": "销售日报",  # @某人 才会显示标题
                                            "text": "##### @13590885469 松哥, 今日( " + str(str(datetime.now()).split(' ')[0]) + ")数据已经送达, 请您查收一下(￣▽￣)。"
                                        },
                            "at": {
                                "atMobiles":[13590885469],  # 指定@某人
                                # "atUserIds": ["sut6m07"],
                                "isAtAll": False  # 是否@所有人[False:否, True:是]
                            }
                        }
                
                    # 发送消息
                    dingdingFunction(RobotWebHookURL[inum], RobotSecret[inum], AppKey, AppSecret).sendMessage(atWhoMessage) 
        else:
            print('存在URL失败！！！')
            
    elif sendTypes == 1: # 发送形式 - 选择性单张 从 1 开始

        export_file()
        print(' > 文件保存成功！！！')

        deleteOldFiles('C:/Users/Zeus/Desktop/autoSend/1_网通/Pictures/')  # 清空文件夹历史文件
        
        send_NO_Picture = int(input('>>>发送第几张图片？'))
        
        print('***单独发送: ' + str(worksheetnames[send_NO_Picture - 1]) + '.PNG')

        excelCatchScreen(fileFullPath, worksheetnames[send_NO_Picture - 1], worksheetnames[send_NO_Picture - 1], savePictuePath)
        
        ddMessage = {  # 发布消息内容
                "msgtype": "markdown",
                "markdown": {"title": "销售日报",  # @某人 才会显示标题
                                "text": sendtitle[send_NO_Picture - 1] +
                                "\n![Image被拦截, 请使用非公司网络查看](" + get_image_url(savePictuePath + worksheetnames[send_NO_Picture - 1] + '.PNG', worksheetnames[send_NO_Picture - 1]) + ")"
                                "\n###### ----------------------------------------"
                                "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                "at": {
                    # "atMobiles": [15817552982],  # 指定@某人
                    "isAtAll": False  # 是否@所有人[False:否, True:是]
                }
            }

        # 发送消息
        dingdingFunction(RobotWebHookURL[send_NO_Picture - 1], RobotSecret[send_NO_Picture - 1], AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
        
    else:
        print('请输入0-1正确的发送方式！！！')
    
    # dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendFile(chatId, fileFullPath)  # 发送文件
  