# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.2.2
@projectName   : pythonProjectPy3.9
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2020/12/11 8:28
'''
import os
import re
import csv
import cv2
import json
import time
import xlrd
import xlwt
import glob
import openpyxl
import datetime
import xlsxwriter
import numpy as np
import pandas as pd
import urllib.request
from termcolor import cprint, colored
from selenium import webdriver
# pypiwin32仅支持window系统
# import win32com.client as win32
import pytesseract  # 用于图片转文字
from PIL import Image  # 用于打开图片和对图片处理
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC

# aaa = input("输入1:这个月，输入2：上个月：")
# aaa = input("输入1:这个月，输入2：上个月：")

# aaa = "1"  # 下载本月数据

aaa = "2"  # 下载上个月数据

# today = int(time.strftime("%d", time.localtime()))
# if today <= 2:
#     aaa = "2"  # 下载上个月数据
# else:
#     aaa = "1"  # 下载本月数据

def get_exce(wei_zhi):
    all_exce = glob.glob(wei_zhi + "*.xls")
    print("该目录下有" + str(len(all_exce)) + "个excel文件：")
    if (len(all_exce) == 0):
        return 0
    else:
        for i in range(len(all_exce)):
            print(all_exce[i])
        return all_exce


def input_a(id_1, value):
    input_box = driver.find_element_by_id(id_1)
    try:
        input_box.send_keys(value)
    except Exception as e:
        print('fail输入', e)


def input_b(id_1, value):
    input_box = driver.find_element_by_xpath(id_1)
    try:
        input_box.send_keys(value)
    except Exception as e:
        print('fail输入', e)


def button_a(xpath):
    global wait
    wait = WebDriverWait(driver, 3)
    button = driver.find_element_by_xpath(xpath)
    try:
        button.click()
    except Exception as e:
        print('fail点击', e)


def button_b(xpath):
    global wait
    wait = WebDriverWait(driver, 3)
    button = driver.find_element_by_xpath(xpath)
    try:
        driver.execute_script("$(arguments[0]).click()", button)
    except Exception as e:
        print('fail点击', e)


def clear_a(id_1):
    input_box = driver.find_element_by_xpath(id_1)
    try:
        input_box.clear()
    except Exception as e:
        print('fail清空输入框', e)


def switch(xpath):
    xf = driver.find_element_by_xpath(xpath)
    try:
        driver.switch_to.frame(xf)  # 切换
    except Exception as e:
        print('切换失败', e)


def xuanfu(xpath):
    element = driver.find_element_by_xpath(xpath)
    try:
        ActionChains(driver).move_to_element(element).perform()  # 鼠标悬浮
    except Exception as e:
        print('悬浮失败', e)


# 获取excel文件下的所有sheet
def get_sheet(fh):
    sheets = fh.sheets()
    return sheets


# 获取sheet下有多少行数据
def get_sheetrow_num(sheet):
    return sheet.nrows


def get_sheet_data(sheet, row, j):
    for i in range(row - 1):
        if (i == 0):
            global biao_tou
            biao_tou = ['类别', '商品SAP编码', '商品名称', '规格', '单位', '店号/区域ID', '店名/区域', '销量', '过账日期', '合同价', '省份']
            continue
        values = sheet.row_values(i)
        values.append(sf[j])
        all_data1.append(values)
    return all_data1


def time_0(time):
    if time > 9:
        time1 = str(time)
    else:
        time1 = '0' + str(time)
    return time1


def last_day_of_month(any_day):
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)
    return next_month - datetime.timedelta(days=next_month.day)


# 清空指定文件夹
def deleteOldFiles(path):
    # 删除指定文件夹下的文件
    # path = "D:/FilesCenter/大客户数据/LBX/LBX/mergeTableLBX/"
    deleteFileList = os.listdir(path)
    all_Xls = glob.glob(path + "*.xls")
    all_Xlsx = glob.glob(path + "*.xlsx")
    all_Csv = glob.glob(path + "*.csv")
    print("该目录下有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中【xls:" + str(len(all_Xls)) + ", xlsx:" + str(len(all_Xlsx)) + ", csv:" + str(
        len(all_Csv)) + "】")
    if (len(all_Xls) != 0 or len(all_Xlsx) != 0 or len(all_Csv) != 0):
        for deletefile in deleteFileList:
            isDeleteFile = os.path.join(path, deletefile)
            if os.path.isfile(isDeleteFile):
                os.remove(isDeleteFile)
        all_DelXls = glob.glob(path + "*.xls")
        all_DelXlsx = glob.glob(path + "*.xlsx")
        all_DelCsv = glob.glob(path + "*.csv")
        if (len(all_DelXls) == 0 and len(all_DelXlsx) == 0
                and len(all_DelCsv) == 0):
            print("已清空文件夹！！！")
        else:
            print("存在未删除文件")
    else:
        print("不存在excel文件")


# Excel格式转换：.csv ---> .xls
def getFormat(openPath, savePath):
    deleteOldFiles(savePath)  # 清空文件夹
    fileList = os.listdir(openPath)  # 该文件夹下所有的文件（包括文件夹）
    print("转换" + str(fileList) + "文件格式")
    for file in fileList:  # 遍历所有文件
        fileName = os.path.splitext(file)[0]  # 获取文件名
        fileType = os.path.splitext(file)[1]  # 获取文件扩展名
        try:
            data = xlrd.open_workbook(openPath + fileName + fileType)  # 读取文件
            sheet1Data = data.sheet_by_index(0)  # 选择第一个子页
            workbook = xlwt.Workbook(encoding='utf-8')
            booksheet = workbook.add_sheet('Sheet1', cell_overwrite_ok=True)
            nrows = sheet1Data.nrows
            cols = sheet1Data.ncols
            for i in range(nrows):
                for j in range(cols):
                    booksheet.write(i, j, sheet1Data.cell_value(rowx=i, colx=j))
            workbook.save(savePath + fileName + '.xls')
        except BaseException:
            with open(openPath + fileName + fileType, 'r') as csvfile:
                openCsvFile = csv.reader(csvfile)  # 读取文件
                workbook = openpyxl.Workbook()  # 打开一个文件
                sheet = workbook.create_sheet(index=0)  # 在文件上创建Sheet1
                for rows, lines in enumerate(openCsvFile, start=1):
                    for cols, values in enumerate(lines, start=1):
                        sheet.cell(row=rows, column=cols, value=values)  # 写入内容
                workbook.save(savePath + fileName + '.xls')  # 保存Excel
    print("（csv->xls）已转换格式完成！！！")


# Excel格式转换：.xls ---> .xlsx
def convFormat(openPath, savePath, finalPath, dkhName):
    fileList = os.listdir(openPath)  # 该文件夹下所有的文件（包括文件夹）
    print("转换" + str(fileList) + "文件格式")
    for file in fileList:  # 遍历所有文件
        fileName = os.path.splitext(file)[0]  # 获取文件名
        fileType = os.path.splitext(file)[1]  # 获取文件扩展名
        openFiles = openPath + fileName + fileType
        saveFiles = savePath + fileName + fileType
        finalFiles = finalPath + dkhName + str(month) + fileType
        # excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel = win32.DispatchEx('Excel.Application')
        wb = excel.Workbooks.Open(openFiles)
        wb.SaveAs(saveFiles + "x", FileFormat=51)  # FileFormat = 51转为.xlsx、FileFormat = 56转为.xls
        wb.SaveAs(finalFiles + "x", FileFormat=51)  # FileFormat = 51转为.xlsx、FileFormat = 56转为.xls
        wb.Close()
        excel.Application.Quit()
    print(dkhName + "（xls->xlsx）已转换格式完成！！！")


# 合并老百姓
def mergeXlsTable():
    global aaa
    path = "D:/FilesCenter/大客户数据/LBX/mergeTableLBX/"
    allXlsNum = glob.glob(path + "*.xls")
    print("下载了" + str(len(allXlsNum)) + "个xls文件")

    def mergeTableLBX():
        # 新建一个dataframe文件
        dfLBX = pd.DataFrame(
            columns=["日期", "商品编码", "商品名称", "单位", "规格", "数量", "业务部门", "厂家"])
        return dfLBX

        # 排序文件并命名

    oldFileList = os.listdir(path)  # 该文件夹下所有的文件（包括文件夹）
    createTimeList = []
    for oldfile in oldFileList:
        fileName = os.path.splitext(oldfile)[0]  # 获取文件名
        createTime = int(fileName[20:34])  # 获取时间部分数字
        createTimeList.append(createTime)  # 合并为一个列表

    createTimeListSort = sorted(
        createTimeList,
        reverse=False)  # reverse=False升序【从小到大，也即是先下载的在最前面，也即是1号到N号】

    # 按下载顺序（日期顺序）重命名文件
    for indexfile, valuefile in enumerate(createTimeListSort):  # 下角标：start = 1
        for file in oldFileList:  # 遍历所有文件
            '''
            if os.path.isdir(Olddir):   #如果是文件夹则跳过
                continue
            '''
            Olddir = os.path.join(path, file)  # 原来的文件路径
            fileMergeName = 'storeProductSummary_' + str(valuefile)
            indexFile = indexfile + 1
            fileName = os.path.splitext(file)[0]  # 获取文件名
            fileType = os.path.splitext(file)[1]  # 获取文件扩展名
            fileNewName = str(year) + '-' + str(month).zfill(2) + '-' + str(
                indexFile).zfill(2)
            if fileMergeName == fileName:
                Newdir = os.path.join(path, fileNewName + fileType)
                os.rename(Olddir, Newdir)  # 文件重命名
    # 老百姓合并数据
    dataLBX = mergeTableLBX()
    newFileList = os.listdir(path)  # 该文件夹下所有的文件（包括文件夹）
    for refile in newFileList:
        refileName = os.path.splitext(refile)[0]  # 获取文件名
        refileType = os.path.splitext(refile)[1]  # 获取文件扩展名
        openFile = pd.read_excel(str(path) + str(refileName) + str(refileType))
        openFile.insert(0, "日期", '')
        openFile["日期"] = refileName
        dataLBX = dataLBX.append(openFile)
    if aaa == '1':
        dataLBX.to_excel('D:/FilesCenter/大客户数据/LBX/dataLBX' + str(year) + '-' + str(month).zfill(2) + '-' + str(day).zfill(2) +
                         '.xlsx', index=False)
    else:
        dataLBX.to_excel('D:/FilesCenter/大客户数据/LBX/dataLBX' + str(year) + '-' + str(month).zfill(2) + '-' + str(lastDay).zfill(2) +
                         '.xlsx', index=False)
    print("老百姓数据合并输出完成！！！")


# 合并漱玉
def mergeTableSY(openPath, savePath, transitPath, finalPath):
    # Excel格式转换：.xls ---> .xlsx
    def convFormatSY(open0Path, save0Path):
        deleteOldFiles(save0Path)  # 清空文件夹
        fileList = os.listdir(open0Path)  # 该文件夹下所有的文件（包括文件夹）
        print("转换" + str(fileList) + "文件格式（xls->xlsx）")
        for file in fileList:  # 遍历所有文件
            fileName = os.path.splitext(file)[0]  # 获取文件名
            fileType = os.path.splitext(file)[1]  # 获取文件扩展名
            openFiles = open0Path + fileName + fileType
            saveFiles = save0Path + fileName + fileType
            # excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel = win32.DispatchEx('Excel.Application')
            wb = excel.Workbooks.Open(openFiles)
            wb.SaveAs(saveFiles + "x", FileFormat=51)  # FileFormat = 51转为.xlsx、FileFormat = 56转为.xls
            wb.Close()
            excel.Application.Quit()
        print("已转换格式完成！！！")

    convFormatSY(openPath, savePath)  # 转换文件格式

    # 新建表
    def dfNewSY():
        dfSY = pd.DataFrame(columns=["销售日期", "公司编码", "公司名称", "实际区域", "门店名称", "货号", "品名", "规格", "生产单位", "数量", "零售总额"])
        return dfSY

    # 合并漱玉数据
    dataSY = dfNewSY()
    fileList = os.listdir(savePath)  # 该文件夹下所有的文件（包括文件夹）
    print("合并" + str(fileList) + "文件")
    for refile in fileList:
        fileName = os.path.splitext(refile)[0]  # 获取文件名
        fileType = os.path.splitext(refile)[1]  # 获取文件扩展名
        openFile = pd.read_excel(str(savePath) + str(fileName) + str(fileType))
        dataSY = dataSY.append(openFile)
    dataSY.to_excel(transitPath + 'dataSY' + str(year) + '-' + str(month) + '-' + str(lastDay) + '.xlsx', index=False)
    dataSY.to_excel(finalPath + 'SY' + str(month) + '.xlsx', index=False)
    print("漱玉数据合并输出完成！！！")


if aaa == "1":  # 判断时间
    year = int(time.strftime("%Y", time.localtime()))  # 年
    month = int(time.strftime("%m", time.localtime()))  # 月
    day = int(time.strftime("%d", time.localtime()))  # 日
    daySub1 = day - 1
    time0 = str(year) + "-" + time_0(month) + "-" + str('01')
    time15 = str(year) + "-" + time_0(month) + "-" + str('15')
    time16 = str(year) + "-" + time_0(month) + "-" + str('16')
    time1 = str(year) + "-" + time_0(month) + "-" + time_0(day - 1)
    time2 = str(year) + "-" + time_0(month) + "-" + time_0(day - 2)
    lasttime = last_day_of_month(datetime.date(year, month, day))
    lastDay = int(day)
    dateDSL = time0 + ' - ' + time2  # 本月大参林下载日期范围
else:
    year = int(time.strftime("%Y", time.localtime()))  # 年
    # year = 2021
    month = int(time.strftime("%m", time.localtime())) - 1  # 上月
    # month = 12
    day = int(time.strftime("%d", time.localtime()))  # 日
    # day = 31
    time0 = str(year) + "-" + time_0(month) + "-" + str('01')
    time15 = str(year) + "-" + time_0(month) + "-" + str('15')
    time16 = str(year) + "-" + time_0(month) + "-" + str('16')
    lasttime = last_day_of_month(datetime.date(year, month, day))
    lastDay = int(str(lasttime)[8:10])  # 上个月最后一天
    dateDSL = time0 + ' - ' + str(lasttime)  # 上月大参林下载日期范围
    time1 = str(lasttime)
    time2 = str(lasttime)


def HW():  # 海王
    print('\n' + ">>>【海王】数据爬取中,稍等片刻")
    global sf
    global driver
    global all_data1
    all_data1 = []
    sf = ['安徽', '北京', '长春', '成都', '大连', '电商', '福州', '广州', '河南', '湖北', '湖南', '深圳总部',
          '杭州', '江苏', '辽宁', '宁波', '青岛', '潍坊', '上海', '沈阳', '深圳', '天津', '泰州']  # 海王省份列表
    wei_zhi = "D:\\FilesCenter\\大客户数据\\HW\\"  # 海王下载路径
    # szLocation = 'D:/FilesCenter/大客户数据/HW-深圳/'
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {
        'profile.default_content_settings.popups': 0,
        'download.default_directory': wei_zhi
    }  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe', options=options)  # 打开浏览器
    driver.implicitly_wait(5)  # 隐式等待
    with open('D://FilesCenter//大客户数据//anquan.txt', 'r', encoding='utf8') as f:
        listCookies = json.loads(f.read())  # 读取cookies
    driver.get('http://srm.nepstar.cn')  # 进入登录界面
    driver.delete_all_cookies()  # 删旧cookies
    for cookie in listCookies:
        if 'expiry' in cookie:
            del cookie['expiry']
        driver.add_cookie(cookie)  # 新增cookies
    time.sleep(3)
    driver.get('http://srm.nepstar.cn/ELSServer_HWXC/default2.jsp?account=110829_1001&loginChage=N&telphone1=18279409642')
    # 读取完cookie刷新页面
    time.sleep(3)
    button_a('//*[@id="treeMenu"]/li[7]/a')  # 数据查询
    time.sleep(0.5)
    button_a('//*[@id="salesOutSourcingInfoManage"]')  # 销售查询
    time.sleep(0.5)
    # xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')  # 先通过xpath定位到iframe
    xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[2]/iframe')  # 先通过xpath定位到iframe
    driver.switch_to.frame(xf)  # 再将定位对象传给switch_to.frame()方法
    button_a('/html/body/div[1]/div[2]/form/div[1]/div/div/span')  # 商品编码
    time.sleep(5)
    driver.switch_to.parent_frame()
    # xf = driver.find_element_by_xpath('/html/body/div[4]/div/table/tbody/tr[2]/td/div/iframe')
    xf = driver.find_element_by_xpath('/html/body/div[5]/div/table/tbody/tr[2]/td/div/iframe')
    driver.switch_to.frame(xf)  # 切换
    button_a('/html/body/div/main/div[1]/div[1]/div[1]/table/thead/tr/th[2]/div/span/input')  # 全选
    time.sleep(0.5)
    button_a('/html/body/div/main/div[2]/button[1]')  # 确定
    time.sleep(0.5)
    driver.switch_to.parent_frame()
    xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[2]/iframe')
    driver.switch_to.frame(xf)  # 切换
    input_b('/html/body/div[1]/div[2]/form/div[2]/div/div/input', time0)  # 查询日期
    time.sleep(0.5)
    input_b('/html/body/div[1]/div[2]/form/div[3]/div/div/input', time2)  # 延迟两天
    time.sleep(0.5)
    button_a('/html/body/div[1]/div[2]/form/div[4]/div/div/div/p')  # 联采合同
    time.sleep(0.5)
    button_a('/html/body/div[1]/div[2]/form/div[4]/div/div/div/div/ul/li[3]')
    time.sleep(0.5)
    for i in range(2, 25):
        if i == 13:
            cprint("跳过深圳总部", 'cyan', attrs=['bold', 'reverse', 'blink'])
        # elif i == 15:
        #     cprint("无权限访问，跳过辽宁", 'cyan', attrs=['bold', 'reverse', 'blink'])
        else:
            button_a('/html/body/div[1]/div[2]/form/div[5]/div/div/div/p')  # 选地区
            time.sleep(0.5)
            ix = '/html/body/div[1]/div[2]/form/div[5]/div/div/div/div/ul/li[' + str(i) + ']'
            button_a(ix)  # 选好地区
            time.sleep(0.5)
            button_a('/html/body/div[1]/div[2]/form/button[2]')  # 查询
            time.sleep(7)
            button_a('/html/body/div[1]/div[3]/div[1]/div[1]/table/thead/tr/th[2]/div/span')  # 全选
            time.sleep(1)
            button_a('/html/body/div[1]/div[1]/nav/div/div/ul/li[1]/a')  # 查看明细
            time.sleep(1)
            driver.switch_to.parent_frame()  # 回退
            try:
                # driver.find_element_by_xpath('/html/body/div[1]/div[1]/nav/div/div/ul/li[1]/a')  # 查看明细是否存在
                xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')  # 切入明细
                driver.switch_to.frame(xf)
                time.sleep(5)
                button_a('/html/body/div/div[1]/nav/div/div/ul/li[2]/a')  # 导出
                time.sleep(6)
                element = driver.find_element_by_xpath('/html/body/div/div[1]/nav/div/div/ul/li[2]')
                ActionChains(driver).move_to_element(element).perform()  # 鼠标悬浮
                button_a('/html/body/div/div[1]/nav/div/div/ul/li[1]/a')  # 返回
                driver.switch_to.parent_frame()  # 切出
                time.sleep(8)
                all_exce = get_exce(wei_zhi)
                # 得到要合并的所有exce表格数据
                if (all_exce == 0):
                    cprint(sf[i - 2] + "下载出错！！！", 'magenta', attrs=['bold', 'reverse', 'blink'])
                else:
                    for exce in all_exce:
                        fh = xlrd.open_workbook(exce)
                        # 打开文件
                        sheets = get_sheet(fh)
                        # 获取文件下的sheet数量
                        for sheet in range(len(sheets)):
                            row = get_sheetrow_num(sheets[sheet])
                            # 获取一个sheet下的所有的数据的行数
                            all_data1 = get_sheet_data(sheets[sheet], row,i - 2)
                            os.remove(exce)
                print("导出完成！", sf[i - 2])
                # xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[3]/iframe')  # 切入查询
                xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[2]/iframe')  # 切入查询
                time.sleep(1)
                driver.switch_to.frame(xf)  # 切换
                button_a('/html/body/div[1]/div[3]/div[1]/div[1]/table/thead/tr/th[2]/div/span')  # 全不选

            except BaseException:
                cprint(sf[i - 2] + "->列无数据", 'magenta', attrs=['bold', 'reverse', 'blink'])
                xf = driver.find_element_by_xpath('/html/body/div[2]/div/nav[2]/div[2]/iframe')  # 切入查询
                driver.switch_to.frame(xf)  # 切换
                button_a('/html/body/div[1]/div[3]/div[1]/div[1]/table/thead/tr/th[2]/div/span')  # 全不选
            # finally:

    dictCookies = driver.get_cookies()  # 获取cookies
    jsonCookies = json.dumps(dictCookies)
    with open('D://FilesCenter//大客户数据//anquan.txt', 'w') as f:
        f.write(jsonCookies)  # 保存新cookies

    biao_tou = ['类别', '商品SAP编码', '商品名称', '规格', '单位', '店号/区域ID', '店名/区域', '销量', '过账日期', '合同价', '省份']
    all_data1.insert(0, biao_tou)  # 表头写入
    # 下面开始文件数据的写入
    new_excel = "D:\\FilesCenter\\大客户数据\\HW\\" + "HAIWANG" + str(month) + ".xlsx"  # 新建的excel文件名字
    fh1 = xlsxwriter.Workbook(new_excel)  # 新建一个excel表
    new_sheet = fh1.add_worksheet()  # 新建一个sheet表
    for i in range(len(all_data1)):
        for j in range(len(all_data1[i])):
            c = all_data1[i][j]
            new_sheet.write(i, j, c)
    fh1.close()  # 关闭该excel表


def LBX():  # 老百姓
    global driver
    print('\n' + ">>>【老百姓】数据爬取中,稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\LBX\\mergeTableLBX\\"  # 老百姓下载路径
    deleteOldFiles(wei_zhi)  # 清空老百姓文件夹文件
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe', options=options)  # 打开浏览器
    driver.implicitly_wait(30)  # 隐式等待
    driver.get('http://srm.lbxdrugs.com:6066/srm/coframe/auth/login/login.jsp')  # 进入网址
    input_b('/html/body/div/div[1]/div/form/div[1]/span/span/input', '900022800005')  # 账号
    input_b('/html/body/div/div[1]/div/form/div[2]/span/span/input', 'zz106126')  # 密码
    button_a('/html/body/div/div[1]/div/form/div[4]/input')  # 登录
    try:  # 点击关闭通知框
        time.sleep(2)
        button_b('/html/body/div[3]/span')
    except:
        pass
    # driver = webdriver.Chrome(executable_path = 'C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',options=options)#打开浏览器
    # driver.implicitly_wait(30)#隐式等待
    # driver.get('http://srm.lbxdrugs.com:6066/srm/coframe/auth/login/login.jsp')#进入网址
    # input_b('/html/body/div/div[1]/div/form/div[1]/span/span/input','900022800005')#账号
    # input_b('/html/body/div/div[1]/div/form/div[2]/span/span/input','zz3737')#密码
    # button_a('/html/body/div/div[1]/div/form/p[3]')#登录
    # time.sleep(3)
    # button_b('/html/body/div[2]/div/div[5]/span')
    time.sleep(1)
    button_a('/html/body/div[2]/div[2]/div[1]/div/ul/li[1]/dl/dt/a')  # 销售管理
    time.sleep(0.5)
    button_a('/html/body/div[2]/div[2]/div[1]/div/ul/li[1]/dl/dd[1]/ul/li/a')  # 门店销售明细
    time.sleep(1)
    switch('/html/body/div[2]/div[2]/div[2]/div[2]/table/tbody/tr/td[2]/div[2]/div[2]/iframe')  # 切入
    input_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[2]/span/span/input', '')  # 清空省公司文本框
    time.sleep(0.5)
    xuanfu('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[2]/span')
    button_b('/html/body/div[1]/fieldset/table/tbody/tr/td[2]/span/span/span/span[2]/span')  # 省公司
    time.sleep(1)
    # button_b('/html/body/div[3]/div/div[1]/div[1]/table/tbody/tr/td[1]/input')  # 全选
    button_b('/html/body/div[7]/div/div[1]/div[1]/table/tbody/tr/td[1]/input')  # 全选
    time.sleep(0.5)
    button_b('/html/body/div[1]/fieldset/table/tbody/tr/td[2]/span/span/span/span[2]/span')  # 省公司
    time.sleep(0.5)
    button_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[4]/span/span/span/span[1]')  # 点击商品名称里面的x
    time.sleep(1)
    input_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[4]/span/span/input','')  # 清空商品名称文本框
    time.sleep(0.5)
    xuanfu('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[4]/span/span')
    time.sleep(1)
    button_b('/html/body/div[1]/fieldset/table/tbody/tr[1]/td[4]/span/span/span/span[2]/span')  # 商品名称
    driver.switch_to.parent_frame()  # 切出
    switch('/html/body/div[5]/div/div[2]/div[2]/iframe')  # 切入
    time.sleep(5)
    button_b('/html/body/div[3]/div/div[2]/div[2]/div[2]/table/tbody/tr[2]/td[3]/div/div[1]/input')  # 双击
    time.sleep(2)
    button_b('/html/body/div[3]/div/div[2]/div[2]/div[2]/table/tbody/tr[2]/td[3]/div/div[1]/input')  # 全选
    time.sleep(2)
    button_b('/html/body/a[1]/span')  # 确定
    driver.switch_to.parent_frame()  # 切出
    if aaa == "1":
        for dayi in range(1, day):
            lbsTime = str(year) + "-" + str(month).zfill(2) + "-" + str(dayi).zfill(2)
            switch('/html/body/div[2]/div[2]/div[2]/div[2]/table/tbody/tr/td[2]/div[2]/div[2]/iframe')  # 切换明细
            driver.find_element_by_xpath('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input').clear()  # 清空开始时间
            time.sleep(0.5)
            input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input', lbsTime)  # 输入开始日期
            time.sleep(0.5)
            driver.find_element_by_xpath('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span/input').clear()  # 清空结尾时间
            time.sleep(0.5)
            input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span/input',lbsTime)  # 输入日期
            button_b('/html/body/div[2]/a[1]/span')  # 查询
            time.sleep(2)
            button_b('/html/body/div[2]/a[2]/span')  # 导出
            time.sleep(4)
            driver.switch_to.parent_frame()  # 切出
            time.sleep(0.5)
    else:
        for dayi in range(1, lastDay + 1):
            lbsTime = str(year) + "-" + str(month).zfill(2) + "-" + str(dayi).zfill(2)
            switch('/html/body/div[2]/div[2]/div[2]/div[2]/table/tbody/tr/td[2]/div[2]/div[2]/iframe')  # 切换明细
            driver.find_element_by_xpath('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input').clear()  # 清空开始时间
            time.sleep(0.5)
            input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input',lbsTime)  # 输入开始日期
            time.sleep(0.5)
            driver.find_element_by_xpath('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span/input').clear()  # 清空结尾时间
            time.sleep(0.5)
            input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span/input',lbsTime)  # 输入日期
            time.sleep(0.5)
            button_b('/html/body/div[2]/a[1]/span')  # 查询
            time.sleep(3)
            button_b('/html/body/div[2]/a[2]/span')  # 导出
            time.sleep(3)
            driver.switch_to.parent_frame()  # 切出
            time.sleep(0.5)
    time.sleep(20)
    mergeXlsTable()  # 合并表格
    # driver.quit()
    '''
    switch('/html/body/div[2]/div/div[2]/div/div/div[2]/div[2]/div/div/iframe')#切换明细
    input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[1]/span[1]/input',time1)#输入开始日期
    time.sleep(0.5)
    input_b('/html/body/div[1]/fieldset/table/tbody/tr[2]/td[2]/span[3]/span[1]/input',time1)#输入日期
    time.sleep(0.5)
    button_b('/html/body/div[2]/a[1]/span')#查询
    time.sleep(5)
    button_b('/html/body/div[2]/a[2]/span')#导出
    time.sleep(20)
    #driver.quit()
    '''


def YF():  # 益丰
    global driver
    print('\n' + ">>>【益丰】数据爬取中, 稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\YF\\"  # 益丰下载路径
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe', options=options)  # 打开浏览器
    driver.implicitly_wait(20)  # 隐式等待
    driver.get('http://vendor.yfdyf.cn/sup/login')  # 进入网址
    input_b('/html/body/form/input[1]', '中山市中智')  # 账号
    time.sleep(0.5)
    input_b('/html/body/form/input[2]', 'zz106126')  # 密码
    # input_b('/html/body/form/input[2]','zsszz')#密码
    time.sleep(0.5)
    button_a('/html/body/form/input[3]')  # 登录
    time.sleep(3)
    # button_a('/html/body/div[1]/div[1]/div/div[2]/ul/li[3]/a')#供应商查询
    button_a('/html/body/div[1]/div[1]/div/div[2]/ul/li[3]/a/span')  # 供应商查询
    time.sleep(1)
    button_a('/html/body/div[1]/div[2]/div[1]/div[1]/div[2]/div/div[2]/div/ul/li[3]/a')  # 销售记录
    time.sleep(3)
    switch('/html/body/div[1]/div[2]/div[1]/div[3]/div/div[2]/div/div[4]/iframe')  # 切入
    time.sleep(2)
    if aaa == "1":  # 判断时间
        js_begin = 'document.getElementById("beginTima").removeAttribute("readonly");'
        driver.execute_script(js_begin)
        # 用js方法输入日期
        js_begin_value = 'document.getElementById("beginTima").value="' + str(time0) + '"'  # 输入每月第一天
        driver.execute_script(js_begin_value)
        time.sleep(2)
        js_end_value = 'document.getElementById("endTima").value="' + str(time1) + '"'  # 前一天
        driver.execute_script(js_end_value)
        js_end = 'document.getElementById("endTima").removeAttribute("readonly");'
        driver.execute_script(js_end)
    else:
        js_begin = 'document.getElementById("beginTima").removeAttribute("readonly");'
        driver.execute_script(js_begin)
        # 用js方法输入日期
        js_begin_value = 'document.getElementById("beginTima").value="' + str(time0) + '"'  # 输入上月第一天
        driver.execute_script(js_begin_value)
        time.sleep(2)
        js_end = 'document.getElementById("endTima").removeAttribute("readonly");'
        driver.execute_script(js_end)
        js_end_value = 'document.getElementById("endTima").value="' + str(time1) + '"'  # 上个月最后一天
        driver.execute_script(js_end_value)
    time.sleep(1)
    button_a('/html/body/form/ul/li[4]/input[1]')  # 查询
    time.sleep(20)
    button_a('/html/body/form/ul/li[4]/input[2]')  # 导出
    time.sleep(3)
    driver.switch_to.parent_frame()  # 切出
    time.sleep(2)
    button_a('/html/body/div[3]/div[4]/table/tbody/tr[2]/td[2]/div/div[2]/div/div[2]/button[1]')  # 确定
    time.sleep(3)  # 睡两秒，看一下效果
    # driver.quit()
    '''
    switch('/html/body/div[1]/div[2]/div[1]/div[3]/div/div[2]/div/div[4]/iframe')#切换明细
    time.sleep(0.5)
    button_a('/html/body/form/ul/li[2]/input[1]')#点击开始日期
    time.sleep(1)
    button_a('/html/body/div/div[3]/table/tbody/tr[2]/td[3]')
    time.sleep(1)
    #input_b('/html/body/form/ul/li[2]/input[1]','')
    driver.switch_to.parent_frame()#切出
    switch('/html/body/div[3]/iframe')#切换日期
    #button_a('/html/body/div/div[6]/input[1]')#鼠标点击清空
    if aaa == "1":
        for i in range(1,8):
            if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[2]/td["+str(i)+"]")[0].text == '1':
                button_a("/html/body/div/div[3]/table/tbody/tr[2]/td["+str(i)+"]")
        driver.switch_to.parent_frame()#切出
        switch('/html/body/div[3]/iframe')#切换日期
        button_a('//td[@onclick="day_Click(2020,'+str(month)+','+str(day-1)+');"]')
        time.sleep(2)
        driver.switch_to.parent_frame()#切出
        switch('/html/body/div[1]/div[2]/div[1]/div[3]/div/div[2]/div/div[4]/iframe')#切换明细
    else:
        button_a('/html/body/div/div[1]/div[2]/a')
        for i in range(1,8):
            if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[2]/td["+str(i)+"]")[0].text == '1':
                button_a("/html/body/div/div[3]/table/tbody/tr[2]/td["+str(i)+"]")
        time.sleep(2)
        driver.switch_to.parent_frame()#切出
        switch('/html/body/div[3]/iframe')#切换明细
        button_a('/html/body/div/div[1]/div[2]/a')
        for i in range(1,8):
            if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[6]/td["+str(i)+"]")[0].text == str(lasttime.day):
                button_a("/html/body/div/div[3]/table/tbody/tr[6]/td["+str(i)+"]")
        driver.switch_to.parent_frame()#切出
        switch('/html/body/div[1]/div[2]/div[1]/div[3]/div/div[2]/div/div[4]/iframe')#切换明细

    #input_b('/html/body/form/ul/li[2]/input[1]',time0)#输入开始日期
    time.sleep(2)
    button_a('/html/body/form/ul/li[3]/input[2]')#导出
    time.sleep(1)
    driver.switch_to.parent_frame()#切出
    button_b('/html/body/div[4]/div[4]/table/tbody/tr[2]/td[2]/div/div[2]/div/div[2]/button[1]')#确定导出
    time.sleep(120)
    #driver.quit()
    '''


def SY():  # 漱玉
    global driver
    print('\n' + ">>>【漱玉】数据爬取中,稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\SY\\downloadSY\\"  # 漱玉下载路径
    deleteOldFiles(wei_zhi)  # 漱玉文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {
        'profile.default_content_settings.popups': 0,
        'download.default_directory': wei_zhi
    }  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(30)  # 隐式等待
    driver.get('http://60.217.250.254:8765/')  # 进入网址
    button_a('/html/body/form/div[3]/div[1]/table/tbody/tr/td[3]/span/a[1]')  # 登录
    time.sleep(3)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[1]/input', '20000618')  # 账号
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[2]/input', 'cjh123')  # 密码
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[2]/div/div/p/input')  # 登录
    time.sleep(2)
    xuanfu('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/a')  # 商品查询
    time.sleep(1)
    button_a('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/ul/li[2]/a')  # 销售查询
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[6]/input', time0)  # 开始时间
    time.sleep(0.5)
    input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[8]/input', time15)  # 结束时间
    time.sleep(0.5)
    button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[9]/input')  # 查询
    time.sleep(2)
    try:
        # 等待alert弹出框可见
        # WebDriverWait(driver,20).until(EC.alert_is_present())
        # 从html页面切换到alert弹框
        alert = driver.switch_to.alert
        # 接受--选择“确定”
        alert.accept
    except BaseException:
        pass
    time.sleep(30)
    button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[2]/table/tbody/tr/td/input')  # 导出
    time.sleep(30)
    button_a('/html/body/form/div[3]/fieldset/table/tbody/tr/td/input[1]')  # 导出
    time.sleep(15)
    if day > 16 or aaa == '2':
        time.sleep(10)
        driver.quit()
        wei_zhi = "D:\\FilesCenter\\大客户数据\\SY\\downloadSY\\"  # 漱玉下载路径
        options = webdriver.ChromeOptions()  # 打开设置
        prefs = {
            'profile.default_content_settings.popups': 0,
            'download.default_directory': wei_zhi
        }  # 设置路径
        options.add_experimental_option('prefs', prefs)
        driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                                  options=options)  # 打开浏览器
        driver.implicitly_wait(10)  # 隐式等待
        driver.get('http://60.217.250.254:8765/')  # 进入网址
        button_a('/html/body/form/div[3]/div[1]/table/tbody/tr/td[3]/span/a[1]')  # 登录
        time.sleep(1)
        input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[1]/input', '20000618')  # 账号
        time.sleep(0.5)
        input_b('/html/body/form/div[3]/div[2]/div/div/fieldset/ul/li[2]/input', 'cjh123')  # 密码
        time.sleep(0.5)
        button_a('/html/body/form/div[3]/div[2]/div/div/p/input')  # 登录
        time.sleep(3)
        xuanfu('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/a')  # 商品查询
        time.sleep(0.5)
        button_a('/html/body/form/div[3]/div[1]/div/div[1]/ul/li[6]/ul/li[2]/a')  # 销售查询
        time.sleep(0.5)
        input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[6]/input', time16)
        time.sleep(0.5)
        input_b('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[8]/input', time1)
        time.sleep(0.5)
        button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[1]/table/tbody/tr/td[9]/input')  # 查询
        time.sleep(30)
        button_a('/html/body/form/div[3]/div[2]/div/div/fieldset[2]/table/tbody/tr/td/input')  # 导出
        time.sleep(30)
        button_a('/html/body/form/div[3]/fieldset/table/tbody/tr/td/input[1]')  # 导出
        time.sleep(30)
        # driver.quit()


def QY():  # 全亿
    global driver
    print('\n' + ">>>【全亿】数据爬取中,稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\QY\\downloadQY\\"  # 全亿下载路径
    deleteOldFiles(wei_zhi)  # 全亿文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {
        'profile.default_content_settings.popups': 0,
        'download.default_directory': wei_zhi
    }  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
                              options=options)  # 打开浏览器
    driver.implicitly_wait(100)  # 隐式等待
    driver.get('http://oa.myquanyi.com/login/Login.jsp?logintype=2')  # 进入网址
    input_b('/html/body/div[1]/div[2]/form/div[2]/table/tbody/tr[1]/td/div/input', 'JT007')  # 账号
    time.sleep(0.5)
    input_b('/html/body/div[1]/div[2]/form/div[2]/table/tbody/tr[3]/td/div/input', '123456')  # 密码
    time.sleep(0.5)
    button_a('/html/body/div[1]/div[2]/form/div[2]/table/tbody/tr[8]/td/input')  # 登录
    time.sleep(3)
    button_a('/html/body/div[2]/div[1]/div/div[3]/ul/li[4]/a/span')  # 销售查询
    time.sleep(2)
    driver.switch_to.parent_frame()  # 切出
    switch('/html/body/div[2]/div[2]/div/iframe')  # 切入流程
    switch('/html/body/div[4]/div[5]/iframe')  # 切入查询
    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[3]/td[2]/div/div/div[1]/span/span/button')  # 公司名称
    time.sleep(0.5)
    driver.switch_to.parent_frame()  # 切出
    driver.switch_to.parent_frame()  # 切出
    switch('/html/body/div[4]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/div/iframe')  # 切入
    switch('/html/body/div[1]/div[2]/div/iframe')  # 切入
    button_a('/html/body/div[1]/form/table[2]/tbody/tr[2]/td/div/div/div/ul/li/ul/li[2]/div/button[2]')  # 全选
    time.sleep(0.5)
    button_a('/html/body/div[2]/table/tbody/tr/td/input[1]')  # 确定
    driver.switch_to.default_content()
    switch('/html/body/div[2]/div[2]/div/iframe')  # 切入流程
    switch('/html/body/div[4]/div[5]/iframe')  # 切入查询
    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[2]/div/button')  # 点击日期低峰日历
    time.sleep(0.5)
    switch('/html/body/div[5]/iframe')  # 切入时间
    if aaa == "1":
        for i in range(1, 8):
            if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[2]/td[" + str(i) + "]")[0].text == '1':
                button_a("/html/body/div/div[3]/table/tbody/tr[2]/td[" + str(i) + "]")
                break
        driver.switch_to.parent_frame()  # 切出
        button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[4]/div/button')  # 点击日期高峰日历
        switch('/html/body/div[5]/iframe')  # 切入时间
        time.sleep(1)
        button_a('/html/body/div/div[6]/input[2]')  # 点击今天
        driver.switch_to.parent_frame()  # 切出
        time.sleep(2)
        button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[4]/div/button')  # 点击日期高峰日历
        time.sleep(2)
        switch('/html/body/div[5]/iframe')  # 切入时间
        # for i in range(2,8):
        #   for j in range(1,8):
        #       if driver.find_element_by_xpath('/html/body/div/div[3]/table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')[0].text == str(day - 1):
        #  解决同一张日历内月初几天与后一月重复问题
        if day <= 20:
            for i in range(2, 6):
                for j in range(1, 8):
                    if driver.find_elements_by_xpath('/html/body/div/div[3]/table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')[0].text == str(day - 1):
                        button_a('/html/body/div/div[3]/table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')  # 点击前一天时间
                        break
        else:
            for i in range(5, 8):
                for j in range(1, 8):
                    if driver.find_elements_by_xpath('/html/body/div/div[3]/table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')[0].text == str(day - 1):
                        button_a('/html/body/div/div[3]/table/tbody/tr[' + str(i) + ']/td[' + str(j) + ']')  # 点击前一天时间
                        break
        # button_a('/html/body/div/div[3]/table/tbody/tr[' + str((day -1)//7 + 2) + ']/td[' + str((day -1)%7) + ']')  # 选中所要的日期
        '''
        if day > 7:
            s1 = 3
        else:
            s1 = 2
        for i in range(1,8):
            for j in range(s1,7):
                if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr["+str(j)+"]/td["+str(i)+"]")[0].text == str(day-1):
                    button_a("/html/body/div/div[3]/table/tbody/tr["+str(j)+"]/td["+str(i)+"]")
                    break
        '''
        driver.switch_to.parent_frame()  # 切出
    else:
        button_a('/html/body/div/div[1]/div[2]/a')  # 点击上个月
        for i in range(1, 8):
            if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[2]/td[" + str(i) + "]")[0].text == '1':
                button_a("/html/body/div/div[3]/table/tbody/tr[2]/td[" + str(i) + "]")
                break
        driver.switch_to.parent_frame()  # 切出

        button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[6]/td[4]/div/button')  # 点击日期高峰日历
        switch('/html/body/div[5]/iframe')  # 切入时间
        button_a('/html/body/div/div[1]/div[2]/a')  # 点击上个月
        for i in range(1, 8):
            for j in range(5, 8):
                if driver.find_elements_by_xpath("/html/body/div/div[3]/table/tbody/tr[" + str(j) + "]/td[" + str(i) + "]")[0].text == str(lasttime.day):
                    button_a("/html/body/div/div[3]/table/tbody/tr[" + str(j) + "]/td[" + str(i) + "]")
                    break
        driver.switch_to.parent_frame()  # 切出
    time.sleep(2)
    button_a('/html/body/form/div[1]/table/tbody/tr/td/table/tbody/tr[7]/td/input')  # 数据下载
    time.sleep(2)
    # 等待alert弹出框可见
    # WebDriverWait(driver,20).until(EC.alert_is_present())
    # 从html页面切换到alert弹框
    alert = driver.switch_to.alert
    # 获取alert的文本内容
    print(alert.text)
    # 接受--选择“确定”
    alert.accept()
    # alert.dismiss()#取消
    time.sleep(5)
    # driver.quit()


def get_track(distance):
    """
    根据偏移量获取移动轨迹
    :param distance: 偏移量
    :return: 移动轨迹
    """
    # 移动轨迹
    track = []
    # 当前位移
    current = 0
    # 减速阈值
    mid = distance * 4 / 5
    # 计算间隔
    t = 0.2
    # 初速度
    v = 0
    while current < distance:
        if current < mid:
            # 加速度为正2
            a = 2
        else:
            # 加速度为负3
            a = -1
        # 初速度v0
        v0 = v
        # 当前速度v = v0 + at
        v = v0 + a * t
        # 移动距离x = v0t + 1/2 * a * t^2
        move = v0 * t + 1 / 2 * a * t * t
        # 当前位移
        current += move
        # 加入轨迹
        track.append(round(move))
    return track


def GJ():  # 高济
    global driver
    print('\n' + ">>>【高济】数据爬取中,稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\GJ\\downloadGJ\\"  # 高济下载路径
    deleteOldFiles(wei_zhi)  # 高济文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    # options.add_argument('ignore-certificate-errors')
    options.add_argument("--user-data-dir="+r"C:/Users/Zeus/AppData/Local/Google/Chrome/User Data/")  # 解决提示不安全内容下数据的下载
    prefs = {
        'profile.default_content_settings.popups': 0,
        'download.default_directory': wei_zhi
    }  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe', options=options)  # 打开浏览器
    driver.set_window_size(1100, 800)
    driver.implicitly_wait(5)  # 隐式等待
    driver.get('https://srm.cowellhealth.com')  # 进入网址
    for cishu in range(1, 6):  # 尝试5次
        try:
            try:
                time.sleep(1)
                input_b('/html/body/div/div/div[1]/div[2]/div[1]/form/div[1]/div/div/span/input', '13590885749')  # 账号
                input_b('/html/body/div/div/div[1]/div[2]/div[1]/form/div[2]/div/div/span/input', 'zz_106126')  # 密码
                time.sleep(1)
                png = driver.find_element_by_xpath('/html/body/div/div/div[1]/div[2]/div[1]/form/div[3]/div/div[1]/img[1]')  # 图片位置
                png_src = png.get_attribute('src')  # 图片链接
                response = urllib.request.urlopen(png_src)  # 登录图片链接
                cat_img = response.read()  # 读取图片
                nparr = np.frombuffer(cat_img, dtype=np.uint8)  # 转格式
                img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)  # 图片处理
                gray = img
                circles1 = cv2.HoughCircles(gray,
                                            cv2.HOUGH_GRADIENT,
                                            1,
                                            130,
                                            param1=10,
                                            param2=50,
                                            minRadius=0,
                                            maxRadius=60)  # 寻找圆形
                circles = circles1[0, :, :]  # 取一个圆形
                circles = np.uint16(np.around(circles))
                print("圆心坐标(" + str(circles[0][0]) + "," + str(circles[0][1]) + ")")
                time.sleep(1)
                # 获取拖拽的圆球
                slideblock = driver.find_element_by_xpath('/html/body/div/div/div[1]/div[2]/div[1]/form/div[3]/div/div[2]/div[3]')  # 拖动按钮
                track = get_track(circles[0][0] - 18)  # 设置速度
                ActionChains(driver).click_and_hold(slideblock).perform()  # 按住
                for x in track:
                    ActionChains(driver).move_by_offset(xoffset=x, yoffset=0).perform()  # 拖动
                time.sleep(0.5)
                ActionChains(driver).release().perform()  # 放手
                time.sleep(1)
                button_a('/html/body/div/div/div[1]/div[2]/div[1]/form/div[4]/div/div/span/div')  # 登录
                time.sleep(1)
                button_a('/html/body/div/div/section/aside/div/ul/li[2]/div[1]')  # 数据中心
                # button_b('/html/body/div/div/div[1]/div[2]/div[1]/form/div[4]/div/div/span/div')#登录
                break
            except BaseException:
                slideblock = driver.find_element_by_xpath('/html/body/div/div/div[1]/div[2]/div[1]/form/div[3]/div/div[2]/div[3]')  # 拖动按钮
                track = get_track(circles[0][0] - 18)  # 设置速度
                ActionChains(driver).click_and_hold(slideblock).perform()  # 按住
                for x in track:
                    ActionChains(driver).move_by_offset(xoffset=150, yoffset=0).perform()  # 拖动
                time.sleep(0.5)
                ActionChains(driver).release().perform()  # 放手
                time.sleep(3)
                button_a('/html/body/div/div/div[1]/div[2]/div[1]/form/div[4]/div/div/span/div')  # 登录
                time.sleep(1)
                button_a('/html/body/div/div/section/aside/div/ul/li[2]/div[1]')  # 数据中心
                break
        except BaseException:
            print('登陆失败', cishu)
            driver.refresh()
            time.sleep(2)
    time.sleep(2)
    button_a('/html/body/div/div/section/aside/div/ul/li[2]/ul/li')  # 数据下载
    if aaa == "1":
        if driver.find_element_by_xpath('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span/span').text != '下载':
            button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span/span')  # 门店销售明细的重新获取
            time.sleep(15)
        driver.refresh()  # 刷新当前页面
        time.sleep(3)
        # button_a("/html/body/div[1]/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/div/div[2]/a")
        # time.sleep(40)
        reNewDateMX = driver.find_element_by_xpath('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/div/div[2]').text  # 获取门店销售明细时间
        reNewDateKC = driver.find_element_by_xpath('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[6]/div/div/div[2]').text  # 获取实时库存时间
        if (str(year) + '-' + str(month).zfill(2) + '-' + str(day).zfill(2)) == str(reNewDateMX[5:16]).strip():
            print("当前数据截止今天" + str(str(year) + '-' + str(month).zfill(2) + '-' + str(day).zfill(2)) + "不需要刷新时间")
            print("获取门店销售明细")
            time.sleep(1)
            button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span')  # 下载销售
            time.sleep(5)
        else:
            print("刷新门店销售明细日期")
            time.sleep(5)
            button_a('/html/body/div[1]/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/div/div[2]/a')  # 门店销售明细的重新获取
            time.sleep(10)
            driver.refresh()  # 刷新当前页面
            reDateAgain = driver.find_element_by_xpath(
                '/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/div/div[2]').text  # 获取门店销售明细时间
            if (str(year) + '-' + str(month).zfill(2) + '-' + str(day).zfill(2)) == str(reDateAgain[5:16]).strip():
                time.sleep(20)
                driver.refresh()  # 刷新当前页面
                print("获取门店销售明细")
                button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span')  # 下载销售
            else:
                print("再次刷新门店销售明细日期")
                try:
                    button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/div/div[2]/a')  # 门店销售明细的重新获取
                except BaseException:
                    pass
            # driver.quit()
        '''
        # 目前不可以导出实时库存，暂时用不上
        time.sleep(3)
        if (str(year) + '-' + str(month).zfill(2) + '-' + str(day).zfill(2)) == str(reNewDateKC[5:16]).strip():
            print("获取实时库存")
            time.sleep(1)
            button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[6]/div/ul/li[1]/span')#下载库存
            time.sleep(10)
        else:
            print("刷新实时库存日期")
            time.sleep(1)
            button_a('/html/body/div[1]/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[6]/div/div/div[2]/a')#实时库存的重新获取
            time.sleep(10)
            driver.quit()
        '''
    else:
        time.sleep(3)
        button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[1]/span[2]/div/input')  # 打开月份
        time.sleep(1)
        if month % 3 == 0:
            button_a('/html/body/div[2]/div/div/div/div/div/div/div[2]/div/div[2]/table/tbody/tr[' + str(month // 3) + ']/td[3]/a')  # 选择月份
        else:
            button_a('/html/body/div[2]/div/div/div/div/div/div/div[2]/div/div[2]/table/tbody/tr[' + str(month // 3 + 1) + ']/td[' + str(
                month % 3) + ']/a')  # 选择月份
        time.sleep(2)
        button_a(
            '/html/body/div[1]/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span/span')  # 下载门店销售明细
        # button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[4]/div/ul/li[1]/span')  # 下载门店销售明细
        # time.sleep(10)
        # button_a('/html/body/div/div/section/section/main/div/div[2]/div[2]/div[2]/div[3]/div[1]/div[2]/div/div[2]/div[6]/div/ul/li[1]/span')  # 下载实时库存
    time.sleep(15)
    driver.quit()


# 二值数组
t2val = {}

def twoValue(image, G):
    for y in range(0, image.size[1]):
        for x in range(0, image.size[0]):
            g = image.getpixel((x, y))
            if g > G:
                t2val[(x, y)] = 1
            else:
                t2val[(x, y)] = 0

def clearNoise(image, N, Z):
    for i in range(0, Z):
        t2val[(0, 0)] = 1
        t2val[(image.size[0] - 1, image.size[1] - 1)] = 1

        for x in range(1, image.size[0] - 1):
            for y in range(1, image.size[1] - 1):
                nearDots = 0
                L = t2val[(x, y)]
                if L == t2val[(x - 1, y - 1)]:
                    nearDots += 1
                if L == t2val[(x - 1, y)]:
                    nearDots += 1
                if L == t2val[(x - 1, y + 1)]:
                    nearDots += 1
                if L == t2val[(x, y - 1)]:
                    nearDots += 1
                if L == t2val[(x, y + 1)]:
                    nearDots += 1
                if L == t2val[(x + 1, y - 1)]:
                    nearDots += 1
                if L == t2val[(x + 1, y)]:
                    nearDots += 1
                if L == t2val[(x + 1, y + 1)]:
                    nearDots += 1

                if nearDots < N:
                    t2val[(x, y)] = 1

def GD():  # 国大
    global driver
    print('\n' + ">>>【国大】数据爬取中,稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\GD\\downloadGD\\"  # 国大下载路径
    deleteOldFiles(wei_zhi)  # 国大文件夹清空
    options = webdriver.ChromeOptions()  # 打开设置
    prefs = {
        'profile.default_content_settings.popups': 0,
        'download.default_directory': wei_zhi
    }  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(
        executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe',
        options=options)  # 打开浏览器
    driver.implicitly_wait(5)  # 隐式等待
    for cishu in range(1, 6):  # 尝试5次
        try:
            driver.get('https://gdedi.sinopharmgroup.net/EDI/Login?ReturnUrl=%2f')
            input_b('//input[@id="loginId"]', '601317')  # 账号
            input_b('//input[@id="password"]', 'zb@888888')  # 密码
            # 获取验证码
            # button_a('/html/body/form/div/div[2]/p[3]/span[2]/a')  # 看不清楚，换一张
            driver.save_screenshot('picture.png')  # 全屏截图
            page_snap_obj = Image.open('picture.png')
            img = driver.find_element_by_xpath('/html/body/form/div/div[2]/p[4]/span[2]/img')  # 验证码元素位置
            time.sleep(1)
            location = img.location
            size = img.size  # 获取验证码的大小参数
            left = location['x']
            top = location['y']
            right = left + size['width']
            bottom = top + size['height']
            image_obj = page_snap_obj.crop((left, top, right, bottom))  # 按照验证码的长宽，切割验证码
            img = image_obj.convert("L")  # 转灰度
            pixdata = img.load()
            w, h = img.size
            threshold = 160
            # 遍历所有像素，大于阈值的为黑色
            for y in range(h):
                for x in range(w):
                    if pixdata[x, y] < threshold:
                        pixdata[x, y] = 0
                    else:
                        pixdata[x, y] = 255
            data = img.getdata()
            w, h = img.size
            black_point = 0
            for x in range(1, w - 1):
                for y in range(1, h - 1):
                    mid_pixel = data[w * y + x]  # 中央像素点像素值
                    if mid_pixel < 50:  # 找出上下左右四个方向像素点像素值
                        top_pixel = data[w * (y - 1) + x]
                        left_pixel = data[w * y + (x - 1)]
                        down_pixel = data[w * (y + 1) + x]
                        right_pixel = data[w * y + (x + 1)]
                        # 判断上下左右的黑色像素点总个数
                        if top_pixel < 10:
                            black_point += 1
                        if left_pixel < 10:
                            black_point += 1
                        if down_pixel < 10:
                            black_point += 1
                        if right_pixel < 10:
                            black_point += 1
                        if black_point < 1:
                            img.putpixel((x, y), 255)
                        black_point = 0
            img.save('image.jpg')
            pytesseract.pytesseract.tesseract_cmd = r"D:\Software_manage\tesseract-v5.0.0\tesseract.exe"  # 设置pyteseract路径
            result = pytesseract.image_to_string(Image.open('image.jpg'))  # 识别图片里面的文字
            dropSpecialString = re.sub(u"([^\u4e00-\u9fa5\u0030-\u0039\u0041-\u005a\u0061-\u007a])", "", result)  # 去除识别出来的特殊字符
            num = dropSpecialString[0:4]  # 只获取前4个字符
            print("本次识别验证码为：", num)
            input_b('//input[@id="imagecode"]', num)  # 输入验证码
            time.sleep(0.5)
            button_a('//*[@onclick="SumitLogin();"]')  # 登陆
            time.sleep(2)
            button_a('/html/body/div[2]/div[2]/div/ul/li/ul/li[3]/div/span[4]')  # 门店零售
            break
        except BaseException:
            print('登陆失败', cishu)
            pass
    '''
    while True:
        try:
            if cishu > 0:
                try:
                    clear_a('//input[@id="imagecode"]')
                    displayArea=driver.find_element_by_xpath('//img[@id="imgCode"]')#检查截图页面是否存在
                    time.sleep(3)
                    left = displayArea.location['x']#x起点
                    top = displayArea.location['y']#y起点
                    right = displayArea.location['x'] + displayArea.size['width']#宽
                    bottom = displayArea.location['y'] + displayArea.size['height']#高
                    driver.get_screenshot_as_file('screenshot.png')#截图
                    time.sleep(0.5)
                    im = Image.open('screenshot.png')#选择图片
                    time.sleep(0.5)
                    im = im.crop((left, top, right, bottom))  # 对浏览器截图进行裁剪
                    time.sleep(0.5)
                    #im = im.convert("P")
                    #im.show()

                    img=im
                    w,h=img.size
                    for x in range(w):
                        for y in range(h):
                            r,g,b,a=img.getpixel((x,y))
                            if 190<=r<=255 and 170<=g<=255 and 0<=b<=140:
                                img.putpixel((x,y),(0,0,0))
                            if 0<=r<=90 and 210<=g<=255 and 0<=b<=90:
                                img.putpixel((x,y),(0,0,0))
                    img=img.convert('L').point([0]*150+[1]*(256-150),'1')
                    time.sleep(0.5)

                    image2 = img.convert("L")
                    twoValue(image2, 100)
                    clearNoise(image2, 3, 3)
                    size = image2.size
                    time.sleep(0.5)
                    image3 = Image.new("1", size)
                    draw = ImageDraw.Draw(image3)
                    time.sleep(0.5)
                    for x in range(0, size[0]):
                        for y in range(0, size[1]):
                            draw.point((x, y), t2val[(x, y)])
                    #image3.show()
                    num = pytesseract.image_to_string(image3)
                    input_b('//input[@id="imagecode"]',num)#验证码
                    time.sleep(0.5)
                    button_a('//*[@onclick="SumitLogin();"]')#登陆
                    time.sleep(5)
                    #button_a('//*[@node-id="9103103"]')#门店零售
                    button_a('/html/body/div[2]/div[2]/div/ul/li/ul/li[5]/div/span[4]')#门店零售
                    break
                except:
                    print('登陆失败',cishu)
                    cishu=cishu-1
            else:
                break
        except:
            pass
    '''
    # button_a('//*[@node-id="9103103"]')
    '''
    switch('/html/body/div[3]/div/div/div[2]/div[2]/div/iframe')#切换
    driver.switchTo().fame("xpath")
    driver.switchTo().fame("selector")
    '''
    # button_a('/html/body/div[2]/div[2]/div/ul/li/ul/li[5]/div/span[3]')  # 门店零售
    # driver.switch_to.parent_frame()#切出
    # button_a('/html/body/form/fieldset/table/tbody/tr[1]/td[1]/span/span/span')#开始日期
    # time.sleep(0.5)
    # button_a('//*[@abbr="'+ str(year) +','+ str(month) +',1"]')
    # button_a('/html/body/form/fieldset/table/tbody/tr[1]/td[1]/span/span/span')#日期
    # time.sleep(0.5)
    # button_a('//*[@abbr="'+ str(year) +','+ str(month) +','+ str(day-1) +'"]')
    time.sleep(2)
    switch('/html/body/div[3]/div/div/div[2]/div[2]/div/iframe')  # 切入
    if aaa == "1":  # 判断时间
        js_begin = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-text.validatebox-text").removeAttribute("readonly");'
        driver.execute_script(js_begin)
        # 用js方法输入日期
        js_begin1_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-text.validatebox-text").value="' + time0 + '"'  # 表面输入本月第一天
        driver.execute_script(js_begin1_value)
        js_begin2_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-value").value="' + time0 + '"'  # 实际有效的输入本月第一天
        driver.execute_script(js_begin2_value)
        time.sleep(2)
        js_end = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-text.validatebox-text").removeAttribute("readonly");'
        driver.execute_script(js_end)
        js_end1_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-text.validatebox-text").value="' + time1 + '"'  # 表面输入前一天
        driver.execute_script(js_end1_value)
        js_end2_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-value").value="' + time1 + '"'  # 实际有效的输入前一天
        driver.execute_script(js_end2_value)
    else:
        js_begin = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-text.validatebox-text").removeAttribute("readonly");'
        driver.execute_script(js_begin)
        # 用js方法输入日期
        js_begin1_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-text.validatebox-text").value="' + time0 + '"'  # 表面输入上月第一天
        driver.execute_script(js_begin1_value)
        js_begin2_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(1) > span > input.combo-value").value="' + time0 + '"'  # 实际有效的输入上月第一天
        driver.execute_script(js_begin2_value)
        time.sleep(2)
        js_end = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-text.validatebox-text").removeAttribute("readonly");'
        driver.execute_script(js_end)
        js_end1_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-text.validatebox-text").value="' + time1 + '"'  # 表面输入上个月最后一天
        driver.execute_script(js_end1_value)
        js_end2_value = 'document.querySelector("#searchArea > tbody > tr:nth-child(1) > td:nth-child(2) > span > input.combo-value").value="' + time1 + '"'  # 实际有效的输入上个月最后一天
        driver.execute_script(js_end2_value)
    time.sleep(2)
    # button_a('/html/body/div[1]/div/div[1]/table/tbody/tr/td[1]/a/span/span')
    # print("稍等片刻......")
    # time.sleep(30)
    button_a('/html/body/div[1]/div/div[1]/table/tbody/tr/td[3]/a/span/span'
             )  # 导出数据
    time.sleep(3)
    print("文件导出成功！！！")
    driver.switch_to.parent_frame()  # 切出
    time.sleep(10)
    '''
    if aaa == "1":
        #button_a('//*[@class="l-btn-text icon-search l-btn-icon-left"]')#查询
        #time.sleep(60)
        #button_a('//*[@class="l-btn-text icon-page_excel l-btn-icon-left"]')#导出
        time.sleep(3)
        switch('/html/body/div[3]/div/div/div[2]/div[2]/div/iframe')  # 切入
        time.sleep(2)
        button_a('/html/body/div[1]/div/div[1]/table/tbody/tr/td[3]/a/span/span')  # 导出数据
        time.sleep(2)
        print("输出成功！！！")
        time.sleep(1)
        driver.switch_to.parent_frame() # 切出
        time.sleep(30)
    else:
        button_a('/html/body/form/fieldset/table/tbody/tr[1]/td[1]/span/span/span')#开始日期
        time.sleep(0.5)
        button_a('/html/body/div[3]/div/div[1]/div/div[1]/div[1]')
        time.sleep(0.5)
        button_a('//*[@abbr="'+ str(year) +','+ str(month) +','+ str(1) +'"]')
        time.sleep(0.5)
        button_a('/html/body/form/fieldset/table/tbody/tr[1]/td[2]/span/span/span')#结束日期
        time.sleep(0.5)
        button_a('/html/body/div[4]/div/div[1]/div/div[1]/div[1]')
        time.sleep(0.5)
        button_a('//*[@abbr="'+ str(year) +','+ str(month) +','+ str(lasttime.day) +'"]')
        time.sleep(0.5)
        #button_a('//*[@class="l-btn-text icon-search l-btn-icon-left"]')#查询
        #time.sleep(60)
        button_a('//*[@class="l-btn-text icon-page_excel l-btn-icon-left"]')#导出
        time.sleep(30)
        pass
    #driver.quit()
    '''


def DSL():  # 大参林
    global driver
    print('\n' + ">>>【大参林】数据爬取中,稍等片刻")
    wei_zhi = "D:\\FilesCenter\\大客户数据\\DSL\\"  # 大参林下载路径
    # deleteOldFiles(wei_zhi)  # 清空大参林文件夹文件
    options = webdriver.ChromeOptions()  # 打开设置
    options.add_argument("--user-data-dir=" + r"C:/Users/Zeus/AppData/Local/Google/Chrome/User Data/")  # 解决提示不安全内容下数据的下载
    prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': wei_zhi}  # 设置路径
    options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(executable_path='C:\\Program Files (x86)\\Google\\Chrome\\Application\\chromedriver.exe', options=options)  # 打开浏览器
    driver.set_window_size(1100, 800)
    driver.implicitly_wait(20)  # 隐式等待
    # driver.get('https://data.dslyy.com/chain/flow/downFlow')  # 接收不了验证码用这个网址
    driver.get('http://120.79.137.132/chain/login')  # 进入网址
    time.sleep(3)
    input_b('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[1]/div/div[1]/input', '13267854059')  # 手机号【数运中心邹德豪】
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[2]/div/div/div[2]/button/span')  # 获取验证码
    verificationCode = input(">> 输入手机验证码：")  # 输入接收的验证码
    input_b('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[2]/div/div/div[1]/div/input', verificationCode)  # 手机验证码
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div/div/div[2]/div/form/div/div[3]/div/button')  # 登录
    try:  # 选择公司下的账号
        time.sleep(2)
        button_a('/html/body/div[6]/div[2]/div/div/div[2]/div/div/div[1]/table/tbody/tr[1]/td/div/div/a[1]')
    except:
        pass
    time.sleep(2)
    button_a('/html/body/div[1]/div/div[1]/div[1]/div/ul/li[4]/div/span')  # 流向服务
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[1]/div[1]/div/ul/li[4]/ul/li[1]/span')  # 数据查询
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div/div/div/div[1]/div/span')  # 授权编号
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[1]/div/div/div/div[2]/ul[2]/li[2]')  # 门店销售
    time.sleep(1)
    input_b('//*[@id="app"]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[2]/div/div/div/div[1]/div/input', dateDSL)  # 查询日期
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[3]/div/div/button/span')  # 确认查询
    time.sleep(1)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div[1]/div[3]/div/div/button/span')  # 确认查询
    time.sleep(2)
    try:
        button_a('/html/body/div[12]/div[2]/div/div/div/div/div[3]/button[2]/span')  # 弹出框 确定
    except BaseException:
        pass
    time.sleep(3)
    button_a('/html/body/div[1]/div/div[1]/div[1]/div/ul/li[4]/ul/li[2]/span')  # 流向下载
    print(">  等待刷新查询中,请稍等片刻")
    time.sleep(200)  # 等待查询完成
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/form/div/div/div/div/button/span')  # 刷新
    time.sleep(2)
    button_a('/html/body/div[1]/div/div[2]/div[2]/div/div[2]/div/div/div[1]/div/div[2]/table/tbody/tr[1]/td[10]/div/button/span')  # 下载结果
    # 需要等待一段时间才能有下载文件的反应


# 0: 爬取数据 其它: 整理Excel到同一个文件夹下
operateType = input(">>>0Reptile-1TidyFiles:")

if operateType == '0':

    try:
        LBX()  # 老百姓
    except Exception as e:
        print("老百姓导出出错", e)

    try:
        YF()  # 益丰
    except Exception as e:
        print("益丰导出出错", e)

    try:
        HW()  # 海王
    except Exception as e:
        print("海王导出出错", e)

    try:
        SY()  # 漱玉
    except Exception as e:
        print("漱玉导出出错", e)

    try:
        QY()  # 全亿
    except Exception as e:
        print("全亿导出出错", e)

    try:
        GD()  # 国大
    except Exception as e:
        print("国大导出出错", e)

    try:
        GJ()  # 高济
    except Exception as e:
        print("高济导出出错", e)

    try:
        DSL()  # 大参林
    except Exception as e:
        print("大参林导出出错", e)

elif operateType == '1':

    try:  # 格式转换与文件迁移

        print(">>>【漱玉】格式转换中")
        mergeTableSY('D:\\FilesCenter\\大客户数据\\SY\\downloadSY\\', 'D:\\FilesCenter\\大客户数据\\SY\\xlsxFormatSY\\',
                     'D:\\FilesCenter\\大客户数据\\SY\\', 'D:\\FilesCenter\\EverydayUpDB\\')  # 漱玉

        print('\n' + ">>>【全亿】格式转换中")
        convFormat('D:\\FilesCenter\\大客户数据\\QY\\downloadQY\\', 'D:\\FilesCenter\\大客户数据\\QY\\', 'D:\\FilesCenter\\EverydayUpDB\\', 'QY')  # 全亿

        print('\n' + ">>>【国大】格式转换中")
        getFormat('D:\\FilesCenter\\大客户数据\\GD/downloadGD\\', 'D:\\FilesCenter\\大客户数据\\GD\\xlsFormatGD\\')  # 国大(csv -> xls)
        convFormat('D:\\FilesCenter\\大客户数据\\GD\\xlsFormatGD\\', 'D:\\FilesCenter\\大客户数据\\GD\\', 'D:\\FilesCenter\\EverydayUpDB\\', 'GD')  # 国大

        print('\n' + ">>>【高济】格式转换中")
        convFormat('D:\\FilesCenter\\大客户数据\\GJ\\downloadGJ\\', 'D:\\FilesCenter\\大客户数据\\GJ\\', 'D:\\FilesCenter\\EverydayUpDB\\', 'GJ')  # 高济

    except Exception as e:
        print("格式转换或文件迁移出错", e)

else:
    print('别无他选, 请输入0或1')