# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.3.2
@projectName   : pythonProjectPy3.9
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2021/01/27 20:36
'''
# ================================================================================
'''
目标：力争实现Excel全部功能
'''
# ================================================================================
# 导入模块
from openpyxl.comments import Comment
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle, GradientFill, Color
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows

# ================================================================================
# 一、创建工作簿/工作表
# ================================================================================
workbook = Workbook()  # write_only=True仅写模式,guess_types=True当读取单元格时,guess_types将启用或禁用（默认）类型推断(比如0.34 -> 34%)
# 获取当前活跃的worksheet
worksheet = workbook.active

# 创建子页及子页名称
worksheet.title = 'Sheet.NO1'  # 子页名称（默认第一个）
worksheet1 = workbook.create_sheet("Mysheet")  # insert at the end (default)
worksheet2 = workbook.create_sheet("Mysheet", 0)  # insert at first position
worksheet3 = workbook.create_sheet("Mysheet", -1)  # insert at the penultimate position

# 工作表在创建时会自动为其命名。 它们按顺序编号（表、表1、表2等）
worksheet.title = "New Title"

# 为工作表命名后，可以将其作为工作簿的键
worksheet3 = workbook["New Title"]

# 打印全部子页名称
print(workbook.sheetnames)

# 遍历工作表
for sheet in workbook:
    print(sheet.title)

# 在单个工作簿中创建工作表的副本
'''
仅复制单元格（包括值,样式,超链接和注释）和某些工作表属性（包括尺寸,格式和属性）.
不复制所有其他工作簿/工作表属性-例如图像,图表.
您不能在工作簿之间复制工作表.如果工作簿以只读或仅写模式打开,则您也无法复制工作表.
'''
source = workbook.active
target = workbook.copy_worksheet(source)

# ================================================================================
# 二、单元格访问 cell iter_roworksheet iter_cols
# ================================================================================
# 2.1 访问单个单元格
worksheet_A4 = workbook['A4']
# 返回 A4 处的单元格，如果它尚不存在，则创建一个
worksheet['A4'] = 4
# Worksheet.cell() 使用行和列表示法对单元格的访问
d = worksheet.cell(row=4, column=2, value=10)

# 2.2 访问多个单元格
cell_range = worksheet['A1':'C2']
# 行或列的范围可以类似地获取
colC = worksheet['C']
col_range = worksheet['C:D']
row10 = worksheet[10]
row_range = worksheet[5:10]

# Worksheet.iter_roworksheet() 访问多个单元格
for row in worksheet.iter_roworksheet(min_row=1, max_col=3, max_row=2):
    for cell in row:
        print(cell)
'''
<Cell Sheet1.A1>
<Cell Sheet1.B1>
<Cell Sheet1.C1>
<Cell Sheet1.A2>
<Cell Sheet1.B2>
<Cell Sheet1.C2>
'''

# Worksheet.iter_cols() 访问多个单元格【注意：出于性能原因，该方法在只读模式下不可用】
for col in worksheet.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)
'''
<Cell Sheet1.A1>
<Cell Sheet1.A2>
<Cell Sheet1.B1>
<Cell Sheet1.B2>
<Cell Sheet1.C1>
<Cell Sheet1.C2>
'''

# 2.3 循环访问文件的所有行或列
# 2.3.1 Worksheet.roworksheet
worksheet = workbook.active
worksheet['C9'] = 'hello world'
tuple(worksheet.roworksheet)
'''
((<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>),
(<Cell Sheet.A2>, <Cell Sheet.B2>, <Cell Sheet.C2>),
(<Cell Sheet.A3>, <Cell Sheet.B3>, <Cell Sheet.C3>),
(<Cell Sheet.A4>, <Cell Sheet.B4>, <Cell Sheet.C4>),
(<Cell Sheet.A5>, <Cell Sheet.B5>, <Cell Sheet.C5>),
(<Cell Sheet.A6>, <Cell Sheet.B6>, <Cell Sheet.C6>),
(<Cell Sheet.A7>, <Cell Sheet.B7>, <Cell Sheet.C7>),
(<Cell Sheet.A8>, <Cell Sheet.B8>, <Cell Sheet.C8>),
(<Cell Sheet.A9>, <Cell Sheet.B9>, <Cell Sheet.C9>))
'''

# 2.3.2 Worksheet.columns 【注意：出于性能原因，该方法在只读模式下不可用】
tuple(worksheet.columns)
'''
((<Cell Sheet.A1>,
<Cell Sheet.A2>,
<Cell Sheet.A3>,
<Cell Sheet.A4>,
<Cell Sheet.A5>,
<Cell Sheet.A6>,

<Cell Sheet.B7>,
<Cell Sheet.B8>,
<Cell Sheet.B9>),
(<Cell Sheet.C1>,
<Cell Sheet.C2>,
<Cell Sheet.C3>,
<Cell Sheet.C4>,
<Cell Sheet.C5>,
<Cell Sheet.C6>,
<Cell Sheet.C7>,
<Cell Sheet.C8>,
<Cell Sheet.C9>))
'''

# 2.4 获取值 Worksheet.values 访问工作表中的所有行，但仅返回单元格值
for row in worksheet.values:
    for value in row:
        print(value)

# Worksheet.iter_roworksheet() Worksheet.iter_cols() values_only
for row in worksheet.iter_roworksheet(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)

# 2.5 数据存储
# cell
c = worksheet.cell(row=1, column=2).value
c.value = 'hello, world'
print(c.value)
# 'hello, world'

d = worksheet.cell(row=2, column=1).value
d.value = 3.14
print(d.value)
# 3.14

# ================================================================================
# 三、设置样式 Font PatternFill Alignment Side Protection GradientFill
# ================================================================================
'''
一、单元格样式简介
openpyxl处理Excel文件中单元格样式，总共有六个属性类。
font(字体类----可设置字号、字体颜色、下划线等)
fill(填充类----可设置单元格填充颜色等)
border（边框类----可以设置单元格各种类型的边框）
alignment(位置类----可以设置单元格内数据各种对齐方式)
number_format(格式类----可以设置单元格内各种类型的数据格式)
protection(保护类----可以设置单元格写保护等)
'''
# 子页全局背景颜色
worksheet.sheet_properties.tabColor = "1072BA"  # 默认白色

number_format = 'General'

# 以下是默认值
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# 字体
font = Font(name='Calibri',  # 字体名称
            size=11,  # 字体大小
            bold=False,  # 加粗-True/False
            italic=False,  # 斜体-True/False
            vertAlign=None,  # 对其方向-
            underline='none',  # 下划线-single/double
            strike=False,
            color='FF000000')
# 文本方向
align = Alignment(horizontal='center', vertical='center')
# 方向参数：general bottom top left right center
alignment = Alignment(horizontal='general',
                      vertical='bottom',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)

# 保护类
protection = Protection(locked=True, hidden=False)

# 单元格边框
# thin 单线; double 双线
bottom_thin = Side(border_style="double", color="000000")  # 下
left_thin = Side(border_style="thin", color="000000")  # 左
top_double = Side(border_style="double", color="000000")  # 上
right_double = Side(border_style="thin", color="000000")  # 右
border = Border(bottom=left_thin, left=left_thin, top=left_thin, right=left_thin)  # 边框样式
# thin double thick
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                diagonal=Side(border_style='thin', color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style='thin', color='FF000000'),
                vertical=Side(border_style='thin', color='FF000000'),
                horizontal=Side(border_style='thin', color='FF000000')
                )

# 颜色填充
title_fill = PatternFill(fill_type='solid', fgColor="F8F8FF")
# fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
interval_fill = GradientFill(stop=("000000", "FFFFFF"))  # 连续区间底色填充
content_font = Font(name='宋体', size=14, bold=False, color="000000")
content_fill = PatternFill(fill_type='solid', fgColor="F8F8FF")

# 复制样式
from openpyxl.styles import Font
from copy import copy

ft1 = Font(name='Arial', size=14)
ft2 = copy(ft1)
ft2.name = "Tahoma"
ft1.name
'Arial'
ft2.name
'Tahoma'
ft2.size  # copied from the
14.0

# 内容填充及字段属性设置【注意：表格cell的最低行列值都是从1开始,不是0】
for rowi in range(3, 8):
    for colj in range(1, 10):
        # _ = worksheet.cell(row=4, column=2, value=10)
        worksheet.cell(row=rowi, column=colj).value = rowi + colj

# 插入图像
from openpyxl.drawing.image import Image

image = Image('C:/Users/Zeus/Desktop/新logo.png')
worksheet.add_image(image, 'A1')  # A1为插入图片起始单元格

from openpyxl import Workbook
from openpyxl.drawing.image import Image

workbook = Workbook()
worksheet = workbook.active
worksheet['A1'] = 'You should see three logos below'
# create an image
img = Image('logo.png')
# add to worksheet and anchor next to cells
worksheet.add_image(img, 'A1')
workbook.save('logo.xlsx')

# 局部设置样式
title_font = Font(name='黑体', size=16, bold=True, color="000000")
for row in worksheet.roworksheet:
    for cell in row:
        cell.alignment = align
        if cell.row == 1 or cell.row == 2:
            cell.border = border
            cell.font = title_font
            cell.fill = title_fill
        else:
            cell.border = border
            cell.font = content_font
            cell.fill = content_fill

# 样式也可以应用于列和行
col = worksheet.column_dimensions['A']
col.font = Font(bold=True)
row = worksheet.row_dimensions[1]
row.font = Font(underline="single")

# 传统的索引颜色以及主题和色调
c = Color(indexed=32)  # 按索引选颜色
c = Color(theme=6, tint=0.5)  # 使用主题

# 使用数字格式
import datetime
from openpyxl import Workbook

workbook = Workbook()
worksheet = workbook.active
# set date using a Python datetime
worksheet['A1'] = datetime.datetime(2010, 7, 21)

worksheet['A1'].number_format
'yyyy-mm-dd h:mm:ss'

worksheet["A2"] = 0.123456
worksheet["A2"].number_format = "0.00"  # Display to 2dp

# 编辑页面设置
worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
worksheet.page_setup.paperSize = worksheet.PAPERSIZE_TABLOID
worksheet.page_setup.fitToHeight = 0
worksheet.page_setup.fitToWidth = 1

# 创建命名样式
highlight = NamedStyle(name="highlight")
workbook.add_named_style(highlight)
worksheet['A1'].style = highlight

# ================================================================================
# 四、使用富文本 TextBlock InlineFont CellRichText
# ================================================================================

from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# 简单例子
rich_string1 = CellRichText(
    'This is a test ',
    TextBlock(InlineFont(b=True), 'xxx'),
    'yyy'
)

# InlineFont() 默认参数
inline_font = InlineFont(rFont='Calibri',  # Font name
                         sz=22,  # in 1/144 in. (1/2 point) units, must be integer
                         charset=None,  # character set (0 to 255), less required with UTF-8
                         family=None,  # Font family
                         b=True,  # Bold (True/False)
                         i=None,  # Italics (True/False)
                         strike=None,  # strikethrough
                         outline=None,
                         shadow=None,
                         condense=None,
                         extend=None,
                         color=None,
                         u=None,
                         vertAlign=None,
                         scheme=None,
                         )

# 简单地使用现有对象初始化对象
from openpyxl.cell.text import Font

font = Font(name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='00FF0000')
inline_font = InlineFont(font)

# 自行创建对象，并在以后使用它们。这使得使用富文本更清晰、更容易：InlineFont
big = InlineFont(sz="30.0")
medium = InlineFont(sz="20.0")
small = InlineFont(sz="10.0")
bold = InlineFont(b=True)
b = TextBlock
rich_string2 = CellRichText(
    b(big, 'M'),
    b(medium, 'i'),
    b(small, 'x'),
    b(medium, 'e'),
    b(big, 'd')
)

# 例如
red = InlineFont(color='FF000000')
rich_string1 = CellRichText(['When the color ', TextBlock(red, 'red'), ' is used, you can expect ', TextBlock(red, 'danger')])

# 编辑富文本
# 由于编辑具有格式的大型文本块可能很棘手，因此 as_list（）方法返回字符串列表以使索引变得容易
l = rich_string1.as_list()
l
['When the color ', 'red', ' is used, you can expect ', 'danger']
l.index("danger")
3
rich_string1[3].text = "fun"
str(rich_string1)
'When the color red is used, you can expect fun'

# 单元格的富文本赋值
from openpyxl import Workbook

workbook = Workbook()
worksheet = workbook.active
worksheet['A1'] = rich_string1
worksheet['A2'] = 'Simple string'

# ================================================================================
# 四、条件格式 色标ColorScale 图标集IconSet 数据栏DataBars
# ================================================================================
# 创建格式规则的基本语法
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle

dxf = DifferentialStyle(font=Font(bold=True), fill=PatternFill(start_color='EE1111', end_color='EE1111'))
rule = Rule(type='cellIs', dxf=dxf, formula=["10"])

# 内置格式
'''
内置条件格式为：
    色标
    图标集
    数据栏
内置格式包含一系列格式设置，这些设置将类型与整数组合在一起进行比较。
可能的类型包括：“数字”、“百分比”、“最大值”、“最小值”、“公式”、“百分位数”。
'''
# 色标 ColorScale
'''
您可以使用 2 或 3 种颜色的色阶。
2 个色标产生从一种颜色到另一种颜色的渐变;3 个色阶对 2 个渐变使用额外的颜色。
'''
# 创建 ColorScale 规则的完整语法
from openpyxl.formatting.rule import ColorScale, FormatObject
from openpyxl.styles import Color

first = FormatObject(type='min')
last = FormatObject(type='max')
# colors match the format objects:
colors = [Color('AA0000'), Color('00AA00')]
cs2 = ColorScale(cfvo=[first, last], color=colors)
# a three color scale would extend the sequences
mid = FormatObject(type='num', val=40)
colors.insert(1, Color('00AA00'))
cs3 = ColorScale(cfvo=[first, mid, last], color=colors)
# create a rule with the color scale
from openpyxl.formatting.rule import Rule

rule = Rule(type='colorScale', colorScale=cs3)

# 有一个方便的功能来创建色阶规则
from openpyxl.formatting.rule import ColorScaleRule

rule = ColorScaleRule(start_type='percentile', start_value=10, start_color='FFAA0000',
                      mid_type='percentile', mid_value=50, mid_color='FF0000AA',
                      end_type='percentile', end_value=90, end_color='FF00AA00')

# 图标集 IconSet
'''
图标集类型：
‘3Arroworksheet’, ‘3ArroworksheetGray’, ‘3Flags’, ‘3TrafficLights1’, ‘3TrafficLights2’, 
‘3Signs’, ‘3Symbols’, ‘3Symbols2’, ‘4Arroworksheet’, ‘4ArroworksheetGray’, ‘4RedToBlack’, 
‘4Rating’, ‘4TrafficLights’, ‘5Arroworksheet’, ‘5ArroworksheetGray’, ‘5Rating’, ‘5Quarters’
'''
# 创建图标集规则的完整语法
from openpyxl.formatting.rule import IconSet, FormatObject

first = FormatObject(type='percent', val=0)
second = FormatObject(type='percent', val=33)
third = FormatObject(type='percent', val=67)
iconset = IconSet(iconSet='3TrafficLights1', cfvo=[first, second, third], showValue=None, percent=None, reverse=None)
# assign the icon set to a rule
from openpyxl.formatting.rule import Rule

rule = Rule(type='iconSet', iconSet=iconset)

# 用于创建图标集规则的便捷功能
from openpyxl.formatting.rule import IconSetRule

rule = IconSetRule('5Arroworksheet', 'percent', [10, 20, 30, 40, 50], showValue=None, percent=None, reverse=None)

# 数据栏 DataBars
# 创建数据栏规则的完整语法
from openpyxl.formatting.rule import DataBar, FormatObject

first = FormatObject(type='min')
second = FormatObject(type='max')
data_bar = DataBar(cfvo=[first, second], color="638EC6", showValue=None, minLength=None, maxLength=None)
# assign the data bar to a rule
from openpyxl.formatting.rule import Rule

rule = Rule(type='dataBar', dataBar=data_bar)

# 用于创建数据栏规则的便捷功能
from openpyxl.formatting.rule import DataBarRule

rule = DataBarRule(start_type='percentile', start_value=10, end_type='percentile', end_value='90',
                   color="FF638EC6", showValue="None", minLength=None, maxLength=None)

# 标准条件格式
'''
标准条件格式为：
    平均 Average
    百分之几 Percent
    唯一或重复 Unique or duplicate
    数值 Value
    排 Rank
'''
from openpyxl import Workbook
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, IconSetRule

workbook = Workbook()
worksheet = workbook.active

# 制作数据
for irow in range(1, 50):
    for icol in range(1, 20):
        worksheet.cell(row=irow, column=icol, value=irow * icol)

# 图标
worksheet.conditional_formatting.add('K1:S49',
                                     IconSetRule('5Arroworksheet', 'percent', [10, 20, 30, 40, 50], showValue=None, percent=None,
                                                 reverse=None))

# Create fill
redFill = PatternFill(start_color='EE1111',
                      end_color='EE1111',
                      fill_type='solid')

# Add a two-color scale
# Takes colors in excel 'RRGGBB' style.
worksheet.conditional_formatting.add('A1:A10',
                                     ColorScaleRule(start_type='min', start_color='AA0000',
                                                    end_type='max', end_color='00AA00')
                                     )

# Add a three-color scale
worksheet.conditional_formatting.add('B1:B10',
                                     ColorScaleRule(start_type='percentile', start_value=10, start_color='AA0000',
                                                    mid_type='percentile', mid_value=50, mid_color='0000AA',
                                                    end_type='percentile', end_value=90, end_color='00AA00')
                                     )

# Add a conditional formatting based on a cell comparison
# addCellIs(range_string, operator, formula, stopIfTrue, workbook, font, border, fill)
# Format if cell is less than 'formula'
worksheet.conditional_formatting.add('C2:C10',
                                     CellIsRule(operator='lessThan', formula=['C$1'], stopIfTrue=True, fill=redFill))

# Format if cell is between 'formula'
worksheet.conditional_formatting.add('D2:D10',
                                     CellIsRule(operator='between', formula=['1', '5'], stopIfTrue=True, fill=redFill))

# Format using a formula
worksheet.conditional_formatting.add('E1:E10',
                                     FormulaRule(formula=['ISBLANK(E1)'], stopIfTrue=True, fill=redFill))

# Aside from the 2-color and 3-color scales, format rules take fonts, borders and fills for styling:
myFont = Font()
myBorder = Border()
worksheet.conditional_formatting.add('E1:E10',
                                     FormulaRule(formula=['E1=0'], font=myFont, border=myBorder, fill=redFill))

# Highlight cells that contain particular text by using a special formula
red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill)
rule = Rule(type="containsText", operator="containsText", text="highlight", dxf=dxf)
rule.formula = ['NOT(ISERROR(SEARCH("highlight",A1)))']
worksheet.conditional_formatting.add('A1:F40', rule)
workbook.save("test.xlsx")

# 设置整行的格式 按行填充数据
worksheet.append(['Software', 'Developer', 'Version'])
worksheet.append(['Excel', 'Microsoft', '2016'])
worksheet.append(['openpyxl', 'Open source', '2.6'])
worksheet.append(['OpenOffice', 'Apache', '4.1.4'])
worksheet.append(['Word', 'Microsoft', '2010'])

# 使用微软公式
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(fill=red_fill)
r = Rule(type="expression", dxf=dxf, stopIfTrue=True)
r.formula = ['$A2="Microsoft"']
workbook.conditional_formatting.add("A1:C10", r)

# ================================================================================
# 五、插入和删除行和列、移动单元格范围、单元格合并拆分
# ================================================================================
'''
使用相关的工作表方法插入行或列：
    openpyxl.worksheet.worksheet.Worksheet.insert_roworksheet（）
    openpyxl.worksheet.worksheet.Worksheet.insert_cols（）
    openpyxl.worksheet.worksheet.Worksheet.delete_roworksheet（）
    openpyxl.worksheet.worksheet.Worksheet.delete_cols（）
'''

# 在第七行插入
worksheet.insert_roworksheet(7)

# 删除列
worksheet.delete_cols(6, 3)

# 移动单元格区域 单元格向上移动一行，向右移动两行 列。单元格将覆盖任何现有单元格
worksheet.move_range("D4:F10", roworksheet=-1, cols=2)
# 单元格内包含数学公式 使用translate=True
worksheet.move_range("G4:H10", roworksheet=1, cols=1, translate=True)

# 合并单元格
worksheet.merge_cells('A1:I2')  # 合并单元格,保留左上角值,其他删除
worksheet.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)  # 合并单元格第二种方法
worksheet.cell(row=1, column=1).value = '合并的单元格内容'
# 拆分单元格
worksheet.unmerge_cells('A1:I2')  # 拆分合并的单元格,值值保留在左上角
worksheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4)  # 拆分合并的单元格第二种方法

# ================================================================================
# 六、其他工作表属性
# ================================================================================
'''
工作表的可用属性
    “启用格式条件计算” “enableFormatConditionsCalculation”
    “过滤模式” “filterMode”
    “已发布” “published”
    “同步水平” “syncHorizontal”
    “同步参考” “syncRef”
    “同步垂直” “syncVertical”
    “过渡评估” “transitionEvaluation”
    “过渡条目” “transitionEntry”
    “标签颜色” “tabColor”
页面设置属性的可用字段
    “自动分页” “适合页面” “autoPageBreaks” “fitToPage”
大纲的可用字段
    “应用样式” “applyStyles”
    “摘要下面” “summaryBelow”
    “总结对” “summaryRight”
    “显示大纲符号” “showOutlineSymbols”
'''
from openpyxl.workbook import Workbook
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

workbook = Workbook()
worksheet = workbook.active

worksheetprops = worksheet.sheet_properties
worksheetprops.tabColor = "1072BA"
worksheetprops.filterMode = False
worksheetprops.pageSetUpPr = PageSetupProperties(fitToPage=True, autoPageBreaks=False)
worksheetprops.outlinePr.summaryBelow = False
worksheetprops.outlinePr.applyStyles = True
worksheetprops.pageSetUpPr.autoPageBreaks = True

# 工作表视图
from openpyxl.workbook import Workbook

workbook = Workbook()
worksheet = workbook.active

worksheet.sheet_view.zoom = 85  # Sets 85% zoom
worksheet.sheet_view.showFormulas = True
worksheet.sheet_view.tabSelected = True

# 折叠行列
import openpyxl

workbook = openpyxl.Workbook()
worksheet = workbook.create_sheet()
# outline_level显示折叠轮廓线(默认1), hidden直接隐藏掉折叠的行列(默认True)
worksheet.row_dimensions.group(1, 3, outline_level=1, hidden=True)
worksheet.column_dimensions.group('A', 'D', outline_level=1, hidden=False)
workbook.save('group.xlsx')

# ================================================================================
# 七、验证单元格
# ================================================================================
# 例子
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Create the workbook and worksheet we'll be working with
workbook = Workbook()
worksheet = workbook.active

# Create a data-validation object with list validation
dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)

# Optionally set a custom error message
dv.error = 'Your entry is not in the list'
dv.errorTitle = 'Invalid Entry'

# Optionally set a custom prompt message
dv.prompt = 'Please select from the list'
dv.promptTitle = 'List Selection'

# Add the data-validation object to the worksheet
worksheet.add_data_validation(dv)

# Create some cells, and add them to the data-validation object
c1 = worksheet["A1"]
c1.value = "Dog"
dv.add(c1)
c2 = worksheet["A2"]
c2.value = "An invalid value"
dv.add(c2)

# Or, apply the validation to a range of cells
dv.add('B1:B1048576')  # This is the same as for the whole of column B

# Check with a cell is in the validator
"B4" in dv
True

# 保存工作簿时，将忽略没有任何单元格区域的验证.

# 其他验证示例
# 任何整数
dv = DataValidation(type="whole")

# 任何大于 100 的整数
dv = DataValidation(type="whole",
                    operator="greaterThan",
                    formula1=100)

# 任何十进制数
dv = DataValidation(type="decimal")

# 0 到 1 之间的任何十进制数
dv = DataValidation(type="decimal",
                    operator="between",
                    formula1=0,
                    formula2=1)

# 任何日期或时间
dv = DataValidation(type="date")
dv = DataValidation(type="time")

# 任何最多 15 个字符的字符串
dv = DataValidation(type="textLength",
                    operator="lessThanOrEqual",
                    formula1=15)

# 单元格范围验证
from openpyxl.utils import quote_sheetname

dv = DataValidation(type="list",
                    formula1="{0}!$B$1:$B$10".format(quote_sheetname('sheetname'))
                    )

# 自定义规则
dv = DataValidation(type="custom",
                    formula1="SOMEFORMULA")

# ================================================================================
# 八、工作表表格
# ================================================================================
# 创建表
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

workbook = Workbook()
worksheet = workbook.active

data = [
    ['Apples', 10000, 5000, 8000, 6000],
    ['Pears', 2000, 3000, 4000, 5000],
    ['Bananas', 6000, 6000, 6500, 6000],
    ['Oranges', 500, 300, 200, 700],
]

# add column headings. NB. these must be strings
worksheet.append(["Fruit", "2011", "2012", "2013", "2014"])
for row in data:
    worksheet.append(row)

tab = Table(displayName="Table1", ref="A1:E5")

# Add a default style with striped roworksheet and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                       showLastColumn=False, showRoworksheettripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

'''
Table must be added using worksheet.add_table() method to avoid duplicate names.
Using this method ensures table name is unque through out defined names and all other table name. 
'''
worksheet.add_table(tab)
workbook.save("table.xlsx")

# 使用表
# worksheet.tables是特定工作表中所有表的类似字典的对象
worksheet.tables
# {"Table1",  <openpyxl.worksheet.table.Table object>}

# 按名称或范围获取表
worksheet.tables["Table1"]
worksheet.tables["A1:D10"]

# 循环访问工作表中的所有表
for table in worksheet.tables.values():
    print(table)

# 获取工作表中所有表的表名称和范围
worksheet.tables.items()
[("Table1", "A1:D10")]

# 删除表
del worksheet.tables["Table1"]

# 工作表中的表数
len(worksheet.tables)

# 手动添加列标题
# 在只写模式下，您只能添加没有标题的表
table.headerRowCount = False

# 手动初始化列标题
headings = ["Fruit", "2011", "2012", "2013", "2014"]  # all values must be strings
table._initialise_columns()
for column, value in zip(table.tableColumns, headings):
    column.name = value

# 过滤 器
# 筛选器将自动添加到包含标题行的表中。无法创建带有没有筛选器的标题行的表。

# 表作为打印区域
from openpyxl import load_workbook

workbook = load_workbook("QueryTable.xlsx")
worksheet = workbook.active
table_range = worksheet.tables["InvoiceData"]
worksheet.print_area = table_range.ref  # Ref is the cell range the table currently covers

# ================================================================================
# 九、使用筛选器和排序
# ================================================================================
from openpyxl import Workbook
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
)

workbook = Workbook()
worksheet = workbook.active

data = [
    ["Fruit", "Quantity"],
    ["Kiwi", 3],
    ["Grape", 15],
    ["Apple", 3],
    ["Peach", 3],
    ["Pomegranate", 3],
    ["Pear", 3],
    ["Tangerine", 3],
    ["Blueberry", 3],
    ["Mango", 3],
    ["Watermelon", 3],
    ["Blackberry", 3],
    ["Orange", 3],
    ["Raspberry", 3],
    ["Banana", 3]
]

for r in data:
    worksheet.append(r)

filters = worksheet.auto_filter
filters.ref = "A1:B15"
col = FilterColumn(colId=0)  # for column A
col.filters = Filters(filter=["Kiwi", "Apple", "Mango"])  # add selected values
filters.filterColumn.append(col)  # add filter to the worksheet

worksheet.auto_filter.add_sort_condition("B2:B15")

# 这会将相关说明添加到文件中，但实际上不会过滤或排序。
workbook.save("filtered.xlsx")

# 高级过滤器 CustomFilter, DateGroupItem, DynamicFilter, ColorFilter, IconFilter and Top10ColorFilter, IconFilter and Top10
# 所有筛选器都与条件格式交互

# 自定义过滤器
# 筛选器运算符 and_'equal', 'lessThan', 'lessThanOrEqual', 'notEqual', 'greaterThanOrEqual', 'greaterThan'

# 筛选值< 10 和 > 90
from openpyxl.worksheet.filters import CustomFilter, CustomFilters

flt1 = CustomFilter(operator="lessThan", val=10)
flt2 = CustomFilter(operator='greaterThan', val=90)

cfs = CustomFilters(customFilter=[flt1, flt2])
col = FilterColumn(colId=2, customFilters=cfs)  # apply to **third** column in the range
filters.filter.append(col)

# 要组合过滤器
cfs.and_ = True

# 运算符和通配符实现的。equalnotEqual
# 对于“以 a 开头”，请使用 ;对于“以 a 结尾”，请使用 ;对于“包含 a”，请使用 。a**a*a*

# 日期组项
from openpyxl.worksheet.filters import DateGroupItem

df1 = DateGroupItem(month=3, dateTimeGrouping="month")
col = FilterColumn(colId=1)  # second column
col.filters.dateGroupItem.append(df1)

df2 = DateGroupItem(year=1984, dateTimeGrouping="year")  # add another element
col.filters.dateGroupItem.append(df2)
filters.filter.append(col)

# ================================================================================
# 十、打印设置
# ================================================================================
# 编辑打印选项
from openpyxl.workbook import Workbook

workbook = Workbook()
worksheet = workbook.active
worksheet.print_options.horizontalCentered = True
worksheet.print_options.verticalCentered = True

# 页眉和页脚 oddHeader evenHeader evenFooter firstHeader firstFooter
from openpyxl.workbook import Workbook

workbook = Workbook()
worksheet = workbook.active
worksheet.oddHeader.left.text = "Page &[Page] of &N"
worksheet.oddHeader.left.size = 14
worksheet.oddHeader.left.font = "Tahoma,Bold"
worksheet.oddHeader.left.color = "CC3366"

# 添加打印标题
from openpyxl.workbook import Workbook

workbook = Workbook()
worksheet = workbook.active
worksheet.print_title_cols = 'A:B'  # the first two cols
worksheet.print_title_roworksheet = '1:1'  # the first row

# 添加打印区域
from openpyxl.workbook import Workbook

workbook = Workbook()
worksheet = workbook.active
worksheet.print_area = 'A1:F10'

# 更改页面布局和大小
workbook = Workbook()
worksheet = workbook.active
worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A5

# ================================================================================
# 十一、数据透视表
# ================================================================================
# 例子
from openpyxl import load_workbook

workbook = load_workbook("campaign.xlsx")
worksheet = workbook["Results"]
pivot = worksheet._pivots[0]  # any will do as they share the same cache
pivot.cache.refreshOnLoad = True

# ================================================================================
# 十二、评论
# ================================================================================
# 向单元格添加批注
from openpyxl import Workbook
from openpyxl.comments import Comment

workbook = Workbook()
worksheet = workbook.active
comment = worksheet["A1"].comment
comment = Comment('This is the comment text', 'Comment Author')
comment.text
'This is the comment text'
comment.author
'Comment Author'

# 如果将相同的注释分配给多个单元格，则 openpyxl 将自动创建副本
from openpyxl import Workbook
from openpyxl.comments import Comment

workbook = Workbook()
worksheet = workbook.active
comment = Comment("Text", "Author")
worksheet["A1"].comment = comment
worksheet["B2"].comment = comment
worksheet["A1"].comment is comment
True
worksheet["B2"].comment is comment
False

# 加载和保存注释
# 指定注释维度。注释维度为 以像素为单位
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.utils import units

workbook = Workbook()
worksheet = workbook.active
comment = Comment("Text", "Author")
comment.width = 300
comment.height = 50
worksheet["A1"].comment = comment
workbook.save('commented_book.xlsx')

# 如果需要，包含用于转换的帮助程序函数 从其他测量值（如毫米或点到像素）：openpyxl.utils.units
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.utils import units

workbook = Workbook()
worksheet = workbook.active
comment = Comment("Text", "Author")
comment.width = units.points_to_pixels(300)
comment.height = units.points_to_pixels(50)
worksheet["A1"].comment = comment

# ================================================================================
# 十三、日期和时间
# ================================================================================
# 使用 ISO 8601 格式
import openpyxl

workbook = openpyxl.Workbook()
workbook.iso_dates = True

# 获取工作簿的日期系统
import openpyxl

workbook = openpyxl.Workbook()
# workbook.epoch = openpyxl.utils.datetime.CALENDAR_MAC_1904
if workbook.epoch == openpyxl.utils.datetime.CALENDAR_WINDOworksheet_1900:
    print("This workbook is using the 1900 date system.")
'This workbook is using the 1900 date system.'

# ================================================================================
# 十四、简单形式
# ================================================================================
# 用公式
from openpyxl import Workbook

workbook = Workbook()
worksheet = workbook.active
# add a simple formula
worksheet["A1"] = "=SUM(1, 1)"
workbook.save("formula.xlsx")

# Openpyxl 从不计算公式，但可以检查公式的名称
from openpyxl.utils import FORMULAE

"HEX2DEC" in FORMULAE
True

# 数组公式
from openpyxl import Workbook
from openpyxl.worksheet.formula import ArrayFormula

workbook = Workbook()
worksheet = workbook.active
worksheet["E2"] = ArrayFormula("E2:E11", "=SUM(C2:C11*D2:D11)")

# ================================================================================
# 十五、定义的名称
# ================================================================================
# 访问全局定义
defn = workbook.defined_names["my_range"]
# the destinations attribute contains a list of ranges in the definitions
dests = defn.destinations  # returns a generator of (worksheet title, cell range) tuples

cells = []
for title, coord in dests:
    worksheet = workbook[title]
    cells.append(worksheet[coord])

# 访问工作表定义
worksheet = workbook["Sheet"]
defn = worksheet.defined_names["private_range"]

# 创建全局定义
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname, absolute_coordinate

workbook = Workbook()
worksheet = workbook.active
# make sure sheetnames and cell references are quoted correctly
ref = "{quote_sheetname(worksheet.title)}!{absolute_coordinate('A1:A5')}"

defn = DefinedName("global_range", attr_text=ref)
workbook.defined_names["global_range"] = defn

# key and `name` must be the same, the `.add()` method makes this easy
workbook.defined_names.add('new_range')

# 创建工作表定义
# create a local named range (only valid for a specific sheet)
worksheet = workbook["Sheet"]
worksheet.title = "My Sheet"
# make sure sheetnames  and cell referencesare quoted correctly
ref = f"{quote_sheetname(worksheet.title)}!{absolute_coordinate('A6')}"

defn = DefinedName("private_range", attr_text=ref)
worksheet.defined_names.add(defn)
print(worksheet.defined_names["private_range"].attr_text)

# 动态命名区域
from openpyxl import load_workbook

workbook = load_workbook("Example.xlsx")
worksheet = workbook.active
area = worksheet.defined_names["TestArea"]  # Globally defined named ranges can be used too
worksheet.print_area = area.value  # value is the cell range the defined name currently covers

# ================================================================================
# 十六、自定义文档属性
# ================================================================================
'''
可以将一个或多个自定义文档属性对象添加到工作簿中。 这些需要唯一的名称（字符串），可以是以下 6 种类型之一：
    字符串属性 StringProperty
    英特属性 IntProperty
    浮子属性 FloatProperty
    日期时间属性 DateTimeProperty
    布尔属性 BoolProperty
    链接属性 LinkProperty
链接属性始终与定义的名称范围相关联。
这些属性是工作簿的全局属性，可从 custom_doc_props 属性访问。
'''
# 样品使用
# 遍历所有自定义属性（“custom_doc_props”）
for prop in workbook.custom_doc_props.props:
    print(f"{prop.name}: {prop.value}")

# 添加新属性
from openpyxl.packaging.custom import (
    BoolProperty,
    DateTimeProperty,
    FloatProperty,
    IntProperty,
    LinkProperty,
    StringProperty,
    CustomPropertyList,
)

props = CustomPropertyList()
props.append(StringProperty(name="PropName1", value="Something"))

# 删除属性
workbook.custom_doc_props.append(StringProperty(name="PropName6", value="Something"))
# check the property
prop = workbook.custom_doc_props["PropName6"]
# delete the string property:
del prop["PropName6"]
# save the file
workbook.save('outfile.xlsx')

# ================================================================================
# 十七、保护 Protection
# ================================================================================
# 工作簿保护
# openpyxl.workbook.protection.WorkbookProtection.workbookPassword（）
workbook.security.workbookPassword = '...'
workbook.security.lockStructure = True

# openpyxl.workbook.protection.WorkbookProtection.revisionsPassword()
workbook.security.revisionsPassword = '...'

# 如果您需要设置原始密码值而不使用 默认哈希算法 - 例如
hashed_password = ...
workbook.security.set_workbook_password(hashed_password, already_hashed=True)

# 工作簿保护
workbook.security.workbookPassword = ''  # 防止其他用户查看隐藏的工作表,添加,移动,删除或隐藏工作表以及重命名工作表
worksheet.security.revisionsPassword = ''  # 防止从共享工作簿中删除更改跟踪和更改历史记录
workbook.security.set_workbook_password(hashed_password='', already_hashed=True)  # 如果需要在不使用默认哈希算法的情况下设置原始密码值，则提供特定的设置器功能

# 工作表保护
# 如果未指定密码,则用户无需指定密码即可禁用配置的工作表保护。否则,他们必须提供密码才能更改已配置的保护。
workbook.protection.sheet = True  # 设置属性来锁定工作表的各个方面
workbook.protection.password = ''  # 属性设置密码

# 工作表保护 openpxyl.worksheet.protection.SheetProtection.sheet
worksheet = workbook.active
worksheet.protection.sheet = True
worksheet.protection.enable()
worksheet.protection.disable()

# openpxyl.worksheet.protection.SheetProtection.password()
worksheet = workbook.active
worksheet.protection.password = '...'

# ================================================================================
# 十八、Pandas和NumPy合作
# ================================================================================
# 使用 Pandas Dataframes
# openpyxl.utils.dataframe.dataframe_to_roworksheet（）
from openpyxl.utils.dataframe import dataframe_to_rows

workbook = Workbook()
worksheet = workbook.active

for r in dataframe_to_rows('df', index=True, header=True):
    worksheet.append(r)

# 将dataframe转换为突出显示标题和索引的工作表
workbook = Workbook()
worksheet = workbook.active

for r in dataframe_to_rows('df', index=True, header=True):
    worksheet.append(r)

for cell in worksheet['A'] + worksheet[1]:
    cell.style = 'Pandas'

workbook.save("pandas_openpyxl.xlsx")

# 想转换数据，则可以使用只写模式
from openpyxl.cell.cell import WriteOnlyCell

workbook = Workbook(write_only=True)
worksheet = workbook.create_sheet()

cell = WriteOnlyCell(worksheet)
cell.style = 'Pandas'


def format_first_row(row, cell):
    for c in row:
        cell.value = c
        yield cell


roworksheet = dataframe_to_rows('df')
first_row = format_first_row(next(roworksheet), cell)
worksheet.append(first_row)

for row in roworksheet:
    row = list(row)
    cell.value = row[0]
    row[0] = cell
    worksheet.append(row)

workbook.save("openpyxl_stream.xlsx")

# 将工作表转换为dataframe
import pandas as pd

df = pd.DataFrame(worksheet.values)

# 如果工作表确实有标题或索引，例如由 Pandas 创建的标题或索引， 然后需要做更多的工作
from itertools import islice

data = worksheet.values
cols = next(data)[1:]
data = list(data)
idx = [r[0] for r in data]
data = (islice(r, 1, None) for r in data)
df = pd.DataFrame(data, index=idx, columns=cols)

# ================================================================================
# 十九、优化模式
# ================================================================================
# 只读模式
# openpyxl.worksheet._read_only.ReadOnlyWorksheet
from openpyxl import load_workbook

workbook = load_workbook(filename='large_file.xlsx', read_only=True)
worksheet = workbook['big_data']

for row in worksheet.roworksheet:
    for cell in row:
        print(cell.value)

# Close the workbook after reading
workbook.close()

# 返回的单元格不是常规的 openpyxl.cell.cell.cell.Cell，
# openpyxl.cell._read_only.ReadOnlyCell
# 工作表尺寸
worksheet.reset_dimensions()

# 只写模式
# openpyxl.worksheet._write_only.WriteOnlyWorksheet
from openpyxl import Workbook

workbook = Workbook(write_only=True)
worksheet = workbook.create_sheet()

# now we'll fill it with 100 roworksheet x 200 columns
for irow in range(100):
    worksheet.append(['%d' % i for i in range(200)])
# save the file
workbook.save('new_big_file.xlsx')  # doctest: +SKIP

# 如果要具有带有样式或注释的单元格，请使用openpyxl.cell.WriteOnlyCell()
from openpyxl import Workbook

workbook = Workbook(write_only=True)
worksheet = workbook.create_sheet()
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

cell = WriteOnlyCell(worksheet, value="hello world")
cell.font = Font(name='Courier', size=36)
cell.comment = Comment(text="A comment", author="Author's Name")
worksheet.append([cell, 3.14, None])
workbook.save('write_only_file.xlsx')

# ================================================================================
# 二十、文件操作
# ================================================================================
# 20.1 保存到文件
workbook.save(filename='C:/Users/Zeus/Desktop/openpyxl.xlsx')  # 储存格式：xlsx、xlsm或zip

'''
# 读取与储存文件
# workbook = load_workbook(filename='large_file.xlsx', read_only=True) # 仅读模式
workbook = load_workbook('document.xlsx')
# 当读取单元格时,guess_types将启用或禁用（默认）类型推断.
# read_only仅读取模式,不可改动
# data_only控制具有公式的单元格是否具有公式(默认值)或Excel上次读取工作表时存储的值.
# keep_vba控制是否保留任何Visual Basic元素(默认).如果保留它们,则它们仍不可编辑.

print(workbook.get_sheet_names())
workbook.template = True  # template = True,以将工作簿另存为模板;
workbook.save('document_template.xltx')  # 注意为xltx格式
workbook = load_workbook('document_template.xltx')
workbook.template = False  # 将此属性设置为 template =False(默认)以另存为文档
workbook.save('document.xlsx', as_template=False)
'''

# 20.2 另存为流
'''
将文件保存到流中，例如在使用 Web 应用程序时 例如金字塔，烧瓶或Django.
'''
from tempfile import NamedTemporaryFile
from openpyxl import Workbook

workbook = Workbook()
with NamedTemporaryFile() as tmp:
    workbook.save(tmp.name)
    tmp.seek(0)
    stream = tmp.read()

# 5.3 从文件加载 openpyxl.load_workbook()
from openpyxl import load_workbook

workbook = load_workbook(filename='empty_book.xlsx')
sheet_ranges = workbook['range names']
print(sheet_ranges['D18'].value)
'''
有几个标志可以在load_workbook中使用。
data_only 控制具有公式的单元格是否具有 公式（默认值）或 Excel 上次读取工作表时存储的值。
keep_vba 控制是否保留任何 Visual Basic 元素或 不是（默认）。如果它们被保留，它们仍然不可编辑。
'''

# ============================================================================================
# 二十一、图表 详见官网
# ================================================================================
'''
1. 面积图
     ---2D区域图
     ---3D区域图
2. 条形图和柱形图---BarChart()
     ---垂直、水平和堆叠条形图
     ---3D条形图
3. 气泡图
4. 折线图
     ---折线图
     ---3D折线图
5. 散点图
6. 饼状图
     ---饼状图
     ---投影饼图
     ---3D饼图
7. 甜甜圈图
8. 雷达图
9. 股票图
10.表面图
'''

# 条形图
'''
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series

workbook = Workbook()
worksheet = workbook.active
for i in range(15):
    worksheet.append([i])

values = Reference(worksheet, min_row=1, max_row=10, min_col=1, max_col=1)  # 选取区域内容
chart = BarChart()
chart.add_data(values)
worksheet.add_chart(chart, "E15")  # 大小为15 x 7.5厘米(约5列乘14行),可以通过设置图表的anchor,width和height属性来更改
workbook.save("C:/Users/Zeus/Desktop/SampleChart.xlsx")
'''

# 散点图
'''
坐标-最小最大范围设置
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

workbook = Workbook()
worksheet = workbook.active

worksheet.append(['X', '1/X'])
for x in range(-10, 11):
    if x:
        worksheet.append([x, 1.0 / x])

chart1 = ScatterChart()
chart1.title = "Full Axes"
chart1.x_axis.title = 'x'
chart1.y_axis.title = '1/x'
chart1.legend = None

chart1.x_axis.scaling.min = -10
chart1.x_axis.scaling.max = 10
chart1.y_axis.scaling.min = -1.5
chart1.y_axis.scaling.max = 1.5

chart2 = ScatterChart()
chart2.title = "Clipped Axes"
chart2.x_axis.title = 'x'
chart2.y_axis.title = '1/x'
chart2.legend = None

chart2.x_axis.scaling.min = 0
chart2.x_axis.scaling.max = 11
chart2.y_axis.scaling.min = 0
chart2.y_axis.scaling.max = 1.5

x = Reference(worksheet, min_col=1, min_row=2, max_row=22)
y = Reference(worksheet, min_col=2, min_row=2, max_row=22)
s = Series(y, xvalues=x)
chart1.append(s)
chart2.append(s)

worksheet.add_chart(chart1, anchor="C1")
worksheet.add_chart(chart2, anchor="C15")

workbook.save("C:/Users/Zeus/Desktop/minmax.xlsx")
'''

# 坐标-对数缩放
'''
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
import math

workbook = Workbook()
worksheet = workbook.active

worksheet.append(['X', 'Gaussian'])
for i, x in enumerate(range(-10, 11)):
    worksheet.append([x, "=EXP(-(($A${row}/6)^2))".format(row=i + 2)])

chart1 = ScatterChart()
chart1.title = "No Scaling"
chart1.x_axis.title = 'x'
chart1.y_axis.title = 'y'
chart1.legend = None

chart2 = ScatterChart()
chart2.title = "X Log Scale"
chart2.x_axis.title = 'x (log10)'
chart2.y_axis.title = 'y'
chart2.legend = None
chart2.x_axis.scaling.logBase = 10

chart3 = ScatterChart()
chart3.title = "Y Log Scale"
chart3.x_axis.title = 'x'
chart3.y_axis.title = 'y (log10)'
chart3.legend = None
chart3.y_axis.scaling.logBase = 10

chart4 = ScatterChart()
chart4.title = "Both Log Scale"
chart4.x_axis.title = 'x (log10)'
chart4.y_axis.title = 'y (log10)'
chart4.legend = None
chart4.x_axis.scaling.logBase = 10
chart4.y_axis.scaling.logBase = 10

chart5 = ScatterChart()
chart5.title = "Log Scale Base e"
chart5.x_axis.title = 'x (ln)'
chart5.y_axis.title = 'y (ln)'
chart5.legend = None
chart5.x_axis.scaling.logBase = math.e
chart5.y_axis.scaling.logBase = math.e

x = Reference(worksheet, min_col=1, min_row=2, max_row=22)
y = Reference(worksheet, min_col=2, min_row=2, max_row=22)
s = Series(y, xvalues=x)
chart1.append(s)
chart2.append(s)
chart3.append(s)
chart4.append(s)
chart5.append(s)

worksheet.add_chart(chart1, "C1")
worksheet.add_chart(chart2, "I1")
worksheet.add_chart(chart3, "C15")
worksheet.add_chart(chart4, "I15")
worksheet.add_chart(chart5, "F30")

workbook.save("C:/Users/Zeus/Desktop/log.xlsx")
'''

# 坐标-方向设置
'''
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series

workbook = Workbook()
worksheet = workbook.active

worksheet["A1"] = "Archimedean Spiral"
worksheet.append(["T", "X", "Y"])
for i, t in enumerate(range(100)):
    worksheet.append([t / 16.0, "=$A${row}*COS($A${row})".format(row=i + 3),
               "=$A${row}*SIN($A${row})".format(row=i + 3)])

chart1 = ScatterChart()
chart1.title = "Default Orientation"
chart1.x_axis.title = 'x'
chart1.y_axis.title = 'y'
chart1.legend = None

chart2 = ScatterChart()
chart2.title = "Flip X"
chart2.x_axis.title = 'x'
chart2.y_axis.title = 'y'
chart2.legend = None
chart2.x_axis.scaling.orientation = "maxMin"
chart2.y_axis.scaling.orientation = "minMax"

chart3 = ScatterChart()
chart3.title = "Flip Y"
chart3.x_axis.title = 'x'
chart3.y_axis.title = 'y'
chart3.legend = None
chart3.x_axis.scaling.orientation = "minMax"
chart3.y_axis.scaling.orientation = "maxMin"

chart4 = ScatterChart()
chart4.title = "Flip Both"
chart4.x_axis.title = 'x'
chart4.y_axis.title = 'y'
chart4.legend = None
chart4.x_axis.scaling.orientation = "maxMin"
chart4.y_axis.scaling.orientation = "maxMin"

x = Reference(worksheet, min_col=2, min_row=2, max_row=102)
y = Reference(worksheet, min_col=3, min_row=2, max_row=102)
s = Series(y, xvalues=x)
chart1.append(s)
chart2.append(s)
chart3.append(s)
chart4.append(s)

worksheet.add_chart(chart1, "D1")
worksheet.add_chart(chart2, "J1")
worksheet.add_chart(chart3, "D15")
worksheet.add_chart(chart4, "J15")

workbook.save("C:/Users/Zeus/Desktop/orientation.xlsx")
'''

# 坐标-添加第二坐标
'''
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, Reference, Series

workbook = Workbook()
worksheet = workbook.active

roworksheet = [
    ['Aliens', 2, 3, 4, 5, 6, 7],
    ['Humans', 10, 40, 50, 20, 10, 50],
]

for row in roworksheet:
    worksheet.append(row)

c1 = BarChart()
v1 = Reference(worksheet, min_col=1, min_row=1, max_col=7)
c1.add_data(v1, titles_from_data=True, from_roworksheet=True)

c1.x_axis.title = 'Days'
c1.y_axis.title = 'Aliens'
c1.y_axis.majorGridlines = None
c1.title = 'Survey results'

# Create a second chart
c2 = LineChart()
v2 = Reference(worksheet, min_col=1, min_row=2, max_col=7)
c2.add_data(v2, titles_from_data=True, from_roworksheet=True)
c2.y_axis.axId = 200
c2.y_axis.title = "Humans"

# Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
c1.y_axis.crosses = "max"
c1 += c2

worksheet.add_chart(c1, "D4")

workbook.save("C:/Users/Zeus/Desktop/secondary.xlsx")
'''
