# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.3.2
@projectName   : pythonProjectPy3.9
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2021/01/27 20:36
'''
# ============================================================================================
'''
目标：实现Excel全部功能
一、单元格样式简介
openpyxl处理Excel文件中单元格样式，总共有六个属性类。
font(字体类----可设置字号、字体颜色、下划线等)
fill(填充类----可设置单元格填充颜色等)
border（边框类----可以设置单元格各种类型的边框）
alignment(位置类----可以设置单元格内数据各种对齐方式)
number_format(格式类----可以设置单元格内各种类型的数据格式)
protection(保护类----可以设置单元格写保护等)
'''
# ============================================================================================
from openpyxl.comments import Comment
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle, GradientFill, Color
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows

'''表格创建'''
workbook = Workbook()  # write_only=True仅写模式,guess_types=True当读取单元格时,guess_types将启用或禁用（默认）类型推断(比如0.34 -> 34%)
worksheet = workbook.active

# 添加评论
# comment = worksheet["A1"].comment
# comment = Comment('This is the comment text', 'Comment Author')
# comment.text  # 'This is the comment text'
# comment.author  # 'Comment Author'

# 创建子页及子页名称
worksheet.title = 'Sheet.NO1'  # 子页名称（默认第一个）
worksheet = workbook.create_sheet('Create.Sheet2', 1)  # # 创建名为Sheet2的子页放在第2
worksheet = workbook.create_sheet('Create.Sheet1', 0)  # 创建名为Sheet1的子页放在第1

# worksheetcopy = workbook["New Title"]  # 给工作表命名后,即可将其作为工作簿的键
# print(workbook.sheetnames) # 打印全部子页名称

# 遍历工作表
# for sheet in workbook:
#     print(sheet.title)

# 在单个工作簿中创建工作表的副本
# 仅复制单元格（包括值,样式,超链接和注释）和某些工作表属性（包括尺寸,格式和属性）.不复制所有其他工作簿/工作表属性-例如图像,图表.
# 您不能在工作簿之间复制工作表.如果工作簿以只读或仅写模式打开,则您也无法复制工作表.
# source = workbook.active
# target = workbook.copy_worksheet(source)

# 工作簿保护
# workbook.security.workbookPassword = ''  # 防止其他用户查看隐藏的工作表,添加,移动,删除或隐藏工作表以及重命名工作表
# worksheet.security.revisionsPassword = ''  # 防止从共享工作簿中删除更改跟踪和更改历史记录
# workbook.security.set_workbook_password(hashed_password = '', already_hashed=True) # 如果需要在不使用默认哈希算法的情况下设置原始密码值，则提供特定的设置器功能

# 工作表保护
# 如果未指定密码,则用户无需指定密码即可禁用配置的工作表保护。否则,他们必须提供密码才能更改已配置的保护。
# workbook.protection.sheet = True # 设置属性来锁定工作表的各个方面
# workbook.protection.password = ''  # 属性设置密码


'''全局设置样式'''
# 子页全局背景颜色
worksheet.sheet_properties.tabColor = "1072BA"  # 默认白色

# 编辑页面设置
worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
worksheet.page_setup.paperSize = worksheet.PAPERSIZE_TABLOID
worksheet.page_setup.fitToHeight = 0
worksheet.page_setup.fitToWidth = 1

highlight = NamedStyle(name="highlight")
number_format = 'General'

font = Font(name='Calibri',
            size=11,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='FF000000')
align = Alignment(horizontal='center', vertical='center')
# general bottom top left right center
alignment = Alignment(horizontal='general',
                      vertical='bottom',
                      text_rotation=0,
                      wrap_text=False,
                      shrink_to_fit=False,
                      indent=0)

protection = Protection(locked=True, hidden=False)
# thin 单线; double 双线
bottom_thin = Side(border_style="double", color="000000")  # 下
left_thin = Side(border_style="thin", color="000000")  # 左
top_double = Side(border_style="double", color="000000")  # 上
right_double = Side(border_style="thin", color="000000")  # 右
# border = Border(bottom=left_thin, left=left_thin, top=left_thin, right=left_thin)  # 边框样式
# thin double thick
border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),
                diagonal=Side(border_style='thin', color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style='thin', color='FF000000'),
                vertical=Side(border_style='thin', color='FF000000'),
                horizontal=Side(border_style='thin', color='FF000000')
                )
# title_font = Font(name='黑体', size=16, bold=True, color="000000")
title_font = Font(name='Calibri',  # 字体名称
                  size=11,  # 字体大小
                  bold=False,  # 加粗-True/False
                  italic=False,  # 斜体-True/False
                  vertAlign=None,  # 对其方向-
                  underline='none',  # 下划线-single/double
                  strike=False,
                  color='FF000000')

title_fill = PatternFill(fill_type='solid', fgColor="F8F8FF")
# fill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
interval_fill = GradientFill(stop=("000000", "FFFFFF"))  # 连续区间底色填充
content_font = Font(name='宋体', size=14, bold=False, color="000000")
content_fill = PatternFill(fill_type='solid', fgColor="F8F8FF")

worksheet.merge_cells('A1:I2')  # 合并单元格,保留左上角值,其他删除
# worksheet.merge_cells(start_row=2, start_column=1, end_row=4, end_column=4)  # 合并单元格第二种方法
# worksheet.unmerge_cells('A1:I2')  # 拆分合并的单元格,值值保留在左上角
# worksheet.unmerge_cells(start_row=2, start_column=1, end_row=4, end_column=4) #拆分合并的单元格第二种方法
worksheet.cell(row=1, column=1).value = '合并的单元格内容'

'''内容填充及字段属性设置'''
# 注意：表格cell的最低行列值都是从1开始,不是0
for rowi in range(3, 8):
    for colj in range(1, 10):
        # _ = ws.cell(row=4, column=2, value=10)
        worksheet.cell(row=rowi, column=colj).value = rowi + colj

# 使用数学公式
# worksheet["A1"] = "=SUM(10, 10)"
# 插入图片
# image = Image('C:/Users/Zeus/Desktop/新logo.png')
# worksheet.add_image(imgage, 'A1')  # A1为插入图片起始单元格
# 折叠行列
# worksheet.row_dimensions.group(1, 3, outline_level=1, hidden=True) # outline_level显示折叠轮廓线(默认1),hidden直接隐藏掉折叠的行列(默认True)
# worksheet.column_dimensions.group('A', 'D', outline_level=1, hidden=False)

# ============================================================================================
'''数据可视化
1. 面积图
     ---2D区域图
     ---3D区域图
2. 条形图和柱形图---BarChart()
     ---垂直、水平和堆叠条形图
     ---3D条形图
3. 气泡图
4. 折线图
     ---折线图
     ---3D折线图
5. 散点图
6. 饼状图
     ---饼状图
     ---投影饼图
     ---3D饼图
7. 甜甜圈图
8. 雷达图
9. 股票图
10.表面图
'''

# ---条形图
# from openpyxl import Workbook
# from openpyxl.chart import BarChart, Reference, Series
#
# wb = Workbook()
# ws = wb.active
# for i in range(15):
#     ws.append([i])
#
# values = Reference(ws, min_row=1, max_row=10, min_col=1, max_col=1)  # 选取区域内容
# chart = BarChart()
# chart.add_data(values)
# ws.add_chart(chart, "E15")  # 大小为15 x 7.5厘米(约5列乘14行),可以通过设置图表的anchor,width和height属性来更改
# wb.save("C:/Users/Zeus/Desktop/SampleChart.xlsx")

# ---散点图
# 坐标-最小最大范围设置
# from openpyxl import Workbook
# from openpyxl.chart import ScatterChart, Reference, Series
#
# wb = Workbook()
# ws = wb.active
#
# ws.append(['X', '1/X'])
# for x in range(-10, 11):
#     if x:
#         ws.append([x, 1.0 / x])
#
# chart1 = ScatterChart()
# chart1.title = "Full Axes"
# chart1.x_axis.title = 'x'
# chart1.y_axis.title = '1/x'
# chart1.legend = None
#
# chart1.x_axis.scaling.min = -10
# chart1.x_axis.scaling.max = 10
# chart1.y_axis.scaling.min = -1.5
# chart1.y_axis.scaling.max = 1.5
#
# chart2 = ScatterChart()
# chart2.title = "Clipped Axes"
# chart2.x_axis.title = 'x'
# chart2.y_axis.title = '1/x'
# chart2.legend = None
#
# chart2.x_axis.scaling.min = 0
# chart2.x_axis.scaling.max = 11
# chart2.y_axis.scaling.min = 0
# chart2.y_axis.scaling.max = 1.5
#
# x = Reference(ws, min_col=1, min_row=2, max_row=22)
# y = Reference(ws, min_col=2, min_row=2, max_row=22)
# s = Series(y, xvalues=x)
# chart1.append(s)
# chart2.append(s)
#
# ws.add_chart(chart1, anchor="C1")
# ws.add_chart(chart2, anchor="C15")
#
# wb.save("C:/Users/Zeus/Desktop/minmax.xlsx")

# 坐标-对数缩放
# from openpyxl import Workbook
# from openpyxl.chart import ScatterChart, Reference, Series
# import math
#
# wb = Workbook()
# ws = wb.active
#
# ws.append(['X', 'Gaussian'])
# for i, x in enumerate(range(-10, 11)):
#     ws.append([x, "=EXP(-(($A${row}/6)^2))".format(row=i + 2)])
#
# chart1 = ScatterChart()
# chart1.title = "No Scaling"
# chart1.x_axis.title = 'x'
# chart1.y_axis.title = 'y'
# chart1.legend = None
#
# chart2 = ScatterChart()
# chart2.title = "X Log Scale"
# chart2.x_axis.title = 'x (log10)'
# chart2.y_axis.title = 'y'
# chart2.legend = None
# chart2.x_axis.scaling.logBase = 10
#
# chart3 = ScatterChart()
# chart3.title = "Y Log Scale"
# chart3.x_axis.title = 'x'
# chart3.y_axis.title = 'y (log10)'
# chart3.legend = None
# chart3.y_axis.scaling.logBase = 10
#
# chart4 = ScatterChart()
# chart4.title = "Both Log Scale"
# chart4.x_axis.title = 'x (log10)'
# chart4.y_axis.title = 'y (log10)'
# chart4.legend = None
# chart4.x_axis.scaling.logBase = 10
# chart4.y_axis.scaling.logBase = 10
#
# chart5 = ScatterChart()
# chart5.title = "Log Scale Base e"
# chart5.x_axis.title = 'x (ln)'
# chart5.y_axis.title = 'y (ln)'
# chart5.legend = None
# chart5.x_axis.scaling.logBase = math.e
# chart5.y_axis.scaling.logBase = math.e
#
# x = Reference(ws, min_col=1, min_row=2, max_row=22)
# y = Reference(ws, min_col=2, min_row=2, max_row=22)
# s = Series(y, xvalues=x)
# chart1.append(s)
# chart2.append(s)
# chart3.append(s)
# chart4.append(s)
# chart5.append(s)
#
# ws.add_chart(chart1, "C1")
# ws.add_chart(chart2, "I1")
# ws.add_chart(chart3, "C15")
# ws.add_chart(chart4, "I15")
# ws.add_chart(chart5, "F30")
#
# wb.save("C:/Users/Zeus/Desktop/log.xlsx")

# 坐标-方向设置
# from openpyxl import Workbook
# from openpyxl.chart import ScatterChart, Reference, Series
#
# wb = Workbook()
# ws = wb.active
#
# ws["A1"] = "Archimedean Spiral"
# ws.append(["T", "X", "Y"])
# for i, t in enumerate(range(100)):
#     ws.append([t / 16.0, "=$A${row}*COS($A${row})".format(row=i + 3),
#                "=$A${row}*SIN($A${row})".format(row=i + 3)])
#
# chart1 = ScatterChart()
# chart1.title = "Default Orientation"
# chart1.x_axis.title = 'x'
# chart1.y_axis.title = 'y'
# chart1.legend = None
#
# chart2 = ScatterChart()
# chart2.title = "Flip X"
# chart2.x_axis.title = 'x'
# chart2.y_axis.title = 'y'
# chart2.legend = None
# chart2.x_axis.scaling.orientation = "maxMin"
# chart2.y_axis.scaling.orientation = "minMax"
#
# chart3 = ScatterChart()
# chart3.title = "Flip Y"
# chart3.x_axis.title = 'x'
# chart3.y_axis.title = 'y'
# chart3.legend = None
# chart3.x_axis.scaling.orientation = "minMax"
# chart3.y_axis.scaling.orientation = "maxMin"
#
# chart4 = ScatterChart()
# chart4.title = "Flip Both"
# chart4.x_axis.title = 'x'
# chart4.y_axis.title = 'y'
# chart4.legend = None
# chart4.x_axis.scaling.orientation = "maxMin"
# chart4.y_axis.scaling.orientation = "maxMin"
#
# x = Reference(ws, min_col=2, min_row=2, max_row=102)
# y = Reference(ws, min_col=3, min_row=2, max_row=102)
# s = Series(y, xvalues=x)
# chart1.append(s)
# chart2.append(s)
# chart3.append(s)
# chart4.append(s)
#
# ws.add_chart(chart1, "D1")
# ws.add_chart(chart2, "J1")
# ws.add_chart(chart3, "D15")
# ws.add_chart(chart4, "J15")
#
# wb.save("C:/Users/Zeus/Desktop/orientation.xlsx")

# 坐标-添加第二坐标
# from openpyxl import Workbook
# from openpyxl.chart import LineChart, BarChart, Reference, Series
#
# wb = Workbook()
# ws = wb.active
#
# rows = [
#     ['Aliens', 2, 3, 4, 5, 6, 7],
#     ['Humans', 10, 40, 50, 20, 10, 50],
# ]
#
# for row in rows:
#     ws.append(row)
#
# c1 = BarChart()
# v1 = Reference(ws, min_col=1, min_row=1, max_col=7)
# c1.add_data(v1, titles_from_data=True, from_rows=True)
#
# c1.x_axis.title = 'Days'
# c1.y_axis.title = 'Aliens'
# c1.y_axis.majorGridlines = None
# c1.title = 'Survey results'
#
# # Create a second chart
# c2 = LineChart()
# v2 = Reference(ws, min_col=1, min_row=2, max_col=7)
# c2.add_data(v2, titles_from_data=True, from_rows=True)
# c2.y_axis.axId = 200
# c2.y_axis.title = "Humans"
#
# # Display y-axis of the second chart on the right by setting it to cross the x-axis at its maximum
# c1.y_axis.crosses = "max"
# c1 += c2
#
# ws.add_chart(c1, "D4")
#
# wb.save("C:/Users/Zeus/Desktop/secondary.xlsx")

# ============================================================================================
'''局部设置样式'''
for row in worksheet.rows:
    for cell in row:
        cell.alignment = align
        if cell.row == 1 or cell.row == 2:
            cell.border = border
            cell.font = title_font
            cell.fill = title_fill
        else:
            cell.border = border
            cell.font = content_font
            cell.fill = content_fill

# col = worksheet.column_dimensions['A']
# col.font = Font(bold=True)
# row = worksheet.row_dimensions[1]
# row.font = Font(underline="single")
# c = Color(indexed=32)  # 按索引选颜色
# c = Color(theme=6, tint=0.5)  # 使用主题
# workbook.add_named_style(highlight)

'''储存表格'''
workbook.save(filename='C:/Users/Zeus/Desktop/openpyxl.xlsx')  # 储存格式：xlsx、xlsm或zip

# # 读取与储存文件
# # workbook = load_workbook(filename='large_file.xlsx', read_only=True) # 仅读模式
# workbook = load_workbook('document.xlsx')
# # 当读取单元格时,guess_types将启用或禁用（默认）类型推断.
# # read_only仅读取模式,不可改动
# # data_only控制具有公式的单元格是否具有公式(默认值)或Excel上次读取工作表时存储的值.
# # keep_vba控制是否保留任何Visual Basic元素(默认).如果保留它们,则它们仍不可编辑.
# print(workbook.get_sheet_names())
# workbook.template = True  # template = True,以将工作簿另存为模板;
# workbook.save('document_template.xltx')  # 注意为xltx格式
# workbook = load_workbook('document_template.xltx')
# workbook.template = False  # 将此属性设置为 template =False(默认)以另存为文档
# workbook.save('document.xlsx', as_template=False)
