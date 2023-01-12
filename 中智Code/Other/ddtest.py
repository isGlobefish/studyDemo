# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.2.2
@projectName   : pythonProjectPy3.9 
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2020/12/10 16:00
'''
# coding: utf-8
import json
import requests
import time
import socket
import sys


def dingding_robot(data):
    # token地址 需要修改
    dingding_robot_token = "https://oapi.dingtalk.com/robot/send?access_token=d8ebca52d6ac9546895670f1bxxxxx"
    headers = {'content - type': 'application / json'}
    r = requests.post(dingding_robot_token, headers=headers, data=json.dumps(data))
    r.encoding = 'utf - 8'
    return r.text


# get ip
def get_host_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('114.114.114.114', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()
    return ip


if __name__ == "__main__":
    ip = get_host_ip()
    role = sys.argv[1]
    data = {
        "msgtype": "markdown",
        # 告警内容，需要调整
        "markdown": {
            "title": "Nginx Keepalived监控",
            "text": "### %s Change to %s\n\n" % (ip, role) +
                    "> #### Nginx Keepalived状态监控\n\n" +
                    "> #### VIP: 192.168.1.1 \n\n" +
                    "> #### 时间: %s\n" % time.strftime("%Y-%m-%d %X") +
                    "> #### 服务名：keepalived \n\n" +
                    "> #### 状态：切换 \n\n" +
                    "> ##### Nginx VIP监控\n\n后端机器IP:192.168.1.2/3 Master为192.168.1.2\n\n状态切换请检查备机"
        },

    }
    res = dingding_robot(data)
    print(res)

import xlrd
import pandas as pd

data = xlrd.open_workbook('E:/大客户数据/DSL/1607996733135流向明细.csv')

data = pd.read_csv('E:/大客户数据/DSL/1607996733135流向明细.csv', encoding='utf-8', sep=',')

data = xlrd.open_workbook('E:/大客户数据/GD/downloadGD/导出三级流向数据_20201215091044.csv')
sheet1Data = data.sheet_by_index(0)

import csv

data = csv.reader('E:/大客户数据/DSL/1607996733135流向明细.csv')

birth_data = []
with open('E:/大客户数据/DSL/1607996733135流向明细.csv', 'rb') as csvfile:
    csv_reader = csv.reader(csvfile)  # 使用csv.reader读取csvfile中的文件
    birth_header = next(csv_reader)  # 读取第一行每一列的标题
    for row in csv_reader:  # 将csv 文件中的数据保存到birth_data中
        birth_data.append(row)

import csv

with open('E:/大客户数据/DSL/1607996733135流向明细.csv') as myFile:
    lines = csv.reader(myFile)
    with open('C:/Users/Long/Desktop/123.txt', 'w') as d1:
        for line in lines:
            d1.write(str(line) + '\n')

import csv
import xlwt


def csv_to_xlsx():
    with open('E:/大客户数据/DSL/1607996733135流向明细.csv', 'r') as f:
        read = csv.reader(f)
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('Sheet1')  # 创建一个Sheet1表格
        for rowi, line in enumerate(read):
            for coli, content in enumerate(line):
                sheet.write(rowi, coli, content)  # 一个一个将单元格数据写入
        sheet.save('C:/Users/Long/Desktop/12.xls')  # 保存Excel


if __name__ == '__main__':
    csv_to_xlsx()




# Excel格式转换：.csv ---> .xls
import csv
import openpyxl

def csv_to_xlsx():
    with open('C:/Users/Long/Desktop/导出三级流向数据_20201215091044.csv', 'r') as csvfile:
        openCsvFile = csv.reader(csvfile)
        workbook = openpyxl.Workbook()  # 打开一个文件
        sheet = workbook.create_sheet(index=0)  # 在文件上创建Sheet1
        # sheet = workbook.create_sheet(name='Sheet1')  # 在文件上创建Sheet1
        for rows, lines in enumerate(openCsvFile, start=1):
            for cols, values in enumerate(lines, start=1):
                sheet.cell(row=rows, column=cols, value=values)  # 写入内容
        workbook.save('C:/Users/Long/Desktop/12.xlsx')  # 保存Excel


csv_to_xlsx()
