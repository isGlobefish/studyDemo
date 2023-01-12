'''
Author: zoodehao
Date: 2021-06-08 17:17:07
LastEditTime: 2021-12-03 13:10:11
FilePath: \Py3.9\2_大客户【整合】发送版.py
Description: 逝者如斯夫, 不舍昼夜.
'''
import os
import re
import hmac
import xlwt
import glob
import xlrd
import time
import json
import base64
import hashlib
import pymysql
import calendar
import requests
import win32com
import pythoncom
import numpy as np
import urllib.parse
import pandas as pd
import urllib.request
import win32api as ap
import openpyxl as opxl
from termcolor import cprint
from datetime import datetime
from matplotlib import colors
import matplotlib.pyplot as plt
import win32com.client as win32
from PIL import ImageGrab, Image
from time import strftime, gmtime
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from win32com.client import Dispatch, DispatchEx
from qiniu import Auth, put_file, etag, BucketManager
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection, NamedStyle, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule,IconSetRule, Rule, IconSet, FormatObject


# ----------------------------------------------------------------------------
# 全局日期控制区
# ----------------------------------------------------------------------------
# 本年 月
globalYear  = datetime.now().year
globalMonth = datetime.now().month
# globalMonth = datetime.now().month - 1
globalDay   = datetime.now().day
# globalDay   = 30
lastDay     = calendar.monthrange(globalYear, globalMonth)[1]  # 本月天数
last1Day    = calendar.monthrange(globalYear, (globalMonth - 1))[1]  # 上月天数(环比)
last2Day    = calendar.monthrange((globalYear - 1), globalMonth)[1]  # 去年同期天数(同比)


# 针对某时候某个网址无法导出数据设置使用
# 老百姓 益丰 延后1天; 大参林 海王 延后 2天
delay1 = datetime.now().day - 1
delay2 = datetime.now().day - 2
lbxday, yfday, gjday = delay1, delay1, delay1
hwday, dslday  = delay2, delay2
# lbxday, yfday, gjday = 30, 30, 30
# hwday, dslday  = 30, 30

# 季度（只作用 老百姓）
quarterlyDay    = 31 + 30 + lbxday
quarterlyDayAll = 92
quarterlyStart  = 10
quarterlyEnd    = 12
quarterly       = 'Q4'


# 特殊情况
# 假如当月是1月， 环比却是 上一年 12月
# ？如何？
huanbiYear = datetime.now().year
huanbiMonth = datetime.now().month

# 单品 - 列表形式
# materialList = ['菊花', '茯苓']
materialList = ['三七']


# 加载数据
def load_data():
    global target
    global bigCdata
    global aliasdata
    global yfsingle
    global colordata
    global lbxCofor
    # ----------------------------------------------------------------------------
    # 全局数据准备区
    # ----------------------------------------------------------------------------
    # 加载目标数据
    target = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/2_大客户/目标/大客户_数据源.xlsx', sheet_name=0, header=0)
    # 已经优化到大客户数据库里面的一列
    # gjcompany = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/2_大客户/目标/大客户_数据源.xlsx', sheet_name=1, header=0)
    lbxCofor = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/2_大客户/目标/大客户_数据源.xlsx', sheet_name=3, header=0)


    # 连接数据库
    conn = pymysql.connect(host='192.168.20.241',
                            port    = 3306,
                            user    = 'root',
                            passwd  = 'Powerbi#1217',
                            db      = 'dkh',
                            charset = 'utf8')

    cursorDkh      = conn.cursor()
    cursorMaterial = conn.cursor()
    cursorColor = conn.cursor()

    # 大客户主数据
    executeDkh = """SELECT date, YEAR(date) AS NIAN, MONTH(date) AS YUE, DAY(date) AS TIAN, customer, sfa_client_desc, desc_1, SUM(amount) AS AMOUNT, SUM(bz_sales_Money) AS BZLSJE
                    FROM dkhfact
                    WHERE customer IN ('老百姓', '海王', '益丰', '大参林', '高济') AND YEAR(date) IN (""" + str(globalYear - 1) + """, """ + str(globalYear) + """)
                    GROUP BY date, sfa_client_desc, desc_1, customer"""

    # 单品数据
    executeMaterial = """SELECT YEAR(date) AS NIAN, MONTH(date) AS YUE, DAY(date) AS TIAN, sfa_client_desc, desc_1, materiel_alias, customer, SUM(amount) as amount
                         FROM dkhfact
                         WHERE ((YEAR(date) = """ + str(globalYear - 1) + """ AND MONTH(date) = """ + str(globalMonth) + """) OR (YEAR(date) = """ + str(globalYear) + """ AND MONTH(date) IN (""" + str(globalMonth - 1) + """, """ + str(globalMonth) + """))) AND customer IN ('老百姓', '海王', '益丰', '大参林')
                         GROUP BY date, sfa_client_desc, desc_1, materiel_alias, customer"""

    # 老百姓彩盒
    executeCofor = """SELECT YEAR(date) AS NIAN, MONTH(date) AS YUE, DAY(date) AS TIAN, sfa_client_desc, materiel_desc, norms, customer, SUM(amount) as AMOUNT, SUM(bz_sales_Money) AS MON
                      FROM dkhfact
                      WHERE customer = '老百姓' AND norms LIKE '%48%' and YEAR(date) = """ + str(globalYear) + """
                      GROUP BY date, sfa_client_desc, materiel_desc, norms, customer"""
                                                       
    cursorDkh.execute(executeDkh)
    cursorMaterial.execute(executeMaterial)
    cursorColor.execute(executeCofor)
    dataDkh      = cursorDkh.fetchall()  # 大客户主数据
    dataMaterial = cursorMaterial.fetchall()  # 单品数据
    dataColor    = cursorColor.fetchall()  # 老百姓彩盒数据
    conn.commit()
    cursorDkh.close()
    cursorMaterial.close()
    cursorColor.close()
    conn.close()

    bigCdata = pd.DataFrame(dataDkh, columns=['日期', '年', '月', '日', '体系', '大区', '区域', '数量', '标准金额'])
    aliasdata = pd.DataFrame(dataMaterial, columns=['年', '月', '日', '大区', '区域', '物料', '体系', '数量'])
    colordata = pd.DataFrame(dataColor, columns=['年', '月', '日', '区域', '物料名称', '规格', '体系', '数量', '标准金额'])


    # ----------------------------------------------------------------------------
    # 数据清洗区
    # ----------------------------------------------------------------------------
    # 耗时太久 待优化
    # for irow in range(gjcompany.shape[0]):
    #     for jrow in range(bigCdata.shape[0]):
    #         if bigCdata.loc[jrow, '体系'] == '高济':
    #             if bigCdata.loc[jrow, '区域'] == gjcompany.loc[irow, '公司全称']:
    #                 bigCdata.loc[jrow, '区域'] = gjcompany.loc[irow, '平台']


# ----------------------------------------------------------------------------
# 自定义功能区
# ----------------------------------------------------------------------------
# 清空指定文件夹
def deleteOldFiles(path):
    deleteFileList = os.listdir(path)
    all_PNG = glob.glob(path + "*.PNG")
    print("该目录下文件有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中, PNG: " + str(len(all_PNG)) + "个")
    if len(all_PNG) != 0:
        for deletefile in deleteFileList:
            isDeleteFile = os.path.join(path, deletefile)
            if os.path.isfile(isDeleteFile):
                os.remove(isDeleteFile)
        all_DelPNG = glob.glob(path + "*.*")
        if len(all_DelPNG) == 0:
            print("已清空文件夹！！！")
        else:
            print("存在未删除文件, 请检查是否存在非PNG格式文件")
    else:
        print("不存在PNG文件")
           

# 截图 - 格式类似"A1:J10"
def excelCatchScreen(file_name, sheet_name, name, save_path):
    pythoncom.CoInitialize()  # excel多线程相关
    Application = win32com.client.gencache.EnsureDispatch("Excel.Application")  # 启动excel
    Application.Visible = False  # 可视化
    Application.DisplayAlerts = False  # 是否显示警告
    wb = Application.Workbooks.Open(file_name, ReadOnly=False)  # 打开excel
    # ws = wb.Sheets(sheet_name)  # 选择Sheet
    ws = wb.Worksheets(sheet_name)  # 选择Sheet
    ws.Activate()  # 激活当前工作表
    userange = ws.UsedRange
    # 注意：要从A1开始的表格
    screen_area = 'A1:' + str(opxl.utils.get_column_letter(userange.Columns.Count)) + str(userange.Rows.Count)
    ws.Range(screen_area).CopyPicture()  # 复制图片区域
    time.sleep(1)
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
    Application.Selection.ShapeRange.Name = name  # 将刚刚选择的Shape重命名, 避免与已有图片混淆
    ws.Shapes(name).Copy()  # 选择图片
    time.sleep(1)
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    img_name = name + ".PNG"
    img.save(save_path + img_name)  # 保存图片
    # time.sleep(1)
    # wb.Save()
    # time.sleep(1)
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    time.sleep(1)
    Application.Quit()  # 退出excel
    pythoncom.CoUninitialize()


# 柱状图
def bar_function(x_value, y_value, img_path, img_name, year, month, system_name, sub1_history_month_bar, lastest_month_bar, all_bar, sum_can, day_avg_can):
    plt.rcParams['font.sans-serif'] = 'KaiTi'  # 显示中文不乱码
    plt.rcParams['axes.unicode_minus'] = False  # 显示负数不乱码
    plt.figure(figsize=(30, 11), dpi=100)  # 设置图形大小
    # 绘制条形图
    plt.bar(x_value, y_value, width=0.8, color=['#A9D18E' if i <= sub1_history_month_bar else '#2F5597' if i >= lastest_month_bar else '#F4B183' for i in range(all_bar)], alpha=1)
    plt.grid(axis='y', color='#A6A6A6', linestyle='--', linewidth=1, alpha=0.25)
    plt.suptitle(system_name + '--' + str(month) + '月每日纯销（罐）', fontsize=32, color='#009E47')
    plt.title(str(year) + '年' + str(month) + '月累计纯销（' + r'$\bf{' + str(sum_can) + '}$罐） 日均纯销（' + r'$\bf{' + str(day_avg_can) + '}$罐)', fontsize=28, color='#FF0000')
    plt.xlim(-1, all_bar)  # x轴起步, 柱子数
    # 设置x, y轴字体大小
    plt.xticks(fontsize=16, rotation=90)
    plt.yticks(fontsize=18)
    if system_name == '老百姓':
        add = 213
    elif system_name == '海王': 
        add = 60
    elif system_name == '益丰': 
        add = 60
    elif system_name == '高济': 
        add = 82
    else: # 大参林
        add = 150
    # 柱子显示数值
    for x, y in zip(x_value, y_value):
        if (system_name == '老百姓' and (x == '05日' or x == '15日' or x == '25日')) or (system_name == '海王' and (x == '08日' or x == '18日' or x == '28日')):
            plt.text(x, y + add, '%.0f' % y, ha='center', va='bottom', fontsize=12, color='red', bbox=dict(facecolor='#FFFF00', alpha=0.3))
        else:
            plt.text(x, y + add, '%.0f' % y, ha='center', va='bottom', fontsize=12, bbox=dict(facecolor='#FFE393', alpha=0.1))
    # 存储图片
    plt.savefig(img_path + img_name + '.PNG')
    plt.close()
  
    
# 定义钉钉功能
class dingdingFunction(object):
    def __init__(self, roboturl, robotsecret, appkey, appsecret):
        """
        :param roboturl: 群机器人WebHook_url
        :param robotsecret: 安全设置的加签秘钥
        :param appkey: 企业开发平台小程序AppKey
        :param appsecret: 企业开发平台小程序AppSecret
        """
        self.roboturl = roboturl
        self.robotsecret = robotsecret
        self.appkey = appkey
        self.appsecret = appsecret
        timestamp = round(time.time() * 1000)  # 时间戳
        secret_enc = robotsecret.encode('utf-8')
        string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
        string_to_sign_enc = string_to_sign.encode('utf-8')
        hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
        self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(timestamp, sign)  # 最终url,url+时间戳+签名

    # 发送文件
    def getAccess_token(self):
        url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (AppKey, AppSecret)
        headers = {
            'Content-Type': "application/x-www-form-urlencoded"
        }
        data = {'appkey': self.appkey,
                'appsecret': self.appsecret}
        r = requests.request('GET', url, data=data, headers=headers)
        access_token = r.json()["access_token"]
        return access_token

    def getMedia_id(self, filespath):
        access_token = self.getAccess_token()  # 拿到接口凭证
        url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
        files = {'media': open(filespath, 'rb')}
        data = {'access_token': access_token,
                'type': 'file'}
        response = requests.post(url, files=files, data=data)
        json = response.json()
        return json["media_id"]

    def sendFile(self, chatid, filespath):
        access_token = self.getAccess_token()
        media_id = self.getMedia_id(filespath)
        url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
        header = {
            'Content-Type': 'application/json'
        }
        data = {'access_token': access_token,
                'chatid': chatid,
                'msg': {
                    'msgtype': 'file',
                    'file': {'media_id': media_id}
                }}
        r = requests.request('POST', url, data=json.dumps(data), headers=header)
        print(r.json()["errmsg"])

    # 发送消息
    def sendMessage(self, content):
        """
        :param content: 发送内容
        """
        header = {
            "Content-Type": "application/json",
            "Charset": "UTF-8"
        }
        sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
        sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
        request = urllib.request.Request(url=self.webhook_url, data=sendContent, headers=header)  # 发送请求
        opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
        print(opener.read().decode())  # 打印返回的结果   


# 上传本地图片获取网上图片URL
# def get_image_url(imagePath, pictureName):
#     if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
#         filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
#     elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
#         filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
#     else:
#         print("请检查图片格式！！！")
#     # 七牛云密钥管理：https://portal.qiniu.com/user/key
#     # 【账号：144714959@qq.com  密码：thebtx1997】
#     access_key = "DZnCErimkn2yQrn4aYel3JX7vPXKRonlvDFoVh1e"
#     secret_key = "FBEHIFyMG28nWZrn316df-ny5bmIz_LanRWtabCi"
#     keyq = Auth(access_key, secret_key)
#     bucket = "qiniu730173201"  # 七牛云盘名
#     # 删除
#     butm = BucketManager(keyq)
#     reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
#     # 上传
#     time.sleep(5)
#     token = keyq.upload_token(bucket, filename)  # 上传新图片
#     reformUp, informUp = put_file(token, filename, imagePath)
#     if reformUp is not None:
#         print('已经成功上传 {}'.format(filename))
#     else:
#         print(filename + '上传失败！！！')
#     time.sleep(5)
#     baseURL = "http://zzsy.zeus.cn/"  # 中智二级域名
#     subURL = baseURL + '/' + filename
#     pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
#     time.sleep(6)
#     return pictureURL

def get_image_url(imagePath, pictureName):
    if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    else:
        print("请检查图片格式！！！")
    # 七牛云密钥管理：https://portal.qiniu.com/user/key
    # 【账号：13267854059  密码：z****】
    access_key = "fjlWDEbF1fqBU98UsdDJRcSSKODT9Gq7tA3gu8eY"
    secret_key = "thiWFpO881GfhlaAz1Wkk2yEcvV3ue2OHnY_5D9V"
    keyq = Auth(access_key, secret_key)
    bucket = "zues3737img"  # 七牛云盘名
    # 删除
    butm = BucketManager(keyq)
    reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
    # 上传
    time.sleep(1)
    token = keyq.upload_token(bucket, filename)  # 上传新图片
    reformUp, informUp = put_file(token, filename, imagePath)
    if reformUp is not None:
        print('已成功上传 {}'.format(filename))
        time.sleep(1)
        baseURL = "https://cjh3737.zeus.cn/"  # 加速域名
        # subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/background/white/ignore-error/1?imageView2/0/quality/100!/sharpen/1/interlace/1/ignore-error/1'
        subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/shapen/50/interlace/1/ignore-error/1'
        pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
        time.sleep(1)
        return pictureURL
    else:
        print(filename + '上传失败！！！')
    


# 生成大客户文件
def bigC_graph():
    # ----------------------------------------------------------------------------
    # 加载格式框架
    # ----------------------------------------------------------------------------
    filepath1 = 'C:/Users/Zeus/Desktop/autoSend/2_大客户/目标/大客户_格式框架.xlsx'
    # workbook = opxl.load_workbook(filepath, data_only=True)
    workbook  = opxl.load_workbook(filepath1)
    # ----------------------------------------------------------------------------
    # ①年度简析
    # ----------------------------------------------------------------------------
    workbooksheet1 = workbook['①年度简析']

    sumtype1 = ['数量', '标准金额']
    system1 = ['老百姓', '海王', '益丰', '大参林', '高济']
    daytype1 = [lbxday, hwday, yfday,dslday, gjday]

    for icol in range(2, 23):
        syssum = 0
        for irow, isys, iday in zip([6, 7, 8, 9, 10], system1, daytype1):  # 非总计部分
            if icol in [2, 3]: 
                    workbooksheet1.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= iday)][sumtype1[icol % 2]]) / 10000  
            elif icol in [4, 5]: 
                    workbooksheet1.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= iday)][sumtype1[icol % 2]]) / 10000 * lastDay / iday 
            elif icol in [6, 7]: 
                    workbooksheet1.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth) ][sumtype1[icol % 2]]) / 10000 
            elif icol in [8, 9]: 
                    workbooksheet1.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1)) ][sumtype1[icol % 2]]) / 10000 
            elif icol in [10, 11]: 
                    workbooksheet1.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] <= (globalMonth - 1)) ][sumtype1[icol % 2]]) / 10000 + sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= iday)][sumtype1[icol % 2]]) / 10000 
            elif icol in [12, 13]: 
                    workbooksheet1.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] <= globalMonth) ][sumtype1[icol % 2]]) / 10000
            elif icol in [20, 21]:  # 1-n月累计同比
                isyssum = sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] <= (globalMonth - 1)) ][sumtype1[icol % 2]]) + sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= iday)][sumtype1[icol % 2]]) * lastDay / iday
                workbooksheet1.cell(row=irow, column=icol).value = isyssum / sum(bigCdata[(bigCdata['体系'] == isys) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] <= globalMonth) ][sumtype1[icol % 2]]) - 1
                syssum += isyssum
        # 补充1-n月累计同比 - 总计部分
        if icol == 20 or icol == 21:
            workbooksheet1.cell(row=11, column=icol).value = syssum / sum(bigCdata[(bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] <= globalMonth) ][sumtype1[icol % 2]]) - 1

    # 总计 - 2-13列
    for icol in range(2, 14):
        workbooksheet1.cell(row=11, column=icol).value = '=SUM(' + get_column_letter(icol) + '6:' + get_column_letter(icol) + '10)'   

    # 同环比 - 公式填充部分
    for irow in range(6, 12):
            workbooksheet1.cell(row=irow, column=14).value = '=D' + str(irow) + '/F' + str(irow) + '-1'
            workbooksheet1.cell(row=irow, column=15).value = '=E' + str(irow) + '/G' + str(irow) + '-1'
            workbooksheet1.cell(row=irow, column=16).value = '=O' + str(irow)
            workbooksheet1.cell(row=irow, column=17).value = '=D' + str(irow) + '/H' + str(irow) + '-1'
            workbooksheet1.cell(row=irow, column=18).value = '=E' + str(irow) + '/I' + str(irow) + '-1'
            workbooksheet1.cell(row=irow, column=19).value = '=R' + str(irow)
            workbooksheet1.cell(row=irow, column=22).value = '=U' + str(irow)
    # ----------------------------------------------------------------------------
    # ②汇总分析
    # ----------------------------------------------------------------------------
    workbooksheet2 = workbook['②汇总分析']

    for icol in range(2, 21):
        for irow in [1, 4, 7, 10, 13]:
            if icol == 2:
                workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月'
            elif icol == 3:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = '(1-' + str(lbxday) + '日)'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=③老百姓!$K$34'
                elif irow == 4:
                    workbooksheet2.cell(row=irow , column=icol).value = '(1-' + str(hwday) + '日)'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑤海王!$I$26'
                elif irow == 7:
                    workbooksheet2.cell(row=irow , column=icol).value = '(1-' + str(yfday) + '日)'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑦益丰!$I$15'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = '(1-' + str(dslday) + '日)'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑨大参林!$J$14'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = '(1-' + str(gjday) + '日)'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=1高济!$I$13'
            elif icol == 4:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=④老百姓!$M$32'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=⑥海王!$M$26'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=⑧益丰!$M$15'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=⑩大参林!$N$14'
                else:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=2高济!$M$13'
            elif icol == 6:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)]['数量'])
                elif irow == 4:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑤海王!$F$26'
                elif irow == 7:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑦益丰!$F$15'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑨大参林!$G$14'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = '=1高济!$F$13'
            elif icol == 7:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=③老百姓!$A$4'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑤海王!$A$4'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑦益丰!$A$4'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=LEFT(⑨大参林!$A$4,2)'
                else:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=1高济!$A$4'
            elif icol == 8:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=④老百姓!$N$32'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=⑥海王!$N$26'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=⑧益丰!$N$15'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=⑩大参林!$O$14'
                else:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=2高济!$N$13'
            elif icol == 9:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=③老百姓!$A$5'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑤海王!$A$5'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑦益丰!$A$5'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=LEFT(⑨大参林!$A$5,2)'
                else:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=1高济!$A$5'
            elif icol == 11:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=③老百姓!$A$6'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑤海王!$A$6'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑦益丰!$A$6'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=LEFT(⑨大参林!$A$6,2)'
                else:
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=1高济!$A$6'
            elif icol == 12:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)]['标准金额']) / 10000
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear) + '年' + str(globalMonth) + '月'
                elif irow == 4:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑤海王!$G$26'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear) + '年' + str(globalMonth) + '月'
                elif irow == 7:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑦益丰!$G$15'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear) + '年' + str(globalMonth) + '月'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑨大参林!$H$14'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear) + '年' + str(globalMonth) + '月'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = '=1高济!$G$13'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear) + '年' + str(globalMonth) + '月'
            elif icol == 13:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$Q$6'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$Q$7'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$Q$8'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$Q$9'
                else:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$Q$10'
            elif icol == 14:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月' +  str(lbxday) + '日'
                elif irow == 4:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月' +  str(hwday) + '日'
                elif irow == 7:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月' +  str(yfday) + '日'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月' +  str(dslday) + '日'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月' +  str(gjday) + '日'
            elif icol == 16:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = '=③老百姓!$B$34'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=④老百姓!$D$32'
                elif irow == 4:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑤海王!$B$26'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑥海王!$D$26'
                elif irow == 7:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑦益丰!$B$15'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑧益丰!$D$15'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑨大参林!$C$14'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑩大参林!$E$14'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = '=1高济!$B$13'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=2高济!$D$13'
            elif icol == 17:
                if irow == 1:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$R$6'
                elif irow == 4:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$R$7'
                elif irow == 7:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$R$8'
                elif irow == 10:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$R$9'
                else:
                    workbooksheet2.cell(row=irow + 2, column=icol).value = '=①年度简析!$R$10'
            elif icol == 18:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = str(quarterlyStart) + '-' + str(quarterlyEnd) + '月' + quarterly
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear - 1) + '年' +  str(globalMonth) + '月'
                elif irow == 4:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear - 1) + '年' +  str(globalMonth) + '月'
                elif irow == 7:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear - 1) + '年' +  str(globalMonth) + '月'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear - 1) + '年' +  str(globalMonth) + '月'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = str(globalMonth) + '月'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = str(globalYear - 1) + '年' +  str(globalMonth) + '月'
            elif icol == 20:
                if irow == 1:
                    workbooksheet2.cell(row=irow, column=icol).value = '=③老百姓!$J$34'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=④老百姓!$H$32'
                elif irow == 4:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑤海王!$H$26'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑥海王!$H$26'
                elif irow == 7:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑦益丰!$H$15'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑧益丰!$H$15'
                elif irow == 10:
                    workbooksheet2.cell(row=irow, column=icol).value = '=⑨大参林!$I$14'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=⑩大参林!$I$14'
                else:
                    workbooksheet2.cell(row=irow, column=icol).value = '=1高济!$H$13'
                    workbooksheet2.cell(row=irow + 1, column=icol).value = '=2高济!$H$13'
    # ----------------------------------------------------------------------------
    # ③老百姓
    # ----------------------------------------------------------------------------
    workbooksheet3 = workbook['③老百姓']

    sumtype3 = ['数量', '标准金额']
    area3 = ['天津', '江苏', '陕西', '百佳惠', '通辽泽强', '邻加医', '惠仁堂', '上海', '湖北', '隆泰源', '湘南', '湘北', '湘中', '百姓缘', '南通普泽', '百信缘', '山东', '河南', '万仁', '广西', '三品堂', '新千秋', '仁德', '广东', '江苏海鹏', '华康', '浙江', '河北', '山西百汇', '赤峰人川']

    for icol in range(2, 14):
        for irow, iarea in enumerate(area3, start=4):
            if icol in [2, 3]:
                workbooksheet3.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == lbxday)][sumtype3[icol % 2]])
            elif icol in [4, 5]:
                workbooksheet3.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype3[icol % 2]]) / last1Day
            elif icol in [6, 7]:
                if icol == 6:
                    workbooksheet3.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)][sumtype3[icol % 2]])
                if icol == 7:
                    workbooksheet3.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)][sumtype3[icol % 2]]) / 10000
            elif icol in [8, 9]:
                if icol == 8:
                    workbooksheet3.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & ((bigCdata['月'] >= quarterlyStart) & (bigCdata['月'] <= quarterlyEnd))][sumtype3[icol % 2]])
                elif icol == 9:
                    workbooksheet3.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & ((bigCdata['月'] >= quarterlyStart) & (bigCdata['月'] <= quarterlyEnd))][sumtype3[icol % 2]]) / 10000
            elif icol == 10:
                workbooksheet3.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '老百姓') & (target['月季'] == quarterly) & (target['区域'] == iarea)]['目标'])
            elif icol == 11:
                workbooksheet3.cell(row=irow, column=icol).value = '=I' + str(irow) + '/J' + str(irow)
            elif icol == 12:
                workbooksheet3.cell(row=irow, column=icol).value = '=I' + str(irow) + '/J' + str(irow) + '*' + str(quarterlyDayAll) + '/' + str(quarterlyDay)
            else:
                workbooksheet3.cell(row=irow, column=icol).value = '=J' + str(irow) + '-I' + str(irow)

    # 总计部分
    workbooksheet3.cell(row=34, column=2).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == lbxday)]['数量'])
    workbooksheet3.cell(row=34, column=3).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == lbxday)]['标准金额'])
    workbooksheet3.cell(row=34, column=4).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))]['数量']) / last1Day
    workbooksheet3.cell(row=34, column=5).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))]['标准金额']) / last1Day
    workbooksheet3.cell(row=34, column=6).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)]['数量'])
    workbooksheet3.cell(row=34, column=7).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)]['标准金额'])
    workbooksheet3.cell(row=34, column=8).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] >= quarterlyStart) & (bigCdata['月'] <= quarterlyEnd)]['数量'])
    workbooksheet3.cell(row=34, column=9).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] >= quarterlyStart) & (bigCdata['月'] <= quarterlyEnd)]['标准金额']) / 10000
    workbooksheet3.cell(row=34, column=10).value = sum(target[(target['体系'] == '老百姓') & (target['月季'] == quarterly)]['目标'])
    workbooksheet3.cell(row=34, column=11).value = '=I' + str(34) + '/J' + str(34)
    workbooksheet3.cell(row=34, column=12).value = '=I' + str(34) + '/J' + str(34) + '*' + str(quarterlyDayAll) + '/' + str(quarterlyDay)
    workbooksheet3.cell(row=34, column=13).value = '=J' + str(34) + '-I' + str(34)
    # ----------------------------------------------------------------------------
    # ④老百姓
    # ----------------------------------------------------------------------------
    workbooksheet4 = workbook['④老百姓']

    sumtype4 = ['数量', '标准金额']
    area4 = ['天津老百姓', '上海老百姓', '陕西老百姓', '湖北老百姓', '安徽百姓缘', '湖南老百姓', '江苏老百姓', '常州万仁', '山东老百姓', '河南老百姓', '广西老百姓', '河北老百姓', '浙江老百姓', '扬州百信缘', '无锡三品堂', '通辽泽强', '泰州隆泰源', '南通普泽', '临沂仁德', '兰州惠仁堂', '金坛新千秋', '江苏海鹏', '江苏百佳惠', '广东老百姓', '安徽邻加医', '镇江华康', '山西百汇', '赤峰人川']

    for icol in range(2, 16):
        for irow, iarea in enumerate(area4, start=4):
            if icol in [2, 3]:
                workbooksheet4.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)][sumtype4[icol % 2]]) / 10000
            elif icol in [4, 5]:
                workbooksheet4.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)][sumtype4[icol % 2]]) * lastDay / lbxday / 10000
            elif icol in [6, 7]:
                workbooksheet4.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype4[icol % 2]]) / 10000
            elif icol in [8, 9]:
                workbooksheet4.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype4[icol % 2]]) / 10000

    # 总计部分
    for icol in range(2, 10):
        if icol in [2, 3]:
            workbooksheet4.cell(row=32, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)][sumtype4[icol % 2]]) / 10000
        elif icol in [4, 5]:
            workbooksheet4.cell(row=32, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= lbxday)][sumtype4[icol % 2]]) * lastDay / lbxday / 10000
        elif icol in [6, 7]:
            workbooksheet4.cell(row=32, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype4[icol % 2]]) / 10000
        elif icol in [8, 9]:
            workbooksheet4.cell(row=32, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype4[icol % 2]]) / 10000

    # 同环比
    for irow in range(4, 33):
        workbooksheet4.cell(row=irow, column=10).value = '=IFERROR(D' + str(irow) + '/F' + str(irow) + '-1,"-")'
        workbooksheet4.cell(row=irow, column=11).value = '=IFERROR(E' + str(irow) + '/G' + str(irow) + '-1,"-")'
        workbooksheet4.cell(row=irow, column=12).value = '=K' + str(irow)
        workbooksheet4.cell(row=irow, column=13).value = '=IFERROR(D' + str(irow) + '/H' + str(irow) + '-1,"-")'
        workbooksheet4.cell(row=irow, column=14).value = '=IFERROR(E' + str(irow)+ '/I' + str(irow) + '-1,"-")'
        workbooksheet4.cell(row=irow, column=15).value = '=N' + str(irow)
    # ----------------------------------------------------------------------------
    # ⑤海王
    # ----------------------------------------------------------------------------
    workbooksheet5 = workbook['⑤海王']

    sumtype5 = ['数量', '标准金额']
    area5 = ['连云港', '上海', '福州', '蚌埠', '惠州', '常州', '无锡', '潍坊', '青岛', '深圳', '南京', '湖北', '佛山', '大连', '成都', '沈阳', '湖南', '苏州', '中珠', '东莞', '广州', '天津']
    huanan5 = ['福州', '惠州', '深圳', '佛山', '中珠', '东莞', '广州']

    # 非总计部分
    for icol in range(2, 11):
        for irow, iarea in enumerate(area5, start=4):
            if icol in [2, 3]:
                workbooksheet5.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == hwday)][sumtype5[icol % 2]])
            elif icol in [4, 5]:
                workbooksheet5.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype5[icol % 2]]) / last1Day
            elif icol in [6, 7]:
                if icol == 6:
                    workbooksheet5.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype5[icol % 2]])
                elif icol == 7:
                    workbooksheet5.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype5[icol % 2]]) / 10000
            elif icol == 8:
                workbooksheet5.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '海王')  & (target['区域'] == iarea) & (target['月季'] == globalMonth)]['目标'])
            elif icol == 9:
                workbooksheet5.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"")'
            else:
                workbooksheet5.cell(row=irow, column=icol).value = '=IFERROR(H' + str(irow) + '-G' + str(irow) + ',"")'

    # 总计部分 - 华南总计部分
    for icol in range(2, 11):
        huanansum = 0
        if icol in [2, 3]:
            workbooksheet5.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == hwday)][sumtype5[icol % 2]])
            for ihua in huanan5:
                ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == hwday)][sumtype5[icol % 2]])
                huanansum += ihuanan
            workbooksheet5.cell(row=27, column=icol).value = huanansum
        elif icol in [4, 5]:
            workbooksheet5.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype5[icol % 2]]) / last1Day
            for ihua in huanan5:
                ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype5[icol % 2]]) / last1Day
                huanansum += ihuanan
            workbooksheet5.cell(row=27, column=icol).value = huanansum
        elif icol in [6, 7]:
            if icol == 6:
                workbooksheet5.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype5[icol % 2]])
                for ihua in huanan5:
                    ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype5[icol % 2]])
                    huanansum += ihuanan
                workbooksheet5.cell(row=27, column=icol).value = huanansum
            elif icol == 7:
                workbooksheet5.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype5[icol % 2]]) / 10000
                for ihua in huanan5:
                    ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype5[icol % 2]]) / 10000
                    huanansum += ihuanan
                workbooksheet5.cell(row=27, column=icol).value = huanansum
        elif icol == 8:
            workbooksheet5.cell(row=26, column=icol).value = sum(target[(target['体系'] == '海王') & (target['月季'] == globalMonth)]['目标'])
            for ihua in huanan5:
                ihuanan =  sum(target[(target['体系'] == '海王')  & (target['区域'] == ihua) & (target['月季'] == globalMonth)]['目标'])
                huanansum += ihuanan
            workbooksheet5.cell(row=27, column=icol).value = huanansum
        elif icol == 9:
            workbooksheet5.cell(row=26, column=icol).value = '=G' + str(26) + '/H' + str(26)
            workbooksheet5.cell(row=27, column=icol).value = '=G' + str(27) + '/H' + str(27)
        else:
            workbooksheet5.cell(row=26, column=icol).value = '=H' + str(26)+ '-G' + str(26)
            workbooksheet5.cell(row=27, column=icol).value = '=H' + str(27)+ '-G' + str(27)
    # ----------------------------------------------------------------------------
    # ⑥海王
    # ----------------------------------------------------------------------------
    workbooksheet6 = workbook['⑥海王']

    sumtype6 = ['数量', '标准金额']
    area6 = ['连云港', '上海', '福州', '蚌埠', '惠州', '常州', '无锡', '潍坊', '青岛', '深圳', '南京', '湖北', '佛山', '大连', '成都', '沈阳', '湖南', '苏州', '中珠', '东莞', '广州', '天津']
    huanan6 = ['福州', '惠州', '深圳', '佛山', '中珠', '东莞', '广州']

    # 非总计部分
    for icol in range(2, 16):
        for irow, iarea in enumerate(area6, start=4):
            if icol in [2, 3]:
                    workbooksheet6.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype6[icol % 2]]) / 10000
            elif icol in [4, 5]:
                    workbooksheet6.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype6[icol % 2]]) / 10000 * lastDay / hwday
            elif icol in [6, 7]:
                    workbooksheet6.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype6[icol % 2]]) / 10000
            elif icol in [8, 9]:
                    workbooksheet6.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype6[icol % 2]]) / 10000
            elif icol == 10:
                    workbooksheet6.cell(row=irow, column=icol).value = '=D' + str(irow) + '/F' + str(irow) + '-1'
            elif icol == 11:
                workbooksheet6.cell(row=irow, column=icol).value = '=E' + str(irow) + '/G' + str(irow) + '-1'
            elif icol == 12:
                workbooksheet6.cell(row=irow, column=icol).value = '=K' + str(irow)
            elif icol == 13:
                workbooksheet6.cell(row=irow, column=icol).value = '=D' + str(irow) + '/H' + str(irow) + '-1'
            elif icol == 14:
                workbooksheet6.cell(row=irow, column=icol).value = '=E' + str(irow) + '/I' + str(irow) + '-1'
            else:
                workbooksheet6.cell(row=irow, column=icol).value = '=N' + str(irow)

    # 总计部分 - 华南总计部分
    for icol in range(2, 16):
        huanansum = 0
        if icol in [2, 3]:
                workbooksheet6.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype6[icol % 2]]) / 10000
                for ihua in huanan6:
                    ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype6[icol % 2]]) / 10000
                    huanansum += ihuanan
                workbooksheet6.cell(row=27, column=icol).value = huanansum
        elif icol in [4, 5]:
                workbooksheet6.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype6[icol % 2]]) / 10000 * lastDay / hwday
                for ihua in huanan6:
                    ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= hwday)][sumtype6[icol % 2]]) / 10000 * lastDay / hwday
                    huanansum += ihuanan
                workbooksheet6.cell(row=27, column=icol).value = huanansum
        elif icol in [6, 7]:
                workbooksheet6.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype6[icol % 2]]) / 10000
                for ihua in huanan6:
                    ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype6[icol % 2]]) / 10000
                    huanansum += ihuanan
                workbooksheet6.cell(row=27, column=icol).value = huanansum
        elif icol in [8, 9]:
                workbooksheet6.cell(row=26, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype6[icol % 2]]) / 10000
                for ihua in huanan6:
                    ihuanan =  sum(bigCdata[(bigCdata['体系'] == '海王') & (bigCdata['区域'] == ihua) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype6[icol % 2]]) / 10000
                    huanansum += ihuanan
                workbooksheet6.cell(row=27, column=icol).value = huanansum
        elif icol == 10:
            workbooksheet6.cell(row=26, column=icol).value = '=D' + str(26) + '/F' + str(26) + '-1'
            workbooksheet6.cell(row=27, column=icol).value = '=D' + str(27) + '/F' + str(27) + '-1'
        elif icol == 11:
            workbooksheet6.cell(row=26, column=icol).value = '=E' + str(26) + '/G' + str(26) + '-1'
            workbooksheet6.cell(row=27, column=icol).value = '=E' + str(27) + '/G' + str(27) + '-1'
        elif icol == 12:
            workbooksheet6.cell(row=26, column=icol).value = '=K' + str(26)
            workbooksheet6.cell(row=27, column=icol).value = '=K' + str(27)
        elif icol == 13:
            workbooksheet6.cell(row=26, column=icol).value = '=D' + str(26) + '/H' + str(26) + '-1'
            workbooksheet6.cell(row=27, column=icol).value = '=D' + str(27) + '/H' + str(27) + '-1'
        elif icol == 14:
            workbooksheet6.cell(row=26, column=icol).value = '=E' + str(26) + '/I' + str(26) + '-1'
            workbooksheet6.cell(row=27, column=icol).value = '=E' + str(27) + '/I' + str(27) + '-1'
        else:
            workbooksheet6.cell(row=26, column=icol).value = '=N' + str(26)
            workbooksheet6.cell(row=27, column=icol).value = '=N' + str(27)
    # ----------------------------------------------------------------------------
    # ⑦益丰
    # ----------------------------------------------------------------------------
    workbooksheet7 = workbook['⑦益丰']

    sumtype7 = ['数量', '标准金额']
    area7 = ['湘南', '长沙', '江苏', '益荔康信', '江西', '河北新兴', '江西天顺', '湘北', '湖北', '乡亲', '上海']

    # 非总计部分
    for icol in range(2, 11):
        for irow, iarea in enumerate(area7, start=4):
            if icol in [2, 3]:
                workbooksheet7.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == yfday)][sumtype7[icol % 2]])
            elif icol in [4, 5]:
                workbooksheet7.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype7[icol % 2]]) / last1Day
            elif icol == 6:
                workbooksheet7.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype7[icol % 2]])
            elif icol == 7:
                workbooksheet7.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype7[icol % 2]]) / 10000
            elif icol == 8:
                workbooksheet7.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '益丰') & (target['区域'] == iarea)  & (target['月季'] == globalMonth)]['目标'])
            elif icol == 9:
                workbooksheet7.cell(row=irow, column=icol).value = '=G' + str(irow) + '/H' + str(irow)
            else:
                workbooksheet7.cell(row=irow, column=icol).value = '=H' + str(irow) + '-G' + str(irow)

    # 总计部分
    for icol in range(2, 11):
        if icol in [2, 3]:
            workbooksheet7.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == yfday)][sumtype7[icol % 2]])
        elif icol in [4, 5]:
            workbooksheet7.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype7[icol % 2]]) / last1Day
        elif icol == 6:
            workbooksheet7.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype7[icol % 2]])
        elif icol == 7:
            workbooksheet7.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype7[icol % 2]]) / 10000
        elif icol == 8:
            workbooksheet7.cell(row=15, column=icol).value = sum(target[(target['体系'] == '益丰') & (target['月季'] == globalMonth)]['目标'])
        elif icol == 9:
            workbooksheet7.cell(row=15, column=icol).value = '=G' + str(15) + '/H' + str(15)
        else:
            workbooksheet7.cell(row=15, column=icol).value = '=H' + str(15) + '-G' + str(15)
    # ----------------------------------------------------------------------------
    # ⑧益丰
    # ----------------------------------------------------------------------------
    workbooksheet8 = workbook['⑧益丰']

    sumtype8 = ['数量', '标准金额']
    area8 = ['湘南', '长沙', '江苏', '益荔康信', '江西', '河北新兴', '江西天顺', '湘北', '湖北', '乡亲', '上海']

    # 非总计部分
    for icol in range(2, 16):
        for irow, iarea in enumerate(area8, start=4):
            if icol in [2, 3]:
                workbooksheet8.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype8[icol % 2]]) / 10000
            elif icol in [4, 5]:
                workbooksheet8.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype8[icol % 2]]) / 10000 * lastDay / yfday
            elif icol in [6, 7]:
                workbooksheet8.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype8[icol % 2]]) / 10000
            elif icol in [8, 9]:
                workbooksheet8.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype8[icol % 2]]) / 10000
            elif icol == 10:
                workbooksheet8.cell(row=irow, column=icol).value = '=D' + str(irow) + '/F' + str(irow) + '-1'
            elif icol == 11:
                workbooksheet8.cell(row=irow, column=icol).value = '=E' + str(irow) + '/G' + str(irow) + '-1'
            elif icol == 12:
                workbooksheet8.cell(row=irow, column=icol).value = '=K' + str(irow)
            elif icol == 13:
                workbooksheet8.cell(row=irow, column=icol).value = '=D' + str(irow) + '/H' + str(irow) + '-1'
            elif icol == 14:
                workbooksheet8.cell(row=irow, column=icol).value = '=E' + str(irow) + '/I' + str(irow) + '-1'
            else:
                workbooksheet8.cell(row=irow, column=icol).value = '=N' + str(irow)

    # 总计部分
    for icol in range(2, 16):
        if icol in [2, 3]:
            workbooksheet8.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype8[icol % 2]]) / 10000
        elif icol in [4, 5]:
            workbooksheet8.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)][sumtype8[icol % 2]]) / 10000 * lastDay / yfday
        elif icol in [6, 7]:
            workbooksheet8.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype8[icol % 2]]) / 10000
        elif icol in [8, 9]:
            workbooksheet8.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype8[icol % 2]]) / 10000
        elif icol == 10:
            workbooksheet8.cell(row=15, column=icol).value = '=D' + str(15) + '/F' + str(15) + '-1'
        elif icol == 11:
            workbooksheet8.cell(row=15, column=icol).value = '=E' + str(15) + '/G' + str(15) + '-1'
        elif icol == 12:
            workbooksheet8.cell(row=15, column=icol).value = '=K' + str(15)
        elif icol == 13:
            workbooksheet8.cell(row=15, column=icol).value = '=D' + str(15) + '/H' + str(15) + '-1'
        elif icol == 14:
            workbooksheet8.cell(row=15, column=icol).value = '=E' + str(15) + '/I' + str(15) + '-1'
        else:
            workbooksheet8.cell(row=15, column=icol).value = '=N' + str(15)
    # ----------------------------------------------------------------------------
    # ⑨大参林
    # ----------------------------------------------------------------------------
    workbooksheet9 = workbook['⑨大参林']

    sumtype9 = ['数量', '标准金额']
    area9 = ['广西大区', '华东大区', '粤中大区', '粤西大区', '河南大区', '鸡西', '保定', '南通', '西安', '新兴大区']

    # 非总计部分
    for icol in range(3, 12):
        for irow, iarea in enumerate(area9, start=4):
            if icol in [3, 4]:
                if irow in range(4, 9):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == dslday)][sumtype9[(icol - 1) % 2]])
                elif irow in range(9, 13):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == dslday)][sumtype9[(icol - 1) % 2]])
                else:
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == dslday)][sumtype9[(icol - 1) % 2]])
            elif icol in [5, 6]:
                if irow in range(4, 9):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / last1Day
                elif irow in range(9, 13):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / last1Day
                else:
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / last1Day
            elif icol == 7:
                if irow in range(4, 9):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]])
                elif irow in range(9, 13):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]])
                else:
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]])
            elif icol == 8:
                if irow in range(4, 9):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
                elif irow in range(9, 13):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
                else:
                    workbooksheet9.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
            elif icol == 9:
                if irow in range(4, 9):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '大参林') & (target['大区'] == iarea) & (target['月季'] == globalMonth)]['目标'])
                elif irow in range(9, 13):
                    workbooksheet9.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '大参林') & (target['区域'] == iarea) & (target['月季'] == globalMonth)]['目标'])
                else:
                    workbooksheet9.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '大参林') & (target['大区'] == '新兴大区') & (target['月季'] == globalMonth)]['目标'])
            elif icol == 10:
                workbooksheet9.cell(row=irow, column=icol).value = '=H' + str(irow) + '/I' + str(irow) 
            elif icol == 11:
                workbooksheet9.cell(row=irow, column=icol).value = '=I' + str(irow)  + '-H' + str(irow) 
                
    # 总计部分
    for icol in range(3, 12):
        if icol in [3, 4]:
            workbooksheet9.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == dslday)][sumtype9[(icol - 1) % 2]])
        elif icol in [5, 6]:
            workbooksheet9.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / last1Day
        elif icol == 7:
            workbooksheet9.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]])
        elif icol == 8:
            workbooksheet9.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
        elif icol == 9:
            workbooksheet9.cell(row=14, column=icol).value = sum(target[(target['体系'] == '大参林') & (target['月季'] == globalMonth)]['目标'])
        elif icol == 10:
            workbooksheet9.cell(row=14, column=icol).value = '=H' + str(14) + '/I' + str(14) 
        elif icol == 11:
            workbooksheet9.cell(row=14, column=icol).value = '=I' + str(14)  + '-H' + str(14) 
    # ----------------------------------------------------------------------------
    # ⑩大参林
    # ----------------------------------------------------------------------------
    workbooksheet10 = workbook['⑩大参林']

    sumtype10 = ['数量', '标准金额']
    area10 = ['广西大区', '华东大区', '粤中大区', '粤西大区', '河南大区', '鸡西', '保定', '南通', '西安', '新兴大区']

    # 非总计部分
    for icol in range(3, 17):
        for irow, iarea in enumerate(area10, start=4):
            if icol in [3, 4]:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
                elif irow in range(9, 13):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
                else:
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
            elif icol in [5, 6]:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000 * lastDay / dslday
                elif irow in range(9, 13):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000 * lastDay / dslday
                else:
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000 * lastDay / dslday
            elif icol in [7, 8]:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / 10000
                elif irow in range(9, 13):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / 10000
                else:
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / 10000
            elif icol in [9, 10]:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype9[(icol - 1) % 2]]) / 10000
                # elif irow in range(9, 13):
                #     workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype9[(icol - 1) % 2]]) / 10000
                # else:
                #    workbooksheet10.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & ((bigCdata['区域'] == '鸡西') | (bigCdata['区域'] == '保定') | (bigCdata['区域'] == '南通') |(bigCdata['区域'] == '西安')) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype9[(icol - 1) % 2]]) / 10000
            elif icol == 11:
                workbooksheet10.cell(row=irow, column=icol).value = '=E' + str(irow) + '/G' + str(irow) + '-1'
            elif icol == 12:
                workbooksheet10.cell(row=irow, column=icol).value = '=F' + str(irow) + '/H' + str(irow) + '-1'
            elif icol == 13:
                workbooksheet10.cell(row=irow, column=icol).value = '=L' + str(irow)
            elif icol == 14:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = '=E' + str(irow) + '/I' + str(irow) + '-1'
            elif icol == 15:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = '=F' + str(irow) + '/J' + str(irow) + '-1'
            else:
                if irow in range(4, 9):
                    workbooksheet10.cell(row=irow, column=icol).value = '=O' + str(irow)

    # 总计部分
    for icol in range(3, 17):
        if icol in [3, 4]:
            workbooksheet10.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000
        elif icol in [5, 6]:
            workbooksheet10.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)][sumtype9[(icol - 1) % 2]]) / 10000 * lastDay / dslday
        elif icol in [7, 8]:
            workbooksheet10.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype9[(icol - 1) % 2]]) / 10000
        elif icol in [9, 10]:
            workbooksheet10.cell(row=14, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype9[(icol - 1) % 2]]) / 10000
        elif icol == 11:
            workbooksheet10.cell(row=14, column=icol).value = '=E' + str(14) + '/G' + str(14) + '-1'
        elif icol == 12:
            workbooksheet10.cell(row=14, column=icol).value = '=F' + str(14) + '/H' + str(14) + '-1'
        elif icol == 13:
            workbooksheet10.cell(row=14, column=icol).value = '=L' + str(14)
        elif icol == 14:
            workbooksheet10.cell(row=14, column=icol).value = '=E' + str(14) + '/I' + str(14) + '-1'
        elif icol == 15:
            workbooksheet10.cell(row=14, column=icol).value = '=F' + str(14) + '/J' + str(14) + '-1'
        else:
            workbooksheet10.cell(row=14, column=icol).value = '=O' + str(14)
    # ----------------------------------------------------------------------------
    # 1高济
    # ----------------------------------------------------------------------------
    workbooksheet011 = workbook['1高济']

    sumtype011 = ['数量', '标准金额']
    # area011 = ['滁州华巨百姓缘', '肇庆邦健', '新疆济康', '汕头凯德', '清远百姓', '大连阳光', '云浮天天邦健', '宜兴天健', '惠州百姓', '广西梧州百姓', '福建宜又佳', '江西开心人', '四川海棠',
    # '揭阳新汇康', '阜阳第一大', '河南美锐', '广西宝和堂', '河源天天邦健', '安徽广济', '芜湖中山', '佛山邦健', '台山人民', '潮州新千禧', '恩平银星', '成都瑞康', '南通天天乐', '三门峡华为',
    # '河南天伦', '茂名南粤', '福建百泰', '河南百家好一生', '河北仁泰', '南阳隆泰仁', '成都成毅康缘', '惠州卫康', '成都倍成', '江门邦健', '成都华杏', '黄骅神农百草堂', '广元芝心', '湖北心连心',
    # '焦作蓝十字', '安徽敬贤堂', '章丘健民', '成都东升', '济源万康', '广安东升', '雅安和康', '开封百氏康', '河南同和堂', '新世纪（北京）', '南充嘉宝堂', '广州高济', '成都兴福', '泸州东升', '平顶山隆泰仁']
    area011 = ['华南', '华东', '新疆', '西南', '辽宁', '华中', '福建', '华北', '试点']
    
    # 非总计部分
    for icol in range(2, 11):
        for irow, iarea in enumerate(area011, start=4):
            if icol in [2, 3]:
                workbooksheet011.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == gjday)][sumtype011[icol % 2]])
            elif icol in [4, 5]:
                workbooksheet011.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype011[icol % 2]]) / last1Day
            elif icol == 6:
                workbooksheet011.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype011[icol % 2]])
            elif icol == 7:
                workbooksheet011.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype011[icol % 2]]) / 10000
            elif icol == 8:
                workbooksheet011.cell(row=irow, column=icol).value = sum(target[(target['体系'] == '高济')  & (target['大区'] == iarea) & (target['月季'] == globalMonth)]['目标'])
            elif icol == 9:
                workbooksheet011.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
            else:
                workbooksheet011.cell(row=irow, column=icol).value = '=IFERROR(H' + str(irow) + '-G' + str(irow) + ',"-")'

    # 总计部分
    for icol in range(2, 11):
        if icol in [2, 3]:
            workbooksheet011.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == gjday)][sumtype011[icol % 2]])
        elif icol in [4, 5]:
            workbooksheet011.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype011[icol % 2]]) / last1Day
        elif icol == 6:
            workbooksheet011.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype011[icol % 2]])
        elif icol == 7:
            workbooksheet011.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype011[icol % 2]]) / 10000
        elif icol == 8:
            workbooksheet011.cell(row=13, column=icol).value = sum(target[(target['体系'] == '高济') & (target['月季'] == globalMonth)]['目标'])
        elif icol == 9:
            workbooksheet011.cell(row=13, column=icol).value = '=IFERROR(G' + str(13) + '/H' + str(13) + ',"-")'
        else:
            workbooksheet011.cell(row=13, column=icol).value = '=IFERROR(H' + str(13)+ '-G' + str(13) + ',"-")'
    # ----------------------------------------------------------------------------
    # 2高济
    # ----------------------------------------------------------------------------
    workbooksheet012 = workbook['2高济']

    sumtype012 = ['数量', '标准金额']
    area012 = ['华南', '华东', '新疆', '西南', '辽宁', '华中', '福建', '华北', '试点']
    
    # 非总计部分
    for icol in range(2, 16):
        for irow, iarea in enumerate(area012, start=4):
            if icol in [2, 3]:
                    workbooksheet012.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype012[icol % 2]]) / 10000
            elif icol in [4, 5]:
                    workbooksheet012.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype012[icol % 2]]) / 10000 * lastDay / gjday
            elif icol in [6, 7]:
                    workbooksheet012.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype012[icol % 2]]) / 10000
            elif icol in [8, 9]:
                    workbooksheet012.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['大区'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype012[icol % 2]]) / 10000
            elif icol == 10:
                    workbooksheet012.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/F' + str(irow) + '-1,"-")'
            elif icol == 11:
                workbooksheet012.cell(row=irow, column=icol).value = '=IFERROR(E' + str(irow) + '/G' + str(irow) + '-1,"-")'
            elif icol == 12:
                workbooksheet012.cell(row=irow, column=icol).value = '=K' + str(irow)
            elif icol == 13:
                workbooksheet012.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/H' + str(irow) + '-1,"-")'
            elif icol == 14:
                workbooksheet012.cell(row=irow, column=icol).value = '=IFERROR(E' + str(irow) + '/I' + str(irow) + '-1,"-")'
            else:
                workbooksheet012.cell(row=irow, column=icol).value = '=N' + str(irow)

    # 总计部分
    for icol in range(2, 16):
        if icol in [2, 3]:
                workbooksheet012.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype012[icol % 2]]) / 10000
        elif icol in [4, 5]:
                workbooksheet012.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= gjday)][sumtype012[icol % 2]]) / 10000 * lastDay / gjday
        elif icol in [6, 7]:
                workbooksheet012.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))][sumtype012[icol % 2]]) / 10000
        elif icol in [8, 9]:
                workbooksheet012.cell(row=13, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '高济') & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)][sumtype012[icol % 2]]) / 10000
        elif icol == 10:
            workbooksheet012.cell(row=13, column=icol).value = '=IFERROR(D' + str(13) + '/F' + str(13) + '-1,"-")'
        elif icol == 11:
            workbooksheet012.cell(row=13, column=icol).value = '=IFERROR(E' + str(13) + '/G' + str(13) + '-1,"-")'
        elif icol == 12:
            workbooksheet012.cell(row=13, column=icol).value = '=K' + str(13)
        elif icol == 13:
            workbooksheet012.cell(row=13, column=icol).value = '=IFERROR(D' + str(13) + '/H' + str(13) + '-1,"-")'
        elif icol == 14:
            workbooksheet012.cell(row=13, column=icol).value = '=IFERROR(E' + str(13) + '/I' + str(13) + '-1,"-")'
        else:
            workbooksheet012.cell(row=13, column=icol).value = '=N' + str(13)
    # ----------------------------------------------------------------------------
    # 3益丰
    # ----------------------------------------------------------------------------
    workbooksheet013 = workbook['3益丰']

    area013 = ['湘南', '长沙', '江苏', '益荔康信', '江西', '河北新兴', '江西天顺', '湘北', '湖北', '乡亲', '上海']

    # 非总计部分
    for icol in range(2, 11):
        for irow, iarea in enumerate(area013, start=4):
            if icol == 2:
                workbooksheet013.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == yfday)]['标准金额'])
            elif icol == 3:
                workbooksheet013.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)]['标准金额']) / 10000
            elif icol == 4:
                workbooksheet013.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))]['标准金额']) / 10000
            elif icol == 5:
                workbooksheet013.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)]['标准金额']) / 10000
            elif icol == 6:
                workbooksheet013.cell(row=irow, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['区域'] == iarea) & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth)]['标准金额']) / 10000 / sum(target[(target['体系'] == '益丰') & (target['区域'] == iarea)  & (target['月季'] == globalMonth)]['目标'])
            elif icol == 7:
                workbooksheet013.cell(row=irow, column=icol).value = '=IFERROR((C' + str(irow) + '/' + str(yfday) + ')/(D' + str(irow) + '/' + str(last1Day) + ')-1,"-")'
            elif icol == 8:
                workbooksheet013.cell(row=irow, column=icol).value = '=G' + str(irow)
            elif icol == 9:
                workbooksheet013.cell(row=irow, column=icol).value = '=IFERROR((C' + str(irow) + '/' + str(yfday) + ')/(E' + str(irow) + '/' + str(last2Day) + ')-1,"-")'
            else:
                workbooksheet013.cell(row=irow, column=icol).value = '=I' + str(irow)

    # 总计部分
    for icol in range(2, 11):
        if icol == 2:
            workbooksheet013.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] == yfday)]['标准金额'])
        elif icol == 3:
            workbooksheet013.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth)]['标准金额']) / 10000
        elif icol == 4:
            workbooksheet013.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == (globalMonth - 1))]['标准金额']) / 10000
        elif icol == 5:
            workbooksheet013.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == (globalYear - 1)) & (bigCdata['月'] == globalMonth)]['标准金额']) / 10000
        elif icol == 6:
            workbooksheet013.cell(row=15, column=icol).value = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth)]['标准金额']) / 10000 / sum(target[(target['体系'] == '益丰') & (target['月季'] == globalMonth)]['目标'])
        elif icol == 7:
            workbooksheet013.cell(row=15, column=icol).value = '=IFERROR((C' + str(15) + '/' + str(yfday) + ')/(D' + str(15) + '/' + str(last1Day) + ')-1,"-")'
        elif icol == 8:
            workbooksheet013.cell(row=15, column=icol).value = '=G' + str(15)
        elif icol == 9:
            workbooksheet013.cell(row=15, column=icol).value = '=IFERROR((C' + str(15) + '/' + str(yfday) + ')/(E' + str(15) + '/' + str(last2Day) + ')-1,"-")'
        else:
            workbooksheet013.cell(row=15, column=icol).value = '=I' + str(15)
        # ----------------------------------------------------------------------------
        # 4老百姓彩盒
        # ----------------------------------------------------------------------------
        workbooksheet014 = workbook['4老百姓彩盒']

        sumtype014 = ['数量', '标准金额']
        area014 = ['湘北', '湘中', '湘南', '百姓缘', '天津', '河南', '湖北', '陕西', '广西', '山东', '浙江']

        for icol in range(2, 10):
            for irow, iarea in enumerate(area014, start=4):
                if icol in [2, 3]:
                    workbooksheet014.cell(row=irow, column=icol).value = sum(colordata[(colordata['区域'] == iarea) & (colordata['年'] == globalYear) & (colordata['月'] == globalMonth) & (colordata['日'] == lbxday)][sumtype014[icol % 2]])
                elif icol in [4, 5]:
                    workbooksheet014.cell(row=irow, column=icol).value = sum(colordata[(colordata['区域'] == iarea) & (colordata['年'] == globalYear) & (colordata['月'] == globalMonth) & (colordata['日'] <= lbxday)][sumtype014[icol % 2]])
                elif icol in [6]:
                    workbooksheet014.cell(row=irow, column=icol).value = sum(lbxCofor[lbxCofor['区域'] == iarea]['目标'])
                elif icol in [7]:
                    workbooksheet014.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/F' + str(irow) + ',"-")'
                elif icol in [8, 9]:
                    workbooksheet014.cell(row=irow, column=icol).value = sum(colordata[(colordata['区域'] == iarea) & (colordata['年'] == globalYear) & (colordata['月'] <= globalMonth)][sumtype014[icol % 2]])
            # 总计部分
            if icol in [2, 3]:
                workbooksheet014.cell(row=15, column=icol).value = sum(colordata[(colordata['年'] == globalYear) & (colordata['月'] == globalMonth) & (colordata['日'] == lbxday)][sumtype014[icol % 2]])
            elif icol in [4, 5]:
                workbooksheet014.cell(row=15, column=icol).value = sum(colordata[(colordata['年'] == globalYear) & (colordata['月'] == globalMonth) & (colordata['日'] <= lbxday)][sumtype014[icol % 2]])
            elif icol in [6]:
                workbooksheet014.cell(row=15, column=icol).value = sum(lbxCofor['目标'])
            elif icol in [7]:
                workbooksheet014.cell(row=15, column=icol).value = '=IFERROR(D' + str(15) + '/F' + str(15) + ',"-")'
            elif icol in [8, 9]:
                workbooksheet014.cell(row=15, column=icol).value = sum(colordata[(colordata['年'] == globalYear) & (colordata['月'] <= globalMonth)][sumtype014[icol % 2]])
    # ----------------------------------------------------------------------------
    # 带时间文字填充
    # ----------------------------------------------------------------------------
    # 第 1 页
    workbooksheet1.cell(row=1, column=1).value = '大客户' + str(globalYear) + '年度简析'
    workbooksheet1.cell(row=3, column=2).value = str(globalYear) + '年' + str(globalMonth) + '月'
    workbooksheet1.cell(row=4, column=2).value = '(1-' + str(max(lbxday, hwday, yfday, dslday, gjday)) + '日)'
    workbooksheet1.cell(row=4, column=4).value = str(globalMonth) + '月（预估）'
    workbooksheet1.cell(row=3, column=6).value = str(globalYear - 1) + '年' + str(globalMonth) + '月'
    workbooksheet1.cell(row=3, column=8).value = str(globalYear) + '年' + str(globalMonth - 1) + '月'
    workbooksheet1.cell(row=4, column=10).value = str(globalYear) + '年(1.1-' + str(globalMonth) + '.' + str(max(lbxday, hwday, yfday, dslday, gjday)) + ')'
    workbooksheet1.cell(row=4, column=12).value = str(globalYear - 1)  + '年1-' + str(globalMonth) + '月'
    workbooksheet1.cell(row=3, column=14).value = str(globalMonth) + '月'
    workbooksheet1.cell(row=3, column=20).value =  '1-' + str(globalMonth) + '月累计纯销'

    # 第 3 页
    lbxSchedule = sum(bigCdata[(bigCdata['体系'] == '老百姓') & (bigCdata['年'] == globalYear) & (bigCdata['月'] >= quarterlyStart) & (bigCdata['月'] <= quarterlyEnd)]['标准金额']) / 10000 / sum(target[(target['体系'] == '老百姓') & (target['月季'] == quarterly)]['目标'])
    workbooksheet3.cell(row=1, column=1).value = '老百姓-' + str(globalMonth) + '月(' + str(quarterlyStart) + '.1-' + str(globalMonth) + '.' + str(lbxday) + '日)达成率' + str("{:.1f}%".format(lbxSchedule * 100)) + '(' + quarterly + '时间进度' + str("{:.0f}%".format(quarterlyDay / quarterlyDayAll * 100)) + ')'
    workbooksheet3.cell(row=2, column=2).value =  str(globalMonth) + '月' + str(lbxday) + '日'
    workbooksheet3.cell(row=2, column=4).value =  str(globalYear) + '年' + str(globalMonth - 1) + '月份日均纯销'
    workbooksheet3.cell(row=2, column=6).value =  str(globalMonth) + '月'
    workbooksheet3.cell(row=2, column=8).value =  str(quarterlyStart) + '-' + str(quarterlyEnd) + '月'
    workbooksheet3.cell(row=2, column=10).value = str(quarterlyStart) + '-' + str(quarterlyEnd) + '月'
    workbooksheet3.cell(row=35, column=1).value = '1. 第' + quarterly[1] + '季度按' + str(quarterlyDayAll) + '天方案执行;'

    # 第 4 页
    workbooksheet4.cell(row=1, column=2).value = str(globalMonth) + '月'
    workbooksheet4.cell(row=2, column=2).value = '(1-' + str(lbxday) + '日)'
    workbooksheet4.cell(row=1, column=4).value = str(globalMonth) + '月纯销'
    workbooksheet4.cell(row=1, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月纯销'
    workbooksheet4.cell(row=1, column=8).value = str(globalYear - 1) + '年' + str(globalMonth) + '月纯销'
    workbooksheet4.cell(row=1, column=10).value = str(globalMonth) + '月'

    # 第 5 页
    workbooksheet5.cell(row=1, column=1).value = '海王-' + str(globalMonth) + '月销售（1-' + str(hwday) + '日)-时间进度' + str("{:.0f}%".format(hwday / lastDay * 100))
    workbooksheet5.cell(row=2, column=2).value = str(globalMonth) + '月' + str(hwday) + '日'
    workbooksheet5.cell(row=2, column=4).value = str(globalYear) + '年' + str(globalMonth - 1) + '月日均纯销'
    workbooksheet5.cell(row=2, column=6).value = str(globalMonth) + '月'
    workbooksheet5.cell(row=2, column=8).value = str(globalMonth) + '月目标(万元)'

    # 第 6 页
    workbooksheet6.cell(row=1, column=2).value = str(globalMonth) + '月'
    workbooksheet6.cell(row=2, column=2).value = '(1-' + str(hwday) + '日)'
    workbooksheet6.cell(row=1, column=4).value = str(globalMonth) + '月纯销'
    workbooksheet6.cell(row=1, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月纯销'
    workbooksheet6.cell(row=1, column=8).value = str(globalYear - 1) + '年' + str(globalMonth) + '月纯销'
    workbooksheet6.cell(row=1, column=10).value = str(globalMonth) + '月'

    # 第 7 页
    yfSchedule = sum(bigCdata[(bigCdata['体系'] == '益丰') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= yfday)]['标准金额']) / 100 / sum(target[(target['体系'] == '益丰') & (target['月季'] == globalMonth)]['目标'])
    workbooksheet7.cell(row=1, column=1).value = ' 益丰-' + str(globalMonth) + '月（1-' + str(yfday) + '日）全月目标达成率' +  str("{:.0f}%".format(yfSchedule))
    workbooksheet7.cell(row=2, column=2).value = str(globalMonth) + '月' + str(yfday) + '日'
    workbooksheet7.cell(row=2, column=4).value = str(globalYear) + '年' + str(globalMonth - 1) + '月日均纯销'
    workbooksheet7.cell(row=2, column=6).value = str(globalMonth)  + '月'
    workbooksheet7.cell(row=2, column=8).value = str(globalMonth)  + '月目标'

    # 第 8 页
    workbooksheet8.cell(row=1, column=2).value = str(globalMonth) + '月'
    workbooksheet8.cell(row=2, column=2).value = '(1-' + str(yfday) + '日)'
    workbooksheet8.cell(row=1, column=4).value = str(globalMonth) + '月纯销'
    workbooksheet8.cell(row=1, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月纯销'
    workbooksheet8.cell(row=1, column=8).value = str(globalYear - 1) + '年' + str(globalMonth) + '月纯销'
    workbooksheet8.cell(row=1, column=10).value = str(globalMonth) + '月'

    # 第 9 页
    dslSchedule = sum(bigCdata[(bigCdata['体系'] == '大参林') & (bigCdata['年'] == globalYear) & (bigCdata['月'] == globalMonth) & (bigCdata['日'] <= dslday)]['标准金额']) / 100 / sum(target[(target['体系'] == '大参林') & (target['月季'] == globalMonth)]['目标'])
    workbooksheet9.cell(row=1, column=1).value = ' 大参林-' + str(globalMonth) + '月(1-' + str(dslday) + '日)全月目标达成率' +  str("{:.0f}%".format(dslSchedule))
    workbooksheet9.cell(row=2, column=3).value = str(globalMonth) + '月' + str(dslday) + '日'
    workbooksheet9.cell(row=2, column=5).value = str(globalYear) + '年' + str(globalMonth - 1) + '月日均纯销'
    workbooksheet9.cell(row=2, column=7).value = str(globalMonth)  + '月'
    workbooksheet9.cell(row=2, column=9).value = str(globalMonth)  + '月目标'

    # 第 10 页
    workbooksheet10.cell(row=1, column=3).value = str(globalMonth) + '月'
    workbooksheet10.cell(row=2, column=3).value = '(1-' + str(dslday) + '日)'
    workbooksheet10.cell(row=1, column=5).value = str(globalMonth) + '月纯销'
    workbooksheet10.cell(row=1, column=7).value = str(globalYear) + '年' + str(globalMonth - 1) + '月纯销'
    workbooksheet10.cell(row=1, column=9).value = str(globalYear - 1) + '年' + str(globalMonth) + '月纯销'
    workbooksheet10.cell(row=1, column=11).value = str(globalMonth) + '月'
    
    # 第 011 页
    workbooksheet011.cell(row=1, column=1).value = '高济-' + str(globalMonth) + '月销售（1-' + str(gjday) + '日)-时间进度' + str("{:.0f}%".format(gjday / lastDay * 100))
    workbooksheet011.cell(row=2, column=2).value = str(globalMonth) + '月' + str(gjday) + '日'
    workbooksheet011.cell(row=2, column=4).value = str(globalYear) + '年' + str(globalMonth - 1) + '月日均纯销'
    workbooksheet011.cell(row=2, column=6).value = str(globalMonth) + '月'
    workbooksheet011.cell(row=2, column=8).value = str(globalMonth) + '月目标(万元)'

    # 第 012 页
    workbooksheet012.cell(row=1, column=2).value = str(globalMonth) + '月'
    workbooksheet012.cell(row=2, column=2).value = '(1-' + str(gjday) + '日)'
    workbooksheet012.cell(row=1, column=4).value = str(globalMonth) + '月纯销'
    workbooksheet012.cell(row=1, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月纯销'
    workbooksheet012.cell(row=1, column=8).value = str(globalYear - 1) + '年' + str(globalMonth) + '月纯销'
    workbooksheet012.cell(row=1, column=10).value = str(globalMonth) + '月'

    # 第 013 页
    workbooksheet013.cell(row=1, column=1).value = ' 益丰-' + str(globalMonth) + '月（1-' + str(yfday) + '日）日均金额同环比'
    workbooksheet013.cell(row=2, column=2).value = str(globalMonth) + '月' + str(yfday) + '日'
    workbooksheet013.cell(row=2, column=3).value = str(globalMonth) + '月（1-' + str(yfday) + '日）'
    workbooksheet013.cell(row=2, column=4).value = str(globalYear) + '年' + str(globalMonth - 1) + '月'
    workbooksheet013.cell(row=2, column=5).value = str(globalYear - 1) + '年' + str(globalMonth) + '月'
    workbooksheet013.cell(row=2, column=7).value = str(globalMonth) + '月日均环比'
    workbooksheet013.cell(row=2, column=9).value = str(globalMonth) + '月日均同比'

    # 第 014 页
    workbooksheet014.cell(row=1, column=1).value = '老百姓-' + str(globalMonth) + '月(1-' + str(lbxday) + '日)彩盒各区域销售情况'
    workbooksheet014.cell(row=2, column=2).value = str(globalMonth) + '月' + str(lbxday) + '日'
    workbooksheet014.cell(row=2, column=4).value = str(globalMonth) + '月'
    workbooksheet014.cell(row=2, column=6).value = str(globalMonth) + '月目标达成'
    workbooksheet014.cell(row=2, column=8).value = '1-' + str(globalMonth) + '月'
    # ----------------------------------------------------------------------------
    # 存储表格
    # ----------------------------------------------------------------------------
    workbook.save('C:/Users/Zeus/Desktop/autoSend/2_大客户/大客户_正式发送文件.xlsx')

    
# 生成单品文件
def material_graph(savePath, materialdata, materialAilas):

    # 加载益丰黄芪 鱼腥草单品数据
    yfsingle =  pd.read_excel('C:/Users/Zeus/Desktop/autoSend/2_大客户/目标/大客户_数据源.xlsx', sheet_name=2, header=0)

    # ----------------------------------------------------------------------------
    # 加载格式框架
    # ----------------------------------------------------------------------------
    filepath2 = 'C:/Users/Zeus/Desktop/autoSend/2_大客户/目标/单品_格式框架.xlsx'
    work_book = opxl.load_workbook(filepath2)
    # ----------------------------------------------------------------------------
    # Ⅰ老百姓单品
    # ----------------------------------------------------------------------------
    workbooksheet11 = work_book['Ⅰ老百姓']

    area11 = ['天津', '江苏', '陕西', '百佳惠', '通辽泽强', '邻加医', '惠仁堂', '上海', '湖北', '隆泰源', '湘南', '湘北', '湘中', '百姓缘', '南通普泽', '百信缘', '山东', '河南', '万仁', '广西', '三品堂', '新千秋', '仁德', '广东', '江苏海鹏', '华康', '浙江', '河北', '山西百汇', '赤峰人川']

    # 非总计部分
    for icol in range(2, 9):
        for irow, iarea in enumerate(area11, start=5):
            if icol == 2:
                workbooksheet11.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == lbxday)]['数量'])
            elif icol == 3:
                workbooksheet11.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= lbxday)]['数量'])
            elif icol == 4:
                workbooksheet11.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= lbxday)]['数量']) * lastDay / lbxday
            elif icol == 5:
                workbooksheet11.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
            elif icol == 6:
                workbooksheet11.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
            elif icol == 7:
                workbooksheet11.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/E' + str(irow) + '-1,"-")'
            elif icol == 8:
                workbooksheet11.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/F' + str(irow) + '-1,"-")'

    # 总计部分
    for icol in range(2, 9):
        if icol == 2:
                workbooksheet11.cell(row=35, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == lbxday)]['数量'])
        elif icol == 3:
                workbooksheet11.cell(row=35, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= lbxday)]['数量'])
        elif icol == 4:
                workbooksheet11.cell(row=35, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= lbxday)]['数量']) * lastDay / lbxday
        elif icol == 5:
                workbooksheet11.cell(row=35, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
        elif icol == 6:
                workbooksheet11.cell(row=35, column=icol).value = sum(materialdata[(materialdata['体系'] == '老百姓') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
        elif icol == 7:
                workbooksheet11.cell(row=35, column=icol).value = '=IFERROR(D' + str(35) + '/E' + str(35) + '-1,"-")'
        elif icol == 8:
                workbooksheet11.cell(row=35, column=icol).value = '=IFERROR(D' + str(35) + '/F' + str(35) + '-1,"-")'
    # ----------------------------------------------------------------------------
    # Ⅱ海王单品
    # ----------------------------------------------------------------------------
    workbooksheet12 = work_book['Ⅱ海王']

    area12 = ['连云港', '上海', '福州', '蚌埠', '惠州', '常州', '无锡', '潍坊', '青岛', '深圳', '南京', '湖北', '佛山', '大连', '成都', '沈阳', '湖南', '苏州', '中珠', '东莞', '广州']

    # 非总计部分
    for icol in range(2, 9):
        for irow, iarea in enumerate(area12, start=5):
            if icol == 2:
                workbooksheet12.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == hwday)]['数量'])
            elif icol == 3:
                workbooksheet12.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= hwday)]['数量'])
            elif icol == 4:
                workbooksheet12.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= hwday)]['数量']) * lastDay / hwday
            elif icol == 5:
                workbooksheet12.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
            elif icol == 6:
                workbooksheet12.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
            elif icol == 7:
                workbooksheet12.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/E' + str(irow) + '-1,"-")'
            elif icol == 8:
                workbooksheet12.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/F' + str(irow) + '-1,"-")'

    # 总计部分
    for icol in range(2, 9):
        if icol == 2:
                workbooksheet12.cell(row=26, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == hwday)]['数量'])
        elif icol == 3:
                workbooksheet12.cell(row=26, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= hwday)]['数量'])
        elif icol == 4:
                workbooksheet12.cell(row=26, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= hwday)]['数量']) * lastDay / hwday
        elif icol == 5:
                workbooksheet12.cell(row=26, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
        elif icol == 6:
                workbooksheet12.cell(row=26, column=icol).value = sum(materialdata[(materialdata['体系'] == '海王') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
        elif icol == 7:
                workbooksheet12.cell(row=26, column=icol).value = '=IFERROR(D' + str(26) + '/E' + str(26) + '-1,"-")'
        elif icol == 8:
                workbooksheet12.cell(row=26, column=icol).value = '=IFERROR(D' + str(26) + '/F' + str(26) + '-1,"-")'
    # ----------------------------------------------------------------------------
    # Ⅲ益丰单品
    # ----------------------------------------------------------------------------
    workbooksheet13 = work_book['Ⅲ益丰']

    area13 = ['湘南', '长沙', '江苏', '益荔康信', '江西', '河北新兴', '江西天顺', '湘北', '湖北', '乡亲', '上海']

    # 非总计部分
    for icol in range(2, 11):
        for irow, iarea in enumerate(area13, start=5):
            if icol == 2:
                workbooksheet13.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == yfday)]['数量'])
            elif icol == 3:
                workbooksheet13.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= yfday)]['数量'])
            elif icol == 4:
                workbooksheet13.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= yfday)]['数量']) * lastDay / yfday
            elif icol == 5:
                workbooksheet13.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
            elif icol == 6:
                workbooksheet13.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
            elif icol == 7:
                workbooksheet13.cell(row=irow, column=icol).value = sum(yfsingle[(yfsingle['区域'] == iarea) & (yfsingle['物料'] == materialAilas)]['标准任务'])
            elif icol == 8:
                workbooksheet13.cell(row=irow, column=icol).value = '=IFERROR(C' + str(irow) + '/G' + str(irow) + ',"-")'
            elif icol == 9:
                workbooksheet13.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/E' + str(irow) + '-1,"-")'
            elif icol == 10:
                workbooksheet13.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '/F' + str(irow) + '-1,"-")'

    # 总计部分
    for icol in range(2, 11):
        if icol == 2:
                workbooksheet13.cell(row=16, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == yfday)]['数量'])
        elif icol == 3:
                workbooksheet13.cell(row=16, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= yfday)]['数量'])
        elif icol == 4:
                workbooksheet13.cell(row=16, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= yfday)]['数量']) * lastDay / yfday
        elif icol == 5:
                workbooksheet13.cell(row=16, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
        elif icol == 6:
                workbooksheet13.cell(row=16, column=icol).value = sum(materialdata[(materialdata['体系'] == '益丰') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
        elif icol == 7:
                workbooksheet13.cell(row=16, column=icol).value = sum(yfsingle[(yfsingle['物料'] == materialAilas)]['标准任务'])
        elif icol == 8:
                workbooksheet13.cell(row=16, column=icol).value =  '=IFERROR(C' + str(16) + '/G' + str(16) + ',"-")'
        elif icol == 9:
                workbooksheet13.cell(row=16, column=icol).value = '=IFERROR(D' + str(16) + '/E' + str(16) + '-1,"-")'
        elif icol == 10:
                workbooksheet13.cell(row=16, column=icol).value = '=IFERROR(D' + str(16) + '/F' + str(16) + '-1,"-")'
    # ----------------------------------------------------------------------------
    # Ⅳ大参林单品
    # ----------------------------------------------------------------------------
    workbooksheet14 = work_book['Ⅳ大参林']

    area14 = ['广西大区', '华东大区', '粤中大区', '粤西大区', '河南大区', '鸡西', '保定', '南通', '西安', '新兴大区']

    # 非总计部分
    for icol in range(3, 10):
        for irow, iarea in enumerate(area14, start=5):
            if icol == 3:
                if irow in [5, 6, 7, 8, 9]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == dslday)]['数量'])
                elif irow in [10, 11, 12, 13]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == dslday)]['数量'])
                else:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & ((materialdata['区域'] == '鸡西') | (materialdata['区域'] == '保定') | (materialdata['区域'] == '南通') | (materialdata['区域'] == '西安')) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == dslday)]['数量'])
            elif icol == 4:
                if irow in [5, 6, 7, 8, 9]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量'])
                elif irow in [10, 11, 12, 13]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量'])
                else:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & ((materialdata['区域'] == '鸡西') | (materialdata['区域'] == '保定') | (materialdata['区域'] == '南通') | (materialdata['区域'] == '西安')) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量'])
            elif icol == 5:
                if irow in [5, 6, 7, 8, 9]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量']) * lastDay / dslday
                elif irow in [10, 11, 12, 13]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量']) * lastDay / dslday
                else:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & ((materialdata['区域'] == '鸡西') | (materialdata['区域'] == '保定') | (materialdata['区域'] == '南通') | (materialdata['区域'] == '西安')) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量']) * lastDay / dslday
            elif icol == 6:
                if irow in [5, 6, 7, 8, 9]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
                # elif irow in [10, 11, 12, 13]:
                #     workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
                # else:
                #     workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & ((materialdata['区域'] == '鸡西') | (materialdata['区域'] == '保定') | (materialdata['区域'] == '南通') | (materialdata['区域'] == '西安')) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
            elif icol == 7:
                if irow in [5, 6, 7, 8, 9]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['大区'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
                elif irow in [10, 11, 12, 13]:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['区域'] == iarea) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
                else:
                    workbooksheet14.cell(row=irow, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & ((materialdata['区域'] == '鸡西') | (materialdata['区域'] == '保定') | (materialdata['区域'] == '南通') | (materialdata['区域'] == '西安')) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
            elif icol == 8:
                if irow in [5, 6, 7, 8, 9]:
                    workbooksheet14.cell(row=irow, column=icol).value = '=IFERROR(E' + str(irow) + '/F' + str(irow) + '-1,"-")'
            elif icol == 9:
                workbooksheet14.cell(row=irow, column=icol).value = '=IFERROR(E' + str(irow) + '/G' + str(irow) + '-1,"-")'

    # 总计部分
    for icol in range(3, 10):
        if icol == 3:
            workbooksheet14.cell(row=15, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] == dslday)]['数量'])
        elif icol == 4:
            workbooksheet14.cell(row=15, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量'])
        elif icol == 5:
            workbooksheet14.cell(row=15, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == globalMonth) & (materialdata['日'] <= dslday)]['数量']) * lastDay / dslday
        elif icol == 6:
            workbooksheet14.cell(row=15, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == (globalYear - 1)) & (materialdata['月'] == globalMonth)]['数量'])
        elif icol == 7:
            workbooksheet14.cell(row=15, column=icol).value = sum(materialdata[(materialdata['体系'] == '大参林') & (materialdata['物料'] == materialAilas) & (materialdata['年'] == globalYear) & (materialdata['月'] == (globalMonth - 1))]['数量'])
        elif icol == 8:
            workbooksheet14.cell(row=15, column=icol).value = '=IFERROR(E' + str(15) + '/F' + str(15) + '-1,"-")'
        elif icol == 9:
            workbooksheet14.cell(row=15, column=icol).value = '=IFERROR(E' + str(15) + '/G' + str(15) + '-1,"-")'
    # ----------------------------------------------------------------------------
    # 带时间文字填充
    # ----------------------------------------------------------------------------
    # 第 11 页
    workbooksheet11.cell(row=1, column=1).value = '老百姓' + str(globalMonth) + '月(1-' + str(lbxday) + '日)-'
    workbooksheet11.cell(row=1, column=5).value = materialAilas
    workbooksheet11.cell(row=2, column=2).value = str(globalMonth) + '月' + str(lbxday) + '日'
    workbooksheet11.cell(row=2, column=3).value = str(globalMonth) + '月'
    workbooksheet11.cell(row=2, column=5).value = str(globalYear - 1) + '年' + str(globalMonth) + '月'
    workbooksheet11.cell(row=2, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月'
    workbooksheet11.cell(row=2, column=7).value = str(globalMonth) + '月'

    # 第 12 页
    workbooksheet12.cell(row=1, column=1).value = '海王' + str(globalMonth) + '月(1-' + str(hwday) + '日)-'
    workbooksheet12.cell(row=1, column=5).value = materialAilas
    workbooksheet12.cell(row=2, column=2).value = str(globalMonth) + '月' + str(hwday) + '日'
    workbooksheet12.cell(row=2, column=3).value = str(globalMonth) + '月'
    workbooksheet12.cell(row=2, column=5).value = str(globalYear - 1) + '年' + str(globalMonth) + '月'
    workbooksheet12.cell(row=2, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月'
    workbooksheet12.cell(row=2, column=7).value = str(globalMonth) + '月'

    # 第 13 页
    workbooksheet13.cell(row=1, column=1).value = '益丰' + str(globalMonth) + '月(1-' + str(yfday) + '日)-'
    workbooksheet13.cell(row=1, column=5).value = materialAilas
    workbooksheet13.cell(row=2, column=2).value = str(globalMonth) + '月' + str(yfday) + '日'
    workbooksheet13.cell(row=2, column=3).value = str(globalMonth) + '月'
    workbooksheet13.cell(row=2, column=5).value = str(globalYear - 1) + '年' + str(globalMonth) + '月'
    workbooksheet13.cell(row=2, column=6).value = str(globalYear) + '年' + str(globalMonth - 1) + '月'
    workbooksheet13.cell(row=2, column=7).value = str(globalMonth) + '月'

    # 第 14 页
    workbooksheet14.cell(row=1, column=1).value = '大参林' + str(globalMonth) + '月(1-' + str(dslday) + '日)-'
    workbooksheet14.cell(row=1, column=6).value = materialAilas
    workbooksheet14.cell(row=2, column=3).value = str(globalMonth) + '月' + str(dslday) + '日'
    workbooksheet14.cell(row=2, column=4).value = str(globalMonth) + '月'
    workbooksheet14.cell(row=2, column=6).value = str(globalYear - 1) + '年' + str(globalMonth) + '月'
    workbooksheet14.cell(row=2, column=7).value = str(globalYear) + '年' + str(globalMonth - 1) + '月'
    workbooksheet14.cell(row=2, column=8).value = str(globalMonth) + '月'
    # ----------------------------------------------------------------------------
    # 存储表格
    # ----------------------------------------------------------------------------
    work_book.save(savePath + materialAilas + '单品_正式发送文件.xlsx')


# 四大客户 2018 - 至今 销售日月均状图
def bar_graph():
    # 连接数据库
    conn = pymysql.connect(host='192.168.20.241',
                            port    = 3306,
                            user    = 'root',
                            passwd  = 'Powerbi#1217',
                            db      = 'dkh',
                            charset = 'utf8')

    cursorGraph    = conn.cursor()

    # 柱状图数据
    executeGraph = """WITH AA AS (SELECT YEAR(date) AS NIAN, MONTH(date) AS YUE, customer, SUM(amount) AS SUM1
                    FROM dkhfact
                    WHERE customer IN ('老百姓', '海王', '益丰', '大参林', '高济') AND ((YEAR(date) != """ + str(globalYear) + """ AND MONTH(date) <= 12) OR (YEAR(date) = """ + str(globalYear) + """ AND MONTH(date) < """ + str(globalMonth) + """))
                    GROUP BY date, customer)
                    (SELECT AA.NIAN, AA.YUE, '' AS TIAN, AA.customer, SUM(SUM1) AS SUM2
                    FROM AA
                    GROUP BY AA.NIAN ASC, AA.YUE ASC, AA.customer)
                    UNION ALL
                    (SELECT YEAR(date) AS NIAN, MONTH(date) AS YUE, DAY(date) AS TIAN, customer, SUM(amount) AS SUM2
                    FROM dkhfact
                    WHERE customer IN ('老百姓', '海王', '益丰', '大参林', '高济') AND YEAR(date) = """ + str(globalYear) + """ AND MONTH(date) = """ + str(globalMonth) + """
                    GROUP BY date ASC, customer)"""

    cursorGraph.execute(executeGraph)
    dataGraph    = cursorGraph.fetchall()  # 柱状图数据
    conn.commit()
    cursorGraph.close()
    conn.close()

    graphdata = pd.DataFrame(dataGraph , columns=['年', '月', '日', '体系', '数量']) 
    # ----------------------------------------------------------------------------
    # 数据清洗区
    # ----------------------------------------------------------------------------
    graphdata.insert(0, '年月', '')
    graphdata.insert(4, '月均', '')
    for irow in range(graphdata.shape[0]):
        if graphdata.loc[irow, '年'] == globalYear and graphdata.loc[irow, '月'] == globalMonth:
            graphdata.loc[irow, '年月'] = str(int(graphdata.loc[irow, '日'])).zfill(2) + '日'
            graphdata.loc[irow, '月均'] = graphdata.loc[irow, '数量']
        else:
            graphdata.loc[irow, '年月'] = str(graphdata.loc[irow, '年']) + '年' + str(graphdata.loc[irow, '月']).zfill(2) + '日'
            graphdata.loc[irow, '月均'] = graphdata.loc[irow, '数量'] / calendar.monthrange(int(graphdata.loc[irow, '年']), int(graphdata.loc[irow, '月']))[1]

    for isys, iday in zip(graphdata['体系'].drop_duplicates().tolist(), [dslday, hwday, lbxday, yfday, gjday]):
        # print(isys)
        sumBox = sum(graphdata[(graphdata['体系'] == isys) & (graphdata['年'] == globalYear) & (graphdata['年'] == globalYear) & (graphdata['月'] == globalMonth)]['数量'])
        avgBox = sumBox / iday
        if isys == '高济':
            bar_function(graphdata[graphdata['体系'] == isys]['年月'].to_list(), graphdata[graphdata['体系'] == isys]['月均'].to_list(), 'C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/', isys + '-柱状图', globalYear, globalMonth, isys, (globalYear - 2019) * 12 -1, (globalYear - 2019) * 12 + globalMonth -1, (globalYear - 2019) * 12 + globalMonth + lastDay-1, "{:.0f}".format(sumBox), "{:.0f}".format(avgBox))
        else:
            bar_function(graphdata[graphdata['体系'] == isys]['年月'].to_list(), graphdata[graphdata['体系'] == isys]['月均'].to_list(), 'C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/', isys + '-柱状图', globalYear, globalMonth, isys, (globalYear - 2018) * 12 -1, (globalYear - 2018) * 12 + globalMonth -1, (globalYear - 2018) * 12 + globalMonth + lastDay-1, "{:.0f}".format(sumBox), "{:.0f}".format(avgBox))



if __name__ == '__main__':
    
    cprint("运行前, 请确保日期修正, 截图中, 请勿操作复制粘贴", 'cyan', attrs=['bold', 'reverse', 'blink'])
    operaterTypes = int(input('>>>0Files-1Img-2SendALL-3SendDSL:'))
    
    if operaterTypes == 0:  # 保存文件
        print(' > 加载数据中ing, 请耐心等候')
        load_data()

        print(' > 生成大客户文件中ing')
        bigC_graph()

        print(' > 生成单品文件中ing')
        for imaterial in materialList:
            material_graph('C:/Users/Zeus/Desktop/autoSend/2_大客户/', aliasdata, imaterial)
         
    elif operaterTypes == 1:  # 保存图片
        deleteOldFiles('C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/')  # 清空文件夹历史文件
        
        print(' > 保存图片大客户ing')
        bigStartTime = datetime.now()
        book = opxl.load_workbook('C:/Users/Zeus/Desktop/autoSend/2_大客户/大客户_正式发送文件.xlsx')
        worksheetnames = book.sheetnames 
        for sheetname, picturename in zip(worksheetnames, worksheetnames):
            try:
                excelCatchScreen('C:/Users/Zeus/Desktop/autoSend/2_大客户/大客户_正式发送文件.xlsx', sheetname, picturename, 'C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/')
            except BaseException:
                print(picturename + '大客户截图出错！！！')
        bigEndTime = datetime.now()
        print(" >>保存大客户耗时：" + strftime("%H:%M:%S", gmtime((bigEndTime - bigStartTime).seconds)))
        
        print(' > 保存图片单品ing')
        singleStartTime = datetime.now()
        for jmaterial in materialList:
            book = opxl.load_workbook('C:/Users/Zeus/Desktop/autoSend/2_大客户/' + str(jmaterial) + '单品_正式发送文件.xlsx')
            worksheetnames = book.sheetnames
            for sheetname, picturename in zip(worksheetnames, worksheetnames):
                try:
                    excelCatchScreen('C:/Users/Zeus/Desktop/autoSend/2_大客户/' + str(jmaterial) + '单品_正式发送文件.xlsx', sheetname, picturename + jmaterial, 'C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/')
                except BaseException:
                    print(picturename + jmaterial + '截图出错！！！')
        singleEndTime = datetime.now()
        print(" >>保存单品耗时：" + strftime("%H:%M:%S", gmtime((singleEndTime - singleStartTime).seconds)))

        print(' > 保存图片柱状图ing')
        graphStartTime = datetime.now()
        bar_graph()
        graphEndTime = datetime.now()
        print(" >>保存柱状图耗时：" + strftime("%H:%M:%S", gmtime((graphEndTime - graphStartTime).seconds)))
    elif operaterTypes == 2:  # 发送图片

        AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序 - AppKey
        AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序 - AppSecret

        RobotWebHookURL = 'https://oapi.dingtalk.com/robot/send?access_token=c720958939172ef981876139356031776df95d9df51d0848c312779383916a0b'  # 数据汇报：大客户
        # RobotWebHookURL = 'https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c'  # 1机器人测试群

        RobotSecret = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'  # 群机器人加签秘钥secret(默认草晶华小助手)

        sendtitle = ["###### **① 大客户年度简析**",
                     "###### **② 汇总分析**",
                     "###### **③ 老百姓-各区域目标达成情况**",
                     "###### **④ 老百姓-同环比**",
                     "###### **⑤ 老百姓-各区域彩盒销售情况**",
                     "###### **⑥ 海王-各分部目标达成情况**",
                     "###### **⑦ 海王-同环比**",
                     "###### **⑧ 益丰-各公司目标达成情况**",
                     "###### **⑨ 益丰-同环比**",
                     "###### **⑩ 益丰-目标+同环比**",
                     "###### **①① 大参林-各大区目标达成情况**",
                     "###### **①② 大参林-同环比**",
                     "###### **①③ 高济-各平台目标达成情况**",
                     "###### **①④ 高济-同环比**",
                     "###### **①⑤ 老百姓-月均销售分布**",
                     "###### **①⑥ 海王-月均销售分布**",
                     "###### **①⑦ 益丰-月均销售分布**",
                     "###### **①⑧ 大参林-月均销售分布**",
                     "###### **①⑨ 高济-月均销售分布**",
                     "###### **①⑩ 老百姓-（三七）单品销售情况**"]
        
        imgName = ['①年度简析', '②汇总分析', '③老百姓', '④老百姓', '4老百姓彩盒', '⑤海王', '⑥海王', '⑦益丰', '⑧益丰', '3益丰', '⑨大参林', '⑩大参林', '1高济', '2高济', '老百姓-柱状图', '海王-柱状图', '益丰-柱状图', '大参林-柱状图', '高济-柱状图', 'Ⅰ老百姓三七']
        
        sendTypes = int(input('>>>0All-1Single:'))
    
        if sendTypes == 0: # 发送形式 - 全部
            
            pictureURL = [] 
            for img in imgName:
                try:
                    getURL = get_image_url('C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/' + img + '.PNG', img)
                    pictureURL.append(getURL)
                except BaseException:
                    print(img + '图片URL出错！！！')
                    
            if len(pictureURL) == len(imgName):
                for ititle, iurl, iname in zip(sendtitle, pictureURL, imgName):
                    ddMessage = {  # 发布消息内容
                        "msgtype": "markdown",
                        "markdown": {"title": iname,  # @某人 才会显示标题
                                        "text": ititle +
                                        "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
                                        "\n###### ----------------------------------------"
                                        "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                        "at": {
                            # "atMobiles": [15817552982],  # 指定@某人
                            "isAtAll": False  # 是否@所有人[False:否, True:是]
                        }
                    }

                    # 发送消息
                    dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
                    time.sleep(3)
            else:
                print('存在图片URL失败, 请检查！！！')
                
        elif sendTypes == 1: # 发送形式 - 选择性单张 从 1 开始
            
            send_NO_Picture = int(input('发送第几张图片？'))
            
            print('***单独发送: ' + str(imgName[send_NO_Picture - 1]) + '.PNG')
            
            ddMessage = {  # 发布消息内容
                    "msgtype": "markdown",
                    "markdown": {"title": imgName[send_NO_Picture - 1],  # @某人 才会显示标题
                                    "text": sendtitle[send_NO_Picture - 1] + 
                                    "\n![Image被拦截, 请使用非公司网络查看](" + get_image_url('C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/' + imgName[send_NO_Picture - 1] + '.PNG', imgName[send_NO_Picture - 1]) + ")"
                                    "\n###### ----------------------------------------"
                                    "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                    "at": {
                        # "atMobiles": [15817552982],  # 指定@某人
                        "isAtAll": False  # 是否@所有人[False:否, True:是]
                    }
                }

            # 发送消息
            dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
            
        else:
            print('请输入0-1正确的发送方式！！！')
        
        # dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendFile(chatId, fileFullPath)  # 发送文件
    elif operaterTypes == 3:  #  大客户-大参林内部沟通群
        
        AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序AppKey
        AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序AppSecret
        

        RobotWebHookURL = 'https://oapi.dingtalk.com/robot/send?access_token=17e23d6778fc4dcc164c26d6ff4d11161c74c6fff86b2a3e09a41613f189d22d'  # 大客户-大参林内部沟通群
        # RobotWebHookURL = 'https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c'  # 1机器人测试群

        RobotSecret = 'SEC702212eb5235b0173d67e234bd3d0e36315f4a780704f114925509098afad0e7'  # 大客户-大参林内部沟通群
        # RobotSecret = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'  # 群机器人加签秘钥secret(默认草晶华小助手)

        sendtitle = ["###### **① 大参林-各大区目标达成情况**",
                     "###### **② 大参林-同环比**"]
        
        imgDsl = ['⑨大参林', '⑩大参林']
        
        pictureURL = [] 
        for img in imgDsl:
            try:
                getURL = get_image_url('C:/Users/Zeus/Desktop/autoSend/2_大客户/Pictures/' + img + '.PNG', img)
                pictureURL.append(getURL)
            except BaseException:
                print(img + '图片URL出错！！！')
                
        if len(pictureURL) == len(imgDsl):
            for ititle, iurl, iname in zip(sendtitle, pictureURL, imgDsl):
                ddMessage = {  # 发布消息内容
                    "msgtype": "markdown",
                    "markdown": {"title": iname,  # @某人 才会显示标题
                                    "text": ititle +
                                    "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
                                    "\n###### ----------------------------------------"
                                    "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                    "at": {
                        # "atMobiles": [15817552982],  # 指定@某人
                        "isAtAll": False  # 是否@所有人[False:否, True:是]
                    }
                }

                # 发送消息
                dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
            
        else:
                print('存在图片URL失败, 请检查！！！')
        
    else:
        print('请输入0-1-2-3正确的发送方式！！！')