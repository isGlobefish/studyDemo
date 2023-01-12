# ----------------------------------------------------------------------------
# 需求说明: 每日21:00程序运行, 自动生成一个多子页的Excel表格, 自动截图发送到指定钉钉群
# ----------------------------------------------------------------------------
import re
import os
import glob
import time
import hmac
import json
import pyhdb
import base64
import hashlib
import requests
import calendar
import win32com
import pythoncom
import numpy as np
import pandas as pd
import urllib.parse
import urllib.request
import openpyxl as opxl
from decimal import Decimal
from termcolor import cprint
from datetime import datetime
from PIL import ImageGrab, Image
from time import strftime, gmtime
from win32com.client import Dispatch, DispatchEx
from qiniu import Auth, put_file, etag, BucketManager
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection, NamedStyle, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule,IconSetRule, Rule, IconSet, FormatObject


# ----------------------------------------------------------------------------
# 加载格式框架
# ----------------------------------------------------------------------------
framepath = 'C:/Users/zoodehao/Desktop/AutoSend/3_销管/Target/销管_格式框架.xlsx'
# workbook = opxl.load_workbook(filepath, data_only=True)
workbook = opxl.load_workbook(framepath)


# ----------------------------------------------------------------------------
# 全局日期设置
# ----------------------------------------------------------------------------
globalYear  = datetime.now().year
globalMonth = datetime.now().month
globalDay   = datetime.now().day


# ----------------------------------------------------------------------------
# 全局【目标 匹配表】数据准备
# ----------------------------------------------------------------------------
# 发货回款各大区目标
target = pd.read_excel('C:/Users/zoodehao/Desktop/AutoSend/3_销管/Target/销管_数据源.xlsx', sheet_name=0, header=0)
# 黄芪区域匹配维度
huangQi = pd.read_excel('C:/Users/zoodehao/Desktop/AutoSend/3_销管/Target/销管_数据源.xlsx', sheet_name=1, header=0)

# # 出货流水账ZSD043
# dataZSD043 = pd.read_excel('C:/Users/zoodehao/Desktop/销管_数据源.xlsx', sheet_name=0, header=0, usecols=['实际发货日期', '月份', '大客户', '黄芪区域', '物料描述', '交货数量', '二级送方', '省份', '送达方名称', '含税总价', '业绩结算金额'])
# # 回款明细ZFI26
# dataZFI26 = pd.read_excel('C:/Users/zoodehao/Desktop/销管_数据源.xlsx', sheet_name=1, header=0, usecols=['核销日期', '月份', '大客户', '确定二级架构', '送达方省份', '送达方名称', '本次核销金额', '核定价目表结算金额'])
# # 未发货
# dispatch = pd.read_excel('C:/Users/zoodehao/Desktop/销管_数据源.xlsx', sheet_name=3, header=0, usecols=['二级组织描述', '三级组织描述', '凭证日期', '售达方名称', '订单总金额', '订单原因文本'])
# # 欠货
# goods = pd.read_excel('C:/Users/zoodehao/Desktop/销管_数据源.xlsx', sheet_name=4, header=0, usecols=['品名', '数量', '欠货金额', '开始欠货时间', '预计出货时间'])

# 黄芪区域 字典   
huangQiDict = {}
for irow in range(huangQi.shape[0]):
    huangQiDict[huangQi.loc[irow, '省份']] = huangQi.loc[irow, '区域省份']


# ----------------------------------------------------------------------------
# HANA数据库获取数据
# ----------------------------------------------------------------------------
# 内网: IP: 192.168.20.183  账号: HANA1107318 密码: Zeus@1107311
# 公网: IP: 119.145.248.183  账号: HANAHBPOUTTEST 密码: Zeus@test147
def get_HANA_Connection(): 
    connectionObj = pyhdb.connect(
        host     = '192.168.20.183',
        port     = 30015,
        user     = 'HANA1107318',
        password = 'Zeus@1107311'
    )
    return connectionObj

# 1. 获取本年发货流水账ZSD043
def get_matZSD043(ZSD043):
    cursorZSD043 = ZSD043.cursor()
    cursorZSD043.execute("""WITH AA AS(SELECT LFDAT, TO_NUMBER(MONTH (LFDAT)) AS YUE, TO_NUMBER(RIGHT(LFDAT, 2)) AS TIAN, BEZEI, ZZPART2_T_S, KUNNR,NAME1, MAKTX, TO_NUMBER(SUM (LFIMG1)) AS AMOUNT, TO_NUMBER(SUM (KWERT)) AS HSZJ, TO_NUMBER(SUM (KWERT2)) AS YJJSJE
                            FROM "ECC_BI"."ZOUTBOUND"
                            WHERE ZZPART1_T_S LIKE N'%草晶华销售事业部%' AND LFDAT != '' AND BEZEI2 != '领用订单' AND YEAR(LFDAT) = :1 
                            GROUP BY LFDAT, BEZEI, ZZPART2_T_S,KUNNR, NAME1, MAKTX)
                            SELECT AA.LFDAT, AA.YUE, AA.TIAN, AA.BEZEI, AA.ZZPART2_T_S,
                                    (CASE P1.KATR5
                                        WHEN '01' THEN '大客户'
                                        WHEN '02' THEN '大客户'
                                        WHEN '03' THEN '大客户'
                                        WHEN '04' THEN '大客户'
                                        WHEN '07' THEN '大客户'
                                        WHEN '08' THEN '大客户'
                                        ELSE ''
                                        END)  AS DIM_DKH, AA.KUNNR, AA.NAME1, AA.MAKTX, AA.AMOUNT, AA.HSZJ / 10000, AA.YJJSJE / 10000
                            FROM AA
                            LEFT JOIN "ECC_BI"."KNA1" AS P1 ON AA.KUNNR = P1.KUNNR""",
                    [str(globalYear)])
    matZSD043 = cursorZSD043.fetchall()
    return matZSD043

# 2. 获取本年核销明细表ZFI26
def get_matZFI26(ZFI26):
    cursorZFI26 = ZFI26.cursor()
    cursorZFI26.execute("""WITH AA AS (SELECT P1.BUDAT, TO_NUMBER(P1.MONAT) AS YUE, TO_NUMBER(RIGHT(BUDAT, 2)) AS TIAN, P1.WW020, P1.ZZ021, P1.WW006, P1.KNDNR, P1.KUNWE, P2.NAME1 AS P2NAME, P1.VV002, P1.VV001
                           FROM "ECC_BI"."ZTFI008" AS P1
                           LEFT JOIN "ECC_BI"."KNA1" AS P2 ON P1.KUNWE = P2.KUNNR
                           WHERE P1.WW020 = '0000005702' AND VRGAR = 'A3' AND ZSJLH = '11' AND LEFT(P1.MATNR, 1) != 'Z' AND YEAR(P1.BUDAT) = :1),  -- 送达方维度
                           BB AS (SELECT AA.BUDAT, AA.YUE, AA.TIAN, AA.WW020, AA.ZZ021, P3.NAME1 AS P3NAME, AA.WW006, AA.KNDNR, AA.KUNWE, AA.P2NAME, AA.VV002, AA.VV001
                           FROM AA
                           LEFT JOIN "ECC_BI"."KNA1" AS P3 ON AA.ZZ021 = P3.KUNNR)  -- 客户维度
                           SELECT BB.BUDAT, BB.YUE, BB.TIAN, BB.P3NAME, BB.WW006,
                               (CASE P4.KATR5
                                   WHEN '01' THEN '大客户'
                                   WHEN '02' THEN '大客户'
                                   WHEN '03' THEN '大客户'
                                   WHEN '04' THEN '大客户'
                                   WHEN '07' THEN '大客户'
                                   WHEN '08' THEN '大客户'
                                   ELSE ''
                                   END)  AS DIM_DKH, BB.KNDNR, P4.NAME1 AS P4NAME, BB.KUNWE, BB.P2NAME, BB.VV002 / 10000, BB.VV001 / 10000
                           FROM BB
                           LEFT JOIN "ECC_BI"."KNA1" AS P4 ON BB.KNDNR = P4.KUNNR""",
                   [str(globalYear)])
    matZFI26 = cursorZFI26.fetchall()
    return matZFI26

# 3. 获取本年未发货明细表ZSD006
def get_matZSD006(ZSD006):
    cursorZSD006 = ZSD006.cursor()
    cursorZSD006.execute("""SELECT (CASE ZZPART2_T
                                    WHEN 'CJH-华东大区' THEN '南部'
                                    WHEN 'CJH-华南大区' THEN '南部'
                                    WHEN 'CJH-华中大区' THEN '南部'
                                    WHEN 'CJH-南部直管省区' THEN '南部'
                                    WHEN 'CJH-鲁皖大区' THEN '北部'
                                    WHEN 'CJH-西北大区' THEN '北部'
                                    WHEN 'CJH-西南大区' THEN '北部'
                                    WHEN 'CJH-华北大区' THEN '北部'
                                    WHEN 'CJH-北部直管省区' THEN '北部'
                                    END)  AS SOUTH_NORTH, ZZPART2_T, ZZPART3_T, AUDAT, NAME1, KZWI1, DATE_SQL, TDLINE3
                            FROM "ECC_BI"."ZUNSHIPPPED"
                            WHERE AUART != '' AND VKORG = '2008' AND AUART = 'ZOR' AND USER_LINE != '释放' AND YEAR(AUDAT) = :1
                            ORDER BY SOUTH_NORTH DESC, ZZPART2_T DESC, ZZPART3_T DESC, AUDAT ASC,  NAME1 DESC, KZWI1 DESC""",
                   [str(globalYear)])
    matZSD006 = cursorZSD006.fetchall()
    return matZSD006

# 4. 获取本年欠货表ZSD082
def get_matZSD082(ZSD082):
    cursorZSD082 = ZSD082.cursor()
    cursorZSD082.execute("""WITH AA AS (SELECT MATNR, MAKTX
                            FROM "ECC_BI"."MAKT"
                            WHERE MANDT = 800 AND SPRAS = 1),
                            BB AS ( SELECT ZZPART1_T_S, DATE_SQL, MATNR, LFIMG3, WJHJE, ERDAT
                            FROM "ECC_BI"."ZEXHSUSTED"
                            WHERE YEAR(DATE_SQL) = :1)
                            SELECT BB.ZZPART1_T_S, AA.MAKTX, BB.LFIMG3, BB.WJHJE, BB.ERDAT, BB.DATE_SQL
                            FROM BB
                            LEFT JOIN AA ON BB.MATNR = AA.MATNR""",
                   [str(globalYear)])
    matZSD082 = cursorZSD082.fetchall()
    return matZSD082

conn = get_HANA_Connection()

# 1. 发货流水账ZSD043
dataZSD043 = pd.DataFrame(get_matZSD043(conn), columns=['实际发货日期', '月份', '天', '省份', '二级送方', '大客户', '送达方编码', '送达方名称', '物料描述',  '交货数量', '含税总价', '业绩结算金额'])
# 2. 核销明细表ZFI26
dataZFI26 = pd.DataFrame(get_matZFI26(conn), columns=['核销日期', '月份', '天', '确定二级架构', '送达方省份', '大客户', '客户编码', '客户名称', '送达方编码', '送达方名称', '本次核销金额', '核定价目表结算金额'])
# 3. 未发货明细表ZSD006
dispatch = pd.DataFrame(get_matZSD006(conn), columns=['南北部', '二级组织描述', '三级组织描述', '凭证日期', '售达方名称', '订单总金额', '抽数日期', '未发货原因'])
# 4. 欠货表ZSD082
misgoods = pd.DataFrame(get_matZSD082(conn), columns=['事业部', '品名', '数量', '欠货金额', '开始欠货时间', '预计出货时间'])

dataZSD043.insert(3, '黄芪区域', '')
dataZSD043['黄芪区域'] = dataZSD043.apply(lambda x: huangQiDict.setdefault(x['省份'], 0), axis=1)


# colorfulbox = ['芡实薏米破壁草本', '玫瑰桑椹破壁草本', '菊花枸杞破壁草本', '山楂玫瑰桑椹破壁草本', '杏茯甘草破壁草本', '人参黄精玛咖破壁草本', '大麦若叶破壁草本', '藤茶菊苣破壁草本', '酸枣仁百合破壁草本', '葛根枳椇子破壁草本', '五子桑菊破壁草本']
# for irow in range(dataZSD043.shape[0]):
#     if dataZSD043.loc[irow, '物料描述'].split('-')[0] in colorfulbox:
#         dataZSD043.loc[irow, '物料描述'] = '彩盒'
#     elif dataZSD043.loc[irow, '物料描述'] == '茯苓鱼腥草破壁草本-4克×10袋':
#         dataZSD043.loc[irow, '物料描述'] = '彩盒'
for irow in range(dataZSD043.shape[0]):
    if '16袋' in dataZSD043.loc[irow, '物料描述']:
        dataZSD043.loc[irow, '物料描述'] = '彩盒'

fahuoMAXdate = max(dataZSD043['实际发货日期'])
huikuanMAXdate = max(dataZFI26["核销日期"])
# print('【发货】最新日期:' + str(max(dataZSD043['实际发货日期'])) + '; 【回款】最新日期:' + str(max(dataZFI26["核销日期"])) + '; 【未发货】最新日期:' + str(max(dispatch["抽数日期"])) + '; 【欠货】最新日期:' + str(max(misgoods["预计出货时间"])))
print('【发货】最新日期:' + str(fahuoMAXdate) + '; 【回款】最新日期:' + str(huikuanMAXdate) + '; 【未发货】最新日期:' + str(max(dispatch["抽数日期"])) + '; 【欠货】最新日期:' + str(max(misgoods["预计出货时间"])))


# ----------------------------------------------------------------------------
# 第一页 1_【发货】
# ----------------------------------------------------------------------------
worksheet1 = workbook.create_sheet('1_【发货】', 0)

# for irow in range(len(dataZSD043["实际发货日期"])):
#     dataZSD043.loc[irow, "实际发货日期"] = str(dataZSD043.loc[irow, "实际发货日期"])[0:10]
# today = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)  # 当天日期
# selectData = dataZSD043[dataZSD043["实际发货日期"] == '2021-04-22']  # 筛选指定日期数据

# selectData = dataZSD043[dataZSD043["实际发货日期"] == max(dataZSD043['实际发货日期'])]  # 筛选指定日期数据
selectData = dataZSD043[dataZSD043["实际发货日期"] == fahuoMAXdate]  # 筛选最新日期数据

# pivotTable = pd.pivot_table(selectData, index=[u"二级送方", u"省份", u"送达方名称"], values=[u"含税总价", u"业绩结算金额"], aggfunc=[np.sum], fill_value=0, margins=True, margins_name="总计")
pivotTable = pd.pivot_table(selectData, index=[u"二级送方", u"省份", u"送达方名称"], values=[u"含税总价", u"业绩结算金额"], aggfunc={u"含税总价": np.sum, u"业绩结算金额": np.sum}, fill_value=0, margins=True, margins_name="总计").sort_index(axis=1, ascending=False)

# 将透视表的多维表头拍平
# pivotTable.columns = [col1 + '_' + str(col2) for (col1, col2) in pivotTable.columns.tolist()]
pivotTable.reset_index(inplace=True)

# 模仿透视-添加分类汇总
# pivotAll = pd.DataFrame(columns=['二级送方', '省份', '送达方名称', 'sum_业绩结算金额', 'sum_含税总价'])
# for irowname in pivotTable['二级送方'].drop_duplicates().tolist():
#     pivotTableNew = pivotTable[pivotTable["二级送方"] == irowname]
#     if irowname != '总计':
#         createDF = pd.DataFrame({'二级送方': irowname,
#                                  '省份': irowname + " 汇总",
#                                  '送达方名称': irowname + " 汇总",
#                                  'sum_业绩结算金额': sum(pivotTableNew.iloc[:, 3]),
#                                  'sum_含税总价': sum(pivotTableNew.iloc[:, 4])},
#                                 columns=['二级送方', '省份', '送达方名称', 'sum_业绩结算金额', 'sum_含税总价'], index=["小计"])
#         pivotTableNew = pivotTableNew.append(createDF)
#     pivotAll = pivotAll.append(pivotTableNew)
pivotAll = pd.DataFrame(columns=['二级送方', '省份', '送达方名称', '含税总价', '业绩结算金额'])
for irowname in pivotTable['二级送方'].drop_duplicates().tolist():
    pivotTableNew = pivotTable[pivotTable["二级送方"] == irowname]
    if irowname != '总计':
        # noinspection PyUnboundLocalVariable
        createDF = pd.DataFrame({'二级送方': irowname,
                                 '省份': irowname + " 汇总",
                                 '送达方名称': irowname + " 汇总",
                                 '含税总价': sum(pivotTableNew.iloc[:, 3]),
                                 '业绩结算金额': sum(pivotTableNew.iloc[:, 4])},
                                columns=['二级送方', '省份', '送达方名称', '含税总价', '业绩结算金额'], index=["小计"])
        pivotTableNew = pivotTableNew.append(createDF)
    pivotAll = pivotAll.append(pivotTableNew)


# 设置行高列宽
for row in range(1, len(pivotAll) + 3):
    if row == 1:
        worksheet1.row_dimensions[row].height = 24
    else:
        worksheet1.row_dimensions[row].height = 17
columnWidth = [22, 10, 40, 20, 20, 12]
for icol, width in enumerate(columnWidth, start=1):
    worksheet1.column_dimensions[opxl.utils.get_column_letter(icol)].width = width

# 第 1 行
worksheet1.merge_cells('A1:F1')
worksheet1['A1'] = "草晶华事业部发货明细表"
title1 = worksheet1['A1']
title1.font = Font(name="微软雅黑", size=14, bold=True, color='FF0000')
title1.alignment = Alignment(horizontal='center', vertical='center')
title1.fill = PatternFill(fill_type='solid', fgColor="548235")

# 第 2 行
subTitle1 = ["大区", "省份", "客户名称", "发货金额（发货价）", "发货金额（结算价）", "备注"]
for icol, content in enumerate(subTitle1, start=1):
    worksheet1[opxl.utils.get_column_letter(icol) + '2'] = content
    subtitle1 = worksheet1[opxl.utils.get_column_letter(icol) + '2']
    subtitle1.font = Font(name="微软雅黑", size=10, bold=True, color='FFFFFF')
    subtitle1.alignment = Alignment(horizontal='center', vertical='center')
    subtitle1.fill = PatternFill(fill_type='solid', fgColor="203764")
    subtitle1.border = Border(left=Side(border_style='thin', color='FF000000'),
                              right=Side(border_style='thin', color='FF000000'),
                              top=Side(border_style='thin', color='FF000000'),
                              bottom=Side(border_style='thin', color='FF000000'))


# 填充数据【注: excel的行列起始值都是1】
for irow in range(pivotAll.iloc[:, 0].size):
    for jcol in range(pivotAll.columns.size + 1):
        if jcol != pivotAll.columns.size:
            if jcol in (0, 1, 2):
                worksheet1.cell(row=irow + 3, column=jcol + 1).value = pivotAll.iloc[irow, jcol]
            else:
                worksheet1.cell(row=irow + 3, column=jcol + 1).value = round(pivotAll.iloc[irow, jcol], 2)
        cell = worksheet1.cell(row=irow + 3, column=jcol + 1)
        if jcol == 0:
            cell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            cell.fill = PatternFill(fill_type='solid', fgColor="F8CBAD")
        elif jcol == 1:
            cell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'))
        else:
            cell.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
            cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 最后1, 2行
        if irow == pivotAll.iloc[:, 0].size - 1:
            worksheet1.cell(row=irow + 3, column=jcol + 1).font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            worksheet1.cell(row=irow + 3, column=jcol + 1).fill = PatternFill(fill_type='solid', fgColor="A9D08C")
            lastcell = worksheet1.cell(row=irow + 4, column=jcol + 1)
            lastcell.fill = PatternFill(fill_type='solid', fgColor="D9D9D9")
            if jcol <= 3:
                lastcell.font = Font(name="微软雅黑", size=10, bold=False, color='FF0000')
                lastcell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                lastcell.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
                lastcell.alignment = Alignment(horizontal='right', vertical='center')
    # 最后1, 2行
    if irow == pivotAll.iloc[:, 0].size - 1:
        worksheet1.merge_cells(start_row=irow + 3, end_row=irow + 3, start_column=1, end_column=3)
        worksheet1.merge_cells(start_row=irow + 4, end_row=irow + 4, start_column=1, end_column=4)
        worksheet1.merge_cells(start_row=irow + 4, end_row=irow + 4, start_column=5, end_column=6)
        worksheet1.cell(row=irow + 4, column=1).value = "注: 以上结算价为伙伴标准结算价, 发货特价差额在结算四季大促活动中扣除; 金额单位:万元"
        # worksheet1.cell(row=irow + 4, column=5).value = "日期: " + str(datetime.now().strftime("%Y/%m/%d"))
        # worksheet1.cell(row=irow + 4, column=5).value = '日期: ' + str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
        worksheet1.cell(row=irow + 4, column=5).value = '最新日期: ' + str(fahuoMAXdate).split('-')[0] + '/' + str(fahuoMAXdate).split('-')[1] + '/' + str(fahuoMAXdate).split('-')[2]


# 合并单元格并填充
for findelement in pivotAll['二级送方'].drop_duplicates().tolist()[:-1]:
    rowInfo = []
    for irow in range(pivotAll.iloc[:, 0].size):
        if pivotAll.iloc[irow, 0] == findelement:
            rowInfo.append(irow)
        firstcolumn = worksheet1.cell(row=irow + 3, column=1)
        firstcolumn.border = Border(left=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
    for jcol in range(2, 7):
        xiaojicell = worksheet1.cell(row=max(rowInfo) + 3, column=jcol)
        xiaojicell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
        xiaojicell.fill = PatternFill(fill_type='solid', fgColor="F8CBAD")
        if jcol == 2:
            xiaojicell.border = Border(
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
        else:
            xiaojicell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                       right=Side(border_style='thin', color='FF000000'),
                                       top=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='thin', color='FF000000'))
    worksheet1.merge_cells(start_row=min(rowInfo) + 3, end_row=max(rowInfo) + 3, start_column=1, end_column=1)
    worksheet1.merge_cells(start_row=max(rowInfo) + 3, end_row=max(rowInfo) + 3, start_column=2, end_column=3)


# ----------------------------------------------------------------------------
# 第二页 2_【回款】
# ----------------------------------------------------------------------------
worksheet2 = workbook.create_sheet('2_【回款】', 1)

# for irow in range(len(dataZFI26["核销日期"])):
#     dataZFI26.loc[irow, "核销日期"] = str(dataZFI26.loc[irow, "核销日期"])[0:10]
# today = datetime.now().strftime("%Y-%m-%d")  # 当天日期

# select_huikuan = dataZFI26[dataZFI26["核销日期"] == max(dataZFI26["核销日期"])]  # 筛选指定日期数据
select_huikuan = dataZFI26[dataZFI26["核销日期"] == huikuanMAXdate]  # 筛选最新日期数据

# pivot_huikua = pd.pivot_table(select_huikuan, index=[u"确定二级架构", u"送达方省份", u"送达方名称"], values=[u"本次核销金额", u"核定价目表结算金额"], aggfunc=[np.sum], fill_value=0, margins=True, margins_name="总计")
pivot_huikua = pd.pivot_table(select_huikuan, index=[u"确定二级架构", u"送达方省份", u"送达方名称"], values=[u"本次核销金额", u"核定价目表结算金额"], aggfunc={u"本次核销金额": np.sum, u"核定价目表结算金额": np.sum}, fill_value=0, margins=True, margins_name="总计")

# 将透视表的多维表头拍平
# pivot_huikua.columns = [col1 + '_' + str(col2) for (col1, col2) in pivot_huikua.columns.tolist()]
pivot_huikua.reset_index(inplace=True)

# 模仿透视-添加分类汇总
# pivot_huikuaAll = pd.DataFrame(columns=['确定二级架构', '送达方省份', '送达方名称', 'sum_本次核销金额', 'sum_核定价目表结算金额'])
# for irowname in pivot_huikua['确定二级架构'].drop_duplicates().tolist():
#     pivot_huikuaNew = pivot_huikua[pivot_huikua["确定二级架构"] == irowname]
#     if irowname != '总计':
#         createDF = pd.DataFrame({'确定二级架构': irowname,
#                                  '送达方省份': irowname + " 汇总",
#                                  '送达方名称': irowname + " 汇总",
#                                  'sum_本次核销金额': sum(pivot_huikuaNew.iloc[:, 3]),
#                                  'sum_核定价目表结算金额': sum(pivot_huikuaNew.iloc[:, 4])},
#                                 columns=['确定二级架构', '送达方省份', '送达方名称', 'sum_本次核销金额', 'sum_核定价目表结算金额'], index=["小计"])
#         pivot_huikuaNew = pivot_huikuaNew.append(createDF)
#     pivot_huikuaAll = pivot_huikuaAll.append(pivot_huikuaNew)
pivot_huikuaAll = pd.DataFrame(columns=['确定二级架构', '送达方省份', '送达方名称', '本次核销金额', '核定价目表结算金额'])
for irowname in pivot_huikua['确定二级架构'].drop_duplicates().tolist():
    pivot_huikuaNew = pivot_huikua[pivot_huikua["确定二级架构"] == irowname]
    if irowname != '总计':
        createDF = pd.DataFrame({'确定二级架构': irowname,
                                 '送达方省份': irowname + " 汇总",
                                 '送达方名称': irowname + " 汇总",
                                 '本次核销金额': sum(pivot_huikuaNew.iloc[:, 3]),
                                 '核定价目表结算金额': sum(pivot_huikuaNew.iloc[:, 4])},
                                columns=['确定二级架构', '送达方省份', '送达方名称', '本次核销金额', '核定价目表结算金额'], index=["小计"])
        pivot_huikuaNew = pivot_huikuaNew.append(createDF)
    pivot_huikuaAll = pivot_huikuaAll.append(pivot_huikuaNew)


# 设置行高列宽
for row in range(1, len(pivot_huikuaAll) + 2):
    if row == 1:
        worksheet2.row_dimensions[row].height = 24
    else:
        worksheet2.row_dimensions[row].height = 17
columnWidth2 = [22, 10, 40, 20, 20, 12]
for icol, width in enumerate(columnWidth2, start=1):
    worksheet2.column_dimensions[opxl.utils.get_column_letter(icol)].width = width

# 第 1 行
worksheet2.merge_cells('A1:F1')
worksheet2['A1'] = "草晶华事业部回款明细表"
title2 = worksheet2['A1']
title2.font = Font(name="微软雅黑", size=14, bold=True, color='FF0000')
title2.alignment = Alignment(horizontal='center', vertical='center')
title2.fill = PatternFill(fill_type='solid', fgColor="1F4E78")

# 第 2 行
subTitle2 = ["大区", "省份", "客户名称", "回款金额（发货价）", "结算金额（结算价）", "备注"]
for icol, content in enumerate(subTitle2, start=1):
    worksheet2[opxl.utils.get_column_letter(icol) + '2'] = content
    subtitle2 = worksheet2[opxl.utils.get_column_letter(icol) + '2']
    subtitle2.font = Font(name="微软雅黑", size=10, bold=True, color='FFFFFF')
    subtitle2.alignment = Alignment(horizontal='center', vertical='center')
    subtitle2.fill = PatternFill(fill_type='solid', fgColor="203764")
    subtitle2.border = Border(left=Side(border_style='thin', color='FF000000'),
                              right=Side(border_style='thin', color='FF000000'),
                              top=Side(border_style='thin', color='FF000000'),
                              bottom=Side(border_style='thin', color='FF000000'))

# 填充数据【注: excel的行列起始值都是1】
for irow in range(pivot_huikuaAll.iloc[:, 0].size):
    for jcol in range(pivot_huikuaAll.columns.size + 1):
        if jcol != pivot_huikuaAll.columns.size:
            if jcol in (0, 1, 2):
                worksheet2.cell(row=irow + 3, column=jcol + 1).value = pivot_huikuaAll.iloc[irow, jcol]
            else:
                worksheet2.cell(row=irow + 3, column=jcol + 1).value = round(pivot_huikuaAll.iloc[irow, jcol], 2)
        cell = worksheet2.cell(row=irow + 3, column=jcol + 1)
        if jcol == 0:
            cell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            cell.fill = PatternFill(fill_type='solid', fgColor="F8CBAD")
        elif jcol == 1:
            cell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'))
        else:
            cell.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
            cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 最后1, 2行
        if irow == pivot_huikuaAll.iloc[:, 0].size - 1:
            worksheet2.cell(row=irow + 3, column=jcol + 1).font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            worksheet2.cell(row=irow + 3, column=jcol + 1).fill = PatternFill(fill_type='solid', fgColor="9BC2E6")
            lastcell = worksheet2.cell(row=irow + 4, column=jcol + 1)
            lastcell.fill = PatternFill(fill_type='solid', fgColor="D9D9D9")
            if jcol <= 3:
                lastcell.font = Font(name="微软雅黑", size=10, bold=False, color='FF0000')
                lastcell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                lastcell.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
                lastcell.alignment = Alignment(horizontal='right', vertical='center')
    # 最后1, 2行
    if irow == pivot_huikuaAll.iloc[:, 0].size - 1:
        worksheet2.merge_cells(start_row=irow + 3, end_row=irow + 3, start_column=1, end_column=3)
        worksheet2.merge_cells(start_row=irow + 4, end_row=irow + 4, start_column=1, end_column=4)
        worksheet2.merge_cells(start_row=irow + 4, end_row=irow + 4, start_column=5, end_column=6)
        # worksheet2.cell(row=irow + 4, column=1).value = "备注：以上结算价为伙伴标准结算价，发货特价差额在结算四季大促活动中扣除"
        worksheet2.cell(row=irow + 4, column=1).value = "注: 金额单位: 万元"
        # worksheet2.cell(row=irow + 4, column=5).value = "日期: " + str(datetime.now().strftime("%Y/%m/%d"))
        # worksheet2.cell(row=irow + 4, column=5).value = '最新日期: ' + str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)
        worksheet2.cell(row=irow + 4, column=5).value = '最新日期: ' + str(huikuanMAXdate)[0:4] + '/' + str(huikuanMAXdate)[4:6] + '/' + str(huikuanMAXdate)[6:8]

# 合并单元格并填充
for findelement in pivot_huikuaAll['确定二级架构'].drop_duplicates().tolist()[:-1]:
    rowInfo = []
    for irow in range(pivot_huikuaAll.iloc[:, 0].size):
        if pivot_huikuaAll.iloc[irow, 0] == findelement:
            rowInfo.append(irow)
        firstcolumn = worksheet2.cell(row=irow + 3, column=1)
        firstcolumn.border = Border(left=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
    for jcol in range(2, 7):
        xiaojicell = worksheet2.cell(row=max(rowInfo) + 3, column=jcol)
        xiaojicell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
        xiaojicell.fill = PatternFill(fill_type='solid', fgColor="F8CBAD")
        if jcol == 2:
            xiaojicell.border = Border(
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'))
        else:
            xiaojicell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                       right=Side(border_style='thin', color='FF000000'),
                                       top=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='thin', color='FF000000'))
    worksheet2.merge_cells(start_row=min(rowInfo) + 3, end_row=max(rowInfo) + 3, start_column=1, end_column=1)
    worksheet2.merge_cells(start_row=max(rowInfo) + 3, end_row=max(rowInfo) + 3, start_column=2, end_column=3)


# ----------------------------------------------------------------------------
# 第三页 3_【大区】
# ----------------------------------------------------------------------------
worksheet3 = workbook['3_【大区】']

for iloc in [11, 17, 18]:
    for jloc in [i for i in range(3, 17, 1)]:
        # 南部小计、北部小计、合计
        if jcol != 4 or jcol != 6 or jcol != 9 or jcol != 12:  # 非完成率部分
            if iloc == 11:
                worksheet3.cell(row=iloc, column=jloc).value = "=SUM(" + str(opxl.utils.get_column_letter(jloc)) + "7:" + str(opxl.utils.get_column_letter(jloc)) + "10)"
            elif iloc == 17:
                worksheet3.cell(row=iloc, column=jloc).value = "=SUM(" + str(opxl.utils.get_column_letter(jloc)) + "12:" + str(opxl.utils.get_column_letter(jloc)) + "16)"
            else:
                worksheet3.cell(row=iloc, column=jloc).value = "=" + str(opxl.utils.get_column_letter(jloc)) + "11+" + str(opxl.utils.get_column_letter(jloc)) + "17"

# 完成率部分
# 11行 - 南部
worksheet3.cell(row=11, column=4).value = Decimal(sum(dataZSD043[((dataZSD043["二级送方"] == 'CJH-华东大区') | (dataZSD043["二级送方"] == 'CJH-华南大区') | (dataZSD043["二级送方"] == 'CJH-华中大区') | (dataZSD043["二级送方"] == 'CJH-南部直管省区')) & (dataZSD043["月份"] <= globalMonth)]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == '发货') & (target["区域"] == '南部') & (target["月份"] <= globalMonth)]['目标']))
worksheet3.cell(row=11, column=9).value = Decimal(sum(dataZSD043[((dataZSD043["二级送方"] == 'CJH-华东大区') | (dataZSD043["二级送方"] == 'CJH-华南大区') | (dataZSD043["二级送方"] == 'CJH-华中大区') | (dataZSD043["二级送方"] == 'CJH-南部直管省区')) & (dataZSD043["月份"] == globalMonth)]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == '发货') & (target["区域"] == '南部') & (target["月份"] == globalMonth)]['目标']))
worksheet3.cell(row=11, column=6).value = Decimal(sum(dataZFI26[((dataZFI26["确定二级架构"] == 'CJH-华东大区') | (dataZFI26["确定二级架构"] == 'CJH-华南大区') | (dataZFI26["确定二级架构"] == 'CJH-华中大区') | (dataZFI26["确定二级架构"] == 'CJH-南部直管省区')) & (dataZFI26["月份"] <= globalMonth)]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == '回款') & (target["区域"] == '南部') & (target["月份"] <= globalMonth)]['目标']))
worksheet3.cell(row=11, column=12).value = Decimal(sum(dataZFI26[((dataZFI26["确定二级架构"] == 'CJH-华东大区') | (dataZFI26["确定二级架构"] == 'CJH-华南大区') | (dataZFI26["确定二级架构"] == 'CJH-华中大区') | (dataZFI26["确定二级架构"] == 'CJH-南部直管省区')) & (dataZFI26["月份"] == globalMonth)]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == '回款') & (target["区域"] == '南部') & (target["月份"] == globalMonth)]['目标']))
# 17行 - 北部
worksheet3.cell(row=17, column=4).value = Decimal(sum(dataZSD043[((dataZSD043["二级送方"] == 'CJH-鲁皖大区') | (dataZSD043["二级送方"] == 'CJH-西北大区') | (dataZSD043["二级送方"] == 'CJH-西南大区') | (dataZSD043["二级送方"] == 'CJH-华北大区') | (dataZSD043["二级送方"] == 'CJH-北部直管省区')) & (dataZSD043["月份"] <= globalMonth)]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == '发货') & (target["区域"] == '北部') & (target["月份"] <= globalMonth)]['目标']))
worksheet3.cell(row=17, column=9).value = Decimal(sum(dataZSD043[((dataZSD043["二级送方"] == 'CJH-鲁皖大区') | (dataZSD043["二级送方"] == 'CJH-西北大区') | (dataZSD043["二级送方"] == 'CJH-西南大区') | (dataZSD043["二级送方"] == 'CJH-华北大区') | (dataZSD043["二级送方"] == 'CJH-北部直管省区')) & (dataZSD043["月份"] == globalMonth)]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == '发货') & (target["区域"] == '北部') & (target["月份"] == globalMonth)]['目标']))
worksheet3.cell(row=17, column=6).value = Decimal(sum(dataZFI26[((dataZFI26["确定二级架构"] == 'CJH-鲁皖大区') | (dataZFI26["确定二级架构"] == 'CJH-西北大区') | (dataZFI26["确定二级架构"] == 'CJH-西南大区') | (dataZFI26["确定二级架构"] == 'CJH-华北大区') | (dataZFI26["确定二级架构"] == 'CJH-北部直管省区')) & (dataZFI26["月份"] <= globalMonth)]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == '回款') & (target["区域"] == '北部') & (target["月份"] <= globalMonth)]['目标']))
worksheet3.cell(row=17, column=12).value = Decimal(sum(dataZFI26[((dataZFI26["确定二级架构"] == 'CJH-鲁皖大区') | (dataZFI26["确定二级架构"] == 'CJH-西北大区') | (dataZFI26["确定二级架构"] == 'CJH-西南大区') | (dataZFI26["确定二级架构"] == 'CJH-华北大区') | (dataZFI26["确定二级架构"] == 'CJH-北部直管省区')) & (dataZFI26["月份"] == globalMonth)]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == '回款') & (target["区域"] == '北部') & (target["月份"] == globalMonth)]['目标']))
# 18行 - 合计
worksheet3.cell(row=18, column=4).value = Decimal(sum(dataZSD043[dataZSD043["月份"] <= globalMonth]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == '发货') & (target["月份"] <= globalMonth)]['目标']))
worksheet3.cell(row=18, column=9).value = Decimal(sum(dataZSD043[dataZSD043["月份"] == globalMonth]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == '发货') & (target["月份"] == globalMonth)]['目标']))
worksheet3.cell(row=18, column=6).value = Decimal(sum(dataZFI26[dataZFI26["月份"] <= globalMonth]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == '回款') & (target["月份"] <= globalMonth)]['目标']))
worksheet3.cell(row=18, column=12).value = Decimal(sum(dataZFI26[dataZFI26["月份"] == globalMonth]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == '回款') & (target["月份"] == globalMonth)]['目标']))


targetrow = [7, 8, 9, 10, 12, 13, 14, 15, 16]
targetlist = ['CJH-华东大区', 'CJH-华南大区', 'CJH-华中大区', 'CJH-南部直管省区', 'CJH-鲁皖大区', 'CJH-西北大区', 'CJH-西南大区', 'CJH-华北大区', 'CJH-北部直管省区']
targetcol = [3, 5]
targettype = ["发货", "回款"]

for icol, type in zip(targetcol, targettype):
    for jrow, list in zip(targetrow, targetlist):
        targetsum = sum(target[(target["类型"] == type) & (target["地区"] == list) & (target["月份"] <= globalMonth)]['目标'])
        worksheet3.cell(row=jrow, column=icol).value = targetsum
        if type == "发货":  # col = 3 发货 4 7 8 9 13 14
            worksheet3.cell(row=jrow, column=icol + 1).value = sum(dataZSD043[dataZSD043["二级送方"] == list]["业绩结算金额"]) / Decimal(targetsum)
            worksheet3.cell(row=jrow, column=icol + 4).value = sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["月份"] == globalMonth)]["含税总价"])
            worksheet3.cell(row=jrow, column=icol + 5).value = sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["月份"] == globalMonth)]["业绩结算金额"])
            worksheet3.cell(row=jrow, column=icol + 6).value = Decimal(sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["月份"] == globalMonth)]["业绩结算金额"])) / Decimal(sum(target[(target["类型"] == type) & (target["地区"] == list) & (target["月份"] == globalMonth)]['目标']))
            # worksheet3.cell(row=jrow, column=icol + 10).value = sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["实际发货日期"] == max(dataZSD043['实际发货日期']))]["含税总价"])
            # worksheet3.cell(row=jrow, column=icol + 11).value = sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["实际发货日期"] == max(dataZSD043['实际发货日期']))]["业绩结算金额"])
            worksheet3.cell(row=jrow, column=icol + 10).value = sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["月份"] == globalMonth) & (dataZSD043["天"] == globalDay)]["含税总价"])
            worksheet3.cell(row=jrow, column=icol + 11).value = sum(dataZSD043[(dataZSD043["二级送方"] == list) & (dataZSD043["月份"] == globalMonth) & (dataZSD043["天"] == globalDay)]["业绩结算金额"])
        else:  # col = 5 回款 6 10 11 12 15 16
            worksheet3.cell(row=jrow, column=icol + 1).value = Decimal(sum(dataZFI26[dataZFI26["确定二级架构"] == list]["核定价目表结算金额"])) / Decimal(targetsum)
            worksheet3.cell(row=jrow, column=icol + 5).value = sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["月份"] == globalMonth)]["本次核销金额"])
            worksheet3.cell(row=jrow, column=icol + 6).value = sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["月份"] == globalMonth)]["核定价目表结算金额"])
            worksheet3.cell(row=jrow, column=icol + 7).value = Decimal(sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["月份"] == globalMonth)]["核定价目表结算金额"])) / Decimal(sum(target[(target["类型"] == type) & (target["地区"] == list) & (target["月份"] == globalMonth)]['目标']))
            # worksheet3.cell(row=jrow, column=icol + 10).value = sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["核销日期"] == max(dataZFI26["核销日期"]))]["本次核销金额"])
            # worksheet3.cell(row=jrow, column=icol + 11).value = sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["核销日期"] == max(dataZFI26["核销日期"]))]["核定价目表结算金额"])
            worksheet3.cell(row=jrow, column=icol + 10).value = sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["月份"] == globalMonth) & (dataZFI26["天"] == globalDay)]["本次核销金额"])
            worksheet3.cell(row=jrow, column=icol + 11).value = sum(dataZFI26[(dataZFI26["确定二级架构"] == list) & (dataZFI26["月份"] == globalMonth) & (dataZFI26["天"] == globalDay)]["核定价目表结算金额"])


# ----------------------------------------------------------------------------
# 第四页 4_【大客户】
# ----------------------------------------------------------------------------
worksheet4 = workbook['4_【大客户】']

worksheet4.cell(row=6, column=2).value = sum(dataZSD043[dataZSD043["大客户"] == "大客户"]["含税总价"])
worksheet4.cell(row=6, column=3).value = sum(dataZSD043[dataZSD043["大客户"] == "大客户"]["业绩结算金额"])
worksheet4.cell(row=6, column=4).value = sum(dataZFI26[dataZFI26["大客户"] == "大客户"]["本次核销金额"])
worksheet4.cell(row=6, column=5).value = sum(dataZFI26[dataZFI26["大客户"] == "大客户"]["核定价目表结算金额"])

worksheet4.cell(row=6, column=6).value = sum(dataZSD043[(dataZSD043["大客户"] == "大客户") & (dataZSD043["月份"] == globalMonth)]["含税总价"])
worksheet4.cell(row=6, column=7).value = sum(dataZSD043[(dataZSD043["大客户"] == "大客户") & (dataZSD043["月份"] == globalMonth)]["业绩结算金额"])
worksheet4.cell(row=6, column=8).value = sum(dataZFI26[(dataZFI26["大客户"] == "大客户") & (dataZFI26["月份"] == globalMonth)]["本次核销金额"])
worksheet4.cell(row=6, column=9).value = sum(dataZFI26[(dataZFI26["大客户"] == "大客户") & (dataZFI26["月份"] == globalMonth)]["核定价目表结算金额"])

# worksheet4.cell(row=6, column=10).value = sum(dataZSD043[(dataZSD043["大客户"] == "大客户") & (dataZSD043["实际发货日期"] == max(dataZSD043['实际发货日期']))]["含税总价"])
# worksheet4.cell(row=6, column=11).value = sum(dataZSD043[(dataZSD043["大客户"] == "大客户") & (dataZSD043["实际发货日期"] == max(dataZSD043['实际发货日期']))]["业绩结算金额"])
# worksheet4.cell(row=6, column=12).value = sum(dataZFI26[(dataZFI26["大客户"] == "大客户") & (dataZFI26["核销日期"] == max(dataZFI26["核销日期"]))]["本次核销金额"])
# worksheet4.cell(row=6, column=13).value = sum(dataZFI26[(dataZFI26["大客户"] == "大客户") & (dataZFI26["核销日期"] == max(dataZFI26["核销日期"]))]["核定价目表结算金额"])
worksheet4.cell(row=6, column=10).value = sum(dataZSD043[(dataZSD043["大客户"] == "大客户") & (dataZSD043["月份"] == globalMonth) & (dataZSD043["天"] == globalDay)]["含税总价"])
worksheet4.cell(row=6, column=11).value = sum(dataZSD043[(dataZSD043["大客户"] == "大客户") & (dataZSD043["月份"] == globalMonth) & (dataZSD043["天"] == globalDay)]["业绩结算金额"])
worksheet4.cell(row=6, column=12).value = sum(dataZFI26[(dataZFI26["大客户"] == "大客户") & (dataZFI26["月份"] == globalMonth) & (dataZFI26["天"] == globalDay)]["本次核销金额"])
worksheet4.cell(row=6, column=13).value = sum(dataZFI26[(dataZFI26["大客户"] == "大客户") & (dataZFI26["月份"] == globalMonth) & (dataZFI26["天"] == globalDay)]["核定价目表结算金额"])


# ----------------------------------------------------------------------------
# 第五页 5_【未发货】
# ----------------------------------------------------------------------------
worksheet5 = workbook.create_sheet('5_【未发货】', 4)

# for irow in range(len(dispatch["凭证日期"])):
#     dispatch.loc[irow, "凭证日期"] = str(dispatch.loc[irow, "凭证日期"])[0:10]

# 获取最新抽数的数据
dispatch = dispatch[dispatch['抽数日期'] == max(dispatch['抽数日期'])]     
South_counts =  len(dispatch[dispatch['南北部'] == '南部']['南北部'])
North_counts =  len(dispatch[dispatch['南北部'] == '北部']['南北部'])
dispatch.drop(columns=['抽数日期'], axis=1, inplace=True)  # 去掉抽数日期这一列
# dispatch['抽数日期'] = ''

# 设置行高列宽
for row in range(1, len(dispatch) + 3):
    if row == 1:
        worksheet5.row_dimensions[row].height = 24
    else:
        worksheet5.row_dimensions[row].height = 16
columnWidth = [15, 17, 17, 17, 40, 17, 65]
for icol, iwidth in enumerate(columnWidth, start=1):
    worksheet5.column_dimensions[opxl.utils.get_column_letter(icol)].width = iwidth

# 第 1 行 - 表头设计
for icol in range(1, 7):
    fcell = worksheet5.cell(row=1, column=icol)
    fcell.border = Border(left=Side(border_style='thin', color='A6A6A6'),
                          right=Side(border_style='thin', color='A6A6A6'),
                          top=Side(border_style='thin', color='A6A6A6'),
                          bottom=Side(border_style='thin', color='A6A6A6'))
worksheet5.merge_cells('A1:G1')
worksheet5['A1'] = "未发货明细表"
title5 = worksheet5['A1']
title5.font = Font(name="微软雅黑", size=14, bold=True, color='FF0000')
title5.alignment = Alignment(horizontal='center', vertical='center')
title5.fill = PatternFill(fill_type='solid', fgColor="FFFFFF") 

# 第 2 行 - 表头设计
subTitle5 = ['区域', '大区', '省份', '订单日期', '客户名称', '订单金额', '未发货原因']
for icol, content in enumerate(subTitle5, start=1):
    worksheet5[opxl.utils.get_column_letter(icol) + '2'] = content
    subtitle5 = worksheet5[opxl.utils.get_column_letter(icol) + '2']
    subtitle5.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
    subtitle5.alignment = Alignment(horizontal='center', vertical='center')
    subtitle5.fill = PatternFill(fill_type='solid', fgColor="FF9F3F")
    subtitle5.border = Border(left=Side(border_style='thin', color='FF000000'),
                              right=Side(border_style='thin', color='FF000000'),
                              top=Side(border_style='thin', color='FF000000'),
                              bottom=Side(border_style='thin', color='FF000000'))

# 填充内容
for irow in range(dispatch.iloc[:, 0].size):
    for jcol in range(dispatch.columns.size):
        worksheet5.cell(row=irow + 3, column=jcol + 1).value = dispatch.iloc[irow, jcol]
        conncell = worksheet5.cell(row=irow + 3, column=jcol + 1)
        if jcol == 0:
            conncell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
        else:
            conncell.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
        conncell.alignment = Alignment(horizontal='center', vertical='center')
        conncell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                 right=Side(border_style='thin', color='FF000000'),
                                 top=Side(border_style='thin', color='FF000000'),
                                 bottom=Side(border_style='thin', color='FF000000'))
        if irow == dispatch.iloc[:, 0].size - 1:
            if jcol == 1:
                worksheet5.cell(row=irow + 4, column=1).value = "合计"
                worksheet5.merge_cells(start_row=irow + 4, end_row=irow + 4, start_column=1, end_column=5)
                worksheet5.cell(row=irow + 4, column=6).value = sum(dispatch["订单总金额"])
            connlastcell = worksheet5.cell(row=irow + 4, column=jcol + 1)
            connlastcell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
            connlastcell.alignment = Alignment(horizontal='center', vertical='center')
            connlastcell.fill = PatternFill(fill_type='solid', fgColor="FFBC79")
            connlastcell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'))

worksheet5.merge_cells(start_row=3, end_row=South_counts + 2, start_column=1, end_column=1)
worksheet5.merge_cells(start_row=South_counts + 3, end_row=South_counts + North_counts + 2, start_column=1, end_column=1)


# ----------------------------------------------------------------------------
# 第六页 6_【彩盒】
# ----------------------------------------------------------------------------
worksheet6 = workbook['6_【彩盒】']

# # 公式填充最后一列
# for lastcol in range(4, 33):
#     worksheet6.cell(row=lastcol, column=9).value = "=SUM(D" + str(lastcol) + ":H" + str(lastcol) + ")"

# material = ['川芎破壁饮片-1g×20袋', '芪枣口服液-10毫升/支×12支/盒', '归芪养血口服液-15ml×10支-30盒', '桔梗破壁饮片-2g×20袋', '太子参破壁饮片-2g×20袋']
# shengqurow = [4, 5, 7, 8, 9, 11, 12, 14, 16, 17, 19, 20, 21, 23, 24, 26, 27, 28, 30]
# shengqu = ['江苏', '上海浙江', '广东', '广西', '福江', '湖北', '湖南', '河北', '山东', '安徽', '陕西', '甘青宁新疆', '山西内蒙', '云南贵州', '四川重庆', '辽宁', '黑龙江', '吉林', '河南']

# for jcol, conncol in enumerate(material, start=4):
#     # 填充非汇总 小计 行
#     for irow, connrow in zip(shengqurow, shengqu):
#         worksheet6.cell(row=irow, column=jcol).value = sum(dataZSD043[(dataZSD043["黄芪区域"] == connrow) & (dataZSD043["物料描述"] == conncol)]["交货数量"])
#     # 公式填充汇总 小计 合计 行
#     worksheet6.cell(row=6, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "4:" + str(opxl.utils.get_column_letter(jcol)) + "5)"
#     worksheet6.cell(row=10, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "7:" + str(opxl.utils.get_column_letter(jcol)) + "9)"
#     worksheet6.cell(row=13, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "11:" + str(opxl.utils.get_column_letter(jcol)) + "12)"
#     worksheet6.cell(row=15, column=jcol).value = "=" + str(opxl.utils.get_column_letter(jcol)) + "6+" + str(opxl.utils.get_column_letter(jcol)) + \
#         "10+" + str(opxl.utils.get_column_letter(jcol)) + "13+" + str(opxl.utils.get_column_letter(jcol)) + "14"  # 南部 汇总
#     worksheet6.cell(row=18, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "16:" + str(opxl.utils.get_column_letter(jcol)) + "17)"
#     worksheet6.cell(row=22, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "19:" + str(opxl.utils.get_column_letter(jcol)) + "21)"
#     worksheet6.cell(row=25, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "23:" + str(opxl.utils.get_column_letter(jcol)) + "24)"
#     worksheet6.cell(row=29, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "26:" + str(opxl.utils.get_column_letter(jcol)) + "28)"
#     worksheet6.cell(row=31, column=jcol).value = "=" + str(opxl.utils.get_column_letter(jcol)) + "18+" + str(opxl.utils.get_column_letter(jcol)) + "22+" + \
#         str(opxl.utils.get_column_letter(jcol)) + "25+" + str(opxl.utils.get_column_letter(jcol)) + "29+" + str(opxl.utils.get_column_letter(jcol)) + "30"  # 北部 汇总
#     worksheet6.cell(row=32, column=jcol).value = "=" + str(opxl.utils.get_column_letter(jcol)) + "15+" + str(opxl.utils.get_column_letter(jcol)) + "31"  # 合计

shengqurow = [5, 6, 8, 9, 10, 12, 13, 15, 17, 18, 20, 21, 22, 24, 25, 27, 28, 29, 31]
shengqu = ['江苏', '上海浙江', '广东', '广西', '福江', '湖北', '湖南', '河北', '山东', '安徽', '陕西', '甘青宁新疆', '山西内蒙', '云南贵州', '四川重庆', '辽宁', '黑龙江', '吉林', '河南']

for irow, connrow in zip(shengqurow, shengqu):
    # 全年
    worksheet6.cell(row=irow, column=4).value = sum(dataZSD043[(dataZSD043['黄芪区域'] == connrow) & (dataZSD043['物料描述'] == '彩盒')]['交货数量'])       
    # 本月
    worksheet6.cell(row=irow, column=5).value = sum(dataZSD043[(dataZSD043['黄芪区域'] == connrow) & (dataZSD043['物料描述'] == '彩盒') & (dataZSD043['月份'] == globalMonth)]['交货数量'])       
    # 当天
    worksheet6.cell(row=irow, column=6).value = sum(dataZSD043[(dataZSD043['黄芪区域'] == connrow) & (dataZSD043['物料描述'] == '彩盒') & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay)]['交货数量'])       

for jcol in [4, 5, 6]:
    # 公式填充汇总 小计 合计 行
    worksheet6.cell(row=7, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "5:" + str(opxl.utils.get_column_letter(jcol)) + "6)"
    worksheet6.cell(row=11, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "8:" + str(opxl.utils.get_column_letter(jcol)) + "10)"
    worksheet6.cell(row=14, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "12:" + str(opxl.utils.get_column_letter(jcol)) + "13)"
    worksheet6.cell(row=16, column=jcol).value = "=" + str(opxl.utils.get_column_letter(jcol)) + "7+" + str(opxl.utils.get_column_letter(jcol)) + \
        "11+" + str(opxl.utils.get_column_letter(jcol)) + "14+" + str(opxl.utils.get_column_letter(jcol)) + "15"  # 南部 汇总
    worksheet6.cell(row=19, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "17:" + str(opxl.utils.get_column_letter(jcol)) + "18)"
    worksheet6.cell(row=23, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "20:" + str(opxl.utils.get_column_letter(jcol)) + "22)"
    worksheet6.cell(row=26, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "24:" + str(opxl.utils.get_column_letter(jcol)) + "25)"
    worksheet6.cell(row=30, column=jcol).value = "=SUM(" + str(opxl.utils.get_column_letter(jcol)) + "27:" + str(opxl.utils.get_column_letter(jcol)) + "29)"
    worksheet6.cell(row=32, column=jcol).value = "=" + str(opxl.utils.get_column_letter(jcol)) + "19+" + str(opxl.utils.get_column_letter(jcol)) + "23+" + \
        str(opxl.utils.get_column_letter(jcol)) + "26+" + str(opxl.utils.get_column_letter(jcol)) + "30+" + str(opxl.utils.get_column_letter(jcol)) + "31"  # 北部 汇总
    worksheet6.cell(row=33, column=jcol).value = "=" + str(opxl.utils.get_column_letter(jcol)) + "16+" + str(opxl.utils.get_column_letter(jcol)) + "32"  # 合计


# ----------------------------------------------------------------------------
# 第七页 7_【欠货】
# ----------------------------------------------------------------------------
worksheet7 = workbook.create_sheet('7_【欠货】', 6)

# 获取最新日期数据
misgoods = misgoods[(misgoods["预计出货时间"] == max(misgoods["预计出货时间"])) & (misgoods["事业部"] == '草晶华销售事业部')]

if misgoods.shape[0] == 0:  # 无欠货情况下
    # 设置行高列宽
    for row in range(1, 4):
        if row == 1:
            worksheet7.row_dimensions[row].height = 30
        else:
            worksheet7.row_dimensions[row].height = 15.6
    columnWidth = [8, 35, 13, 13, 13, 13, 13]
    for icol, width in enumerate(columnWidth, start=1):
        worksheet7.column_dimensions[opxl.utils.get_column_letter(icol)].width = width

    # 第 1 行 - 表头设计
    for icol in range(1, 8):
        fcell = worksheet7.cell(row=1, column=icol)
        fcell.border = Border(left=Side(border_style='thin', color='A6A6A6'),
                            right=Side(border_style='thin', color='A6A6A6'),
                            top=Side(border_style='thin', color='A6A6A6'),
                            bottom=Side(border_style='thin', color='A6A6A6'))
    worksheet7.merge_cells('A1:G1')
    worksheet7['A1'] = "欠货表"
    title7 = worksheet7['A1']
    title7.font = Font(name="微软雅黑", size=14, bold=True, color='FF0000')
    title7.alignment = Alignment(horizontal='center', vertical='center')
    title7.fill = PatternFill(fill_type='solid', fgColor="FFFFFF")

    # 第 2 行 - 表头设计
    subTitle7 = ['序号', '品名/规格', '数量（罐/盒）', '欠货金额', '开始欠货 时间', '预计出货 时间', '备注']
    for icol, content in enumerate(subTitle7, start=1):
        worksheet7[opxl.utils.get_column_letter(icol) + '2'] = content
        subtitle7 = worksheet7[opxl.utils.get_column_letter(icol) + '2']
        subtitle7.font = Font(name="微软雅黑", size=10, bold=True, color='FFFFFF')
        subtitle7.alignment = Alignment(horizontal='center', vertical='center')
        subtitle7.fill = PatternFill(fill_type='solid', fgColor="7030A0")
        subtitle7.border = Border(left=Side(border_style='thin', color='FF000000'),
                                right=Side(border_style='thin', color='FF000000'),
                                top=Side(border_style='thin', color='FF000000'),
                                bottom=Side(border_style='thin', color='FF000000'))

    # 填充内容
    worksheet7.cell(row=4, column=1).value = "合计"
    worksheet7.merge_cells(start_row=4, end_row=4, start_column=1, end_column=2)
    worksheet7.cell(row=4, column=3).value = 0
    worksheet7.cell(row=4, column=4).value = 0
        
    for jcol in range(7):
        if jcol == 0:
            worksheet7.cell(row=3, column=1).value = 1
        elif jcol == 6:
            worksheet7.cell(row=3, column=7).value = '无欠货'
        else:
            worksheet7.cell(row=3, column=jcol + 1).value = '(空白)'
        
        conncell7 = worksheet7.cell(3, column=jcol + 1)
        conncell7.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
        conncell7.alignment = Alignment(horizontal='center', vertical='center')
        conncell7.border = Border(left=Side(border_style='thin', color='FF000000'),
                                right=Side(border_style='thin', color='FF000000'),
                                top=Side(border_style='thin', color='FF000000'),
                                bottom=Side(border_style='thin', color='FF000000'))
        
        connlastcell = worksheet7.cell(row= 4, column=jcol + 1)
        connlastcell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
        connlastcell.alignment = Alignment(horizontal='center', vertical='center')
        connlastcell.fill = PatternFill(fill_type='solid', fgColor="BD92DE")
        connlastcell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
else:  # 有欠货情况下
    # 透视原欠货表
    # goods = pd.pivot_table(misgoods, index=[u'品名'], values=[u'数量', u'欠货金额'], aggfunc=[np.sum], fill_value=0, margins=False, margins_name="总计")
    goods = pd.pivot_table(misgoods, index=[u'品名'], values=[u'数量', u'欠货金额'], aggfunc={u'数量':np.sum, u'欠货金额':np.sum}, fill_value=0, margins=False, margins_name="总计")

    # 将透视表的多维表头拍平
    # goods.columns = [col1 + '_' + str(col2) for (col1, col2) in goods.columns.tolist()]
    goods.reset_index(inplace=True)
    goods.columns = ['品名', '数量', '欠货金额']
    goods.insert(3, '开始欠货时间', '')
    goods.insert(4, '预计出货时间', '')

    # 填充最早欠货日期
    for irow in range(goods.shape[0]):
        goods.loc[irow, '开始欠货时间'] = min(misgoods[misgoods['品名'] == goods.loc[irow, '品名']]['开始欠货时间'])
        goods.loc[irow, '预计出货时间'] = '待定'

    # 设置行高列宽
    for row in range(1, len(goods) + 3):
        if row == 1:
            worksheet7.row_dimensions[row].height = 30
        else:
            worksheet7.row_dimensions[row].height = 15.6
    columnWidth = [8, 35, 13, 13, 13, 13, 13]
    for icol, width in enumerate(columnWidth, start=1):
        worksheet7.column_dimensions[opxl.utils.get_column_letter(icol)].width = width

    # 第 1 行 - 表头设计
    for icol in range(1, 8):
        fcell = worksheet7.cell(row=1, column=icol)
        fcell.border = Border(left=Side(border_style='thin', color='A6A6A6'),
                            right=Side(border_style='thin', color='A6A6A6'),
                            top=Side(border_style='thin', color='A6A6A6'),
                            bottom=Side(border_style='thin', color='A6A6A6'))
    worksheet7.merge_cells('A1:G1')
    worksheet7['A1'] = "欠货表"
    title7 = worksheet7['A1']
    title7.font = Font(name="微软雅黑", size=14, bold=True, color='FF0000')
    title7.alignment = Alignment(horizontal='center', vertical='center')
    title7.fill = PatternFill(fill_type='solid', fgColor="FFFFFF")

    # 第 2 行 - 表头设计
    subTitle7 = ['序号', '品名/规格', '数量（罐/盒）', '欠货金额', '开始欠货 时间', '预计出货 时间', '备注']
    for icol, content in enumerate(subTitle7, start=1):
        worksheet7[opxl.utils.get_column_letter(icol) + '2'] = content
        subtitle7 = worksheet7[opxl.utils.get_column_letter(icol) + '2']
        subtitle7.font = Font(name="微软雅黑", size=10, bold=True, color='FFFFFF')
        subtitle7.alignment = Alignment(horizontal='center', vertical='center')
        subtitle7.fill = PatternFill(fill_type='solid', fgColor="7030A0")
        subtitle7.border = Border(left=Side(border_style='thin', color='FF000000'),
                                right=Side(border_style='thin', color='FF000000'),
                                top=Side(border_style='thin', color='FF000000'),
                                bottom=Side(border_style='thin', color='FF000000'))

    # 填充内容
    for irow in range(goods.iloc[:, 0].size):
        for jcol in range(goods.columns.size + 2):
            worksheet7.cell(row=irow + 3, column=1).value = irow + 1
            if jcol <= 4:
                worksheet7.cell(row=irow + 3, column=jcol + 2).value = goods.iloc[irow, jcol]
            conncell7 = worksheet7.cell(row=irow + 3, column=jcol + 1)
            conncell7.font = Font(name="微软雅黑", size=10, bold=False, color='000000')
            conncell7.alignment = Alignment(horizontal='center', vertical='center')
            conncell7.border = Border(left=Side(border_style='thin', color='FF000000'),
                                    right=Side(border_style='thin', color='FF000000'),
                                    top=Side(border_style='thin', color='FF000000'),
                                    bottom=Side(border_style='thin', color='FF000000'))
            if irow == goods.iloc[:, 0].size - 1:
                if jcol == 1:
                    worksheet7.cell(row=irow + 4, column=1).value = "合计"
                    worksheet7.merge_cells(start_row=irow + 4, end_row=irow + 4, start_column=1, end_column=2)
                    worksheet7.cell(row=irow + 4, column=3).value = sum(goods["数量"])
                    worksheet7.cell(row=irow + 4, column=4).value = sum(goods["欠货金额"])
                connlastcell = worksheet7.cell(row=irow + 4, column=jcol + 1)
                connlastcell.font = Font(name="微软雅黑", size=10, bold=True, color='000000')
                connlastcell.alignment = Alignment(horizontal='center', vertical='center')
                connlastcell.fill = PatternFill(fill_type='solid', fgColor="BD92DE")
                connlastcell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                            right=Side(border_style='thin', color='FF000000'),
                                            top=Side(border_style='thin', color='FF000000'),
                                            bottom=Side(border_style='thin', color='FF000000'))


# ----------------------------------------------------------------------------
# 带时间文字填充
# ----------------------------------------------------------------------------
# 第 3 页
worksheet3.cell(row=2, column=2).value = '草晶华-' + str(globalYear) + '年' + str(globalMonth) + '月发货汇总表'
worksheet3.cell(row=3, column=2).value = '报表更新时间: ' + str(datetime.now().strftime("%Y-%m-%d"))
worksheet3.cell(row=4, column=7).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
worksheet3.cell(row=4, column=13).value = str(globalDay) + '日当天销售金额'
worksheet3.cell(row=5, column=3).value = '1-' + str(globalMonth) + '月发货'
worksheet3.cell(row=5, column=4).value = '1-' + str(globalMonth) + '月发货'
worksheet3.cell(row=5, column=5).value = '1-' + str(globalMonth) + '月回款'
worksheet3.cell(row=5, column=6).value = '1-' + str(globalMonth) + '月回款'

# 第 4 页
worksheet4.cell(row=1, column=1).value = '草晶华-大客户' + str(globalYear) + '年' + str(globalMonth) + '月发货汇总表（含分部）'
worksheet4.cell(row=3, column=2).value = str(globalYear) + '年销售金额'
worksheet4.cell(row=3, column=6).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
worksheet4.cell(row=3, column=10).value = str(globalDay) + '日当天销售金额'
worksheet4.cell(row=7, column=12).value = '日期: ' + str(globalYear) + '/' + str(globalMonth).zfill(2) + '/' + str(globalDay).zfill(2)

# 第 6 页
worksheet6.cell(row=2, column=1).value = '日期范围: ' + str(globalYear) + '年1月1日 至 ' + str(globalYear) + '年' + str(globalMonth) + '月' + str(globalDay) + '日'
worksheet6.cell(row=4,column=5).value = str(globalMonth) + '月'
worksheet6.cell(row=4,column=6).value = str(globalMonth) + '月' + str(globalDay) + '日'


# ----------------------------------------------------------------------------
# 激活或者禁止激活指定子页
# ----------------------------------------------------------------------------
for act in workbook.sheetnames:
    workbook[act].views.sheetView[0].tabSelected = False
workbook.active = workbook['1_【发货】']

 
# ----------------------------------------------------------------------------
# 存储表格
# ----------------------------------------------------------------------------
try:
    workbook.save('C:/Users/zoodehao/Desktop/AutoSend/3_销管/销管_正式发送文件.xlsx')
    print('文件保存成功！！！')
except Exception as e:
    print('文件保存失败！！！', e)


# ----------------------------------------------------------------------------
# 以下是发送板块
# ----------------------------------------------------------------------------

# 清空指定文件夹
def deleteOldFiles(path):
    deleteFileList = os.listdir(path)
    all_PNG = glob.glob(path + "*.PNG")
    print("该目录下文件有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中, PNG: " + str(len(all_PNG)) + "个")
    if len(all_PNG) != 0:
        for deletefile in deleteFileList:
            isDeleteFile = os.path.join(path, deletefile)
            if os.path.isfile(isDeleteFile):
                os.remove(isDeleteFile)
        all_DelPNG = glob.glob(path + "*.*")
        if len(all_DelPNG) == 0:
            print("已清空文件夹！！！")
        else:
            print("存在未删除文件, 请检查是否存在非PNG格式文件")
    else:
        print("不存在PNG文件")


# screenArea——格式类似"A1:J10"
def excelCatchScreen(file_name, sheet_name, name, save_path):
    pythoncom.CoInitialize()  # excel多线程相关
    Application = win32com.client.gencache.EnsureDispatch("Excel.Application")  # 启动excel
    Application.Visible = False  # 是否可视化
    Application.DisplayAlerts = False  # 是否显示警告
    wb = Application.Workbooks.Open(file_name, ReadOnly=False)  # 打开excel
    # ws = wb.Sheets(sheet_name)  # 选择Sheet
    ws = wb.Worksheets(sheet_name)  # 选择Sheet
    ws.Activate()  # 激活当前工作表
    userange = ws.UsedRange
    # 注意：要从A1开始的表格
    screen_area = 'A1:' + str(opxl.utils.get_column_letter(userange.Columns.Count)) + str(userange.Rows.Count)
    ws.Range(screen_area).CopyPicture()  # 复制图片区域
    time.sleep(1)
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
    Application.Selection.ShapeRange.Name = name  # 将刚刚选择的Shape重命名, 避免与已有图片混淆
    ws.Shapes(name).Copy()  # 选择图片
    time.sleep(1)
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    img_name = name + ".PNG"
    img.save(save_path + img_name)  # 保存图片
    # time.sleep(1)
    # wb.Save()
    # time.sleep(1)
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    time.sleep(1)
    Application.Quit()  # 退出excel
    pythoncom.CoUninitialize()


# 定义钉钉功能
class dingdingFunction(object):
    def __init__(self, roboturl, robotsecret, appkey, appsecret):
        """
        :param roboturl: 群机器人WebHook_url
        :param robotsecret: 安全设置的加签秘钥
        :param appkey: 企业开发平台小程序AppKey
        :param appsecret: 企业开发平台小程序AppSecret
        """
        self.roboturl = roboturl
        self.robotsecret = robotsecret
        self.appkey = appkey
        self.appsecret = appsecret
        timestamp = round(time.time() * 1000)  # 时间戳
        secret_enc = robotsecret.encode('utf-8')
        string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
        string_to_sign_enc = string_to_sign.encode('utf-8')
        hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
        self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(timestamp, sign)  # 最终url,url+时间戳+签名

    # 发送文件
    def getAccess_token(self):
        url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (AppKey, AppSecret)
        headers = {
            'Content-Type': "application/x-www-form-urlencoded"
        }
        data = {'appkey': self.appkey,
                'appsecret': self.appsecret}
        r = requests.request('GET', url, data=data, headers=headers)
        access_token = r.json()["access_token"]
        return access_token

    def getMedia_id(self, filespath):
        access_token = self.getAccess_token()  # 拿到接口凭证
        url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
        files = {'media': open(filespath, 'rb')}
        data = {'access_token': access_token,
                'type': 'file'}
        response = requests.post(url, files=files, data=data)
        json = response.json()
        return json["media_id"]

    def sendFile(self, chatid, filespath):
        access_token = self.getAccess_token()
        media_id = self.getMedia_id(filespath)
        url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
        header = {
            'Content-Type': 'application/json'
        }
        data = {'access_token': access_token,
                'chatid': chatid,
                'msg': {
                    'msgtype': 'file',
                    'file': {'media_id': media_id}
                }}
        r = requests.request('POST', url, data=json.dumps(data), headers=header)
        print(r.json()["errmsg"])

    # 发送消息
    def sendMessage(self, content):
        """
        :param content: 发送内容
        """
        header = {
            "Content-Type": "application/json",
            "Charset": "UTF-8"
        }
        sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
        sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
        request = urllib.request.Request(url=self.webhook_url, data=sendContent, headers=header)  # 发送请求
        opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
        print(opener.read().decode())  # 打印返回的结果


# 上传本地图片获取网上图片URL
def get_image_url(imagePath, pictureName):
    if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    else:
        print("请检查图片格式！！！")
    # 七牛云密钥管理：https://portal.qiniu.com/user/key
    # 【账号：13267854059  密码：z****】
    access_key = "fjlWDEbF1fqBU98UsdDJRcSSKODT9Gq7tA3gu8eY"
    secret_key = "thiWFpO881GfhlaAz1Wkk2yEcvV3ue2OHnY_5D9V"
    keyq = Auth(access_key, secret_key)
    bucket = "zues3737img"  # 七牛云盘名
    # 删除
    butm = BucketManager(keyq)
    reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
    # 上传
    time.sleep(1)
    token = keyq.upload_token(bucket, filename)  # 上传新图片
    reformUp, informUp = put_file(token, filename, imagePath)
    if reformUp is not None:
        print('已成功上传 {}'.format(filename))
        time.sleep(1)
        baseURL = "https://cjh3737.zeus.cn/"  # 加速域名
        # subURL = baseURL + '/' + filename + '?imageView2/0/quality/100!/sharpen/1/interlace/1'
        subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/sharpen/50/interlace/1/ignore-error/1'
        pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
        time.sleep(1)
        return pictureURL
    else:
        print(filename + '上传失败！！！')


if __name__ == '__main__':

    deleteOldFiles('C:/Users/zoodehao/Desktop/AutoSend/3_销管/Pictures/')  # 清空文件夹历史文件

    AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序AppKey
    AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序AppSecret
    
    # 机器人测试群: https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c
    # 销管日销售数据核对群: https://oapi.dingtalk.com/robot/send?access_token=ed2b6e7ea36a9d768b7899d0ca8a9788cff317487d9021ede275a7350fb27f1a
    # 销管：日销售数据共享群: https://oapi.dingtalk.com/robot/send?access_token=f55407f99b80521faf30aa8e78035d006e41bddd358073f373551d0b91f36601
    # 草晶华销售事业部（领导群）: https://oapi.dingtalk.com/robot/send?access_token=7b601d4772a4d367af6e9bba268caabaed36ed62e12a5eb6f5da29f451bfcbfc
    
    # RobotWebHookURL1 = 'https://oapi.dingtalk.com/robot/send?access_token=dd024c8278110ff67cc706c1cc44234b3469f2e44fb9b5e1c17eecae713ad94c'  # 1机器人测试群
    # RobotWebHookURL2 = 'https://oapi.dingtalk.com/robot/send?access_token=d477fdc99015272414d38c4241e61091d8bb4f9d627c8947bb4b963ee69096a1'  # 2机器人测试群

    RobotWebHookURL1 = 'https://oapi.dingtalk.com/robot/send?access_token=f55407f99b80521faf30aa8e78035d006e41bddd358073f373551d0b91f36601'  # 销管：日销售数据共享群
    RobotWebHookURL2 = 'https://oapi.dingtalk.com/robot/send?access_token=7b601d4772a4d367af6e9bba268caabaed36ed62e12a5eb6f5da29f451bfcbfc'  # 草晶华销售事业部（领导群）
    
    RobotSecret = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'   # 群机器人加签秘钥secret(默认数运小助手)
    # chatId = 'chat294fd3795ede63fc6a479e5c074f9ba2'

    fileFullPath = 'C:/Users/zoodehao/Desktop/AutoSend/3_销管/销管_正式发送文件.xlsx'
    savePictuePath = 'C:/Users/zoodehao/Desktop/AutoSend/3_销管/Pictures/'

    workbook = opxl.load_workbook(fileFullPath)
    worksheetnames = workbook.sheetnames
    # order = ['①', '②', '③', '④', '⑤', '⑥', '⑦']  # 图片序号
    sendtitle = ["###### **① 草晶华事业部本日发货明细表**",
                 "###### **② 草晶华事业部本日回款明细表**",
                 "###### **③ 草晶华-发货汇总表**",
                 "###### **④ 草晶华-大客户发货汇总表（含分部）**",
                 "###### **⑤ 草晶华-未发货明细表**",
                 "###### **⑥ 草晶华-彩盒出货情况**",
                 "###### **⑦ 草晶华-欠货表**"]
    
    sendTypes = int(input('>>>0sendAll-1sendSingle:')) 
    
    if sendTypes == 0: # 发送形式 - 全部
        
        pictureURL = []
        for sheetname, picturename in zip(worksheetnames, worksheetnames):
            try:
                excelCatchScreen(fileFullPath, sheetname, picturename, savePictuePath)
            except BaseException:
                print(picturename + '截图出错！！！')
            try:
                getURL = get_image_url(savePictuePath + picturename + '.PNG', picturename)
                pictureURL.append(getURL)
            except BaseException:
                print(picturename + '图片URL出错！！！')
                
        for ititle, iurl in zip(sendtitle, pictureURL):
            ddMessage = {  # 发布消息内容
                "msgtype": "markdown",
                "markdown": {"title": "【发货回款】进度",  # @某人 才会显示标题
                                "text": ititle +
                                "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
                                "\n###### ----------------------------------------"
                                "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                "at": {
                    # "atMobiles": [15817552982],  # 指定@某人
                    "isAtAll": False  # 是否@所有人[False:否, True:是]
                }
            }

            # 发送消息
            dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 销管：日销售数据共享群
            dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 草晶华销售事业部（领导群）
        
        # 消息数据
        fahuo_Fa      = sum(dataZSD043[(dataZSD043["月份"] == globalMonth) & (dataZSD043["天"] == globalDay)]["含税总价"])
        fahuo_He      = sum(dataZSD043[(dataZSD043["月份"] == globalMonth) & (dataZSD043["天"] == globalDay)]["业绩结算金额"])
        huikuan_Fa    = sum(dataZFI26[(dataZFI26["月份"] == globalMonth) & (dataZFI26["天"] == globalDay)]["本次核销金额"])
        huikuan_He    = sum(dataZFI26[(dataZFI26["月份"] == globalMonth) & (dataZFI26["天"] == globalDay)]["核定价目表结算金额"])
        fahuo_FaSum   = sum(dataZSD043[dataZSD043["月份"] == globalMonth]["含税总价"])
        fahuo_HeSum   = sum(dataZSD043[dataZSD043["月份"] == globalMonth]["业绩结算金额"])
        huikuan_FaSum = sum(dataZFI26[dataZFI26["月份"] == globalMonth]["本次核销金额"])
        huikuan_HeSum = sum(dataZFI26[dataZFI26["月份"] == globalMonth]["核定价目表结算金额"])
        fahuo_rate    = sum(dataZSD043[dataZSD043["月份"] <= globalMonth]["业绩结算金额"]) / Decimal(sum(target[(target["类型"] == '发货') & (target["月份"] <= globalMonth)]['目标'])) * 100
        huikuan_rate  = sum(dataZFI26[dataZFI26["月份"] <= globalMonth]["核定价目表结算金额"]) / Decimal(sum(target[(target["类型"] == '回款') & (target["月份"] <= globalMonth)]['目标'])) * 100

        summaryMessage = {  # 发布消息内容
            "msgtype": "markdown",
            "markdown": {"title": "发货核销汇总分析",  # @某人 才会显示标题
                            "text": "##### **各位领导：**"
                                    "\n ##### **草晶华事业部**"
                                    "\n ##### **① " + str(globalYear) + "年" + str(globalMonth) + "月" + str(globalDay) + "日**"  
                                    "\n > ##### 发货（发货价）**" + str("{:.2f}".format(fahuo_Fa)) + "** 万元"
                                    "\n > ##### 发货（结算价）**" + str("{:.2f}".format(fahuo_He)) + "** 万元"
                                    "\n > ##### 回款（发货价）**" + str("{:.2f}".format(huikuan_Fa)) + "** 万元"
                                    "\n > ##### 回款（结算价）**" + str("{:.2f}".format(huikuan_He)) + "** 万元"
                                    "\n ##### **② " + str(globalMonth) + "月累计**"
                                    "\n > ##### 发货（发货价）**" + str("{:.2f}".format(fahuo_FaSum)) + "** 万元"
                                    "\n > ##### 发货（结算价）**" + str("{:.2f}".format(fahuo_HeSum)) + "** 万元"
                                    "\n > ##### 回款（发货价）**" + str("{:.2f}".format(huikuan_FaSum)) + "** 万元"
                                    "\n > ##### 回款（结算价）**" + str("{:.2f}".format(huikuan_HeSum)) + "** 万元"
                                    "\n ##### **③ " + str(globalYear) + "年总况**"
                                    "\n > ##### 1-" + str(globalMonth) + "月发货完成率 **" + str("{:.2f}%".format(fahuo_rate)) + "**, 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate)) + "**"}, 
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        
        dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(summaryMessage)  # 销管：日销售数据共享群
        dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(summaryMessage)  # 草晶华销售事业部（领导群）
            
    elif sendTypes == 1: # 发送形式 - 选择性单张 从 1 开始
        
        send_NO_Picture = int(input('发送第几张图片？'))
        
        print('***单独发送: ' + str(worksheetnames[send_NO_Picture - 1]) + '.PNG')

        excelCatchScreen(fileFullPath, worksheetnames[send_NO_Picture - 1], worksheetnames[send_NO_Picture - 1], savePictuePath)
        
        
        ddMessage = {  # 发布消息内容
                "msgtype": "markdown",
                "markdown": {"title": "【发货回款】进度",  # @某人 才会显示标题
                                "text": sendtitle[send_NO_Picture - 1] + 
                                "\n![Image被拦截, 请使用非公司网络查看](" + get_image_url(savePictuePath + worksheetnames[send_NO_Picture - 1] + '.PNG', worksheetnames[send_NO_Picture - 1]) + ")"
                                "\n###### ----------------------------------------"
                                "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                "at": {
                    # "atMobiles": [15817552982],  # 指定@某人
                    "isAtAll": False  # 是否@所有人[False:否, True:是]
                }
            }

        # 发送消息
        dingdingFunction(RobotWebHookURL1, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 销管：日销售数据共享群
        dingdingFunction(RobotWebHookURL2, RobotSecret, AppKey, AppSecret).sendMessage(ddMessage)  # 草晶华销售事业部（领导群）
        
    else:
        print('请输入0-1正确的发送方式！！！')
    
    # dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendFile(chatId, fileFullPath)  # 发送文件
