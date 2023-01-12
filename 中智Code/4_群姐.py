# '''
# Author: zoodehao
# Date: 2021-11-30 21:23:55
# LastEditTime: 2021-11-30 21:23:56
# FilePath: \PyCodec:\Users\zoodehao\Desktop\群姐.py
# Description: 逝者如斯夫, 不舍昼夜.
# '''
# ----------------------------------------------------------------------------
# 需求说明：每日21: 30程序运行, 自动生成一个多子页的Excel表格, 自动截图发送到指定钉钉群
# ----------------------------------------------------------------------------
import os
import re
import xlwt
import glob
import xlrd
import uuid
import time
import hmac
import json
import pyhdb
import base64
import shutil
import pymysql
import hashlib
import requests
import win32com
import calendar
import pythoncom
import numpy as np
import urllib.parse
import pandas as pd
import urllib.request
import openpyxl as opxl
from decimal import Decimal
from termcolor import cprint
from datetime import datetime
from PIL import ImageGrab, Image
from time import strftime, gmtime
from openpyxl.utils import get_column_letter
from win32com.client import Dispatch, DispatchEx
from qiniu import Auth, put_file, etag, BucketManager
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Protection, NamedStyle, GradientFill, Color
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule,IconSetRule, Rule, IconSet, FormatObject


# ----------------------------------------------------------------------------
# 全局日期设置 (直接影响数据获取范围)
# ----------------------------------------------------------------------------
globalYear  = datetime.now().year
globalMonth = datetime.now().month
globalDay   = datetime.now().day


def export_file():
    global dataZSD043
    global dataZFI26
    global hexiao_target
    # ----------------------------------------------------------------------------
    # 加载格式框架
    # ----------------------------------------------------------------------------
    filepath = 'C:/Users/Zeus/Desktop/autoSend/4_群姐/Target/群姐_格式框架.xlsx'
    # workbook = opxl.load_workbook(filepath, data_only=True)
    workbook = opxl.load_workbook(filepath)


    # ----------------------------------------------------------------------------
    # 【目标 匹配表】数据准备
    # ----------------------------------------------------------------------------
    hexiao_target = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/4_群姐/Target/群姐_数据源.xlsx', sheet_name=0, header=0)
    keke_target   = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/4_群姐/Target/群姐_数据源.xlsx', sheet_name=1, header=0)
    area_divide   = pd.read_excel('C:/Users/Zeus/Desktop/autoSend/4_群姐/Target/群姐_数据源.xlsx', sheet_name=2, header=0)


    # ----------------------------------------------------------------------------
    # HANA数据库获取数据
    # ----------------------------------------------------------------------------
    # 获取 Connection 对象
    # 内网: IP: 192.168.20.183  账号: HANA1107318 密码: Zeus@1107311
    # 公网: IP: 119.145.248.183  账号: HANAHBPOUTTEST 密码: Zeus@test147
    def get_HANA_Connection(): 
        connectionObj = pyhdb.connect(
            host     = '192.168.20.183',
            port     = 30015,
            user     = 'HANA1107318',
            password = 'Zeus@1107311'
        )
        return connectionObj

    # 1. 获取本年发货流水账ZSD043
    def get_matZSD043(ZSD043):
        cursorZSD043 = ZSD043.cursor()
        cursorZSD043.execute("""SELECT LFDAT, TO_NUMBER(MONTH (LFDAT)) AS YUE, TO_NUMBER(RIGHT(LFDAT, 2)) AS TIAN, BEZEI, ZZPART3_T_S, ZZPART4_T_S, MAKTX, KBETR, TO_NUMBER(SUM (LFIMG1)) AS AMOUNT, TO_NUMBER(SUM (KWERT) / 10000) AS HSZJ
                                FROM "ECC_BI"."ZOUTBOUND"
                                WHERE ZZPART1_T_S LIKE N'%网通事业部%' AND LFDAT != '' AND AUART IN ('ZOR', 'ZRE') AND YEAR(LFDAT) = :1
                                GROUP BY LFDAT, BEZEI, ZZPART3_T_S, ZZPART4_T_S, MAKTX, KBETR""",
                        [str(globalYear)])
        matZSD043 = cursorZSD043.fetchall()
        return matZSD043

    # 2. 获取本年核销明细表ZFI26
    def get_matZFI26(ZFI26):
        cursorZFI26 = ZFI26.cursor()
        cursorZFI26.execute("""WITH AA AS (SELECT MATNR, MAKTX
                               FROM "ECC_BI"."MAKT"
                               WHERE MANDT = 800 AND SPRAS = 1), -- 物料名称
                               BB AS (SELECT KUNNR, NAME1
                               FROM "ECC_BI"."KNA1"), -- 五级名称
                               CC AS (SELECT BUDAT, TO_NUMBER(MONTH(BUDAT)) AS YUE, TO_NUMBER(RIGHT(BUDAT, 2)) AS TIAN, WW006, ZZ022, ZZ023, MATNR , ZZPBZ, SUM(VV011) AS AMOUNT, SUM(VV002) AS BCHHJE
                               FROM "ECC_BI"."ZTFI008"
                               WHERE ZZ020 = '0000001003' AND VRGAR = 'A3' AND ZSJLH = '11' AND YEAR(BUDAT) = :1
                               GROUP BY BUDAT, ZZ020, ZZ022, ZZ023, WW006, MATNR, ZZPBZ),
                               DD AS (SELECT CC.BUDAT, CC.YUE, CC.TIAN, CC.WW006, CC.ZZ022, CC.ZZ023, CC.MATNR, AA.MAKTX, CC.ZZPBZ, CC.AMOUNT, CC.BCHHJE
                               FROM CC
                               LEFT JOIN AA ON CC.MATNR = AA.MATNR),
                               EE AS (SELECT DD.BUDAT, DD.YUE, DD.TIAN, DD.WW006, DD.ZZ022, BB.NAME1, DD.ZZ023, DD.MATNR, DD.MAKTX, DD.ZZPBZ, DD.AMOUNT, DD.BCHHJE
                               FROM DD
                               LEFT JOIN BB ON DD.ZZ022 = BB.KUNNR)
                               SELECT EE.BUDAT, EE.YUE, EE.TIAN, EE.WW006, EE.ZZ022, EE.NAME1, EE.ZZ023, BB.NAME1, EE.MATNR, EE.MAKTX, EE.ZZPBZ, EE.AMOUNT, EE.BCHHJE / 10000
                               FROM EE
                               LEFT JOIN BB ON EE.ZZ023 = BB.KUNNR""",
                    [str(globalYear)])
        matZFI26 = cursorZFI26.fetchall()
        return matZFI26

    conn = get_HANA_Connection()

    # 1. 发货流水账ZSD043
    dataZSD043 = pd.DataFrame(get_matZSD043(conn), columns=['发货日期', '月份', '天', '省份', '三级组织', '四级组织', '物料描述', '销售价', '交货数量', '含税总价'])
    # 2. 核销明细表ZFI26
    dataZFI26 = pd.DataFrame(get_matZFI26(conn), columns=['核销日期', '月份', '天', '省份', '三级组织ID', '三级组织', '四级组织ID', '四级组织', '物料描述ID', '物料描述', '赠品标识', '核销数量', '核销金额'])

    print('【发货】最新日期:' + str(max(dataZSD043['发货日期'])) + '; 【回款】最新日期:' + str(max(dataZFI26["核销日期"])))


    # ----------------------------------------------------------------------------
    # 数据清洗整理部分
    # ----------------------------------------------------------------------------
    # 重新归属百强连锁旧框架
    areaList = ['LT-南区', 'LT-中区', 'LT-北区', 'BQ-珠江大区', 'BQ-长江大区', 'BQ-黄河大区', 'TG-推广一区', 'TG-推广二区', 'TG-推广三区', 'KA-山东', 'TL-试点区', 'JY-基药大区']
    # 发货
    for irow in range(dataZSD043.shape[0]):
        if dataZSD043.loc[irow, '三级组织'] not in areaList:
            for jrow in range(area_divide.shape[0]):
                if area_divide.loc[jrow, '省份'] == dataZSD043.loc[irow, '省份']:
                    dataZSD043.loc[irow, '三级组织'] = area_divide.loc[jrow, '大区']
    # 核销
    for irow in range(dataZFI26.shape[0]):
        if dataZFI26.loc[irow, '三级组织'] not in areaList:
            for jrow in range(area_divide.shape[0]):
                if area_divide.loc[jrow, '省份'] == dataZFI26.loc[irow, '省份']:
                    dataZFI26.loc[irow, '三级组织'] = area_divide.loc[jrow, '大区']
                    
    # 发货核销 - 河南2021年6月起数据归为试点， 陕西2021年3月起数据归为试点
    # 发货
    for irow in range(dataZSD043.shape[0]):
        if (dataZSD043.loc[irow, '月份'] >= 6 and dataZSD043.loc[irow, '省份'] == '陕西') or ((dataZSD043.loc[irow, '月份'] >= 3 and dataZSD043.loc[irow, '省份'] == '河南')):
            dataZSD043.loc[irow, '三级组织'] = 'TL-试点区'
    # 核销
    for irow in range(dataZFI26.shape[0]):
        if (dataZFI26.loc[irow, '月份'] >= 6 and dataZFI26.loc[irow, '省份'] == '陕西') or ((dataZFI26.loc[irow, '月份'] >= 3 and dataZFI26.loc[irow, '省份'] == '河南')):
            dataZFI26.loc[irow, '三级组织'] = 'TL-试点区'

    # dataZSD043.to_excel('C:/Users/zoodehao/Desktop/发货.xlsx', index=False)
    # dataZFI26.to_excel('C:/Users/zoodehao/Desktop/核销.xlsx', index=False)

    # 克咳片单位换算成件 - 去掉赠品
    dataZSD043['销售价'] = pd.to_numeric(dataZSD043['销售价'], errors='coerce')
    kekeZSD043 = dataZSD043[((dataZSD043['物料描述'] == '克咳片-0.54g×16片') | (dataZSD043['物料描述'] == '克咳片-0.54g×32片')) & (~dataZSD043['销售价'].isin([0.02]))].copy()
    kekeZSD043.reset_index(inplace=True)
    for irow in range(kekeZSD043.shape[0]):
        if kekeZSD043.loc[irow, '物料描述'] == '克咳片-0.54g×16片':
            kekeZSD043.loc[irow, '交货数量'] = kekeZSD043.loc[irow, '交货数量'] / 400
        elif kekeZSD043.loc[irow, '物料描述'] == '克咳片-0.54g×32片':
            kekeZSD043.loc[irow, '交货数量'] = kekeZSD043.loc[irow, '交货数量'] / 200

    kekeZFI26 = dataZFI26[((dataZFI26['物料描述'] == '克咳片-0.54g×16片') | (dataZFI26['物料描述'] == '克咳片-0.54g×32片')) & (dataZFI26['赠品标识'] == '')].copy()
    kekeZFI26.reset_index(inplace=True)
    for irow in range(kekeZFI26.shape[0]):
        if kekeZFI26.loc[irow, '物料描述'] == '克咳片-0.54g×16片':
            kekeZFI26.loc[irow, '核销数量'] = kekeZFI26.loc[irow, '核销数量'] / 400
        elif kekeZFI26.loc[irow, '物料描述'] == '克咳片-0.54g×32片':
            kekeZFI26.loc[irow, '核销数量'] = kekeZFI26.loc[irow, '核销数量'] / 200
    
    # 秋冬大促时间范围 8月-12月 - 去掉赠品  
    qinZSD043 = kekeZSD043[(kekeZSD043['月份'] >= 8) & (kekeZSD043['三级组织'].isin(['TG-推广一区', 'TG-推广二区', 'TG-推广三区', 'TL-试点区'])) & (~kekeZSD043['销售价'].isin([0.01, 0.02, 0.03]))]
    qinZFI26  = kekeZFI26[(kekeZFI26['月份'] >= 8) & (kekeZFI26['三级组织'].isin(['TG-推广一区', 'TG-推广二区', 'TG-推广三区', 'TL-试点区'])) & (kekeZFI26['赠品标识'] == '')]
    
    # qinZSD043.to_excel('C:/Users/zoodehao/Desktop/发货keke.xlsx', index=False)
    # qinZFI26.to_excel('C:/Users/zoodehao/Desktop/核销keke.xlsx', index=False)
    

    # ----------------------------------------------------------------------------
    # 第一页 ①发货汇总表
    # ----------------------------------------------------------------------------
    worksheet1 = workbook['①发货汇总表']

    row1 = [5, 6, 7, 9, 10, 11, 13, 14, 15, 16, 18, 19]
    area1 = ['LT-南区', 'LT-中区', 'LT-北区', 'TG-推广一区', 'TG-推广二区', 'TG-推广三区', 'BQ-珠江大区', 'BQ-长江大区', 'BQ-黄河大区', 'KA-山东', 'TL-试点区', 'JY-基药大区']

    # 非合计部分
    for icol in range(6, 16):
        for irow, iarea in zip(row1, area1):
            if icol == 6:
                worksheet1.cell(row=irow, column=icol).value = sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标'])  
            elif icol == 7:
                if iarea == 'KA-山东':
                    worksheet1.cell(row=irow, column=icol).value = '-'  # 无任务目标
                else:
                    worksheet1.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标']))
            if icol == 8:
                worksheet1.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] <= globalMonth)]['含税总价']) 
            elif icol == 9:
                worksheet1.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额']) 
            elif icol == 10:
                worksheet1.cell(row=irow, column=icol).value = sum(hexiao_target[(hexiao_target['三级组织'] == iarea) & (hexiao_target['月份'] <= globalMonth)]['目标']) 
            elif icol == 11:
                worksheet1.cell(row=irow, column=icol).value = '=IFERROR(I' + str(irow) + '/J' + str(irow) + ',"-")'
            elif icol == 12:
                worksheet1.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] <= globalDay)]['含税总价']) 
            elif icol == 13:
                worksheet1.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] <= globalDay)]['核销金额']) 
            elif icol == 14:
                worksheet1.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay)]['含税总价']) 
            elif icol == 15:
                worksheet1.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay)]['核销金额']) 



    # 合计/汇总部分
    sumrow1 = [8, 12, 17, 20]
    for irow in sumrow1:
        for icol in range(6, 16):
            if irow == 8:
                if icol == 11:
                    worksheet1.cell(row=irow, column=icol).value = '=IFERROR(I' + str(irow) + '/J' + str(irow) + ',"-")'
                elif icol == 7:
                    worksheet1.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'LT-南区') | (dataZFI26['三级组织'] == 'LT-中区') | (dataZFI26['三级组织'] == 'LT-北区')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'LT-南区') | (hexiao_target['三级组织'] == 'LT-中区') | (hexiao_target['三级组织'] == 'LT-北区'))]['目标']))
                else:
                    worksheet1.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '5:' + str(get_column_letter(icol)) + '7)'
            elif irow == 12:
                if icol == 11:
                    worksheet1.cell(row=irow, column=icol).value = '=IFERROR(I' + str(irow) + '/J' + str(irow) + ',"-")'
                elif icol == 7:
                    worksheet1.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'TG-推广一区') | (dataZFI26['三级组织'] == 'TG-推广二区') | (dataZFI26['三级组织'] == 'TG-推广三区')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'TG-推广一区') | (hexiao_target['三级组织'] == 'TG-推广二区') | (hexiao_target['三级组织'] == 'TG-推广三区'))]['目标']))
                else:
                    worksheet1.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '9:' + str(get_column_letter(icol)) + '11)'
            elif irow == 17:
                if icol == 11:
                    worksheet1.cell(row=irow, column=icol).value = '=IFERROR(I' + str(irow) + '/J' + str(irow) + ',"-")'
                elif icol == 7:
                    worksheet1.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'BQ-珠江大区') | (dataZFI26['三级组织'] == 'BQ-长江大区') | (dataZFI26['三级组织'] == 'BQ-黄河大区') | (dataZFI26['三级组织'] == 'KA-山东')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'BQ-珠江大区') | (hexiao_target['三级组织'] == 'BQ-长江大区') | (hexiao_target['三级组织'] == 'BQ-黄河大区') | (hexiao_target['三级组织'] == 'KA-山东'))]['目标']))
                else:
                    worksheet1.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '13:' + str(get_column_letter(icol)) + '16)'
            else:  # 汇总部分 
                if icol == 11:
                    worksheet1.cell(row=irow, column=icol).value = '=IFERROR(I' + str(irow) + '/J' + str(irow) + ',"-")'
                elif icol == 7:
                    worksheet1.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'TL-试点区') | (dataZFI26['三级组织'] == 'JY-基药大区')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'TL-试点区') | (hexiao_target['三级组织'] == 'JY-基药大区'))]['目标']))
                else:
                    worksheet1.cell(row=irow, column=icol).value = '=' + str(get_column_letter(icol)) + '8+' + str(get_column_letter(icol)) + '12+' + str(get_column_letter(icol)) + '17+' + str(get_column_letter(icol)) + '18+' + str(get_column_letter(icol)) + '19'

    worksheet1.column_dimensions['C'].hidden = True  # 隐藏C列
    worksheet1.column_dimensions['D'].hidden = True  # 隐藏D列


    # ----------------------------------------------------------------------------
    # 第二页 ②克咳片
    # ----------------------------------------------------------------------------
    worksheet2 = workbook['②克咳片']

    row2 = [5, 6, 7]
    area2 = ['TG-推广一区', 'TG-推广二区', 'TG-推广三区']

    # 非合计部分
    for icol in range(3, 13):
        for irow, iarea in zip(row2, area2):
            if icol == 3:
                worksheet2.cell(row=irow, column=icol).value = sum(keke_target[keke_target['三级组织'] == iarea]['目标'])
            elif icol == 4:
                worksheet2.cell(row=irow, column=icol).value = Decimal(sum(kekeZSD043[(kekeZSD043['三级组织'] == iarea) & (kekeZSD043['月份'] <= globalMonth)]['交货数量'])) / Decimal(sum(keke_target[keke_target['三级组织'] == iarea]['目标'])) 
            elif icol == 5:
                worksheet2.cell(row=irow, column=icol).value = sum(kekeZSD043[(kekeZSD043['三级组织'] == iarea) & (kekeZSD043['月份'] <= globalMonth)]['交货数量']) 
            elif icol == 6:
                worksheet2.cell(row=irow, column=icol).value = sum(kekeZFI26[(kekeZFI26['三级组织'] == iarea) & (kekeZFI26['月份'] <= globalMonth)]['核销数量']) 
            elif icol == 7:
                worksheet2.cell(row=irow, column=icol).value = sum(keke_target[(keke_target['三级组织'] == iarea) & (keke_target['月份'] <= globalMonth)]['目标']) 
            elif icol == 8:
                worksheet2.cell(row=irow, column=icol).value = '=IFERROR(F' + str(irow) + '/G' + str(irow) + ',"-")'
            elif icol == 9:
                worksheet2.cell(row=irow, column=icol).value = sum(kekeZSD043[(kekeZSD043['三级组织'] == iarea) & (kekeZSD043['月份'] == globalMonth) & (kekeZSD043['天'] <= globalDay)]['交货数量']) 
            elif icol == 10:
                worksheet2.cell(row=irow, column=icol).value = sum(kekeZFI26[(kekeZFI26['三级组织'] == iarea) & (kekeZFI26['月份'] == globalMonth) & (kekeZFI26['天'] <= globalDay)]['核销数量']) 
            elif icol == 11:
                worksheet2.cell(row=irow, column=icol).value = sum(kekeZSD043[(kekeZSD043['三级组织'] == iarea) & (kekeZSD043['月份'] == globalMonth) & (kekeZSD043['天'] == globalDay)]['交货数量']) 
            elif icol == 12:
                worksheet2.cell(row=irow, column=icol).value = sum(kekeZFI26[(kekeZFI26['三级组织'] == iarea) & (kekeZFI26['月份'] == globalMonth) & (kekeZFI26['天'] == globalDay)]['核销数量']) 

    # 合计部分
    sumrow2 = [8]
    for irow in sumrow2:
        for icol in range(3, 13):
            if irow == 8:
                if icol == 8:
                    worksheet2.cell(row=irow, column=icol).value = '=IFERROR(F' + str(irow) + '/G' + str(irow) + ',"-")'
                elif icol == 4:
                    worksheet2.cell(row=irow, column=icol).value = Decimal(sum(kekeZSD043[((kekeZSD043['三级组织'] == 'TG-推广一区') | (kekeZSD043['三级组织'] == 'TG-推广二区') | (kekeZSD043['三级组织'] == 'TG-推广三区')) & (kekeZSD043['月份'] <= globalMonth)]['交货数量'])) / Decimal(sum(keke_target[((keke_target['三级组织'] == 'TG-推广一区') | (keke_target['三级组织'] == 'TG-推广二区') | (keke_target['三级组织'] == 'TG-推广三区'))]['目标']))
                else:
                    worksheet2.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '5:' + str(get_column_letter(icol)) + '7)'


    # ----------------------------------------------------------------------------
    # 第三页 ③推广线
    # ----------------------------------------------------------------------------
    worksheet3 = workbook['③推广线']

    row3 = [5, 6, 7]
    area3 = ['TG-推广一区', 'TG-推广二区', 'TG-推广三区']

    # 非合计部分
    for icol in range(4, 14):
        for irow, iarea in zip(row3, area3):
            if icol == 4:
                worksheet3.cell(row=irow, column=icol).value = sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标'])
            elif icol == 5:
                worksheet3.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标'])) 
            elif icol == 6:
                worksheet3.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] <= globalMonth)]['含税总价']) 
            elif icol == 7:
                worksheet3.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额']) 
            elif icol == 8:
                worksheet3.cell(row=irow, column=icol).value = sum(hexiao_target[(hexiao_target['三级组织'] == iarea) & (hexiao_target['月份'] <= globalMonth)]['目标']) 
            elif icol == 9:
                worksheet3.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
            elif icol == 10:
                worksheet3.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] <= globalDay)]['含税总价']) 
            elif icol == 11:
                worksheet3.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] <= globalDay)]['核销金额']) 
            elif icol == 12:
                worksheet3.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay)]['含税总价']) 
            elif icol == 13:
                worksheet3.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay)]['核销金额']) 

    # 合计部分
    sumrow3 = [8]
    for irow in sumrow3:
        for icol in range(4, 14):
            if irow == 8:
                if icol == 9:
                    worksheet3.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
                elif icol == 5:
                    worksheet3.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'TG-推广一区') | (dataZFI26['三级组织'] == 'TG-推广二区') | (dataZFI26['三级组织'] == 'TG-推广三区')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'TG-推广一区') | (hexiao_target['三级组织'] == 'TG-推广二区') | (hexiao_target['三级组织'] == 'TG-推广三区'))]['目标']))
                else:
                    worksheet3.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '5:' + str(get_column_letter(icol)) + '7)'


    # ----------------------------------------------------------------------------
    # 第四页 ④流通线
    # ----------------------------------------------------------------------------
    worksheet4 = workbook['④流通线']

    row4 = [5, 6, 7, 9]
    area4 = ['LT-南区', 'LT-中区', 'LT-北区', 'TL-试点区']

    # 非合计部分
    for icol in range(4, 14):
        for irow, iarea in zip(row4, area4):
            if icol == 4:
                worksheet4.cell(row=irow, column=icol).value = sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标'])
            elif icol == 5:
                worksheet4.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标'])) 
            elif icol == 6:
                worksheet4.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] <= globalMonth)]['含税总价']) 
            elif icol == 7:
                worksheet4.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额']) 
            elif icol == 8:
                worksheet4.cell(row=irow, column=icol).value = sum(hexiao_target[(hexiao_target['三级组织'] == iarea) & (hexiao_target['月份'] <= globalMonth)]['目标']) 
            elif icol == 9:
                worksheet4.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
            elif icol == 10:
                worksheet4.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] <= globalDay)]['含税总价']) 
            elif icol == 11:
                worksheet4.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] <= globalDay)]['核销金额']) 
            elif icol == 12:
                worksheet4.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay)]['含税总价']) 
            elif icol == 13:
                worksheet4.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay)]['核销金额']) 

    # 汇总部分
    sumrow4 = [8]
    for irow in sumrow4:
        for icol in range(4, 14):
            if irow == 8:
                if icol == 9:
                    worksheet4.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
                elif icol == 5:
                    worksheet4.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'LT-南区') | (dataZFI26['三级组织'] == 'LT-中区') | (dataZFI26['三级组织'] == 'LT-北区')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'LT-南区') | (hexiao_target['三级组织'] == 'LT-中区') | (hexiao_target['三级组织'] == 'LT-北区'))]['目标']))
                else:
                    worksheet4.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '5:' + str(get_column_letter(icol)) + '7)'


    # ----------------------------------------------------------------------------
    # 第五页 ⑤百强连锁
    # ----------------------------------------------------------------------------
    worksheet5 = workbook['⑤百强连锁']

    row5 = [5, 6, 7, 8]
    area5 = ['BQ-珠江大区', 'BQ-长江大区', 'BQ-黄河大区', 'KA-山东']

    # 非合计部分
    for icol in range(4, 14):
        for irow, iarea in zip(row5, area5):
            if icol == 4:
                worksheet5.cell(row=irow, column=icol).value = sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标'])
            elif icol == 5:
                if iarea == 'KA-山东':
                    worksheet5.cell(row=irow, column=icol).value = '-'  # 无任务目标
                else:
                    worksheet5.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[hexiao_target['三级组织'] == iarea]['目标']))
            elif icol == 6:
                worksheet5.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] <= globalMonth)]['含税总价']) 
            elif icol == 7:
                worksheet5.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] <= globalMonth)]['核销金额']) 
            elif icol == 8:
                worksheet5.cell(row=irow, column=icol).value = sum(hexiao_target[(hexiao_target['三级组织'] == iarea) & (hexiao_target['月份'] <= globalMonth)]['目标']) 
            elif icol == 9:
                worksheet5.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
            elif icol == 10:
                worksheet5.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] <= globalDay)]['含税总价']) 
            elif icol == 11:
                worksheet5.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] <= globalDay)]['核销金额']) 
            elif icol == 12:
                worksheet5.cell(row=irow, column=icol).value = sum(dataZSD043[(dataZSD043['三级组织'] == iarea) & (dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay)]['含税总价']) 
            elif icol == 13:
                worksheet5.cell(row=irow, column=icol).value = sum(dataZFI26[(dataZFI26['三级组织'] == iarea) & (dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay)]['核销金额']) 

    # 汇总部分
    sumrow5 = [9]
    for irow in sumrow5:
        for icol in range(4, 14):
            if irow == 9:
                if icol == 9:
                    worksheet5.cell(row=irow, column=icol).value = '=IFERROR(G' + str(irow) + '/H' + str(irow) + ',"-")'
                elif icol == 5:
                    worksheet5.cell(row=irow, column=icol).value = Decimal(sum(dataZFI26[((dataZFI26['三级组织'] == 'BQ-珠江大区') | (dataZFI26['三级组织'] == 'BQ-长江大区') | (dataZFI26['三级组织'] == 'BQ-黄河大区') | (dataZFI26['三级组织'] == 'KA-山东')) & (dataZFI26['月份'] <= globalMonth)]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'BQ-珠江大区') | (hexiao_target['三级组织'] == 'BQ-长江大区') | (hexiao_target['三级组织'] == 'BQ-黄河大区') | (hexiao_target['三级组织'] == 'KA-山东'))]['目标']))
                else:
                    worksheet5.cell(row=irow, column=icol).value = '=SUM(' + str(get_column_letter(icol)) + '5:' + str(get_column_letter(icol)) + '8)'


    # ----------------------------------------------------------------------------
    # 第六页 ⑥秋冬大促
    # ----------------------------------------------------------------------------
    worksheet6 = workbook['⑥秋冬大促']
    
    area6 = ['TL-陕西省区', 'TL-河南省区', 'TG-浙江省份', 'TG-山东省份', 'TG-福建省份', 'TG-江苏省份', 'TG-安徽省份', 'TG-江西省份', 'TG-湖北省份', 'TG-河北（含北京、天津）省份', 'TG-四川省份', 'TG-山西（含蒙西）省份',
            'TG-辽宁（含蒙东）省份', 'TG-黑龙江省份', 'TG-吉林省份', 'TG-甘青宁省份', 'TG-粤西（含海南）省份', 'TG-湖南省份', 'TG-广西省份', 'TG-贵州省份', 'TG-粤东省份', 'TG-云南省份', 'TG-重庆省份']

    # 非汇总部分
    for icol in [5, 6, 7, 8, 9]:
        for irow, iarea in zip(range(4, 27), area6):
            if icol == 5:
                worksheet6.cell(row=irow, column=icol).value = sum(qinZSD043[qinZSD043['四级组织'] == iarea]['交货数量'])
            elif icol ==6:
                worksheet6.cell(row=irow, column=icol).value = sum(qinZFI26[qinZFI26['四级组织'] == iarea]['核销数量'])
            elif icol ==7:
                worksheet6.cell(row=irow, column=icol).value = '=IFERROR(F' + str(irow) + '/D' + str(irow) + ',"-")'
            elif icol ==8:
                worksheet6.cell(row=irow, column=icol).value = ''
            else:
                worksheet6.cell(row=irow, column=icol).value = '=IFERROR(D' + str(irow) + '-F' + str(irow) + ',"-")'
                
    # 汇总部分
    for icol in [5, 6, 7, 9]:
        if icol == 5:
            worksheet6.cell(row=27, column=icol).value = '=SUM(E4:E26)'
        elif icol ==6:
            worksheet6.cell(row=27, column=icol).value = '=SUM(F4:F26)'
        elif icol ==7:
            worksheet6.cell(row=27, column=icol).value = '=IFERROR(F' + str(27) + '/D' + str(27) + ',"-")'
        else:
            worksheet6.cell(row=27, column=icol).value = '=IFERROR(D' + str(27) + '-F' + str(27) + ',"-")'
    
    
    # worksheet6.auto_filter.ref = 'A3:I26'
    # worksheet6.auto_filter.add_filter_column(3, [300,400], blank=False)
    # worksheet6.auto_filter.add_sort_condition('G4:G26', descending=True)
    

    # ----------------------------------------------------------------------------
    # 带时间文字填充
    # ----------------------------------------------------------------------------
    # 第 1 页 - ①发货汇总表
    worksheet1.cell(row=1, column=1).value = '网通事业部 - ' + str(globalYear) + '年1-' + str(globalMonth) + '月发货汇总表'
    worksheet1.cell(row=2, column=2).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)
    worksheet1.cell(row=3, column=8).value = '1-' + str(globalMonth) + '月累计完成情况'
    worksheet1.cell(row=3, column=12).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
    worksheet1.cell(row=3, column=14).value = str(globalDay) + '日当天销售金额'


    # 第 2 页 - ②克咳片
    worksheet2.cell(row=1, column=1).value = '推广线 - ' + str(globalYear) + '年1-' + str(globalMonth) + '月克咳片16s发货汇总表'
    worksheet2.cell(row=2, column=2).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)
    worksheet2.cell(row=3, column=5).value = '1-' + str(globalMonth) + '月累计完成情况'
    worksheet2.cell(row=3, column=9).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
    worksheet2.cell(row=3, column=11).value = str(globalDay) + '日当天销售金额'


    # 第 3 页 - ③推广线
    worksheet3.cell(row=1, column=1).value = '推广线 - ' + str(globalYear) + '年1-' + str(globalMonth) + '月发货汇总表'
    worksheet3.cell(row=2, column=2).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)
    worksheet3.cell(row=3, column=6).value = '1-' + str(globalMonth) + '月累计完成情况'
    worksheet3.cell(row=3, column=10).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
    worksheet3.cell(row=3, column=12).value = str(globalDay) + '日当天销售金额'


    # 第 4 页 - ④流通线
    worksheet4.cell(row=1, column=1).value = '流通线 - ' + str(globalYear) + '年1-' + str(globalMonth) + '月发货汇总表'
    worksheet4.cell(row=2, column=2).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)
    worksheet4.cell(row=3, column=6).value = '1-' + str(globalMonth) + '月累计完成情况'
    worksheet4.cell(row=3, column=10).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
    worksheet4.cell(row=3, column=12).value = str(globalDay) + '日当天销售金额'


    # 第 5 页 - ⑤百强连锁
    worksheet5.cell(row=1, column=1).value = '百强连锁 - ' + str(globalYear) + '年1-' + str(globalMonth) + '月发货汇总表'
    worksheet5.cell(row=2, column=2).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)
    worksheet5.cell(row=3, column=6).value = '1-' + str(globalMonth) + '月累计完成情况'
    worksheet5.cell(row=3, column=10).value = str(globalMonth) + '月（1-' + str(globalDay) + '日）销售金额'
    worksheet5.cell(row=3, column=12).value = str(globalDay) + '日当天销售金额'


    # 第 6 页 - ⑥秋冬大促
    worksheet6.cell(row=2, column=2).value = str(globalYear) + '-' + str(globalMonth).zfill(2) + '-' + str(globalDay).zfill(2)
    worksheet6.cell(row=2, column=3).value = '以下数据包含江苏省32片克咳片:（发货:' + str(int(sum(qinZSD043[(qinZSD043['四级组织'] == 'TG-江苏省份') & (qinZSD043['物料描述'] == '克咳片-0.54g×32片') ]['交货数量']))) + '件，核销:' + str(int(sum(qinZFI26[(qinZFI26['四级组织'] == 'TG-江苏省份') & (qinZFI26['物料描述'] == '克咳片-0.54g×32片')]['核销数量']))) + '件）'
    
    
    # ----------------------------------------------------------------------------
    # 存储表格
    # ----------------------------------------------------------------------------
    workbook.save('C:/Users/Zeus/Desktop/autoSend/4_群姐/群姐_正式发送文件.xlsx')


# ----------------------------------------------------------------------------
# 以下是发送部分
# ----------------------------------------------------------------------------
# 清空指定文件夹
def deleteOldFiles(path):
    deleteFileList = os.listdir(path)
    all_PNG = glob.glob(path + "*.PNG")
    print("该目录下文件有" + '\n' + str(deleteFileList) + ";" + '\n' + "其中, PNG: " + str(len(all_PNG)) + "个")
    if len(all_PNG) != 0:
        for deletefile in deleteFileList:
            isDeleteFile = os.path.join(path, deletefile)
            if os.path.isfile(isDeleteFile):
                os.remove(isDeleteFile)
        all_DelPNG = glob.glob(path + "*.*")
        if len(all_DelPNG) == 0:
            print("已清空文件夹！！！")
        else:
            print("存在未删除文件, 请检查是否存在非PNG格式文件")
    else:
        print("不存在PNG文件")


# screenArea——格式类似"A1:J10"
def excelCatchScreen(file_name, sheet_name, name, save_path):
    pythoncom.CoInitialize()  # excel多线程相关
    Application = win32com.client.gencache.EnsureDispatch("Excel.Application")  # 启动excel
    Application.Visible = False  # 可视化
    Application.DisplayAlerts = False  # 是否显示警告
    wb = Application.Workbooks.Open(file_name, ReadOnly=False)  # 打开excel
    # ws = wb.Sheets(sheet_name)  # 选择Sheet
    ws = wb.Worksheets(sheet_name)  # 选择Sheet
    ws.Activate()  # 激活当前工作表
    userange = ws.UsedRange
    # 注意：要从A1开始的表格
    screen_area = 'A1:' + str(opxl.utils.get_column_letter(userange.Columns.Count)) + str(userange.Rows.Count)
    ws.Range(screen_area).CopyPicture()  # 复制图片区域
    time.sleep(1)
    ws.Paste()  # 粘贴 ws.Paste(ws.Range('B1'))  # 将图片移动到具体位置
    Application.Selection.ShapeRange.Name = name  # 将刚刚选择的Shape重命名, 避免与已有图片混淆
    ws.Shapes(name).Copy()  # 选择图片
    time.sleep(1)
    img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    img_name = name + ".PNG"
    img.save(save_path + img_name)  # 保存图片
    # time.sleep(1)
    # wb.Save()
    # time.sleep(1)
    wb.Close(SaveChanges=0)  # 关闭工作薄，不保存
    # time.sleep(1)
    Application.Quit()  # 退出excel
    pythoncom.CoUninitialize()


# 定义钉钉功能
class dingdingFunction(object):
    def __init__(self, roboturl, robotsecret, appkey, appsecret):
        """
        :param roboturl: 群机器人WebHook_url
        :param robotsecret: 安全设置的加签秘钥
        :param appkey: 企业开发平台小程序AppKey
        :param appsecret: 企业开发平台小程序AppSecret
        """
        self.roboturl = roboturl
        self.robotsecret = robotsecret
        self.appkey = appkey
        self.appsecret = appsecret
        timestamp = round(time.time() * 1000)  # 时间戳
        secret_enc = robotsecret.encode('utf-8')
        string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
        string_to_sign_enc = string_to_sign.encode('utf-8')
        hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
        self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(timestamp, sign)  # 最终url,url+时间戳+签名

    # 发送文件
    def getAccess_token(self):
        url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (AppKey, AppSecret)
        headers = {
            'Content-Type': "application/x-www-form-urlencoded"
        }
        data = {'appkey': self.appkey,
                'appsecret': self.appsecret}
        r = requests.request('GET', url, data=data, headers=headers)
        access_token = r.json()["access_token"]
        return access_token

    def getMedia_id(self, filespath):
        access_token = self.getAccess_token()  # 拿到接口凭证
        url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
        files = {'media': open(filespath, 'rb')}
        data = {'access_token': access_token,
                'type': 'file'}
        response = requests.post(url, files=files, data=data)
        json = response.json()
        return json["media_id"]

    def sendFile(self, chatid, filespath):
        access_token = self.getAccess_token()
        media_id = self.getMedia_id(filespath)
        url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
        header = {
            'Content-Type': 'application/json'
        }
        data = {'access_token': access_token,
                'chatid': chatid,
                'msg': {
                    'msgtype': 'file',
                    'file': {'media_id': media_id}
                }}
        r = requests.request('POST', url, data=json.dumps(data), headers=header)
        print(r.json()["errmsg"])

    # 发送消息
    def sendMessage(self, content):
        """
        :param content: 发送内容
        """
        header = {
            "Content-Type": "application/json",
            "Charset": "UTF-8"
        }
        sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
        sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
        request = urllib.request.Request(url=self.webhook_url, data=sendContent, headers=header)  # 发送请求
        opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
        print(opener.read().decode())  # 打印返回的结果
        

# 上传本地图片获取网上图片URL
def get_image_url(imagePath, pictureName):
    if str(imagePath).split('.')[-1] == 'jpg' or str(imagePath).split('.')[-1] == 'JPG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    elif str(imagePath).split('.')[-1] == 'png' or str(imagePath).split('.')[-1] == 'PNG':
        filename = pictureName + '.' + str(imagePath).split('.')[-1]  # 七牛云网盘文件名
    else:
        print("请检查图片格式！！！")
    # 七牛云密钥管理：https://portal.qiniu.com/user/key
    # 【账号：13267854059  密码：z****】
    access_key = "fjlWDEbF1fqBU98UsdDJRcSSKODT9Gq7tA3gu8eY"
    secret_key = "thiWFpO881GfhlaAz1Wkk2yEcvV3ue2OHnY_5D9V"
    keyq = Auth(access_key, secret_key)
    bucket = "zues3737img"  # 七牛云盘名
    # 删除
    butm = BucketManager(keyq)
    reformDel, informDel = butm.delete(bucket, filename)  # 删除旧图片
    # 上传
    time.sleep(1)
    token = keyq.upload_token(bucket, filename)  # 上传新图片
    reformUp, informUp = put_file(token, filename, imagePath)
    if reformUp is not None:
        print('已成功上传 {}'.format(filename))
        time.sleep(1)
        baseURL = "https://cjh3737.zeus.cn/"  # 加速域名
        # subURL = baseURL + '/' + filename + '?imageView2/0/quality/100!/sharpen/1/interlace/1'
        subURL = baseURL + '/' + filename + '?imageMogr2/format/jpg/quality/100!/sharpen/50/interlace/1/ignore-error/1'
        pictureURL = keyq.private_download_url(subURL)  # 链接图片URL
        time.sleep(1)
        return pictureURL
    else:
        print(filename + '上传失败！！！')


if __name__ == '__main__':

    AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序AppKey
    AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序AppSecret
    
    webhook1 = 'https://oapi.dingtalk.com/robot/send?access_token=a7353d886cc636016a82202e37eaf5e147e82c202e68232ee999a775ada2a35a'  # 销管：网通服务部经理对接群
    webhook2 = 'https://oapi.dingtalk.com/robot/send?access_token=94833bcd65512716c331961a263b76bbcb04caf9052fab7bbc3130685197d80c'  # 网通：推广大区经理群
    webhook3 = 'https://oapi.dingtalk.com/robot/send?access_token=a24cea6459b80e86bce0c4a8a8e748e474e41a92b8f64a420fcd0326730a6171'  # 网通：流通大区管理群
    webhook4 = 'https://oapi.dingtalk.com/robot/send?access_token=df1e274c1104d977ffb4c6ab3b033abfd54ca9ed5da16677d078be14fe428aa5'  # 网通：百强大区经理群
    webhook5 = 'https://oapi.dingtalk.com/robot/send?access_token=cd7751ba01a4774cf729fc33b85a015f1c4c475fc8e95f5190958039ad0f30eb'  # 网通：网通推广省区高效协同工作群
    
    # webhook1 = 'https://oapi.dingtalk.com/robot/send?access_token=94ddc4b6fcaa2d400d03c0a6c5667808216215e45349b09f198338f8521041e9'  # 1机器人测试群
    # webhook2 = 'https://oapi.dingtalk.com/robot/send?access_token=2105d464fe9af82bee335d768a40438fa4a0b3513da612a65bc02c6443bfaaf4'  # 2机器人测试群
    # webhook3 = 'https://oapi.dingtalk.com/robot/send?access_token=6f84a094c5a77820bd12cf80d86633da4032abb39d063106db1ee2b88d10a8d5'  # 3机器人测试群
    # webhook4 = 'https://oapi.dingtalk.com/robot/send?access_token=57d986d58975ab00b8b95ed1ef3bded92639b4b18b213c85afcf253be677810b'  # 4机器人测试群
    # webhook5 = 'https://oapi.dingtalk.com/robot/send?access_token=7a9677def3c33e2c5e9cf3fc7500dadf2769e366f1be3ea6391066607f256b5d'  # 5机器人测试群
    
    secret0 = 'TU-9kMk1BpOtX5ypywP8v0gt7SK3qHatJXsVnSXC-tROFLVqKMX55qRML-PXSX6i'  # 群机器人加签秘钥secret(默认网通数据小助手)
    
    RobotWebHookURL = [webhook1, webhook1, webhook2, webhook2, webhook3, webhook4, webhook5]
    RobotSecret = [secret0, secret0, secret0, secret0, secret0, secret0, secret0]
    # RobotWebHookURL = [webhook1, webhook1, webhook2, webhook2, webhook3, webhook4]
    # RobotSecret = [secret0, secret0, secret0, secret0, secret0, secret0]
    

    fileFullPath = 'C:/Users/Zeus/Desktop/autoSend/4_群姐/群姐_正式发送文件.xlsx' 
    savePictuePath = 'C:/Users/Zeus/Desktop/autoSend/4_群姐/Pictures/'

    workbook = opxl.load_workbook(fileFullPath)
    worksheetnames = workbook.sheetnames
    
    sendtitle = ["###### **① 网通事业部 - 发货汇总表**",
                  "###### **② 推广线 - 克咳片发货汇总表**",
                  "###### **③ 推广线 - 克咳片发货汇总表**",
                  "###### **④ 推广线 - 发货汇总表**",
                  "###### **⑤ 流通线 - 发货汇总表**",
                  "###### **⑥ 百强连锁 - 发货汇总表**",
                  "###### **⑦ 秋冬大促 - 克咳片汇总表**"
                  ]
    
    # sendtitle = ["###### **① 网通事业部 - 发货汇总表**",
    #              "###### **② 推广线 - 克咳片发货汇总表**",
    #              "###### **③ 推广线 - 克咳片发货汇总表**",
    #              "###### **④ 推广线 - 发货汇总表**",
    #              "###### **⑤ 流通线 - 发货汇总表**",
    #              "###### **⑥ 百强连锁 - 发货汇总表**"
    #              ]
    
    export_file()
    print('文件保存成功！！！')
    
    sendTypes = int(input('>>> 0sendAll-1sendSingle:'))
    
    if sendTypes == 0:  # 发送形式 - 全部
        
        deleteOldFiles('C:/Users/Zeus/Desktop/autoSend/4_群姐/Pictures/')  # 清空文件夹历史文件

        imgURL = []
        for sheetname, picturename in zip(worksheetnames, worksheetnames):
            try:
                excelCatchScreen(fileFullPath, sheetname, picturename, savePictuePath)
            except BaseException:
                print(picturename + '截图出错！！！')
            try:
                getURL = get_image_url(savePictuePath + picturename + '.PNG', picturename)
                imgURL.append(getURL)
            except BaseException:
                print(picturename + '图片URL出错！！！')
                
        pictureURL = [imgURL[0], imgURL[1], imgURL[1], imgURL[2], imgURL[3], imgURL[4], imgURL[5]]
        # pictureURL = [imgURL[0], imgURL[1], imgURL[1], imgURL[2], imgURL[3], imgURL[4]]
        
        for inum, iurl, itext in zip(range(7), pictureURL, sendtitle):
            ddMessage = {  # 发布消息内容
                "msgtype": "markdown",
                "markdown": {"title": "网通日报",  # @某人 才会显示标题
                                "text": itext + 
                                "\n![Image被拦截, 请使用非公司网络查看](" + iurl + ")"
                                "\n###### ----------------------------------------"
                                "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                "at": {
                    # "atMobiles": [15817552982],  # 指定@某人
                    "isAtAll": False  # 是否@所有人[False:否, True:是]
                }
            }

            # 发送消息
            dingdingFunction(RobotWebHookURL[inum], RobotSecret[inum], AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
            
        
        # 消息数据
        fahuo_today = sum(dataZSD043[(dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay)]['含税总价'])
        huikuan_today = sum(dataZFI26[(dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay)]['核销金额'])
        
        fahuo_today_LT = sum(dataZSD043[(dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay) &  ((dataZSD043['三级组织'] == 'LT-南区') | (dataZSD043['三级组织'] == 'LT-中区') | (dataZSD043['三级组织'] == 'LT-北区'))]['含税总价'])
        fahuo_today_TG = sum(dataZSD043[(dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay) & ((dataZSD043['三级组织'] == 'TG-推广一区') | (dataZSD043['三级组织'] == 'TG-推广二区') | (dataZSD043['三级组织'] == 'TG-推广三区'))]['含税总价'])
        fahuo_today_BQ = sum(dataZSD043[(dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay) & ((dataZSD043['三级组织'] == 'BQ-珠江大区') | (dataZSD043['三级组织'] == 'BQ-长江大区') | (dataZSD043['三级组织'] == 'BQ-黄河大区') | (dataZSD043['三级组织'] == 'KA-山东'))]['含税总价'])
        fahuo_today_SD = sum(dataZSD043[(dataZSD043['月份'] == globalMonth) & (dataZSD043['天'] == globalDay) & (dataZSD043['三级组织'] == 'TL-试点区')]['含税总价'])
        
        huikuan_today_LT = sum(dataZFI26[(dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay) & ((dataZFI26['三级组织'] == 'LT-南区') | (dataZFI26['三级组织'] == 'LT-中区') | (dataZFI26['三级组织'] == 'LT-北区'))]['核销金额'])
        huikuan_today_TG = sum(dataZFI26[(dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay) & ((dataZFI26['三级组织'] == 'TG-推广一区') | (dataZFI26['三级组织'] == 'TG-推广二区') | (dataZFI26['三级组织'] == 'TG-推广三区'))]['核销金额'])
        huikuan_today_BQ = sum(dataZFI26[(dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay) & ((dataZFI26['三级组织'] == 'BQ-珠江大区') | (dataZFI26['三级组织'] == 'BQ-长江大区') | (dataZFI26['三级组织'] == 'BQ-黄河大区') | (dataZFI26['三级组织'] == 'KA-山东'))]['核销金额'])
        huikuan_today_SD = sum(dataZFI26[(dataZFI26['月份'] == globalMonth) & (dataZFI26['天'] == globalDay) & (dataZFI26['三级组织'] == 'TL-试点区')]['核销金额'])
        
        fahuo_Sum_LT = sum(dataZSD043[(dataZSD043['月份'] <= globalMonth) &  ((dataZSD043['三级组织'] == 'LT-南区') | (dataZSD043['三级组织'] == 'LT-中区') | (dataZSD043['三级组织'] == 'LT-北区'))]['含税总价'])
        fahuo_Sum_TG = sum(dataZSD043[(dataZSD043['月份'] <= globalMonth) & ((dataZSD043['三级组织'] == 'TG-推广一区') | (dataZSD043['三级组织'] == 'TG-推广二区') | (dataZSD043['三级组织'] == 'TG-推广三区'))]['含税总价'])
        fahuo_Sum_BQ = sum(dataZSD043[(dataZSD043['月份'] <= globalMonth) & ((dataZSD043['三级组织'] == 'BQ-珠江大区') | (dataZSD043['三级组织'] == 'BQ-长江大区') | (dataZSD043['三级组织'] == 'BQ-黄河大区') | (dataZSD043['三级组织'] == 'KA-山东'))]['含税总价'])
        
        huikuan_Sum_LT = sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'LT-南区') | (dataZFI26['三级组织'] == 'LT-中区') | (dataZFI26['三级组织'] == 'LT-北区'))]['核销金额'])
        huikuan_Sum_TG = sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'TG-推广一区') | (dataZFI26['三级组织'] == 'TG-推广二区') | (dataZFI26['三级组织'] == 'TG-推广三区'))]['核销金额'])
        huikuan_Sum_BQ = sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'BQ-珠江大区') | (dataZFI26['三级组织'] == 'BQ-长江大区') | (dataZFI26['三级组织'] == 'BQ-黄河大区') | (dataZFI26['三级组织'] == 'KA-山东'))]['核销金额'])
        
        huikuan_rate = Decimal(sum(dataZFI26[dataZFI26['月份'] <= globalMonth]['核销金额'])) / Decimal(sum(hexiao_target[hexiao_target['月份'] <= globalMonth]['目标'])) * 100
        huikuan_rate_LT = Decimal(sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'LT-南区') | (dataZFI26['三级组织'] == 'LT-中区') | (dataZFI26['三级组织'] == 'LT-北区'))]['核销金额'])) / Decimal(sum(hexiao_target[(hexiao_target['月份'] <= globalMonth) & ((hexiao_target['三级组织'] == 'LT-南区') | (hexiao_target['三级组织'] == 'LT-中区') | (hexiao_target['三级组织'] == 'LT-北区'))]['目标'])) * 100
        huikuan_rate_TG = Decimal(sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'TG-推广一区') | (dataZFI26['三级组织'] == 'TG-推广二区') | (dataZFI26['三级组织'] == 'TG-推广三区'))]['核销金额'])) / Decimal(sum(hexiao_target[(hexiao_target['月份'] <= globalMonth) & ((hexiao_target['三级组织'] == 'TG-推广一区') | (hexiao_target['三级组织'] == 'TG-推广二区') | (hexiao_target['三级组织'] == 'TG-推广三区'))]['目标'])) * 100
        huikuan_rate_BQ = Decimal(sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'BQ-珠江大区') | (dataZFI26['三级组织'] == 'BQ-长江大区') | (dataZFI26['三级组织'] == 'BQ-黄河大区') | (dataZFI26['三级组织'] == 'KA-山东'))]['核销金额'])) / Decimal(sum(hexiao_target[(hexiao_target['月份'] <= globalMonth) & ((hexiao_target['三级组织'] == 'BQ-珠江大区') | (hexiao_target['三级组织'] == 'BQ-长江大区') | (hexiao_target['三级组织'] == 'BQ-黄河大区') | (hexiao_target['三级组织'] == 'KA-山东'))]['目标'])) * 100
        
        huikuan_fullyear_rate = Decimal(sum(dataZFI26[dataZFI26['月份'] <= globalMonth]['核销金额'])) / Decimal(sum(hexiao_target['目标'])) * 100
        huikuan_fullyear_rate_LT = Decimal(sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'LT-南区') | (dataZFI26['三级组织'] == 'LT-中区') | (dataZFI26['三级组织'] == 'LT-北区'))]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'LT-南区') | (hexiao_target['三级组织'] == 'LT-中区') | (hexiao_target['三级组织'] == 'LT-北区'))]['目标'])) * 100
        huikuan_fullyear_rate_TG = Decimal(sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'TG-推广一区') | (dataZFI26['三级组织'] == 'TG-推广二区') | (dataZFI26['三级组织'] == 'TG-推广三区'))]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'TG-推广一区') | (hexiao_target['三级组织'] == 'TG-推广二区') | (hexiao_target['三级组织'] == 'TG-推广三区'))]['目标'])) * 100
        huikuan_fullyear_rate_BQ = Decimal(sum(dataZFI26[(dataZFI26['月份'] <= globalMonth) & ((dataZFI26['三级组织'] == 'BQ-珠江大区') | (dataZFI26['三级组织'] == 'BQ-长江大区') | (dataZFI26['三级组织'] == 'BQ-黄河大区') | (dataZFI26['三级组织'] == 'KA-山东'))]['核销金额'])) / Decimal(sum(hexiao_target[((hexiao_target['三级组织'] == 'BQ-珠江大区') | (hexiao_target['三级组织'] == 'BQ-长江大区') | (hexiao_target['三级组织'] == 'BQ-黄河大区') | (hexiao_target['三级组织'] == 'KA-山东'))]['目标'])) * 100
        
        
        summaryMessageALL = {   # 汇总
            "msgtype": "markdown",
            "markdown": {"title": "网通汇总分析",  # @某人 才会显示标题
                            "text": "##### **各位领导：**"
                                    "\n ##### **网通事业部**"
                                    "\n ##### **Ⅰ " + str(globalYear) + "年" + str(globalMonth) + "月" + str(globalDay) + "日**"
                                    "\n ##### 今日发货**" + str("{:.2f}".format(fahuo_today)) + "**万元; 今日回款**" + str("{:.2f}".format(huikuan_today)) + "**万元"
                                    "\n > ##### **① 流通线**"
                                    "\n > ##### 发货**" + str("{:.2f}".format(fahuo_today_LT)) + "**万元; 回款**" + str("{:.2f}".format(huikuan_today_LT)) + "**万元"
                                    "\n > ##### **② 推广线**"
                                    "\n > ##### 发货**" + str("{:.2f}".format(fahuo_today_TG)) + "**万元; 回款**" + str("{:.2f}".format(huikuan_today_TG)) + "**万元"
                                    "\n > ##### **③ 百强线**"
                                    "\n > ##### 发货**" + str("{:.2f}".format(fahuo_today_BQ)) + "**万元; 回款**" + str("{:.2f}".format(huikuan_today_BQ)) + "**万元"
                                    "\n > ##### **④ 试点线**"
                                    "\n > ##### 发货**" + str("{:.2f}".format(fahuo_today_SD)) + "**万元; 回款**" + str("{:.2f}".format(huikuan_today_SD)) + "**万元"
                                    "\n ##### **Ⅱ 前" + str(globalMonth) + "月完成进度**"
                                    "\n ##### 事业部: 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate)) + "**"
                                    "\n > ##### ① 流通线: 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate_LT)) + "**"
                                    "\n > ##### ② 推广线: 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate_TG)) + "**"
                                    "\n > ##### ③ 百强线: 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate_BQ)) + "**"
                                    "\n ##### **Ⅲ " + str(globalYear) + "年总况**"
                                    "\n ##### 事业部: 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate)) + "**"
                                    "\n > ##### ① 流通线: 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate_LT)) + "**"
                                    "\n > ##### ② 推广线: 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate_TG)) + "**"
                                    "\n > ##### ③ 百强线: 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate_BQ)) + "**"
                                    }, 
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        
        dingdingFunction(RobotWebHookURL[1], RobotSecret[1], AppKey, AppSecret).sendMessage(summaryMessageALL)  # 发消息
        
        summaryMessageTG = {  # 推广
            "msgtype": "markdown",
            "markdown": {"title": "推广线汇总分析",  # @某人 才会显示标题
                            "text": "##### **各位领导:**"
                                    "\n ##### **推广线** 截止" + str(globalYear) + "年" + str(globalMonth) + "月" + str(globalDay) + "日:"
                                    "\n > ##### ① 当日发货 **" + str("{:.2f}".format(fahuo_today_TG)) + "** 万元; 回款 **" + str("{:.2f}".format(huikuan_today_TG)) + "** 万元"
                                    "\n > ##### ② 1-" + str(globalMonth) + "月累计发货 **" + str("{:.2f}".format(fahuo_Sum_TG)) + "** 万元; 回款 **" + str("{:.2f}".format(huikuan_Sum_TG)) + "** 万元"
                                    "\n > ##### ③ 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate_TG)) + "**"
                                    "\n > ##### ④ 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate_TG)) + "**"
                                    }, 
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        
        dingdingFunction(RobotWebHookURL[3], RobotSecret[3], AppKey, AppSecret).sendMessage(summaryMessageTG)  # 发消息
        
        summaryMessageLT = {  # 流通
            "msgtype": "markdown",
            "markdown": {"title": "流通线汇总分析",  # @某人 才会显示标题
                            "text": "##### **各位领导:**"
                                    "\n ##### **流通线** 截止" + str(globalYear) + "年" + str(globalMonth) + "月" + str(globalDay) + "日:"
                                    "\n > ##### ① 当日发货 **" + str("{:.2f}".format(fahuo_today_LT)) + "** 万元; 回款 **" + str("{:.2f}".format(huikuan_today_LT)) + "** 万元"
                                    "\n > ##### ② 1-" + str(globalMonth) + "月累计发货 **" + str("{:.2f}".format(fahuo_Sum_LT)) + "** 万元; 回款 **" + str("{:.2f}".format(huikuan_Sum_LT)) + "** 万元"
                                    "\n > ##### ③ 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate_LT)) + "**"
                                    "\n > ##### ④ 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate_LT)) + "**"
                                    }, 
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        
        dingdingFunction(RobotWebHookURL[4], RobotSecret[4], AppKey, AppSecret).sendMessage(summaryMessageLT)  # 发消息
        
        summaryMessageBQ = {  # 百强连锁
            "msgtype": "markdown",
            "markdown": {"title": "百强连锁线汇总分析",  # @某人 才会显示标题
                            "text": "##### **各位领导:**"
                                    "\n ##### **百强连锁线** 截止" + str(globalYear) + "年" + str(globalMonth) + "月" + str(globalDay) + "日:"
                                    "\n > ##### ① 当日发货 **" + str("{:.2f}".format(fahuo_today_BQ)) + "** 万元; 回款 **" + str("{:.2f}".format(huikuan_today_BQ)) + "** 万元"
                                    "\n > ##### ② 1-" + str(globalMonth) + "月累计发货 **" + str("{:.2f}".format(fahuo_Sum_BQ)) + "**万元; 回款 **" + str("{:.2f}".format(huikuan_Sum_BQ)) + "** 万元"
                                    "\n > ##### ③ 1-" + str(globalMonth) + "月回款完成率 **" + str("{:.2f}%".format(huikuan_rate_BQ)) + "**"
                                    "\n > ##### ④ 全年回款完成率 **" + str("{:.2f}%".format(huikuan_fullyear_rate_BQ)) + "**"
                                    }, 
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        
        dingdingFunction(RobotWebHookURL[5], RobotSecret[5], AppKey, AppSecret).sendMessage(summaryMessageBQ)  # 发消息
        
        summaryMessageQin = {  # 秋冬大促
            "msgtype": "markdown",
            "markdown": {"title": "秋冬大促汇总分析",  # @某人 才会显示标题
                            "text": "##### **各位经理:**"
                                    "\n ##### 晚上好, 以上为2021年度克咳片秋冬大促截止至今天的销售完成情况, 请大家查阅！"
                                    }, 
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        
        dingdingFunction(RobotWebHookURL[6], RobotSecret[6], AppKey, AppSecret).sendMessage(summaryMessageQin)  # 发消息
        
        

            
    elif sendTypes == 1: # 发送形式 - 选择性单张 从 1 开始

        deleteOldFiles('C:/Users/Zeus/Desktop/autoSend/4_群姐/Pictures/')  # 清空文件夹历史文件
        
        send_NO_Picture = int(input('>>>发送第几张图片？'))
        
        print('***单独发送: ' + str(worksheetnames[send_NO_Picture - 1]) + '.PNG')

        excelCatchScreen(fileFullPath, worksheetnames[send_NO_Picture - 1], worksheetnames[send_NO_Picture - 1], savePictuePath)
        
        ddMessage = {  # 发布消息内容
                "msgtype": "markdown",
                "markdown": {"title": "网通日报",  # @某人 才会显示标题
                                "text": sendtitle[send_NO_Picture - 1] +
                                "\n![Image被拦截, 请使用非公司网络查看](" + get_image_url(savePictuePath + worksheetnames[send_NO_Picture - 1] + '.PNG', worksheetnames[send_NO_Picture - 1]) + ")"
                                "\n###### ----------------------------------------"
                                "\n###### 发布时间：" + str(datetime.now()).split('.')[0]},  # 发布时间
                "at": {
                    # "atMobiles": [15817552982],  # 指定@某人
                    "isAtAll": False  # 是否@所有人[False:否, True:是]
                }
            }

        # 发送消息
        dingdingFunction(RobotWebHookURL[send_NO_Picture - 1], RobotSecret[send_NO_Picture - 1], AppKey, AppSecret).sendMessage(ddMessage)  # 发图片消息
        
    else:
        print('请输入0-1正确的发送方式！！！')
    
    # dingdingFunction(RobotWebHookURL, RobotSecret, AppKey, AppSecret).sendFile(chatId, fileFullPath)  # 发送文件
