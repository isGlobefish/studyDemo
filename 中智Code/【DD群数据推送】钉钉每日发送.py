# -*- coding: utf-8 -*-
'''
@createTool    : PyCharm-2020.2.2
@projectName   : pythonProject 
@originalAuthor: Made in win10.Sys design by deHao.Zou
@createTime    : 2020/11/18 9:56
'''
import xlwt
import time
import hmac
import json
import base64
import pymysql
import hashlib
import calendar
import requests
import itertools
import numpy as np
import urllib.parse
import pandas as pd
import urllib.request
import openpyxl as opxl
from datetime import datetime
from time import strftime, gmtime
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, PatternFill



# 0:发送本月数据 1:发送上一个月数据
inputTpye = input(">>>0thisMonth-1lastMonth:")


if inputTpye == '1':  # 上个月数据
    Year = str(time.strftime("%Y", time.localtime())).zfill(4)  # 本年
    Month = str(int(time.strftime("%m", time.localtime())) - 1).zfill(2)  # 前一月
    # if today == 1:
    today = str(calendar.monthrange(int(Year), int(Month))[1]).zfill(2)  # 前一月最后一日
    # daySub2 = str(int(calendar.monthrange(int(Year), int(Month))[1]) - 1).zfill(2)  # 前一月倒数第二天
    # elif today == 2:
    # daySub1 = str(int(time.strftime("%d", time.localtime())) - 1).zfill(2)  # 前一日
    # daySub2 = str(calendar.monthrange(int(Year), int(Month))[1]).zfill(2)  # 前一月最后一日
else:  # 本月数据
    Year = str(time.strftime("%Y", time.localtime())).zfill(4)  # 本年
    Month = str(int(time.strftime("%m", time.localtime()))).zfill(2)  # 本月
    today = int(time.strftime("%d", time.localtime()))  # 本日
    # daySub1 = str(int(time.strftime("%d", time.localtime())) - 1).zfill(2)  # 前一日
    # daySub2 = str(int(time.strftime("%d", time.localtime())) - 2).zfill(2)  # 前两日


# 制作表格
def Create_TableImage(Dtable):
    # 设置单元格样式
    def set_style(fontName, height, bold=False, Halign=False, Valign=False, setBorder=False, setbgcolor=False):

        style = xlwt.XFStyle()  # 设置类型

        font = xlwt.Font()  # 为样式创建字体
        font.name = fontName
        font.height = height  # 字体大小，220就是11号字体，大概就是11*20得来
        font.bold = bold
        font.color = 'black'
        font.color_index = 4
        style.font = font

        alignment = xlwt.Alignment()  # 设置字体在单元格的位置
        if Halign == 0:
            alignment.horz = xlwt.Alignment.HORZ_CENTER  # 水平居中
        elif Halign == 1:
            alignment.horz = xlwt.Alignment.HORZ_LEFT  # 水平偏左
        else:
            alignment.horz = xlwt.Alignment.HORZ_RIGHT  # 水平偏右
        if Valign == 0:
            alignment.vert = xlwt.Alignment.VERT_CENTER  # 竖直居中
        elif Valign == 1:
            alignment.vert = xlwt.Alignment.VERT_TOP  # 竖直置顶
        else:
            alignment.vert = xlwt.Alignment.VERT_BOTTOM  # 竖直底部
        # alignment.horz = xlwt.Alignment.HORZ_CENTER  # 水平居中
        # alignment.horz = xlwt.Alignment.HORZ_LEFT  # 水平偏左
        # alignment.horz = xlwt.Alignment.HORZ_RIGHT  # 水平偏右
        # alignment.vert = xlwt.Alignment.VERT_CENTER  # 竖直居中
        # alignment.vert = xlwt.Alignment.VERT_TOP  # 竖直置顶
        # alignment.vert = xlwt.Alignment.VERT_BOTTOM  # 竖直底部
        style.alignment = alignment

        border = xlwt.Borders()  # 给单元格加框线
        if setBorder == 0:
            border.left = xlwt.Borders.THIN  # 左
            border.top = xlwt.Borders.THIN  # 上
            border.right = xlwt.Borders.THIN  # 右
            border.bottom = xlwt.Borders.THIN  # 下
            border.left_colour = 0x40  # 设置框线颜色，0x40是黑色，颜色真的巨多
            border.right_colour = 0x40
            border.top_colour = 0x40
            border.bottom_colour = 0x40
        else:
            pass
        style.borders = border

        pattern = xlwt.Pattern()  # 设置背景颜色
        if setbgcolor == 0:
            pattern.pattern = xlwt.Pattern.SOLID_PATTERN
            pattern.pattern_fore_colour = xlwt.Style.colour_map['turquoise']
        else:
            pass
        style.pattern = pattern
        return style

    try:
        # 创建表格模板并写入数据
        f = xlwt.Workbook()  # 创建工作簿
        sheet1 = f.add_sheet('Sheet1', cell_overwrite_ok=True)

        # 设置列宽
        # for i in range(17):
        #     if i == 2 or i == 5 or i == 14:
        #         sheet1.col(i).width = 256 * 50
        #     elif i == 0:
        #         sheet1.col(i).width = 256 * 30
        #     else:
        #         sheet1.col(i).width = 256 * 13
        colWidth = [11, 9, 9, 9, 16, 16, 10, 10, 10, 40, 10, 15, 10, 13, 8, 10, 14]*2
        for icol, iwidth in enumerate(colWidth):
            sheet1.col(icol).width = 256 * iwidth

        # 设置行高
        for rowi in range(0, 3):
            rowHeight = xlwt.easyxf('font:height 220;')  # 36pt,类型小初的字号
            rowNum = sheet1.row(rowi)
            rowNum.set_style(rowHeight)

        """ 设计表头 """
        colNames = ['销售日期', '客户体系', '大区', '省区', '调整细分区域', '原始自带营运区', '客户编码', '客户名称', '门店编码', '门店名称', '商品编码', '商品名称', '商品简称', '商品规格', '数量', '标准单价', '标准零售金额']
        # 第1行
        for icol, iname in enumerate(colNames):
            sheet1.write_merge(0, 0, icol, icol, iname,
                               set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # 第2行
        sheet1.write_merge(1, 1, 0, 13, "大参林",
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(1, 1, 14, 14,
                           str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "大参林"].reset_index(drop=True).loc[:, '数量'])])),
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(1, 3, 15, 15, '',
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(1, 1, 16, 16,
        #                    str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "大参林"].reset_index(drop=True).loc[:, '零售金额'])])),
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(1, 1, 10, 10, '',
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(1, 1, 16, 16,
                           str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "大参林"].reset_index(drop=True).loc[:, '标准零售金额'])])),
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # 第3行
        sheet1.write_merge(2, 2, 0, 13, "益丰",
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(2, 2, 14, 14,
                           str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "益丰"].reset_index(drop=True).loc[:, '数量'])])),
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(2, 2, 15, 15, '',
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(2, 2, 16, 16,
        #                    str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "益丰"].reset_index(drop=True).loc[:, '零售金额'])])),
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(2, 2, 10, 10, '',
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(2, 2, 16, 16,
                           str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "益丰"].reset_index(drop=True).loc[:, '标准零售金额'])])),
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # 第4行
        sheet1.write_merge(3, 3, 0, 13, "高济",
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(3, 3, 14, 14,
                           str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "高济"].reset_index(drop=True).loc[:, '数量'])])),
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(3, 3, 15, 15, '',
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(3, 3, 15, 15,
        #                    str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "高济"].reset_index(drop=True).loc[:, '零售金额'])])),
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # sheet1.write_merge(3, 3, 10, 10, '',
        #                    set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        sheet1.write_merge(3, 3, 16, 16,
                           str(sum([float(i) for i in list(Dtable[Dtable["客户体系"] == "高济"].reset_index(drop=True).loc[:, '标准零售金额'])])),
                           set_style('等线', 210, True, Halign=0, Valign=0, setBorder=0, setbgcolor=0))
        # 填充表格数据
        for rowi in range(Dtable.shape[0]):
            for colj, content in enumerate(colNames):
                if colj == 0:
                    sheet1.write_merge(rowi + 4, rowi + 4, colj, colj, str(Dtable.loc[rowi, content]))
                else:
                    sheet1.write_merge(rowi + 4, rowi + 4, colj, colj, Dtable.loc[rowi, content])
        # for irow in range(Dtable.shape[0]):
        #     for jcol, content in enumerate(colNames):
        #         if jcol == 0:
        #             sheet1.write_merge(irow + 4, rowi + 4, jcol, jcol, str(Dtable.loc[irow, content]))
        #         else:
        #             sheet1.write_merge(irow + 4, rowi + 4, jcol, jcol, Dtable.loc[irow, content])
    except Exception as e:
        print("制作发送表格出错！！！", e)
    f.save('D:/FilesCenter/大客户数据/Everyday-江西/' + Year + '-' + Month + '-' + str(today).zfill(2) + '江西.xls')  # 保存文件


# 设置表格格式
# def setStyle_saveFile(data, colname, filepath):
#     workbook = opxl.Workbook()
#     worksheet = workbook.active
#     worksheet.title = 'Sheet1'
#     colWidth = [11, 8, 9, 9, 14, 14, 9, 9, 11, 21, 13, 15, 10, 9, 6, 8, 11]
#     for iwidth, conwidth in enumerate(colWidth, start=1):
#         worksheet.column_dimensions[opxl.utils.get_column_letter(iwidth)].width = conwidth
#     for icol in range(data.shape[1]):
#         for irow in range(data.shape[0]):
#             if irow == 0:
#                 for ic, icon in enumerate(colname, start=1):
#                     worksheet.cell(row=1, column=ic).value = icon
#                     workcell2 = worksheet.cell(row=1, column=ic)
#                     workcell2.font = Font(name='微软雅黑', size=10, bold=True)
#             else:
#                 worksheet.cell(row=irow+2, column=icol+1).value = data.iloc[irow, icol]
#                 workcell1 = worksheet.cell(row=irow + 1, column=icol + 1)
#                 workcell1.font = Font(name='微软雅黑', size=10, bold=False)
#     workbook.save(filepath)

# def setStyle_saveFile(data, filepath):
#     workbook = opxl.load_workbook('D:/FilesCenter/大客户数据/send_style.xlsx', keep_links=True)
#     worksheet = workbook['Sheet1']
#     colWidth = [11, 8, 9, 9, 14, 14, 9, 9, 11, 21, 13, 15, 10, 9, 6, 8, 12]
#     for iwidth, conwidth in enumerate(colWidth, start=1):
#         worksheet.column_dimensions[opxl.utils.get_column_letter(iwidth)].width = conwidth
#     for irow in dataframe_to_rows(data, index=False, header=True):
#         worksheet.append(irow)
#     workbook.save(filepath)

# def setStyle_saveFile(data, filepath):
#     workbook = opxl.load_workbook('D:/FilesCenter/大客户数据/send_style.xlsx')
#     worksheet = workbook['Sheet1']
#     colWidth = [11, 8, 9, 9, 14, 14, 9, 9, 11, 21, 13, 15, 10, 9, 6, 8, 12]
#     for iwidth, conwidth in enumerate(colWidth, start=1):
#         worksheet.column_dimensions[opxl.utils.get_column_letter(iwidth)].width = conwidth
#     for irow in dataframe_to_rows(data, index=False, header=True):
#         worksheet.append(irow)
#     workbook.save(filepath)


# 定义钉钉功能
class dingdingFunction(object):
    def __init__(self, roboturl, robotsecret, appkey, appsecret):
        """
        :param roboturl: 群机器人WebHook_url
        :param robotsecret: 安全设置的加签秘钥
        :param appkey: 企业开发平台小程序AppKey
        :param appsecret: 企业开发平台小程序AppSecret
        """
        self.roboturl = roboturl
        self.robotsecret = robotsecret
        self.appkey = appkey
        self.appsecret = appsecret
        timestamp = round(time.time() * 1000)  # 时间戳
        secret_enc = robotsecret.encode('utf-8')
        string_to_sign = '{}\n{}'.format(timestamp, robotsecret)
        string_to_sign_enc = string_to_sign.encode('utf-8')
        hmac_code = hmac.new(secret_enc, string_to_sign_enc, digestmod=hashlib.sha256).digest()
        sign = urllib.parse.quote_plus(base64.b64encode(hmac_code))  # 最终签名
        self.webhook_url = self.roboturl + '&timestamp={}&sign={}'.format(timestamp, sign)  # 最终url,url+时间戳+签名

    # 发送文件
    def getAccess_token(self):
        url = 'https://oapi.dingtalk.com/gettoken?appkey=%s&appsecret=%s' % (AppKey, AppSecret)
        headers = {
            'Content-Type': "application/x-www-form-urlencoded"
        }
        data = {'appkey': self.appkey,
                'appsecret': self.appsecret}
        r = requests.request('GET', url, data=data, headers=headers)
        access_token = r.json()["access_token"]
        return access_token

    def getMedia_id(self, filespath):
        access_token = self.getAccess_token()  # 拿到接口凭证
        url = 'https://oapi.dingtalk.com/media/upload?access_token=' + access_token + '&type=file'
        files = {'media': open(filespath, 'rb')}
        data = {'access_token': access_token,
                'type': 'file'}
        response = requests.post(url, files=files, data=data)
        json = response.json()
        return json["media_id"]

    def sendFile(self, chatid, filespath):
        access_token = self.getAccess_token()
        media_id = self.getMedia_id(filespath)
        url = 'https://oapi.dingtalk.com/chat/send?access_token=' + access_token
        header = {
            'Content-Type': 'application/json'
        }
        data = {'access_token': access_token,
                'chatid': chatid,
                'msg': {
                    'msgtype': 'file',
                    'file': {'media_id': media_id}
                }}
        r = requests.request('POST', url, data=json.dumps(data), headers=header)
        print(r.json()["errmsg"])

    # 发送消息
    def sendMessage(self, content):
        """
        :param content: 发送内容
        """
        header = {
            "Content-Type": "application/json",
            "Charset": "UTF-8"
        }
        sendContent = json.dumps(content)  # 将字典类型数据转化为json格式
        sendContent = sendContent.encode("utf-8")  # 编码为UTF-8格式
        request = urllib.request.Request(url=self.webhook_url, data=sendContent, headers=header)  # 发送请求
        opener = urllib.request.urlopen(request)  # 将请求发回的数据构建成为文件格式
        print(opener.read().decode())  # 打印返回的结果


#     加一个发送文件（包括图片、文本、表格、压缩文件等等）
#     获取全部手机号（匹配人名与手机号）


if __name__ == '__main__':

    startTime = datetime.now()
    print(" > 数据获取中,请稍等片刻")

    conn = pymysql.connect(host    = 'localhost',  # 数据库地址
                           port    = 3306,  # 数据库端口
                           user    = 'root',  # 用户名
                           passwd  = '123456',  # 数据库密码
                           db      = 'dkh',  # 数据库名
                           charset = 'utf8')  # 字符串类型


    cursorSZ = conn.cursor()  # 海王深圳
    cursorDL = conn.cursor()  # 高济大连
    cursorJX = conn.cursor()  # 江西
    cursorHBXX = conn.cursor()  # 河北新兴
    cursorCD = conn.cursor()  # 成都
    cursorQY = conn.cursor()  # 全亿
    cursorGJ = conn.cursor()  # 高济
    cursorLBX = conn.cursor()  # 老百姓
    cursorHRT = conn.cursor()  # 惠仁堂
    cursorPZYFJH = conn.cursor()  # 普泽益丰江海
    cursorHB = conn.cursor()  # 湖北
    cursorQD = conn.cursor()  # 青岛
    cursorHW = conn.cursor()  # 海王
    cursorYF = conn.cursor()  # 益丰
    cursorDSL = conn.cursor()  # 大参林
    cursorGZZSQ = conn.cursor()  # 广中珠韶清
    cursorAH = conn.cursor()  # 安徽
    cursorYX = conn.cursor()  # 粤西


    # ------------------------------------------------------------------------------------------------
    # SQL查询语句区
    # ------------------------------------------------------------------------------------------------
    searchCols = 'date, customer, sfa_client_desc, state, desc_1, client_id, client_desc, client_alias, sfa_id, sfa_desc, DKH_materiel_id, materiel_desc, materiel_alias, norms, amount, bz_UnitPrice, bz_sales_Money'


    executeCodeSZ = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='海王' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """' AND state='深圳';"""


    executeCodeDL = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='高济' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """' AND desc_1='大连阳光大药房有限公司';"""


    executeCodeJX = """SELECT """ + searchCols + """ FROM ddmonth where customer = '益丰' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 = '江西' ) AND (sfa_desc NOT LIKE '%江西新康健民大药房连锁有限公司%') UNION ALL
    SELECT """ + searchCols + """ FROM ddmonth where customer = '大参林' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 = '赣州' OR desc_1 = '南昌') UNION ALL
    SELECT """ + searchCols + """ FROM ddmonth where customer = '高济' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 = '江西开心人大药房连锁有限公司');"""


    executeCodeHBXX = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='益丰' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """' AND desc_1='河北新兴';"""


    executeCodeCD = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='海王' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """' AND desc_1='成都';"""


    executeCodeQY = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='全亿' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """';"""

    # 高济
    executeCodeGJ = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='高济' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """';"""

    # 老百姓
    executeCodeLBX = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='老百姓' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """';"""


    executeCodeHRT = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='老百姓' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """' AND desc_1 LIKE '%惠仁堂%';"""


    executeCodePZYFJH = """SELECT """ + searchCols + """ FROM ddmonth where customer = '老百姓' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 LIKE '%普泽%') UNION ALL
    SELECT """ + searchCols + """ FROM ddmonth where customer = '益丰' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' UNION ALL
    SELECT """ + searchCols + """ FROM ddmonth where customer = '大参林' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 LIKE '%南通%');"""


    executeCodeHB = """SELECT """ + searchCols + """ FROM ddmonth where customer = '老百姓' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 LIKE '%湖北%') UNION ALL
    SELECT """ + searchCols + """ FROM ddmonth where customer = '海王' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 LIKE '%湖北%') UNION ALL
    SELECT """ + searchCols + """ FROM ddmonth where customer = '益丰' AND YEAR(date) = '""" + Year + """' AND MONTH(date) = '""" + Month + """' AND (desc_1 LIKE '%湖北%');"""


    executeCodeQD = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='海王' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """' AND state='青岛';"""

    # 海王
    executeCodeHW = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='海王' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """';"""

    # 益丰
    executeCodeYF = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='益丰' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """';"""

    # 大参林
    executeCodeDSL = """SELECT """ + searchCols + """ FROM ddmonth WHERE customer='大参林' AND YEAR(date)= '""" + Year + """' AND MONTH(date)= '""" + Month + """';"""

    # 广中珠韶清
    executeCodeGZZSQ = """SELECT """ + searchCols + """ FROM dkhfact WHERE YEAR(date) = '""" + Year + """' and MONTH(date) = '""" + Month + """' AND desc_1 IN ('广东加盟', '广东老百姓', '中珠', '广州', '柏康', '清远', '花都', '柏和', '番禺', '韶关', '清远百姓大药房医药连锁有限公司', '广州高济医药连锁有限公司', '益荔康信', '广东国大', '广州连锁') AND customer IN ('老百姓', '海王', '大参林', '高济', '益丰', '国大')"""

    # 高济安徽
    executeCodeAH = """SELECT """ + searchCols + """ FROM dkhfact WHERE YEAR(date) = '""" + Year + """' and MONTH(date) = '""" + Month + """' AND desc_1 IN ('芜湖中山大药房连锁有限公司', '安徽广济大药房连锁有限公司', '阜阳市第一大药房零售连锁有限公司', '滁州华巨百姓缘大药房连锁股份有限公司', '安徽高济敬贤堂药业有限责任公司') AND customer = '高济'"""

    # 粤西
    executeCodeYX = """SELECT """ + searchCols + """ FROM dkhfact WHERE YEAR(date) = '""" + Year + """' and MONTH(date) = '""" + Month + """' AND desc_1 IN ('湛江', '茂名', '阳江', '海口') AND customer = '大参林'"""


    cursorSZ.execute(executeCodeSZ)  # 执行查询
    cursorDL.execute(executeCodeDL)
    cursorJX.execute(executeCodeJX)
    cursorHBXX.execute(executeCodeHBXX)
    cursorCD.execute(executeCodeCD)
    cursorQY.execute(executeCodeQY)
    cursorGJ.execute(executeCodeGJ)
    cursorLBX.execute(executeCodeLBX)
    cursorHRT.execute(executeCodeHRT)
    cursorPZYFJH.execute(executeCodePZYFJH)
    cursorHB.execute(executeCodeHB)
    cursorQD.execute(executeCodeQD)
    cursorHW.execute(executeCodeHW)
    cursorYF.execute(executeCodeYF)
    cursorDSL.execute(executeCodeDSL)
    cursorGZZSQ.execute(executeCodeGZZSQ)
    cursorAH.execute(executeCodeAH)
    cursorYX.execute(executeCodeYX)


    rowNumSZ = cursorSZ.rowcount  # 查询数据条数
    rowNumDL = cursorDL.rowcount
    rowNumJX = cursorJX.rowcount
    rowNumHBXX = cursorHBXX.rowcount
    rowNumCD = cursorCD.rowcount
    rowNumQY = cursorQY.rowcount
    rowNumGJ = cursorGJ.rowcount
    rowNumLBX = cursorLBX.rowcount
    rowNumHRT = cursorHRT.rowcount
    rowNumPZYFJH = cursorPZYFJH.rowcount
    rowNumHB = cursorHB.rowcount
    rowNumQD = cursorQD.rowcount
    rowNumHW = cursorHW.rowcount
    rowNumYF = cursorYF.rowcount
    rowNumDSL = cursorDSL.rowcount
    rowNumGZZSQ = cursorGZZSQ.rowcount
    rowNumAH = cursorAH.rowcount
    rowNumYX = cursorYX.rowcount


    dataSZ = cursorSZ.fetchall()  # 获取全部查询数据
    dataDL = cursorDL.fetchall()
    dataJX = cursorJX.fetchall()
    dataHBXX = cursorHBXX.fetchall()
    dataCD = cursorCD.fetchall()
    dataQY = cursorQY.fetchall()
    dataGJ = cursorGJ.fetchall()
    dataLBX = cursorLBX.fetchall()
    dataHRT = cursorHRT.fetchall()
    dataPZYFJH = cursorPZYFJH.fetchall()
    dataHB = cursorHB.fetchall()
    dataQD = cursorQD.fetchall()
    dataHW = cursorHW.fetchall()
    dataYF = cursorYF.fetchall()
    dataDSL = cursorDSL.fetchall()
    dataGZZSQ = cursorGZZSQ.fetchall()
    dataAH = cursorAH.fetchall()
    dataYX = cursorYX.fetchall()


    conn.commit()  # 提交确认


    cursorSZ.close()  # 关闭光标
    cursorDL.close()
    cursorJX.close()
    cursorHBXX.close()
    cursorCD.close()
    cursorQY.close()
    cursorGJ.close()
    cursorLBX.close()
    cursorHRT.close()
    cursorPZYFJH.close()
    cursorHB.close()
    cursorHW.close()
    cursorYF.close()
    cursorDSL.close()
    cursorGZZSQ.close()
    cursorAH.close()
    cursorYX.close()


    conn.close()  # 关闭连接


    endTime = datetime.now()
    print(" > " + Year + '.' + Month + '.' + str(today).zfill(2) +
          "\n>> 【海王深圳】获取数据" + str(rowNumSZ) + "条;" +
          "\n>> 【高济大连阳光】获取数据" + str(rowNumDL) + "条;" +
          "\n>> 【江西区域】获取数据" + str(rowNumJX) + "条;" +
          "\n>> 【益丰河北新兴】获取数据" + str(rowNumHBXX) + "条;" +
          "\n>> 【海王西南(成都)】获取数据" + str(rowNumCD) + "条;" +
          "\n>> 【全亿】获取数据" + str(rowNumQY) + "条;" +
          "\n>> 【高济】获取数据" + str(rowNumGJ) + "条;" +
          "\n>> 【老百姓】获取数据" + str(rowNumLBX) + "条;" +
          "\n>> 【惠仁堂】获取数据" + str(rowNumHRT) + "条;" +
          "\n>> 【普泽益丰江海】获取数据" + str(rowNumPZYFJH) + "条;" +
          "\n>> 【湖北】获取数据" + str(rowNumHB) + "条;" +
          "\n>> 【青岛】获取数据" + str(rowNumQD)+ "条;" +
          "\n>> 【海王】获取数据" + str(rowNumHW) + "条;" +
          "\n>> 【益丰】获取数据" + str(rowNumYF) + "条;" +
          "\n>> 【大参林】获取数据" + str(rowNumDSL) + "条;" +
          "\n>> 【广中珠韶清】获取数据" + str(rowNumGZZSQ) + "条;" +
          "\n>> 【高济安徽】获取数据" + str(rowNumAH) + "条;" +
          "\n>> 【粤西】获取数据" + str(rowNumYX) + "条." +
          "\n>>> 数据获取耗时：" + strftime("%H:%M:%S", gmtime((endTime - startTime).seconds)))


    # ------------------------------------------------------------------------------------------------
    # 表格指向存储区
    # ------------------------------------------------------------------------------------------------
    print(">> 存储表格中,请稍等片刻")

    startsaveTime = datetime.now()

    colNames = ['销售日期', '客户体系', '大区', '省区', '调整细分区域', '原始自带营运区', '客户编码', '客户名称', '门店编码', '门店名称', '商品编码', '商品名称', '商品简称', '商品规格', '数量', '标准单价', '标准零售金额']
    # searchCols = ['date', 'customer', 'sfa_client_desc', 'state', 'desc_1', 'client_id', 'client_desc', 'client_alias', 'sfa_id', 'sfa_desc', 'DKH_materiel_id', 'materiel_desc', 'materiel_alias', 'norms', 'amount', 'bz_UnitPrice', 'bz_sales_Money']

    # colNames22 = ['调整区域', '商品编码', '商品名称', '商品规格', '销售日期', '门店名称', '门店编码', '数量', '单价', '零售金额', '标准单价', '标准零售金额', '客户体系', '商品简称', '标识符', '大区', '省区', '实际营运区', '客户编码', '客户名称']

    # 1 深圳
    shenzhenData = pd.DataFrame(dataSZ, columns=colNames, dtype=np.float64)
    # shenzhenData['销售日期'] = pd.to_datetime(shenzhenData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathSZ = 'D:/FilesCenter/大客户数据/HW-深圳/' + Year + '-' + Month + '-' + str(today).zfill(2) + '海王深圳.xlsx'
    shenzhenData.to_excel(FilePathSZ, index=False)
    # setStyle_saveFile(shenzhenData, FilePathSZ)

    # 2 大连阳光
    dalianData = pd.DataFrame(dataDL, columns=colNames, dtype=np.float64)
    # dalianData['销售日期'] = pd.to_datetime(dalianData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathDL = 'D:/FilesCenter/大客户数据/GJ-大连阳光/' + Year + '-' + Month + '-' + str(today).zfill(2) + '大连阳光.xlsx'
    dalianData.to_excel(FilePathDL, index=False)
    # setStyle_saveFile(dalianData, FilePathDL)

    # 3 江西
    jiangxiData = pd.DataFrame(dataJX, columns=colNames, dtype=np.float64)
    # jiangxiData['销售日期'] = pd.to_datetime(jiangxiData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathJX = 'D:/FilesCenter/大客户数据/Everyday-江西/' + Year + '-' + Month + '-' + str(today).zfill(2) + '江西.xls'
    Create_TableImage(jiangxiData)  # 创建自定义格式表格

    # 4 河北新兴
    hebeixinxingData = pd.DataFrame(dataHBXX, columns=colNames, dtype=np.float64)
    # hebeixinxingData['销售日期'] = pd.to_datetime(hebeixinxingData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathHBXX = 'D:/FilesCenter/大客户数据/YF-河北新兴/' + Year + '-' + Month + '-' + str(today).zfill(2) + '益丰河北新兴.xlsx'
    hebeixinxingData.to_excel(FilePathHBXX, index=False)
    # setStyle_saveFile(hebeixinxingData, FilePathHBXX)

    # 5 成都
    chengduData = pd.DataFrame(dataCD, columns=colNames, dtype=np.float64)
    # chengduData['销售日期'] = pd.to_datetime(chengduData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathCD = 'D:/FilesCenter/大客户数据/HW-西南/' + Year + '-' + Month + '-' + str(today).zfill(2) + '海王西南(成都).xlsx'
    chengduData.to_excel(FilePathCD, index=False)
    # setStyle_saveFile(chengduData, FilePathCD)

    # 6 全亿
    quanyiData = pd.DataFrame(dataQY, columns=colNames, dtype=np.float64)
    # quanyiData['销售日期'] = pd.to_datetime(quanyiData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathQY = 'D:/FilesCenter/大客户数据/QY-全亿/' + Year + '-' + Month + '-' + str(today).zfill(2) + '全亿.xlsx'
    quanyiData.to_excel(FilePathQY, index=False)
    # setStyle_saveFile(quanyiData, FilePathQY)

    # 7 高济
    gaojiData = pd.DataFrame(dataGJ, columns=colNames, dtype=np.float64)
    # gaojiData['销售日期'] = pd.to_datetime(gaojiData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathGJ = 'D:/FilesCenter/大客户数据/GJ-高济/' + Year + '-' + Month + '-' + str(today).zfill(2) + '高济.xlsx'
    gaojiData.to_excel(FilePathGJ, index=False)
    # setStyle_saveFile(gaojiData, FilePathGJ)

    # 8 老百姓
    laobaixingData = pd.DataFrame(dataLBX, columns=colNames, dtype=np.float64)
    # laobaixingData['销售日期'] = pd.to_datetime(laobaixingData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathLBX = 'D:/FilesCenter/大客户数据/LBX-老百姓/' + Year + '-' + Month + '-' + str(today).zfill(2) + '老百姓.xlsx'
    laobaixingData.to_excel(FilePathLBX, index=False)
    # setStyle_saveFile(laobaixingData, FilePathLBX)

    # 9 惠仁堂
    huirentangData = pd.DataFrame(dataHRT, columns=colNames, dtype=np.float64)
    # huirentangData['销售日期'] = pd.to_datetime(huirentangData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathHRT = 'D:/FilesCenter/大客户数据/LBX-惠仁堂/' + Year + '-' + Month + '-' + str(today).zfill(2) + '惠仁堂.xlsx'
    huirentangData.to_excel(FilePathHRT, index=False)
    # setStyle_saveFile(huirentangData, FilePathHRT)

    # 10 普泽益丰江海
    puzeyifengjianghaiData = pd.DataFrame(dataPZYFJH, columns=colNames, dtype=np.float64)
    # puzeyifengjianghaiData['销售日期'] = pd.to_datetime(puzeyifengjianghaiData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathPZYFJH = 'D:/FilesCenter/大客户数据/LBX-普泽益丰江海/' + Year + '-' + Month + '-' + str(today).zfill(2) + '普泽益丰江海.xlsx'
    puzeyifengjianghaiData.to_excel(FilePathPZYFJH, index=False)
    # setStyle_saveFile(puzeyifengjianghaiData, FilePathPZYFJH)

    # 11 湖北
    hubeiData = pd.DataFrame(dataHB, columns=colNames, dtype=np.float64)
    # hubeiData['销售日期'] = pd.to_datetime(hubeiData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathHB = 'D:/FilesCenter/大客户数据/Everyday-湖北/' + Year + '-' + Month + '-' + str(today).zfill(2) + '湖北.xlsx'
    hubeiData.to_excel(FilePathHB, index=False)
    # setStyle_saveFile(hubeiData, FilePathHB)

    # 12 青岛
    qingdaoData = pd.DataFrame(dataQD, columns=colNames, dtype=np.float64)
    # qingdaoData['销售日期'] = pd.to_datetime(qingdaoData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathQD = 'D:/FilesCenter/大客户数据/HW-青岛/' + Year + '-' + Month + '-' + str(today).zfill(2) + '海王青岛.xlsx'
    qingdaoData.to_excel(FilePathQD, index=False)
    # setStyle_saveFile(qingdaoData, FilePathQD)

    # 13 四大客户
    haiwangData = pd.DataFrame(dataHW, columns=colNames, dtype=np.float64)
    yifengData = pd.DataFrame(dataYF, columns=colNames, dtype=np.float64)
    dashenlinData = pd.DataFrame(dataDSL, columns=colNames, dtype=np.float64)
    # sidakehuData['销售日期'] = pd.to_datetime(sidakehuData['销售日期'], format='%Y/%m/%d').dt.date
    FilePathHW = 'D:/FilesCenter/大客户数据/HW-海王/' + Year + '-' + Month + '-' + str(today).zfill(2) + '海王.xlsx'
    FilePathYF = 'D:/FilesCenter/大客户数据/YF-益丰/' + Year + '-' + Month + '-' + str(today).zfill(2) + '益丰.xlsx'
    FilePathDSL = 'D:/FilesCenter/大客户数据/DSL-大参林/' + Year + '-' + Month + '-' + str(today).zfill(2) + '大参林.xlsx'
    haiwangData.to_excel(FilePathHW, index=False)
    yifengData.to_excel(FilePathYF, index=False)
    dashenlinData.to_excel(FilePathDSL, index=False)
    # setStyle_saveFile(haiwangData, FilePathHW)
    # setStyle_saveFile(yifengData, FilePathYF)
    # setStyle_saveFile(dashenlinData, FilePathDSL)

    # 14 广中珠韶清
    guangzzsqData = pd.DataFrame(dataGZZSQ, columns=colNames, dtype=np.float64)
    FilePathGZZSQ = 'D:/FilesCenter/大客户数据/Everyday-广中珠韶清/' + Year + '-' + Month + '-' + str(today).zfill(2) + '广中珠韶清.xlsx'
    guangzzsqData.to_excel(FilePathGZZSQ, index=False)
    # setStyle_saveFile(guangzzsqData, FilePathGZZSQ)

    # 15 高济安徽
    anhuiData = pd.DataFrame(dataAH, columns=colNames, dtype=np.float64)
    FilePathAH = 'D:/FilesCenter/大客户数据/GJ-安徽/' + Year + '-' + Month + '-' + str(today).zfill(2) + '高济安徽.xlsx'
    anhuiData.to_excel(FilePathAH, index=False)
    # setStyle_saveFile(guangzzsqData, FilePathGZZSQ)

    # 16 粤西
    yuexiData = pd.DataFrame(dataYX, columns=colNames, dtype=np.float64)
    FilePathYX= 'D:/FilesCenter/大客户数据/DSL-粤西/' + Year + '-' + Month + '-' + str(today).zfill(2) + '粤西.xlsx'
    yuexiData.to_excel(FilePathYX, index=False)

    endsaveTime = datetime.now()
    print('>>> 存储表格耗时：' + strftime("%H:%M:%S", gmtime((endsaveTime - startsaveTime).seconds)))


    # ------------------------------------------------------------------------------------------------
    # 内部群 ChatId 与 webhook 区
    # ------------------------------------------------------------------------------------------------
    # 通过jsapi工具 https://wsdebug.dingtalk.com/?spm=a219a.7629140.0.0.7bc84a972WUfGd 获取目标群聊chatId  中智CorpId:ding08a53a0a5cdf47aa
    ChatIdSZ = 'chat36cbc03177f4b48623d505f449852499'  # 深圳
    ChatIdDL = 'chatca0d82c9d6e94531f7d639b33653a8f5'  # 大连阳光
    ChatIdJX = 'chat8971c7614a674dc958a6a4baf91f2632'  # 江西
    ChatIdHBXX = 'chat8e8f0ba98e0e512377716d416bf992a5'  # 河北新兴
    ChatIdXN = 'chat75495b5797bc775a06bd9722f42a8260'  # 西南
    ChatIdLBX = 'chat34ffac1b8839a6a3df8854750c9a8836'  # 老百姓
    ChatIdHRT = 'chat0cb3560baecbad4f68938a3d872f5a0c'  # 惠仁堂
    ChatIdPZYFJH = 'chatcf38482868ce95eed23540f7e3a6820d'  # 普泽益丰江海
    ChatIdHB = 'chat9bee77b820c60cbee4cc49877ec19596'  # 湖北
    ChatIdQD = 'chat8cb11f2f92c9e2e41eb1816830510259'  # 青岛
    ChatIdBigC = 'chat404c83b42c82b08b81e404b4c446a2a1'  # 四大客户
    ChatIdGZZSQ = 'chat25dcc12237e712c2e4f46c388d4ca60a'  # 广中珠韶清
    ChatIdAH = 'chatac431f3df9002bff7154611e54b87614'  # 安徽
    ChatIdYX = 'chat5959c3a77c4b31ed31e100ae5133d2cf'  # 粤西

    AppKey = 'dingjpjkc2vaqjoqgmhz'  # 企业开发平台小程序AppKey
    AppSecret = 'oKNcuSF12oW0j9eBeO53wA6qwmKCVz34NVy1NvtvnjsvKPOdKiozsSZzUypNSWDc'  # 企业开发平台小程序AppSecret

    RobotWebHookURLSZ = 'https://oapi.dingtalk.com/robot/send?access_token=bbffc889ad9051e4ea7d011f46c15146e7ab893f1cccd2b2a9170d7c54e47de0'  # 深圳群
    RobotWebHookURLDL = 'https://oapi.dingtalk.com/robot/send?access_token=262aae91ab9db9d928bc8c933f1136b139582596a34a5e315e7916d362abc5fc'  # 高济大连阳光群
    RobotWebHookURLJX = 'https://oapi.dingtalk.com/robot/send?access_token=ad57e60ff8079e27cde8b59d06c0ef46303e6235eedfcab03ab368757d6a468a'  # 江西群
    RobotWebHookURLHBXX = 'https://oapi.dingtalk.com/robot/send?access_token=d3919b5b65783f24e5bf78b2ce1fb79e7eefa043ca1944eff7b683d8e6fed22f'  # 河北新兴群
    RobotWebHookURLXN = 'https://oapi.dingtalk.com/robot/send?access_token=89cbf9cfca724a54cf8939cf5e10fdb013a4581554716584fa189c00f2779583'  # 西南群
    RobotWebHookURLLBX = 'https://oapi.dingtalk.com/robot/send?access_token=6dabd3db1800f83198f07a228d619e62a46f70dd5d58e7ddc7770f036ea81ce8'  # 老百姓群
    RobotWebHookURLHRT = 'https://oapi.dingtalk.com/robot/send?access_token=88861e201507a056ba9582f15fee29d41826fa0acb21160c20044f11f2ede71a'  # 惠仁堂群
    RobotWebHookURLPZYFJH = 'https://oapi.dingtalk.com/robot/send?access_token=ec14984eae07bd0d68a8664151a053718569eb7671abfbfe7c8ce2fc0df98aa2'  # 普泽益丰江海群
    RobotWebHookURLHB = 'https://oapi.dingtalk.com/robot/send?access_token=7e0ecf1bafb8894a6181e8e6ece961dec1bf5f57e9a7fe54bb528bdc3931c233'  # 湖北群
    RobotWebHookURLQD = 'https://oapi.dingtalk.com/robot/send?access_token=84c1566d140ef723fc240dbd635651b02bc0ad41421d2ce3a284550fc4e3d0be'  # 青岛群
    RobotWebHookURLBigC = 'https://oapi.dingtalk.com/robot/send?access_token=9a83434645a7b38677a714f9c4e72381ec5488be24c4fd2eef1748d0334ef49b'  # 四大客户群
    RobotWebHookURLGZZSQ = 'https://oapi.dingtalk.com/robot/send?access_token=410e2b4752cb211d7f322dc2ca04d4779c52c66fdd4331c4c20f6bfad1834ea4'  # 广中珠韶清
    RobotWebHookURLAH = 'https://oapi.dingtalk.com/robot/send?access_token=7414a3e541547c2ba226aadb549410eb0ef2d7ffd86521ffee63bb7de7d60ba3'  # 安徽
    RobotWebHookURLYX = 'https://oapi.dingtalk.com/robot/send?access_token=81c15297c883d28b6e74f00ac09cb39c966c92a8784744f3ffb072f784048734'  # 粤西

    RobotSecret = 'GbSFeeIHgYNJfXT5WoPT6c6GRmMVRd2wVODyexo7SQIF5HJkucowab6cNMiyR8IV'  # 群机器人加签秘钥secret(默认 草晶华小助手 )


    # ------------------------------------------------------------------------------------------------
    # 发送文本区
    # ------------------------------------------------------------------------------------------------
    def publish_content(publish_date):
        ddMessage = {  # 消息内容
            "msgtype": "markdown",
            "markdown": {"title": "每日数据推送",  # @某人 才会显示标题
                         "text": "###### 大家好！我是 草晶华小助手 机器人。今日数据已经送达, 请大家注意查收哦(⊙o⊙), 如有疑问请及时回馈, 感谢您们的理解与支持。"
                                 "\n> ###### **特别说明: 海王、大参林数据延后2天, 其他客户延后1天！本数据为网上客户流向数据, 与实时销售数据可能存在差异, 考核、业绩、提成等不建议使用该数据, 建议用纯销回流向数据。**"
                                 "\n###### *温馨提示: 手机电脑端均可在线浏览, 建议下载浏览！！！*"
                                 "\n###### ----------------------------------------"
                                 "\n###### 数据来源: 中智药业集团之信息中心数据运营组"
                                 "\n###### 发布时间: " + str(publish_date).split('.')[0]},  # 发布时间
            "at": {
                # "atMobiles": [15817552982],  # 指定@某人
                "isAtAll": False  # 是否@所有人[False:否, True:是]
            }
        }
        return ddMessage


    # ------------------------------------------------------------------------------------------------
    # 发送区
    # ------------------------------------------------------------------------------------------------

    # 深圳海王数据群
    dingdingFunction(RobotWebHookURLSZ, RobotSecret, AppKey, AppSecret).sendFile(ChatIdSZ, FilePathSZ)  # 文件
    dingdingFunction(RobotWebHookURLSZ, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 高济大连阳光流向沟通群
    dingdingFunction(RobotWebHookURLDL, RobotSecret, AppKey, AppSecret).sendFile(ChatIdDL, FilePathDL)  # 文件
    dingdingFunction(RobotWebHookURLDL, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 草晶华江西区域数据发送
    dingdingFunction(RobotWebHookURLJX, RobotSecret, AppKey, AppSecret).sendFile(ChatIdJX, FilePathJX)  # 文件
    dingdingFunction(RobotWebHookURLJX, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 河北新兴流向沟通群
    dingdingFunction(RobotWebHookURLHBXX, RobotSecret, AppKey, AppSecret).sendFile(ChatIdHBXX, FilePathHBXX)  # 文件
    dingdingFunction(RobotWebHookURLHBXX, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 海王西南大区数据群
    dingdingFunction(RobotWebHookURLXN, RobotSecret, AppKey, AppSecret).sendFile(ChatIdXN, FilePathCD)  # 文件
    dingdingFunction(RobotWebHookURLXN, RobotSecret, AppKey, AppSecret).sendFile(ChatIdXN, FilePathQY)  # 文件
    dingdingFunction(RobotWebHookURLXN, RobotSecret, AppKey, AppSecret).sendFile(ChatIdXN, FilePathGJ)  # 文件
    dingdingFunction(RobotWebHookURLXN, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 老百姓数据流向群
    dingdingFunction(RobotWebHookURLLBX, RobotSecret, AppKey, AppSecret).sendFile(ChatIdLBX, FilePathLBX)  # 文件
    dingdingFunction(RobotWebHookURLLBX, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 惠仁堂数据流向群
    dingdingFunction(RobotWebHookURLHRT, RobotSecret, AppKey, AppSecret).sendFile(ChatIdHRT, FilePathHRT)  # 文件
    dingdingFunction(RobotWebHookURLHRT, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 普泽益丰江海数据群
    dingdingFunction(RobotWebHookURLPZYFJH, RobotSecret, AppKey, AppSecret).sendFile(ChatIdPZYFJH, FilePathPZYFJH)  # 文件
    dingdingFunction(RobotWebHookURLPZYFJH, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 湖北老百姓益丰海王数据群
    dingdingFunction(RobotWebHookURLHB, RobotSecret, AppKey, AppSecret).sendFile(ChatIdHB, FilePathHB)  # 文件
    dingdingFunction(RobotWebHookURLHB, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 青岛海王流向沟通群
    dingdingFunction(RobotWebHookURLQD, RobotSecret, AppKey, AppSecret).sendFile(ChatIdQD, FilePathQD)  # 文件
    dingdingFunction(RobotWebHookURLQD, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 大客户每日数据源发送群（文件超10M, 报错）
    dingdingFunction(RobotWebHookURLBigC, RobotSecret, AppKey, AppSecret).sendFile(ChatIdBigC, FilePathLBX)  # 文件
    dingdingFunction(RobotWebHookURLBigC, RobotSecret, AppKey, AppSecret).sendFile(ChatIdBigC, FilePathHW)  # 文件
    dingdingFunction(RobotWebHookURLBigC, RobotSecret, AppKey, AppSecret).sendFile(ChatIdBigC, FilePathYF)  # 文件
    # dingdingFunction(RobotWebHookURLBigC, RobotSecret, AppKey, AppSecret).sendFile(ChatIdBigC, FilePathDSL)  # 文件（文件太大）
    dingdingFunction(RobotWebHookURLBigC, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 广中珠韶清数据群
    dingdingFunction(RobotWebHookURLGZZSQ, RobotSecret, AppKey, AppSecret).sendFile(ChatIdGZZSQ, FilePathGZZSQ)  # 文件
    dingdingFunction(RobotWebHookURLGZZSQ, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 高济安徽区域数据群
    dingdingFunction(RobotWebHookURLAH, RobotSecret, AppKey, AppSecret).sendFile(ChatIdAH, FilePathAH)  # 文件
    dingdingFunction(RobotWebHookURLAH, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息

    # 粤西大参林流向沟通群
    dingdingFunction(RobotWebHookURLYX, RobotSecret, AppKey, AppSecret).sendFile(ChatIdYX, FilePathYX)  # 文件
    dingdingFunction(RobotWebHookURLYX, RobotSecret, AppKey, AppSecret).sendMessage(publish_content(datetime.now()))  # 消息
